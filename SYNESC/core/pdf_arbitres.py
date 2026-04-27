"""
core/pdf_arbitres.py
Génère un PDF — liste des arbitres retenus/libérés.
Colonnes : Club | Nom Prénom | Licence | Niveau | Arme
Pas de tarifs ni de coûts.
"""

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl

NAVY = colors.HexColor("#1F4E79")
BLUE = colors.HexColor("#2E6099")
WHITE = colors.white


def lire_arbitres_depuis_excel(excel_bytes):
    """
    Lit la feuille 'Arbitres' et regroupe par date + statut (col G).
    Lit aussi les responsables depuis la feuille 'Récap Arbitres'.
    Retourne :
      { 'titre': str,
        'responsables': {'type': 'cra'|'superviseurs', 'valeurs': {...}},
        'jours': [{'label', 'retenus', 'liberes'}] }
    """
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)

    if "Arbitres" not in wb.sheetnames:
        raise ValueError("La feuille 'Arbitres' est introuvable dans l'Excel.")

    # Titre et responsables depuis Récap
    titre = ""
    responsables = {}
    if "Récap Arbitres" in wb.sheetnames:
        ws_r = wb["Récap Arbitres"]
        rows_r = list(ws_r.iter_rows(values_only=True))
        if rows_r: titre = str(rows_r[0][0] or "").strip()

        for row in rows_r[1:]:
            v0 = str(row[0] or "").strip()
            v2 = str(row[2] or "").strip()
            if "Responsable CRA" in v0:
                responsables = {"type": "cra", "valeur": v2}
            elif v0 == "Superviseur":
                # Superviseur Grand Est (ligne distincte sous le CRA)
                responsables.setdefault("type", "cra")
                if v2: responsables["superviseur"] = v2
            elif "Superviseur Fleuret" in v0:
                responsables.setdefault("type", "superviseurs")
                responsables.setdefault("valeurs", {})
                responsables["valeurs"]["Fleuret"] = v2
            elif "Superviseur Épée" in v0 or "Superviseur Epee" in v0:
                responsables.setdefault("type", "superviseurs")
                responsables.setdefault("valeurs", {})
                responsables["valeurs"]["Épée"] = v2
            elif "Superviseur Sabre" in v0:
                responsables.setdefault("type", "superviseurs")
                responsables.setdefault("valeurs", {})
                responsables["valeurs"]["Sabre"] = v2

    # Lire la feuille Arbitres
    ws = wb["Arbitres"]
    jours = []
    current_jour = None

    for row in ws.iter_rows(values_only=True):
        v0 = str(row[0] or "").strip()
        if not v0 or v0 == "Club": continue
        if "arbitres du" in v0.lower() or "arbitres —" in v0.lower():
            label = v0.replace("Arbitres du", "").replace("Arbitres —", "").strip()
            current_jour = {"label": label, "retenus": [], "liberes": []}
            jours.append(current_jour)
            continue
        if "tireur" in v0.lower(): break
        if current_jour and row[1]:
            statut = str(row[6] or "Retenu").strip()
            entry  = (str(row[0] or ""), str(row[1] or ""), str(row[2] or ""),
                      str(row[3] or ""), str(row[4] or ""))
            if statut == "Libéré":
                current_jour["liberes"].append(entry)
            else:
                current_jour["retenus"].append(entry)

    return {"titre": titre, "responsables": responsables, "jours": jours}


def generer_pdf_arbitres(excel_bytes):
    """Génère le PDF et retourne un BytesIO."""
    data = lire_arbitres_depuis_excel(excel_bytes)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm)

    W = A4[0] - 30*mm
    styles = getSampleStyleSheet()

    s_titre = ParagraphStyle("t", parent=styles["Normal"],
        fontSize=14, fontName="Helvetica-Bold",
        textColor=WHITE, alignment=TA_CENTER)
    s_jour = ParagraphStyle("j", parent=styles["Normal"],
        fontSize=11, fontName="Helvetica-Bold",
        textColor=WHITE, alignment=TA_LEFT)
    s_ret = ParagraphStyle("r", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#276221"))
    s_lib = ParagraphStyle("l", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#9C0006"))
    s_vide = ParagraphStyle("v", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey, alignment=TA_CENTER)

    # Colonnes : Club | Nom Prénom | Licence | Niveau | Arme
    COL_W = [W*0.23, W*0.32, W*0.14, W*0.12, W*0.19]
    HDR = ["Club", "Nom Prénom", "Licence", "Niveau", "Arme"]

    def _table(arbitres, row_bg, hdr_bg):
        if not arbitres:
            return None
        rows = [HDR] + [[e[0][:24], e[1][:34], e[2], e[3], e[4]] for e in arbitres]
        t = Table(rows, colWidths=COL_W, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#"+hdr_bg)),
            ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 8),
            ("ALIGN",         (0,0), (-1,0), "CENTER"),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 8),
            ("ALIGN",         (0,1), (1,-1), "LEFT"),
            ("ALIGN",         (2,1), (-1,-1), "CENTER"),
            ("ROWBACKGROUNDS",(0,1), (-1,-1),
             [colors.HexColor("#"+row_bg), colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("LINEBELOW",     (0,0), (-1,0), 0.8, colors.HexColor("#999999")),
            ("TOPPADDING",    (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("LEFTPADDING",   (0,0), (-1,-1), 4),
            ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ]))
        return t

    story = []

    # Titre
    t_titre = Table(
        [[Paragraph(data["titre"] or "Liste des Arbitres", s_titre)]],
        colWidths=[W])
    t_titre.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
    ]))
    story.append(t_titre)
    story.append(Spacer(1, 3*mm))

    # Responsables
    resp = data.get("responsables", {})
    ARME_BG_PDF = {"Fleuret": "DDEEFF", "Épée": "DDFFEE", "Sabre": "FFEECC"}

    if resp.get("type") == "cra" and (resp.get("valeur") or resp.get("superviseur")):
        s_resp = ParagraphStyle("resp", parent=styles["Normal"],
            fontSize=10, fontName="Helvetica-Bold", textColor=WHITE)
        rows_resp = []
        if resp.get("valeur"):
            rows_resp.append([Paragraph(f"  Responsable CRA : {resp['valeur']}", s_resp)])
        if resp.get("superviseur"):
            s_sup = ParagraphStyle("sup", parent=styles["Normal"],
                fontSize=10, fontName="Helvetica-Bold", textColor=WHITE)
            rows_resp.append([Paragraph(f"  Superviseur : {resp['superviseur']}", s_sup)])
        if rows_resp:
            t_resp = Table(rows_resp, colWidths=[W])
            t_resp.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), BLUE),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("LINEBELOW",     (0,0), (-1,-2), 0.3, colors.HexColor("#FFFFFF55")),
            ]))
            story.append(t_resp)
            story.append(Spacer(1, 3*mm))

    elif resp.get("type") == "superviseurs" and resp.get("valeurs"):
        sup_rows = []
        for arme, nom in resp["valeurs"].items():
            if nom:
                sup_rows.append([
                    Paragraph(f"Superviseur {arme}",
                        ParagraphStyle("sa", parent=styles["Normal"],
                            fontSize=9, fontName="Helvetica-Bold")),
                    Paragraph(nom,
                        ParagraphStyle("sv", parent=styles["Normal"],
                            fontSize=9, fontName="Helvetica-Bold")),
                ])
        if sup_rows:
            t_sup = Table(sup_rows, colWidths=[W*0.35, W*0.65])
            ts = [
                ("FONTSIZE",      (0,0), (-1,-1), 9),
                ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ]
            for i, (arme, _) in enumerate(resp["valeurs"].items()):
                if list(resp["valeurs"].values())[i]:
                    bg = ARME_BG_PDF.get(arme, "EEEEEE")
                    ts.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#"+bg)))
            t_sup.setStyle(TableStyle(ts))
            story.append(t_sup)
            story.append(Spacer(1, 3*mm))

    for jour in data["jours"]:
        nb_ret = len(jour["retenus"])
        nb_lib = len(jour["liberes"])

        # Bandeau jour
        t_jour = Table([[Paragraph(f"  {jour['label']}", s_jour)]], colWidths=[W])
        t_jour.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), BLUE),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        story.append(t_jour)
        story.append(Spacer(1, 2*mm))

        # Retenus
        story.append(Paragraph(f"✅  Arbitres RETENUS — {nb_ret}", s_ret))
        story.append(Spacer(1, 1*mm))
        t = _table(jour["retenus"], "EBF5E8", "276221")
        story.append(t if t else Paragraph("Aucun arbitre retenu.", s_vide))
        story.append(Spacer(1, 4*mm))

        # Libérés
        story.append(Paragraph(f"❌  Arbitres LIBÉRÉS — {nb_lib}", s_lib))
        story.append(Spacer(1, 1*mm))
        t = _table(jour["liberes"], "FFE8E8", "9C0006")
        story.append(t if t else Paragraph("Aucun arbitre libéré.", s_vide))
        story.append(Spacer(1, 5*mm))
        story.append(HRFlowable(width=W, thickness=0.4,
                                color=colors.HexColor("#CCCCCC")))
        story.append(Spacer(1, 4*mm))

    doc.build(story)
    buf.seek(0)
    return buf
