"""
core/excel_base.py
Helpers de style Excel partagés par toutes les compétitions.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def border_all(color="000000", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def mk_font(bold=False, size=11, color="000000", name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)


def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def mk_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def sc(cell, val=None, bold=False, size=11, fg=None, bg=None,
       h="center", v="center", wrap=False, border=True):
    """Style complet d'une cellule en une seule ligne."""
    if val is not None:
        cell.value = val
    cell.font      = mk_font(bold=bold, size=size, color=fg or "000000")
    if bg:
        cell.fill  = mk_fill(bg)
    cell.alignment = mk_align(h=h, v=v, wrap=wrap)
    if border:
        cell.border = border_all()


def gcl(col_idx):
    """Alias court pour get_column_letter."""
    return get_column_letter(col_idx)
