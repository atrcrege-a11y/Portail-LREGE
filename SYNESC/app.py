"""
app.py
Point d'entrée Flask — routes uniquement.
Toute la logique métier est dans core/ et competitions/.
"""

import io
import os
import re
import uuid
import unicodedata
import zipfile
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, session

from core.config import APP_NAME, APP_VERSION, APP_RELEASE_DATE, CHANGELOG
from core.parser import parse_xml
from competitions import get_competition, COMPETITIONS_META

app = Flask(__name__)
app.secret_key = "lrege-synesc-secret-2025"
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

# ── Store en mémoire par session
_store = {}

def get_store():
    sid = session.get("sid")
    if not sid:
        sid = str(uuid.uuid4())
        session["sid"] = sid
    return _store.setdefault(sid, [])


# ── Routes principales

@app.route("/")
def index():
    return render_template("index.html",
                           app_name=APP_NAME,
                           version=APP_VERSION,
                           release_date=APP_RELEASE_DATE,
                           competitions=COMPETITIONS_META)


@app.route("/api/version")
def api_version():
    return jsonify({
        "version":   APP_VERSION,
        "date":      APP_RELEASE_DATE,
        "app_name":  APP_NAME,
        "changelog": CHANGELOG,
    })


# ── Upload

ARME_LABEL = {"F": "Fleuret", "E": "Épée", "S": "Sabre"}
SEXE_LABEL = {"M": "Hommes", "F": "Dames"}
TYPE_LABEL = {"I": "Individuel", "E": "Équipe"}


def _meta_to_dict(meta, tireurs, arbitres):
    return {
        "uid":        meta.get("uid", ""),
        "arme":       ARME_LABEL.get(meta["arme"], meta["arme"]),
        "categorie":  meta["categorie"],
        "sexe":       SEXE_LABEL.get(meta["sexe"], meta["sexe"]),
        "type":       TYPE_LABEL.get(meta["type"], meta["type"]),
        "nb_tireurs": len(tireurs),
        "nb_arbitres":len(arbitres),
        "date":       meta["date"],
        "date_debut": meta.get("date_debut", ""),
        "date_fin":   meta.get("date_fin", ""),
        "titre":      meta["titre"],
        "titre_long": meta["titre"].strip(),
        "id":         meta["id"],
        "filename":   meta["filename"],
    }


def _process_xml(content, filename, store, errors):
    """Parse un XML et l'ajoute au store. Retourne le dict ajouté ou None."""
    try:
        meta, tireurs, arbitres = parse_xml(content, filename)
        # Dédoublonnage uid unique
        if any(s[0].get("uid") and
               s[0]["id"] == meta["id"] and
               s[0]["categorie"] == meta["categorie"] and
               s[0]["type"] == meta["type"]
               for s in store):
            errors.append(f"{filename} — déjà chargé.")
            return None
        meta["uid"] = str(uuid.uuid4())
        store.append((meta, tireurs, arbitres))
        return _meta_to_dict(meta, tireurs, arbitres)
    except Exception as e:
        errors.append(f"{filename} — {e}")
        return None


@app.route("/api/upload", methods=["POST"])
def upload():
    store  = get_store()
    added  = []
    errors = []

    for f in request.files.getlist("files"):
        fname   = f.filename.lower()
        content = f.read()

        if fname.endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    xml_names = sorted([
                        n for n in zf.namelist()
                        if n.lower().endswith(".xml") and "__MACOSX" not in n
                    ])
                    if not xml_names:
                        errors.append(f"{f.filename} — aucun XML trouvé.")
                        continue
                    for xml_name in xml_names:
                        result = _process_xml(
                            zf.read(xml_name),
                            Path(xml_name).name,
                            store, errors,
                        )
                        if result:
                            added.append(result)
            except zipfile.BadZipFile:
                errors.append(f"{f.filename} — ZIP invalide.")

        elif fname.endswith(".xml"):
            result = _process_xml(content, f.filename, store, errors)
            if result:
                added.append(result)
        else:
            errors.append(f"{f.filename} — format non supporté.")

    return jsonify({"added": added, "errors": errors, "total": len(store)})


@app.route("/api/files")
def list_files():
    store = get_store()
    return jsonify([_meta_to_dict(m, t, a) for m, t, a in store])


@app.route("/api/remove", methods=["POST"])
def remove_file():
    uid   = (request.get_json() or {}).get("uid", "")
    store = get_store()
    before = len(store)
    store[:] = [s for s in store if s[0].get("uid") != uid]
    return jsonify({"removed": before - len(store), "total": len(store)})


@app.route("/api/clear", methods=["POST"])
def clear_files():
    get_store().clear()
    return jsonify({"total": 0})


# ── Aperçu

@app.route("/api/preview")
def preview():
    store = get_store()
    if not store:
        return jsonify({"clubs": [], "categories": []})

    from collections import defaultdict
    clubs_total = defaultdict(lambda: defaultdict(lambda: {"H": 0, "D": 0, "MX": 0}))

    for meta, tireurs, arbitres in store:
        cat_raw   = meta["categorie"].upper()
        type_comp = meta["type"]
        comp_ge = get_competition("grand_est")
        comp_al = get_competition("alsace")
        cat_map_eq = {**comp_al.CAT_MAP_EQUIPE, **comp_ge.CAT_MAP_EQUIPE}
        cat_map_in = {**comp_al.CAT_MAP_INDIV,  **comp_ge.CAT_MAP_INDIV}
        cat_map = cat_map_eq if type_comp == "E" else cat_map_in
        cat_key = cat_map.get(cat_raw)
        if not cat_key:
            continue
        for t in tireurs:
            club = t["club"].strip()
            sexe = "H" if t["sexe"].upper() == "M" else "D"
            # Pour les épreuves mixtes (MF), compter dans une clé MX
            sexe_epreuve = meta.get("sexe", "").upper()
            cle = "MX" if (type_comp == "E" and sexe_epreuve == "MF") else sexe
            clubs_total[club][cat_key][cle] += 1

    # Catégories présentes — union de Grand Est + Alsace
    all_cats = []
    seen = set()
    comp_ge = get_competition("grand_est")
    comp_al = get_competition("alsace")
    all_cat_order = list(dict.fromkeys(
        comp_al.CATS_INDIV + comp_al.CATS_EQUIPE +
        comp_ge.CATS_INDIV + comp_ge.CATS_EQUIPE
    ))
    for cat in all_cat_order:
        if cat not in seen and any(cat in clubs_total[c] for c in clubs_total):
            all_cats.append(cat)
            seen.add(cat)

    rows = []
    for club in sorted(clubs_total.keys()):
        row = {"club": club, "values": {}, "total": 0}
        for cat in all_cats:
            vh  = clubs_total[club][cat]["H"]
            vd  = clubs_total[club][cat]["D"]
            vmx = clubs_total[club][cat]["MX"]
            row["values"][f"{cat}_H"]  = vh  or ""
            row["values"][f"{cat}_D"]  = vd  or ""
            if vmx:
                row["values"][f"{cat}_MX"] = vmx
            row["total"] += vh + vd + vmx
        rows.append(row)

    rows.sort(key=lambda r: -r["total"])
    return jsonify({"clubs": rows, "categories": all_cats})


# ── Génération Excel

@app.route("/api/generate", methods=["POST"])
def generate():
    store = get_store()
    if not store:
        return jsonify({"error": "Aucun fichier chargé"}), 400

    data       = request.get_json() or {}
    titre_long = data.get("titre_long", "").strip()
    lieu       = data.get("lieu", "").strip()
    comp_type  = data.get("comp_type", "grand_est")

    titre = (f"{titre_long} - {lieu}" if titre_long and lieu
             else titre_long or lieu or "SYNESC")

    comp = get_competition(comp_type)
    buf  = comp.generer_excel([(m, t, a) for m, t, a in store], titre_comp=titre)

    # Nom de fichier sans accents
    safe = unicodedata.normalize("NFD", titre)
    safe = "".join(c for c in safe if unicodedata.category(c) != "Mn")
    safe = "".join(c for c in safe if c.isalnum() or c in " -_").strip()
    safe = re.sub(r" {2,}", " ", safe).replace(" ", "_")
    filename = f"{safe}.xlsx"

    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



# ── Génération mail .eml





@app.route("/api/upload_programme", methods=["POST"])
def upload_programme():
    """
    Parse le tableau des horaires depuis un PDF programme.
    Gère 3 formats : tableau structuré, texte ligne par ligne, texte condensé.
    Sans appel API externe.
    """
    import re as _re

    HEURE_RE = _re.compile(r'\b(\d{1,2})[hH](\d{0,2})\b')
    CATS = ["M9","M11","M13","M15","M17","M20",
            "Vétérans","Veteran","Seniors","Senior","Gds Vétérans",
            "V1","V2","V3","V4","V1/V2","V3/V4","V1, V2","V3, V4",
            "Hommes","Dames","Hommes & Dames",
            "M9-M11","M17-M15","M11/M15","M20/Séniors/Vétérans","M13/M17"]

    ARMES_RE = _re.compile(r'(?<![A-Za-z])(fleuret|[eé]p[eé]e|sabre)(?![A-Za-z])', _re.IGNORECASE)
    CAT_RE   = _re.compile(r'\b(M\s*\d+|V\d+|Vétérans?|Veteran|Seniors?)\b', _re.IGNORECASE)

    def normaliser_cat(cat):
        """Normalise un libellé brut vers (cat_norm, arme_code).
        Ex: 'M11 Fleuret H/D' → ('M11', 'F')
        Retourne une chaîne 'M11|F' si arme détectée, sinon juste 'M11'."""
        VET_MAP_ = {"V1":"Vétérans","V2":"Vétérans","V3":"Vétérans","V4":"Vétérans",
                    "V1/V2":"Vétérans","V3/V4":"Vétérans","Veteran":"Vétérans","Senior":"Seniors",
                    "Hommes":"Vétérans","Dames":"Vétérans","Hommes & Dames":"Vétérans"}
        cat = cat.strip().replace("\n"," ")
        # Chercher une arme dans le libellé
        arme_m = ARMES_RE.search(cat)
        arme_code = None
        if arme_m:
            a = arme_m.group(1).lower()
            arme_code = "F" if "fleuret" in a else ("S" if "sabre" in a else "E")
        # Chercher la catégorie
        cat_m = CAT_RE.search(cat)
        if cat_m:
            c = cat_m.group(1).replace(" ", "")  # M 9 → M9
            c = VET_MAP_.get(c, c)
            return f"{c}|{arme_code}" if arme_code else c
        # Correspondance directe
        c = VET_MAP_.get(cat, cat)
        cat_low = cat.lower()
        if any(k in cat_low for k in ["v1","v2","v3","v4","vétéran","veteran","hommes","dames"]):
            return f"Vétérans|{arme_code}" if arme_code else "Vétérans"
        return f"{c}|{arme_code}" if arme_code else c

    def norm_h(h, m=""):
        return f"{int(h)}h{int(m):02d}" if m else f"{int(h)}h00"

    def extraire_heure(txt):
        if not txt: return None
        m = HEURE_RE.search(txt)
        return norm_h(m.group(1), m.group(2)) if m else None

    def est_foulée(txt):
        return bool(txt) and any(k in txt.lower() for k in ["foulée","issue","individuel"])

    def detecter_date(text):
        """Extrait la première date trouvée dans le texte."""
        MOIS = {"janvier":"01","février":"02","mars":"03","avril":"04",
                "mai":"05","juin":"06","juillet":"07","août":"08",
                "septembre":"09","octobre":"10","novembre":"11","décembre":"12"}
        JOURS = ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
        txt = text.lower().replace("\n"," ")

        # Format "Samedi30mars2024" ou "Samedi 30 mars 2024" ou "samedi 30.03.2024"
        for mois_fr, mois_n in MOIS.items():
            for jour in JOURS:
                # Avec ou sans espaces autour du mois
                pat = _re.compile(
                    rf'{jour}\s*(\d{{1,2}})\s*{mois_fr}\s*(\d{{4}})', _re.IGNORECASE)
                m = pat.search(txt)
                if m:
                    dd = m.group(1).zfill(2)
                    yy = m.group(2)
                    return f"{jour} {dd}.{mois_n}.{yy}"

        # Format numérique "samedi 30.03.2024" ou "samedi 30/03/2024"
        DATE_NUM = _re.compile(
            r'(lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)'
            r'\s+(\d{1,2})[./](\d{1,2})[./](\d{4})', _re.IGNORECASE)
        m = DATE_NUM.search(txt)
        if m:
            return f"{m.group(1).lower()} {m.group(2).zfill(2)}.{m.group(3).zfill(2)}.{m.group(4)}"

        return ""

    def detecter_lieu(text):
        for line in text.splitlines():
            l = line.strip()
            if any(k in l.lower() for k in ["gymnase","complexe","salle","palais","halle","centre sportif"]):
                return l
        return ""

    try:
        f = request.files.get("pdf")
        if not f:
            return jsonify({"error": "Aucun fichier PDF fourni"}), 400
        pdf_bytes = f.read()

        try:
            import pdfplumber
        except ImportError:
            return jsonify({"error": "Module pdfplumber manquant — pip install pdfplumber"}), 500

        categories = []
        lieu = ""
        date_globale = ""
        full_text = ""

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:3]:
                full_text += (page.extract_text() or "") + "\n"

        # Date et lieu globaux
        date_globale = detecter_date(full_text)
        lieu = detecter_lieu(full_text)

        # ── FORMAT 0 : catégories en colonnes (M15 | M20 | Vétérans | ...)
        # Tableau où une ligne contient plusieurs catégories en colonnes
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:2]:
                for table in page.extract_tables():
                    if not table or len(table) < 3: continue
                    # Chercher la ligne de catégories
                    cats_row_idx = None
                    cats_cols = {}   # {col_idx: cat_label}
                    dates_cols = {}  # {col_idx: date_str}
                    for ri, row in enumerate(table):
                        row_vals = [str(c or "").strip() for c in row]
                        # Ligne avec au moins 2 catégories connues
                        cat_count = sum(1 for v in row_vals
                                       if any(v == c for c in CATS))
                        if cat_count >= 2:
                            cats_row_idx = ri
                            for ci, v in enumerate(row_vals):
                                if any(v == c for c in CATS):
                                    cats_cols[ci] = v
                            break
                    if not cats_cols: continue

                    # Ligne des dates (chercher samedi/dimanche)
                    DATE_COL_RE = _re.compile(
                        r'(samedi|dimanche|lundi|mardi|mercredi|jeudi|vendredi)',
                        _re.IGNORECASE)
                    for row in table[:cats_row_idx]:
                        for ci, cell in enumerate(row):
                            if cell and DATE_COL_RE.search(str(cell)):
                                # Cette cellule couvre plusieurs colonnes (merged) —
                                # attribuer la date à toutes les colonnes cat qui suivent
                                dm = detecter_date(str(cell).replace("\n",""))
                                if dm:
                                    for ccat in cats_cols:
                                        if ccat not in dates_cols:
                                            dates_cols[ccat] = dm

                    # Lire Appel / Scratch / Début / Par Équipe
                    appels = {}; scratchs = {}; debuts = {}; equipes = {}
                    for row in table[cats_row_idx+1:]:
                        label = str(row[0] or "").lower().replace("\n","").strip()
                        if "appel" in label:
                            for ci in cats_cols:
                                h = extraire_heure(str(row[ci] or ""))
                                if h: appels[ci] = h
                        elif "scratch" in label:
                            for ci, cell in cats_cols.items():
                                h = extraire_heure(str(row[ci] or ""))
                                if h: scratchs[ci] = h
                        elif "début" in label or "debut" in label:
                            for ci in cats_cols:
                                h = extraire_heure(str(row[ci] or ""))
                                if h: debuts[ci] = h
                        elif "quipe" in label:
                            for ci in cats_cols:
                                cell = str(row[ci] or "")
                                if est_foulée(cell):
                                    equipes[ci] = "à l'issue des épreuves individuelles"
                                elif extraire_heure(cell):
                                    equipes[ci] = extraire_heure(cell)

                    if appels or debuts:
                        for ci, cat in cats_cols.items():
                            eq = equipes.get(ci, "à l'issue des épreuves individuelles")
                            categories.append({
                                "cat":          normaliser_cat(cat),
                                "date":         dates_cols.get(ci, date_globale),
                                "appel":        appels.get(ci),
                                "scratch":      scratchs.get(ci),
                                "debut":        debuts.get(ci),
                                "equipe_appel": eq if _re.match(r'\d+h', str(eq)) else None,
                                "equipe_debut": None,
                                "equipe_info":  eq if not _re.match(r'\d+h', str(eq)) else None,
                            })
                        if categories: break
                if categories: break


        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:4]:
                page_text = page.extract_text() or ""

                # Dates avec position y dans la page
                words = page.extract_words()
                JOURS_SET = {"samedi","dimanche","lundi","mardi","mercredi","jeudi","vendredi"}
                dates_by_y = []  # [(y_top, date_str)]
                for wi, w in enumerate(words):
                    if w['text'].lower() in JOURS_SET:
                        # Reconstituer la ligne autour de ce mot
                        y = w['top']
                        nearby = " ".join(x['text'] for x in words
                                         if abs(x['top'] - y) < 5)
                        d = detecter_date(nearby)
                        if d and not any(abs(yy - y) < 5 for yy, _ in dates_by_y):
                            dates_by_y.append((y, d))

                def date_pour_y(table_y):
                    """Retourne la date dont le y est juste avant le tableau."""
                    best = date_globale
                    for y, d in sorted(dates_by_y):
                        if y <= table_y + 10:
                            best = d
                    return best

                tables_obj = page.find_tables()
                tables_data = page.extract_tables()

                for ti, table in enumerate(tables_data):
                    if not table or len(table) < 2: continue

                    # Position y du tableau
                    table_y = tables_obj[ti].bbox[1] if ti < len(tables_obj) else 0

                    # Chercher la ligne d'en-tête
                    header_idx = None
                    header = []
                    date_tableau = date_pour_y(table_y)

                    for ri, row in enumerate(table):
                        row_str = [str(c or "").lower().strip() for c in row]
                        # Date dans les cellules du tableau (titre fusionné type Vittel)
                        for cell in row:
                            d = detecter_date(str(cell or ""))
                            if d: date_tableau = d
                        if "appel" in row_str and ("début" in row_str or "debut" in row_str or any("assaut" in s for s in row_str)):
                            header_idx = ri
                            header = row_str
                            break

                    if header_idx is None: continue

                    ci_cat  = next((i for i,h in enumerate(header) if "cat" in h), 0)
                    ci_app  = next((i for i,h in enumerate(header) if "appel" in h or "inscript" in h), None)
                    ci_scr  = next((i for i,h in enumerate(header) if "scratch" in h), None)
                    ci_deb  = next((i for i,h in enumerate(header) if "début" in h or "debut" in h or "assaut" in h), None)

                    for row in table[header_idx+1:]:
                        cat_raw = str(row[ci_cat] or "").strip().replace("\n", " ")
                        if not cat_raw or "arbitre" in cat_raw.lower(): continue
                        cat_norm = normaliser_cat(cat_raw)
                        categories.append({
                            "cat":          cat_norm,
                            "date":         date_tableau,
                            "appel":        extraire_heure(str(row[ci_app] or "")) if ci_app is not None else None,
                            "scratch":      extraire_heure(str(row[ci_scr] or "")) if ci_scr is not None else None,
                            "debut":        extraire_heure(str(row[ci_deb] or "")) if ci_deb is not None else None,
                            "equipe_appel": None,
                            "equipe_debut": None,
                            "equipe_info":  "à l'issue des épreuves individuelles",
                        })

        # ── FORMAT 2 : ligne par ligne "M13 ... Appel : 9h00 scratch : 9h15 début : 09h30"
        if not categories:
            # Chercher les blocs "cat + ligne appel/scratch/début"
            CAT_LINE_RE = _re.compile(
                r'^(' + '|'.join(_re.escape(c) for c in CATS) + r')\b',
                _re.IGNORECASE | _re.MULTILINE)
            BLOC_RE = _re.compile(
                r'(?:appel|inscript)[^\d]*(\d{1,2}[hH]\d{0,2})'
                r'.*?(?:scratch)[^\d]*(\d{1,2}[hH]\d{0,2})'
                r'.*?(?:début|debut)[^\d]*(\d{1,2}[hH]\d{0,2})',
                _re.IGNORECASE | _re.DOTALL)

            lines = full_text.splitlines()
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                cat_m = CAT_LINE_RE.match(line)
                if cat_m:
                    cat = cat_m.group(1)
                    # Chercher les horaires dans les 3 lignes suivantes
                    bloc = " ".join(lines[i:i+4])
                    bm = BLOC_RE.search(bloc)
                    if bm:
                        categories.append({
                            "cat":          normaliser_cat(cat),
                            "date":         date_globale,
                            "appel":        extraire_heure(bm.group(1)),
                            "scratch":      extraire_heure(bm.group(2)),
                            "debut":        extraire_heure(bm.group(3)),
                            "equipe_appel": None,
                            "equipe_debut": None,
                            "equipe_info":  "à l'issue des épreuves individuelles",
                        })
                i += 1

        # ── FORMAT 3 : texte condensé "M13 – Inscriptions : 8h30 – Scratch : 8h45 – Début : 9h00"
        if not categories:
            LINE_RE = _re.compile(
                r'(M\d+(?:[/-]M\d+)?|[Vv]étérans?|[Ss]eniors?)'
                r'.*?(?:inscript|appel)[^\d]*(\d{1,2}[hH]\d{0,2})'
                r'.*?scratch[^\d]*(\d{1,2}[hH]\d{0,2})'
                r'.*?(?:début|debut)[^\d]*(\d{1,2}[hH]\d{0,2})',
                _re.IGNORECASE)
            for line in full_text.splitlines():
                m = LINE_RE.search(line)
                if m:
                    categories.append({
                        "cat":          normaliser_cat(m.group(1)),
                        "date":         date_globale,
                        "appel":        extraire_heure(m.group(2)),
                        "scratch":      extraire_heure(m.group(3)),
                        "debut":        extraire_heure(m.group(4)),
                        "equipe_appel": None,
                        "equipe_debut": None,
                        "equipe_info":  "à l'issue des épreuves individuelles",
                    })

        # ── FORMAT 4 : tableau texte avec jours en en-tête et 3 colonnes horaires
        # Ex: "Samedi 2 mai   appel   scratch   assauts"
        #     "M13 Sabre Hommes   14h45   15h   15h15"
        # Tenté EN PREMIER ou si FORMAT 1 donne peu de résultats (< 5)
        MOIS = {"janvier":"01","février":"02","mars":"03","avril":"04",
                "mai":"05","juin":"06","juillet":"07","août":"08",
                "septembre":"09","octobre":"10","novembre":"11","décembre":"12"}
        JOURS_FR = ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]

        LINE4_RE = _re.compile(
            r'^(.+?)\s+'
            r'(\d{1,2}[hH]\d{0,2})\s+'
            r'(\d{1,2}[hH]\d{0,2})\s+'
            r'(\d{1,2}[hH]\d{0,2})\s*$'
        )
        JOUR_ENTETE_RE = _re.compile(
            r'^(' + '|'.join(JOURS_FR) + r')\s+(\d{1,2})\s+(' +
            '|'.join(MOIS.keys()) + r')(?:\s+(\d{4}))?',
            _re.IGNORECASE
        )

        annee_m = _re.search(r'\b(202\d)\b', full_text)
        annee = annee_m.group(1) if annee_m else "2026"

        categories_f4 = []
        current_date_f4 = date_globale or ""

        for line in full_text.splitlines():
            line = line.strip()
            if not line: continue
            jm = JOUR_ENTETE_RE.match(line)
            if jm:
                jour_fr = jm.group(1).lower()
                dd      = jm.group(2).zfill(2)
                mois_fr = jm.group(3).lower()
                mois_n  = MOIS.get(mois_fr, "01")
                an      = jm.group(4) or annee
                current_date_f4 = f"{jour_fr} {dd}.{mois_n}.{an}"
                continue
            if any(k in line.lower() for k in ["appel","scratch","assaut","début","merci"]):
                continue
            m4 = LINE4_RE.match(line)
            if not m4: continue
            cat_raw  = m4.group(1).strip()
            cat_norm = normaliser_cat(cat_raw)
            if not cat_norm or cat_norm in ["None", ""]: continue
            categories_f4.append({
                "cat":          cat_norm,
                "date":         current_date_f4,
                "appel":        extraire_heure(m4.group(2)),
                "scratch":      extraire_heure(m4.group(3)),
                "debut":        extraire_heure(m4.group(4)),
                "equipe_appel": None,
                "equipe_debut": None,
                "equipe_info":  "à l'issue des épreuves individuelles",
            })

        # Préférer FORMAT 4 s'il donne plus de résultats que l'actuel
        if len(categories_f4) > len(categories):
            categories = categories_f4

        if not categories:
            if not full_text.strip():
                return jsonify({"error": "PDF image (scanné) — l'extraction automatique n'est pas possible. Renseignez les horaires manuellement dans le mail."}), 400
            return jsonify({"error": "Format de programme non reconnu. L'extraction automatique n'a pu détecter aucune catégorie."}), 400

        # Déduplication : si plusieurs lignes → même cat+date après normalisation,
        # garder la première (horaire le plus tôt généralement)
        seen = {}
        categories_uniq = []
        for entry in categories:
            key = (entry["cat"], entry["date"])
            if key not in seen:
                seen[key] = True
                categories_uniq.append(entry)

        data = {"lieu": lieu, "categories": categories_uniq}
        session["programme"] = data
        return jsonify(data)

    except Exception as e:
        app.logger.exception("Erreur dans upload_programme")
        return jsonify({"error": f"Erreur inattendue : {str(e)[:200]}"}), 500


@app.route("/api/clear_programme", methods=["POST"])
def clear_programme():
    session.pop("programme", None)
    return jsonify({"ok": True})



@app.route("/api/mail_body", methods=["POST"])
def mail_body():
    """Retourne le sujet et le corps du mail en JSON (pour Gmail)."""
    from core.parser import construire_donnees
    from core.config import BAREME_ARBITRES
    from collections import defaultdict as _dd
    from core.parser import date_avec_jour as _daj

    store = get_store()
    if not store:
        return jsonify({"error": "Aucun fichier chargé"}), 400

    data       = request.get_json() or {}
    titre_long = data.get("titre_long", "").strip()
    lieu       = data.get("lieu", "").strip()
    comp_type  = data.get("comp_type", "grand_est")

    titre = (f"{titre_long} - {lieu}" if titre_long and lieu
             else titre_long or lieu or "SYNESC")

    fichiers_list = [(m, t, a) for m, t, a in store]
    comp_obj = get_competition(comp_type)
    (groupes_indiv, groupes_equipe, arbitres_all,
     _, _, plage_dates, dates_ordonnees) = construire_donnees(
        fichiers_list, comp_obj.CAT_MAP_INDIV, comp_obj.CAT_MAP_EQUIPE,
        par_arme=(comp_type == "alsace"))

    CAT_LBL_I = comp_obj.CAT_LABEL_INDIV
    CAT_LBL_E = comp_obj.CAT_LABEL_EQUIPE
    BAREME_D  = dict(BAREME_ARBITRES)

    arb_par_date = _dd(list)
    seen_arb = _dd(set)
    for a in arbitres_all:
        date_src = a.get("date_source", "")
        lic = a.get("licence", "").strip()
        if lic and lic not in seen_arb[date_src]:
            seen_arb[date_src].add(lic)
            arb_par_date[date_src].append(a)

    licences_tireurs = {}
    for meta, tireurs, _ in fichiers_list:
        for t in tireurs:
            lic = t.get("licence", "").strip()
            if lic:
                licences_tireurs[lic] = t
    licences_arbitres = {a.get("licence","").strip(): a
                         for a in arbitres_all if a.get("licence","").strip()}
    arb_tireurs = [(lic, licences_tireurs[lic], licences_arbitres[lic])
                   for lic in licences_tireurs if lic in licences_arbitres]

    SEP = "─" * 50

    # Horaires depuis le programme PDF (si chargé)
    programme = session.get("programme", {})
    horaires_index = {}   # {(cat_norm, date_norm): {appel, scratch, debut}}
    for h in programme.get("categories", []):
        cat_norm  = h.get("cat", "").strip().upper()
        date_norm = h.get("date", "").strip().lower()
        horaires_index[(cat_norm, date_norm)] = h

    # Alias pour retrouver les horaires : V1/V2/V3/V4 → chercher aussi "VÉTÉRANS"
    VET_ALIASES = {"V1","V2","V3","V4","V1/V2","V3/V4","VETERAN","VETERANS"}
    SEN_ALIASES = {"SENIOR"}

    def _horaire(cat_key, cat_lbl, date_str, arme_code=None):
        """Retourne la ligne horaire si disponible, sinon chaîne vide."""
        date_label = _daj(date_str).lower()
        # Construire les clés à tester
        keys = [cat_lbl.upper(), cat_key.upper()]
        # Si arme connue, chercher d'abord la clé précise cat|arme
        if arme_code:
            keys = [f"{cat_key.upper()}|{arme_code}", f"{cat_lbl.upper()}|{arme_code}"] + keys
        # Fallback Vétérans / Seniors
        if cat_key.upper() in VET_ALIASES or cat_lbl.upper() in VET_ALIASES:
            keys += ["VÉTÉRANS", "VETERANS", "VÉTÉRAN"]
            if arme_code:
                keys += [f"VÉTÉRANS|{arme_code}"]
        if cat_key.upper() in SEN_ALIASES or cat_lbl.upper() in SEN_ALIASES:
            keys += ["SENIORS"]
        # Fallback : chercher cat|* (toutes armes) si rien trouvé
        keys += [k for k in set(k for k,_ in horaires_index.keys())
                 if k.startswith(f"{cat_key.upper()}|") or k.startswith(f"{cat_lbl.upper()}|")]
        for key in keys:
            h = horaires_index.get((key, date_label))
            if h:
                parts = []
                if h.get("appel"):   parts.append(f"Appel {h['appel']}")
                if h.get("scratch"): parts.append(f"Scratch {h['scratch']}")
                if h.get("debut"):   parts.append(f"Début {h['debut']}")
                return f" ({' · '.join(parts)})" if parts else ""
        return ""

    lignes = []
    lignes.append("Bonjour,")
    lignes.append("")
    lignes.append(f"Veuillez trouver ci-joint le tableau de synthèse SYNESC ainsi que le récapitulatif des inscriptions et de la situation arbitrale pour {titre}.")
    if programme.get("lieu"):
        lignes.append(f"Lieu : {programme['lieu']}")
    lignes.append("")

    def _decomposer_poules(nb, taille_min=5, taille_max=8):
        """
        Décompose nb tireurs en poules de taille_min à taille_max.
        Préfère les poules de 7, sinon cherche la combinaison la plus homogène.
        Retourne une liste de tailles de poules, ou None si impossible.
        """
        if nb < taille_min:
            return None
        # Essayer toutes les combinaisons de taille_ref de 7 à 5
        for ref in [7, 6, 8, 5]:
            nb_p = nb // ref
            reste = nb % ref
            if nb_p == 0:
                continue
            if reste == 0:
                return [ref] * nb_p
            # Distribuer le reste en augmentant certaines poules
            if reste > 0:
                # Peut-on absorber le reste en augmentant nb_p poules de 1 ?
                if ref + 1 <= taille_max:
                    # reste poules de (ref+1), nb_p-reste poules de ref
                    if reste <= nb_p:
                        poules = [ref + 1] * reste + [ref] * (nb_p - reste)
                        if all(taille_min <= t <= taille_max for t in poules):
                            return sorted(poules, reverse=True)
                # Peut-on ajouter une poule de plus avec les restes ?
                nb_p2 = nb_p + 1
                taille = nb // nb_p2
                reste2 = nb % nb_p2
                if taille >= taille_min:
                    poules = [taille + 1] * reste2 + [taille] * (nb_p2 - reste2)
                    if all(taille_min <= t <= taille_max for t in poules):
                        return sorted(poules, reverse=True)
        return None

    def _desc_poules(poules):
        """Décrit les poules de façon compacte ex: '2×7 + 1×8' ou '4×6'."""
        from collections import Counter
        c = Counter(poules)
        parts = [f"{v}×{k}" for k, v in sorted(c.items(), reverse=True)]
        return " + ".join(parts)

    def _formule(nb, is_equipe=False):
        """Propose une formule poules + TED selon l'effectif (sans rappel DT)."""
        if nb <= 0: return ""
        if is_equipe:
            if nb == 1: return f"    ↳ Formule proposée : 1 seule équipe — pas de compétition possible"
            elif nb == 2: return f"    ↳ Formule proposée : finale directe"
            elif nb <= 4: return f"    ↳ Formule proposée : 1 poule de {nb} + finale"
            else:
                ted = 4
                while ted < nb: ted *= 2
                return f"    ↳ Formule proposée : poule(s) + TED{ted}"
        else:
            if nb == 1: return "    ↳ Formule proposée : 1 seul tireur — titre attribué sans assaut"
            elif nb == 2: return f"    ↳ Formule proposée : finale directe pour le titre"
            elif nb <= 4: return f"    ↳ Formule proposée : 1 poule de {nb} + TED4"
            elif nb <= 9:
                ted = 4
                while ted < nb: ted *= 2
                return f"    ↳ Formule proposée : 1 poule de {nb} + TED{ted}"
            else:
                ted = 4
                while ted < nb: ted *= 2
                poules = _decomposer_poules(nb)
                if poules:
                    desc = _desc_poules(poules)
                    return f"    ↳ Formule proposée : {desc} ({len(poules)} poule{'s' if len(poules)>1 else ''}) + TED{ted}"
                else:
                    nb_p = 2 if nb <= 10 else round(nb / 7)
                    t = nb // nb_p; r = nb % nb_p
                    desc = (f"{r}×{t+1} + {nb_p-r}×{t}" if r else f"{nb_p}×{t}")
                    return f"    ↳ Formule proposée : {desc} ({nb_p} poules) + TED{ted}"

    # ── Chemin spécifique LORRAINE
    if comp_type == "lorraine":
        from competitions.lorraine import Lorraine as _Lorraine
        lorraine_obj = _Lorraine()
        groupes_arme = lorraine_obj._grouper_par_arme(fichiers_list)
        dates_lorraine = sorted({
            date_str
            for arme_data in groupes_arme.values()
            for date_str in arme_data
        })
        ARME_LBL = {"F": "Fleuret", "E": "Épée", "S": "Sabre"}

        for date_str in dates_lorraine:
            label = _daj(date_str).capitalize()
            lignes += [SEP, f"  {label.upper()}", SEP, ""]
            lignes.append("ÉPREUVES INDIVIDUELLES")

            total_t_lor = 0
            for arme_code in ["F", "E", "S"]:
                arme_data = groupes_arme.get(arme_code, {}).get(date_str, {})
                if not arme_data: continue
                arme_lbl = ARME_LBL.get(arme_code, arme_code)
                lignes.append(f"\n  [{arme_lbl}]")
                for cat, clubs in arme_data.items():
                    nb_h = sum(v.get("H", 0) for v in clubs.values())
                    nb_d = sum(v.get("D", 0) for v in clubs.values())
                    lbl  = CAT_LBL_I.get(cat, cat)
                    hor  = _horaire(cat, lbl, date_str, arme_code=arme_code)
                    total_t_lor += nb_h + nb_d
                    if nb_h:
                        lignes.append(f"    • {lbl} H — {nb_h} tireur(s){hor}")
                        lignes.append(_formule(nb_h))
                    if nb_d:
                        lignes.append(f"    • {lbl} D — {nb_d} tireur(s){hor}")
                        lignes.append(_formule(nb_d))
            lignes.append("")
            lignes.append("  ℹ Le DT reste libre de modifier les formules en fonction des effectifs.")
            lignes.append("")

            arb_jour = arb_par_date.get(date_str, [])
            nb_arb   = len(arb_jour)
            b_i      = 0 if total_t_lor < 4 else (1 if total_t_lor <= 8 else 2)
            lignes.append(f"ARBITRES INSCRITS — {nb_arb} / BESOIN ESTIMÉ — {b_i}")
            if nb_arb >= b_i:
                lignes.append(f"Situation : quota couvert ({nb_arb}/{b_i}).")
            else:
                lignes.append(f"⚠ Situation : il manque {b_i - nb_arb} arbitre(s).")

            clubs_all_lor = set()
            for arme_data in groupes_arme.values():
                for cat_data in arme_data.get(date_str, {}).values():
                    clubs_all_lor |= set(cat_data.keys())
            arb_par_club_lor = _dd(int)
            for a in arb_jour:
                arb_par_club_lor[a.get("club","").strip()] += 1
            deficit = []
            for club in sorted(clubs_all_lor):
                nb_t = sum(
                    v.get("H",0) + v.get("D",0)
                    for arme_data in groupes_arme.values()
                    for cat_data in arme_data.get(date_str, {}).values()
                    for c, v in cat_data.items() if c == club
                )
                bi = 0 if nb_t < 4 else (1 if nb_t <= 8 else 2)
                f  = arb_par_club_lor.get(club, 0)
                if bi > 0 and f < bi:
                    deficit.append((club, bi, f))
            if deficit:
                lignes.append("")
                lignes.append("⚠ Clubs en déficit d'arbitrage :")
                for club, b, f in deficit:
                    lignes.append(f"  • {club} — besoin : {b}, fournis : {f}, manque : {b-f}")
            lignes.append("")

    # ── Chemin standard GRAND EST / ALSACE
    else:
      for date_str in dates_ordonnees:
        label = _daj(date_str).capitalize()
        lignes += [SEP, f"  {label.upper()}", SEP, ""]

        cats_indiv = groupes_indiv.get(date_str, {})
        if cats_indiv:
            lignes.append("ÉPREUVES INDIVIDUELLES")
            for cat, clubs in cats_indiv.items():
                nb_h = sum(v.get("H", 0) for v in clubs.values())
                nb_d = sum(v.get("D", 0) for v in clubs.values())
                lbl  = CAT_LBL_I.get(cat, cat)
                hor  = _horaire(cat, lbl, date_str)
                if nb_h:
                    lignes.append(f"  • {lbl} H — {nb_h} tireur(s){hor}")
                    lignes.append(_formule(nb_h))
                if nb_d:
                    lignes.append(f"  • {lbl} D — {nb_d} tireur(s){hor}")
                    lignes.append(_formule(nb_d))
            lignes.append("")
            lignes.append("  ℹ Le DT reste libre de modifier les formules en fonction des effectifs.")
            lignes.append("")

        cats_equipe = groupes_equipe.get(date_str, {})
        if cats_equipe:
            lignes.append("ÉPREUVES PAR ÉQUIPES")

            # Détecter si les clés sont par arme (ex: "M13|F") ou simples (ex: "M13")
            par_arme_mail = any("|" in k for k in cats_equipe.keys())

            ARMES_MAIL = [("F","Fleuret"), ("E","Épée"), ("S","Sabre")]

            def _ecrire_equipes_bloc(cats_a_afficher):
                """Écrit les lignes équipes pour un sous-ensemble de cats."""
                for cat, clubs in cats_a_afficher:
                    is_mx  = any(v.get("mixte", False) for v in clubs.values())
                    nb_mx  = sum(len(v.get("equipes", set())) for v in clubs.values() if v.get("tireurs_MX", 0) > 0) if is_mx else 0
                    nb_h   = sum(len(v.get("equipes", set())) for v in clubs.values() if v.get("tireurs_H", 0) > 0) if not is_mx else 0
                    nb_d   = sum(len(v.get("equipes", set())) for v in clubs.values() if v.get("tireurs_D", 0) > 0) if not is_mx else 0
                    cat_base    = cat.split("|")[0] if "|" in cat else cat
                    arme_code_eq= cat.split("|")[1] if "|" in cat else None
                    lbl         = CAT_LBL_E.get(cat_base, cat_base)
                    hor         = _horaire(cat_base, lbl, date_str, arme_code=arme_code_eq)
                    if not hor:
                        hor = " (à l'issue des épreuves individuelles)"
                    if is_mx and nb_mx:
                        lignes.append(f"  • Équipe {lbl} Mixte — {nb_mx} équipe(s){hor}")
                        lignes.append(_formule(nb_mx, is_equipe=True))
                    else:
                        if nb_h:
                            lignes.append(f"  • Équipe {lbl} H — {nb_h} équipe(s){hor}")
                            lignes.append(_formule(nb_h, is_equipe=True))
                        if nb_d:
                            lignes.append(f"  • Équipe {lbl} D — {nb_d} équipe(s){hor}")
                            lignes.append(_formule(nb_d, is_equipe=True))

            if par_arme_mail:
                for arme_code, arme_lbl in ARMES_MAIL:
                    cats_arme = [(k, v) for k, v in cats_equipe.items()
                                 if k.endswith(f"|{arme_code}")]
                    if not cats_arme:
                        continue
                    lignes.append(f"  ── {arme_lbl} ──")
                    _ecrire_equipes_bloc(cats_arme)
            else:
                _ecrire_equipes_bloc(list(cats_equipe.items()))

            lignes.append("")

        arb_jour = arb_par_date.get(date_str, [])
        nb_arb = len(arb_jour)
        total_t = sum(v.get("H",0)+v.get("D",0)
                      for cats in groupes_indiv.get(date_str,{}).values()
                      for v in cats.values())
        total_e = sum(1 for cats in groupes_equipe.get(date_str,{}).values()
                      for v in cats.values()
                      if v.get("tireurs_H",0)+v.get("tireurs_D",0)+v.get("tireurs_MX",0)>0)
        b_i = 0 if total_t < 4 else (1 if total_t <= 8 else 2)
        b_e = 0 if total_e == 0 else (1 if total_e <= 2 else (2 if total_e <= 4 else 3))
        besoin = b_i + b_e
        lignes.append(f"ARBITRES INSCRITS — {nb_arb} / BESOIN ESTIMÉ — {besoin}")
        if nb_arb >= besoin:
            lignes.append(f"Situation : quota couvert ({nb_arb}/{besoin}).")
        else:
            lignes.append(f"⚠ Situation : il manque {besoin - nb_arb} arbitre(s).")

        arb_par_club = _dd(int)
        for a in arb_jour:
            arb_par_club[a.get("club","").strip()] += 1
        clubs_all = set()
        for cats in list(groupes_indiv.get(date_str,{}).values()) + list(groupes_equipe.get(date_str,{}).values()):
            clubs_all |= set(cats.keys())
        deficit = []
        for club in sorted(clubs_all):
            nb_t = sum(v.get("H",0)+v.get("D",0)
                       for cats in groupes_indiv.get(date_str,{}).values()
                       for c,v in cats.items() if c==club)
            nb_e = sum(1 for cats in groupes_equipe.get(date_str,{}).values()
                       for c,v in cats.items()
                       if c==club and v.get("tireurs_H",0)+v.get("tireurs_D",0)+v.get("tireurs_MX",0)>0)
            bi = 0 if nb_t<4 else (1 if nb_t<=8 else 2)
            be = 0 if nb_e==0 else (1 if nb_e<=2 else (2 if nb_e<=4 else 3))
            b  = bi + be
            f  = arb_par_club.get(club, 0)
            if b > 0 and f < b:
                deficit.append((club, b, f))
        if deficit:
            lignes.append("")
            lignes.append("⚠ Clubs en déficit d'arbitrage :")
            for club, b, f in deficit:
                lignes.append(f"  • {club} — besoin : {b}, fournis : {f}, manque : {b-f}")
        lignes.append("")

    if arb_tireurs:
        lignes += [SEP, "  ARBITRES ÉGALEMENT INSCRITS EN COMPÉTITION", SEP, ""]
        for lic, tireur, arb in sorted(arb_tireurs, key=lambda x: x[1]["club"]):
            nom = f"{tireur['nom']} {tireur['prenom']}".strip()
            lignes.append(f"  • {nom} ({tireur['club']}) — licence {lic}")
        lignes.append("")

    # ── Bilan financier prévisionnel global
    try:
        tarif_i = getattr(comp_obj, "TARIF_INDIV",  {})
        tarif_e = getattr(comp_obj, "TARIF_EQUIPE", {})
        total_recettes = 0
        for date_str in dates_ordonnees:
            # Recettes individuelles
            for cat, clubs in groupes_indiv.get(date_str, {}).items():
                cat_base = cat.split("|")[0] if "|" in cat else cat
                pu = tarif_i.get(cat_base, 0)
                nb = sum(v.get("H", 0) + v.get("D", 0) for v in clubs.values())
                total_recettes += nb * pu
            # Recettes équipes — même logique que le bilan Excel :
            # sum(len(equipes) clubs avec H>0) + sum(len(equipes) clubs avec D>0)
            for cat, clubs in groupes_equipe.get(date_str, {}).items():
                cat_base = cat.split("|")[0] if "|" in cat else cat
                pu = tarif_e.get(cat_base, 0)
                is_mx = any(v.get("mixte", False) for v in clubs.values())
                if is_mx:
                    nb = sum(len(v.get("equipes", set()))
                             for v in clubs.values() if v.get("tireurs_MX", 0) > 0)
                else:
                    nb_h = sum(len(v.get("equipes", set()))
                               for v in clubs.values() if v.get("tireurs_H", 0) > 0)
                    nb_d = sum(len(v.get("equipes", set()))
                               for v in clubs.values() if v.get("tireurs_D", 0) > 0)
                    nb = nb_h + nb_d
                total_recettes += nb * pu

        # Dépenses arbitres — utilise les statuts Excel si disponibles
        arbitres_statuts = session.get("arbitres_statuts", {})
        BAREME_D = dict(BAREME_ARBITRES)
        source_statuts = "prévisionnel"

        if arbitres_statuts:
            # Statuts réels depuis l'Excel modifié
            total_depenses = sum(v["cout_retenu"] for v in arbitres_statuts.values())
            nb_ret_mail = sum(v["retenu"] for v in arbitres_statuts.values())
            nb_lib_mail = sum(v["libere"] for v in arbitres_statuts.values())
            source_statuts = "réel"
        else:
            # Fallback : tous les arbitres comme retenus
            total_depenses = 0
            for date_str in dates_ordonnees:
                for a in arb_par_date.get(date_str, []):
                    total_depenses += BAREME_D.get(a.get("categorie", ""), 0)
            nb_ret_mail = sum(len(v) for v in arb_par_date.values())
            nb_lib_mail = 0

        solde = total_recettes - total_depenses
        if total_recettes > 0 or total_depenses > 0:
            signe    = "+" if solde >= 0 else "−"
            abs_solde = abs(solde)
            lignes += [SEP, "  BILAN FINANCIER PRÉVISIONNEL", SEP, ""]
            if source_statuts == "réel":
                pass  # statuts réels chargés, on affiche juste la balance
            else:
                lignes.append(f"Arbitres inscrits : {nb_ret_mail} (tous comptés — charger l'Excel pour affiner)")
            lignes.append(f"Balance estimée : {signe} {abs_solde:,.0f} €".replace(",", " "))
            lignes.append("")
            if solde > 200:
                lignes.append(
                    "Cette marge positive permet d'envisager un ajustement du nombre "
                    "d'arbitres si nécessaire pour couvrir les besoins de la compétition."
                )
            elif solde >= 0:
                lignes.append(
                    "La balance est équilibrée. Tout désistement d'arbitre devra être "
                    "traité avec attention pour ne pas dégrader le solde."
                )
            else:
                lignes.append(
                    f"La balance est négative ({signe} {abs_solde:,.0f} €). "
                    "Il est recommandé de revoir le nombre d'arbitres ou de vérifier "
                    "les engagements des clubs.".replace(",", " ")
                )
            lignes.append("")
    except Exception:
        pass  # Ne pas bloquer le mail si le calcul échoue

    lignes += [SEP, "  RAPPELS", SEP, ""]
    lignes.append("• Tireurs absents : les engagements restent dus sauf motif valable.")
    lignes.append('  Inscriptions hors délai majorées — <a href="https://tinyurl.com/hdlrege">https://tinyurl.com/hdlrege</a>')
    lignes.append('• Dérogations au niveau d\'arbitrage : contacter la Présidente de la CRA, Auxane Cholley — <a href="mailto:auxane.cholley@hotmail.fr">auxane.cholley@hotmail.fr</a>')
    lignes.append('• Fichiers résultats (*.cotcot, PDF, FFF) à transmettre à <a href="mailto:atrcrege@gmail.com">atrcrege@gmail.com</a>')
    lignes += ["", "Cordialement,", "La Ligue Régionale d'Escrime Grand Est"]

    sujet = f"{titre} — Synthèse engagements et arbitrage"

    # Générer HTML minimal avec liens actifs
    import html as _html
    import re as _re
    def _to_html(txt):
        parts = _re.split(r'(<a [^>]+>.*?</a>)', txt)
        out = []
        for p in parts:
            if p.startswith('<a '): out.append(p)
            else: out.append(_html.escape(p))
        return "".join(out)

    lines_html = [_to_html(l) for l in lignes]
    corps_html = (
        '<html><body><pre style="font-family:Consolas,monospace;font-size:13px;'
        'white-space:pre-wrap;line-height:1.5">'
        + "\n".join(lines_html)
        + '</pre></body></html>'
    )
    corps_txt = "\n".join(lignes)  # version texte pour la modale copier-coller
    # Nettoyer les balises HTML du texte brut
    import re as _re2
    corps_txt_clean = _re2.sub(r'<[^>]+>', '', corps_txt)

    return jsonify({"sujet": sujet, "corps": corps_txt_clean, "corps_html": corps_html})



@app.route("/api/generate_mail", methods=["POST"])
def generate_mail():
    import email.mime.multipart as _mp
    import email.mime.text as _mt
    import email.mime.base as _mb
    import email.encoders as _enc
    from collections import defaultdict as _dd
    from core.parser import construire_donnees
    from core.config import BAREME_ARBITRES

    store = get_store()
    if not store:
        return jsonify({"error": "Aucun fichier chargé"}), 400

    data       = request.get_json() or {}
    titre_long = data.get("titre_long", "").strip()
    lieu       = data.get("lieu", "").strip()
    comp_type  = data.get("comp_type", "grand_est")
    destinataires = data.get("destinataires", "").strip()

    titre = (f"{titre_long} - {lieu}" if titre_long and lieu
             else titre_long or lieu or "SYNESC")

    fichiers_list = [(m, t, a) for m, t, a in store]

    # Générer le fichier Excel
    comp = get_competition(comp_type)
    excel_buf = comp.generer_excel(fichiers_list, titre_comp=titre)
    safe = unicodedata.normalize("NFD", titre)
    safe = "".join(c for c in safe if unicodedata.category(c) != "Mn")
    safe = "".join(c for c in safe if c.isalnum() or c in " -_").strip()
    safe = re.sub(r" {2,}", " ", safe).replace(" ", "_")
    excel_filename = f"{safe}.xlsx"

    # Construire les données pour le corps du mail
    comp_obj = get_competition(comp_type)
    (groupes_indiv, groupes_equipe, arbitres_all,
     _, _, plage_dates, dates_ordonnees) = construire_donnees(
        fichiers_list, comp_obj.CAT_MAP_INDIV, comp_obj.CAT_MAP_EQUIPE)

    from core.parser import date_avec_jour as _daj

    # Corps du mail : synthèse par jour
    lignes = []
    lignes.append(f"Bonjour,")
    lignes.append(f"")
    lignes.append(f"Veuillez trouver en pièce jointe le tableau de synthèse SYNESC ainsi que le récapitulatif des inscriptions et de la situation arbitrale pour {titre}.")
    lignes.append(f"")

    CAT_LBL_I = comp_obj.CAT_LABEL_INDIV
    CAT_LBL_E = comp_obj.CAT_LABEL_EQUIPE
    BAREME_D  = dict(BAREME_ARBITRES)

    # Arbitres par date et par club
    arb_par_date = _dd(list)
    seen_arb = _dd(set)
    for a in arbitres_all:
        date_src = a.get("date_source", "")
        lic = a.get("licence", "").strip()
        if lic and lic not in seen_arb[date_src]:
            seen_arb[date_src].add(lic)
            arb_par_date[date_src].append(a)

    # Arbitres-tireurs
    licences_tireurs = {}
    for meta, tireurs, _ in fichiers_list:
        for t in tireurs:
            lic = t.get("licence", "").strip()
            if lic:
                licences_tireurs[lic] = t
    licences_arbitres = {a.get("licence","").strip(): a
                         for a in arbitres_all if a.get("licence","").strip()}
    arb_tireurs = [(lic, licences_tireurs[lic], licences_arbitres[lic])
                   for lic in licences_tireurs if lic in licences_arbitres]

    SEP = "─" * 50

    for date_str in dates_ordonnees:
        label = _daj(date_str).capitalize()
        lignes.append(SEP)
        lignes.append(f"  {label.upper()}")
        lignes.append(SEP)
        lignes.append("")

        # Indiv
        cats_indiv = groupes_indiv.get(date_str, {})
        if cats_indiv:
            lignes.append("ÉPREUVES INDIVIDUELLES")
            for cat, clubs in cats_indiv.items():
                nb_h = sum(v.get("H", 0) for v in clubs.values())
                nb_d = sum(v.get("D", 0) for v in clubs.values())
                lbl  = CAT_LBL_I.get(cat, cat)
                if nb_h: lignes.append(f"  • {lbl} H — {nb_h} tireur(s)")
                if nb_d: lignes.append(f"  • {lbl} D — {nb_d} tireur(s)")
            lignes.append("")

        # Équipes
        cats_equipe = groupes_equipe.get(date_str, {})
        if cats_equipe:
            lignes.append("ÉPREUVES PAR ÉQUIPES")
            for cat, clubs in cats_equipe.items():
                is_mx   = any(v.get("mixte", False) for v in clubs.values())
                lbl     = CAT_LBL_E.get(cat, cat)
                if is_mx:
                    nb_mx = sum(len(v.get("equipes", set())) for v in clubs.values() if v.get("tireurs_MX", 0) > 0)
                    if nb_mx: lignes.append(f"  • Équipe {lbl} Mixte — {nb_mx} équipe(s)")
                else:
                    nb_eq_h = sum(len(v.get("equipes", set())) for v in clubs.values() if v.get("tireurs_H", 0) > 0)
                    nb_eq_d = sum(len(v.get("equipes", set())) for v in clubs.values() if v.get("tireurs_D", 0) > 0)
                    if nb_eq_h: lignes.append(f"  • Équipe {lbl} H — {nb_eq_h} équipe(s)")
                    if nb_eq_d: lignes.append(f"  • Équipe {lbl} D — {nb_eq_d} équipe(s)")
            lignes.append("")

        # Arbitres
        arb_jour = arb_par_date.get(date_str, [])
        nb_arb = len(arb_jour)
        lignes.append(f"ARBITRES INSCRITS — {nb_arb}")

        # Besoin total indiv + équipe
        total_tireurs = sum(
            v.get("H", 0) + v.get("D", 0)
            for cats in groupes_indiv.get(date_str, {}).values()
            for v in cats.values()
        )
        total_equipes = sum(
            1 for cats in groupes_equipe.get(date_str, {}).values()
            for v in cats.values()
            if v.get("tireurs_H", 0) + v.get("tireurs_D", 0) + v.get("tireurs_MX", 0) > 0
        )
        besoin_i = 0 if total_tireurs < 4 else (1 if total_tireurs <= 8 else 2)
        besoin_e = 0 if total_equipes == 0 else (1 if total_equipes <= 2 else (2 if total_equipes <= 4 else 3))
        besoin   = besoin_i + besoin_e
        lignes.append(f"BESOIN ESTIMÉ — {besoin}")

        if nb_arb >= besoin:
            lignes.append(f"Situation : quota couvert ({nb_arb}/{besoin}).")
        else:
            manque = besoin - nb_arb
            lignes.append(f"⚠ Situation : il manque {manque} arbitre(s) pour couvrir le quota.")

        # Clubs en déficit
        clubs_all = set()
        for cats in list(groupes_indiv.get(date_str, {}).values()) + list(groupes_equipe.get(date_str, {}).values()):
            clubs_all |= set(cats.keys())
        arb_par_club = _dd(int)
        for a in arb_jour:
            arb_par_club[a.get("club","").strip()] += 1

        deficit = []
        for club in sorted(clubs_all):
            # Besoin de ce club ce jour
            nb_t = sum(
                v.get("H", 0) + v.get("D", 0)
                for cats in groupes_indiv.get(date_str, {}).values()
                for c, v in cats.items() if c == club
            )
            nb_e = sum(
                1 for cats in groupes_equipe.get(date_str, {}).values()
                for c, v in cats.items()
                if c == club and v.get("tireurs_H", 0) + v.get("tireurs_D", 0) + v.get("tireurs_MX", 0) > 0
            )
            b_i = 0 if nb_t < 4 else (1 if nb_t <= 8 else 2)
            b_e = 0 if nb_e == 0 else (1 if nb_e <= 2 else (2 if nb_e <= 4 else 3))
            besoin_club = b_i + b_e
            fournis = arb_par_club.get(club, 0)
            if besoin_club > 0 and fournis < besoin_club:
                deficit.append((club, besoin_club, fournis))

        if deficit:
            lignes.append("")
            lignes.append("⚠ Clubs en déficit d'arbitrage :")
            for club, b, f in deficit:
                lignes.append(f"  • {club} — besoin : {b}, fournis : {f}, manque : {b - f}")

        lignes.append("")

    # Arbitres-tireurs
    if arb_tireurs:
        lignes.append(SEP)
        lignes.append("  ARBITRES ÉGALEMENT INSCRITS EN COMPÉTITION")
        lignes.append(SEP)
        lignes.append("")
        for lic, tireur, arb in sorted(arb_tireurs, key=lambda x: x[1]["club"]):
            nom = f"{tireur['nom']} {tireur['prenom']}".strip()
            lignes.append(f"  • {nom} ({tireur['club']}) — licence {lic}")
        lignes.append("")

    # Rappels
    lignes.append(SEP)
    lignes.append("  RAPPELS")
    lignes.append(SEP)
    lignes.append("")
    lignes.append("• Tireurs absents : les engagements restent dus sauf motif valable.")
    lignes.append('  Inscriptions hors délai majorées — <a href="https://tinyurl.com/hdlrege">https://tinyurl.com/hdlrege</a>')
    lignes.append('• Dérogations au niveau d\'arbitrage : contacter la Présidente de la CRA, Auxane Cholley — <a href="mailto:auxane.cholley@hotmail.fr">auxane.cholley@hotmail.fr</a>')
    lignes.append('• Fichiers résultats (*.cotcot, PDF, FFF) à transmettre à <a href="mailto:atrcrege@gmail.com">atrcrege@gmail.com</a>')
    lignes.append("")
    lignes.append("Cordialement,")
    lignes.append("La Ligue Régionale d'Escrime Grand Est")

    # Corps HTML : on enveloppe le texte préformaté avec les liens actifs
    corps_txt = "\r\n".join(lignes)
    # Convertir en HTML minimal : balises <br> + liens déjà présents
    import html as _html
    def _to_html(txt):
        # Préserver les liens <a href=...> déjà insérés, échapper le reste
        import re
        parts = re.split(r'(<a [^>]+>.*?</a>)', txt)
        out = []
        for p in parts:
            if p.startswith('<a '):
                out.append(p)
            else:
                out.append(_html.escape(p))
        return "".join(out)

    lines_html = [_to_html(l) for l in lignes]
    corps_html = (
        '<html><body><pre style="font-family:Consolas,monospace;font-size:13px;'
        'white-space:pre-wrap;line-height:1.5">'
        + "\n".join(lines_html)
        + '</pre></body></html>'
    )

    # Construire le .eml
    msg = _mp.MIMEMultipart()
    msg["Subject"] = f"{titre} — Synthèse engagements et arbitrage"
    msg["From"]    = "lrege@escrime-grandest.fr"
    msg["To"]      = destinataires or ""
    msg.attach(_mt.MIMEText(corps_html, "html", "utf-8"))

    # Pièce jointe Excel
    part = _mb.MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_buf.getvalue())
    _enc.encode_base64(part)
    part.add_header("Content-Disposition", "attachment",
                    filename=excel_filename)
    msg.attach(part)

    eml_bytes = msg.as_bytes()
    eml_filename = excel_filename.replace(".xlsx", ".eml")

    return send_file(
        io.BytesIO(eml_bytes),
        as_attachment=True,
        download_name=eml_filename,
        mimetype="message/rfc822",
    )


@app.route("/api/pdf_arbitres", methods=["POST"])
def pdf_arbitres():
    """
    Reçoit un Excel SYNESC (avec feuille 'Récap Arbitres'),
    lit les statuts Retenu/Libéré et génère un PDF récapitulatif.
    """
    from core.pdf_arbitres import generer_pdf_arbitres

    f = request.files.get("excel")
    if not f:
        return jsonify({"error": "Aucun fichier Excel reçu"}), 400

    ext = f.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xlsm"):
        return jsonify({"error": "Fichier Excel (.xlsx) attendu"}), 400

    excel_bytes = f.read()

    try:
        pdf_buf = generer_pdf_arbitres(excel_bytes)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Erreur génération PDF : {e}"}), 500

    nom_base = f.filename.rsplit(".", 1)[0]
    return send_file(
        pdf_buf,
        as_attachment=True,
        download_name=f"{nom_base}_arbitres.pdf",
        mimetype="application/pdf",
    )


@app.route("/api/charger_excel_arbitres", methods=["POST"])
def charger_excel_arbitres():
    """
    Reçoit l'Excel SYNESC modifié, lit les statuts Retenu/Libéré
    et les tarifs de la feuille Arbitres, et les stocke en session
    pour que le mail utilise les vrais montants.
    Retourne un résumé : nb retenus, nb libérés, coût retenus.
    """
    import openpyxl
    from io import BytesIO

    f = request.files.get("excel")
    if not f:
        return jsonify({"error": "Aucun fichier Excel reçu"}), 400
    if f.filename.rsplit(".", 1)[-1].lower() not in ("xlsx", "xlsm"):
        return jsonify({"error": "Fichier Excel (.xlsx) attendu"}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True)
        if "Arbitres" not in wb.sheetnames:
            return jsonify({"error": "Feuille 'Arbitres' introuvable"}), 400

        ws = wb["Arbitres"]
        arbitres_statuts = {}  # {date_label: {retenu: nb, libere: nb, cout_retenu: float}}
        current_date = ""

        for row in ws.iter_rows(values_only=True):
            v0 = str(row[0] or "").strip()
            if not v0 or v0 == "Club": continue
            if "arbitres du" in v0.lower() or "arbitres —" in v0.lower():
                current_date = v0
                arbitres_statuts[current_date] = {"retenu": 0, "libere": 0, "cout_retenu": 0.0}
                continue
            if "tireur" in v0.lower(): break
            if current_date and row[1]:
                statut = str(row[6] or "Retenu").strip()
                tarif  = float(row[5] or 0) if row[5] is not None else 0
                if statut == "Libéré":
                    arbitres_statuts[current_date]["libere"] += 1
                else:
                    arbitres_statuts[current_date]["retenu"] += 1
                    arbitres_statuts[current_date]["cout_retenu"] += tarif

        # Totaux globaux
        nb_ret   = sum(v["retenu"]       for v in arbitres_statuts.values())
        nb_lib   = sum(v["libere"]       for v in arbitres_statuts.values())
        cout_ret = sum(v["cout_retenu"]  for v in arbitres_statuts.values())

        # Stocker en session
        session["arbitres_statuts"] = arbitres_statuts

        return jsonify({
            "ok": True,
            "nb_retenus":  nb_ret,
            "nb_liberes":  nb_lib,
            "cout_retenus": cout_ret,
            "detail": {k: v for k, v in arbitres_statuts.items()},
        })

    except Exception as e:
        return jsonify({"error": f"Erreur lecture Excel : {e}"}), 500


@app.route("/api/clear_excel_arbitres", methods=["POST"])
def clear_excel_arbitres():
    session.pop("arbitres_statuts", None)
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(debug=True, port=5002)
