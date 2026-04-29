"""
competitions/grand_est.py
Compétition Championnat Régional Grand Est.
Implémente toutes les règles spécifiques.
"""

from collections import defaultdict
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as _PF

from core.excel_base import sc, mk_font, mk_fill, mk_align, border_all, gcl
from core.config import BAREME_ARBITRES, COLORS
from core.parser import date_avec_jour, parse_date_key
from .base import CompetitionBase


class GrandEst(CompetitionBase):
    """Championnat Régional Grand Est."""

    # ── Catégories
    CAT_MAP_INDIV = {
        "M13": "M13", "M15": "M15",
        "M17": "M17", "M20": "M20",
        "SENIORS": "SENIORS", "SENIOR": "SENIORS",
        "V1": "V1", "V2": "V2", "V3": "V3", "V4": "V4",
        "V1, V2": "V1", "V3, V4": "V3", "V1, V2, V3, V4": "V1",
    }

    CAT_MAP_EQUIPE = {
        "M15": "M15",
        "M17": "M17", "M20": "M20",
        "SENIORS": "SENIORS", "SENIOR": "SENIORS",
        "V1": "VET", "V2": "VET", "V3": "GVET", "V4": "GVET",
        "V1, V2": "VET", "V3, V4": "GVET", "V1, V2, V3, V4": "VET",
    }

    CATS_INDIV  = ["M13", "M15", "M17", "M20", "SENIORS", "V1", "V2", "V3", "V4"]
    CATS_EQUIPE = ["M15", "M17", "M20", "SENIORS", "VET", "GVET"]

    CATS_EXTRANET_INDIV  = ["M13", "M15", "M17", "M20", "SENIORS", "V1", "V2", "V3", "V4"]
    CATS_EXTRANET_EQUIPE = ["M15", "M17", "M20", "SENIORS", "VET", "GVET"]

    CAT_LABEL_INDIV = {
        "M13": "M13", "M15": "M15",
        "M17": "M17", "M20": "M20", "SENIORS": "Séniors",
        "V1": "V1", "V2": "V2", "V3": "V3", "V4": "V4",
    }

    CAT_LABEL_EQUIPE = {
        "M15": "M15",
        "M17": "M17", "M20": "M20", "SENIORS": "Séniors",
        "VET": "Vétérans", "GVET": "Gds Vétérans",
    }

    TARIF_INDIV = {
        "M13": 10, "M15": 10,
        "M17": 15, "M20": 15, "SENIORS": 15,
        "V1": 15, "V2": 15, "V3": 15, "V4": 15,
    }

    TARIF_EQUIPE = {
        "M15": 30,
        "M17": 40, "M20": 40, "SENIORS": 40,
        "VET": 40, "GVET": 40,
    }

    NOM_FEUILLE_INDIV  = "Épreuve Individuelle"
    NOM_FEUILLE_EQUIPE = "Épreuve Équipe"

    # ── Feuille Individuelle

    def feuille_indiv(self, ws, groupes_indiv, titre, dates_ordonnees):
        self._feuille_egesc(ws, groupes_indiv, titre, dates_ordonnees,
                            is_equipe=False, arbitres_all=self._arbitres_all)

    # ── Feuille Équipe

    def feuille_equipe(self, ws, groupes_equipe, titre, dates_ordonnees):
        self._feuille_egesc(ws, groupes_equipe, titre, dates_ordonnees,
                            is_equipe=True, arbitres_all=self._arbitres_all)

    # ── Moteur commun indiv/équipe

    def _feuille_egesc(self, ws, groupes, titre, dates_ordonnees, is_equipe,
                       arbitres_all=None):
        """Génère le tableau récapitulatif indiv ou équipe."""

        ordre_cats = self.CATS_EQUIPE if is_equipe else self.CATS_INDIV
        cat_labels = self.CAT_LABEL_EQUIPE if is_equipe else self.CAT_LABEL_INDIV

        # Catégories présentes par date — uniquement celles avec au moins 1 tireur
        def cat_a_des_tireurs(date_str, cat):
            clubs = groupes.get(date_str, {}).get(cat, {})
            if is_equipe:
                return any(
                    v.get("tireurs_H", 0) + v.get("tireurs_D", 0) + v.get("tireurs_MX", 0) > 0
                    for v in clubs.values()
                )
            return any(
                v.get("H", 0) + v.get("D", 0) > 0
                for v in clubs.values()
            )

        cats_par_date = {
            d: [c for c in ordre_cats
                if c in groupes.get(d, {}) and cat_a_des_tireurs(d, c)]
            for d in dates_ordonnees
        }

        # Tous les clubs
        tous_clubs = sorted({
            club
            for d in groupes.values()
            for cat in d.values()
            for club in cat.keys()
        })

        # ── Plan des colonnes
        col_idx    = 3
        plan_dates = []

        for date_str in dates_ordonnees:
            cats = cats_par_date.get(date_str, [])
            if not cats:
                continue
            dp = {"date": date_str, "label": date_avec_jour(date_str),
                  "col_start": col_idx, "cats": []}
            for cat in cats:
                if is_equipe:
                    dp["cats"].append({
                        "cat": cat,
                        "col_eq_h": col_idx,
                        "col_ti_h": col_idx + 1,
                        "col_eq_d": col_idx + 2,
                        "col_ti_d": col_idx + 3,
                    })
                    col_idx += 4
                else:
                    dp["cats"].append({
                        "cat": cat,
                        "col_h": col_idx,
                        "col_d": col_idx + 1,
                    })
                    col_idx += 2
            dp["col_end"] = col_idx - 1
            plan_dates.append(dp)

        col_tot_h   = col_idx;     col_idx += 1
        col_tot_d   = col_idx;     col_idx += 1
        col_tot_clb = col_idx;     col_idx += 1
        # Après totaux : besoin arb / fournis / statut — une triplet par jour
        for dp in plan_dates:
            dp["col_arb"]     = col_idx;  col_idx += 1  # besoin calculé
            dp["col_fournis"] = col_idx;  col_idx += 1  # arbitres fournis par le club
            dp["col_statut"]  = col_idx;  col_idx += 1  # OK / Manque N
        last_col = col_idx - 1

        DATA_START = 5

        # ── Pré-calcul : nb d'arbitres fournis par club et par jour
        # {date_str: {club: nb_arbitres}}
        from collections import defaultdict as _dd
        arb_par_club_date = _dd(lambda: _dd(int))
        seen_arb = _dd(set)
        for a in (arbitres_all or []):
            date_src = a.get("date_source", "")
            club = a.get("club", "").strip()
            lic  = a.get("licence", "").strip()
            if lic and lic not in seen_arb[(date_src, club)]:
                seen_arb[(date_src, club)].add(lic)
                arb_par_club_date[date_src][club] += 1

        # ── Ligne 1 : titre
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        c = ws.cell(row=1, column=1, value=titre)
        c.font = mk_font(bold=True, size=13)
        c.alignment = mk_align(h="center")
        ws.row_dimensions[1].height = 26

        # ── En-têtes col A:B
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=2)
        sc(ws.cell(row=2, column=1, value="Club"),
           bold=True, bg=COLORS["navy"], fg=COLORS["white"])

        for dp in plan_dates:
            # Ligne 2 : date
            ws.merge_cells(start_row=2, start_column=dp["col_start"],
                           end_row=2, end_column=dp["col_end"])
            c = ws.cell(row=2, column=dp["col_start"],
                        value=dp["label"].capitalize())
            c.font      = mk_font(bold=True, size=10, color=COLORS["white"])
            c.fill      = mk_fill("1F3A6E")
            c.alignment = mk_align(h="center")
            c.border    = border_all()

            for cp in dp["cats"]:
                lbl = cat_labels.get(cp["cat"], cp["cat"])
                if is_equipe:
                    # Détecter si cette catégorie est mixte
                    is_mx_cat = any(
                        v.get("mixte", False)
                        for d_str in dates_ordonnees
                        for v in groupes.get(d_str, {}).get(cp["cat"], {}).values()
                    )
                    ws.merge_cells(start_row=3, start_column=cp["col_eq_h"],
                                   end_row=3, end_column=cp["col_ti_d"])
                    sc(ws.cell(row=3, column=cp["col_eq_h"], value=lbl),
                       bold=True, bg=COLORS["blue"], fg=COLORS["white"])
                    if is_mx_cat:
                        for ci, lbl4, bg4 in [
                            (cp["col_eq_h"], "Eq.MX", "E8D5F5"),
                            (cp["col_ti_h"], "Tir.MX","F5E8FF"),
                            (cp["col_eq_d"], "",      COLORS["light_pink"]),
                            (cp["col_ti_d"], "",      COLORS["light_d"]),
                        ]:
                            sc(ws.cell(row=4, column=ci, value=lbl4),
                               bold=True, bg=bg4, size=9)
                    else:
                        for ci, lbl4, bg4 in [
                            (cp["col_eq_h"], "Eq.H",  COLORS["light_blue"]),
                            (cp["col_ti_h"], "Tir.H", COLORS["light_h"]),
                            (cp["col_eq_d"], "Eq.D",  COLORS["light_pink"]),
                            (cp["col_ti_d"], "Tir.D", COLORS["light_d"]),
                        ]:
                            sc(ws.cell(row=4, column=ci, value=lbl4),
                               bold=True, bg=bg4, size=9)
                else:
                    ws.merge_cells(start_row=3, start_column=cp["col_h"],
                                   end_row=3, end_column=cp["col_d"])
                    sc(ws.cell(row=3, column=cp["col_h"], value=lbl),
                       bold=True, bg=COLORS["blue"], fg=COLORS["white"])
                    sc(ws.cell(row=4, column=cp["col_h"], value="H"),
                       bold=True, bg=COLORS["light_blue"])
                    sc(ws.cell(row=4, column=cp["col_d"], value="D"),
                       bold=True, bg=COLORS["light_pink"])

        # Totaux en-têtes
        for ci, lbl in [
            (col_tot_h,   "Total H"),
            (col_tot_d,   "Total D"),
            (col_tot_clb, "Total Club"),
        ]:
            ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
            sc(ws.cell(row=2, column=ci, value=lbl),
               bold=True, bg=COLORS["accent"], fg=COLORS["white"], wrap=True)
            sc(ws.cell(row=4, column=ci, value=""), bg="EBD5A0", border=True)

        # En-têtes besoin arbitres / fournis / statut — après Total Club, triplet par jour
        BG_ARB = "FFF2CC"
        BG_OK  = "C6EFCE"
        BG_KO  = "FFC7CE"
        for dp in plan_dates:
            jour_court = dp["label"].split()[0].capitalize()
            # Besoin
            ws.merge_cells(start_row=2, start_column=dp["col_arb"],
                           end_row=3, end_column=dp["col_arb"])
            sc(ws.cell(row=2, column=dp["col_arb"], value="Besoin\narb."),
               bold=True, bg=BG_ARB, size=9, fg="7F6000", wrap=True)
            sc(ws.cell(row=4, column=dp["col_arb"], value=""), bg=BG_ARB, border=True)
            # Fournis
            ws.merge_cells(start_row=2, start_column=dp["col_fournis"],
                           end_row=3, end_column=dp["col_fournis"])
            sc(ws.cell(row=2, column=dp["col_fournis"],
                       value=f"Fournis\n{jour_court}"),
               bold=True, bg="D9E8F5", size=9, fg="1F4E79", wrap=True)
            sc(ws.cell(row=4, column=dp["col_fournis"], value=""), bg="D9E8F5", border=True)
            # Statut
            ws.merge_cells(start_row=2, start_column=dp["col_statut"],
                           end_row=3, end_column=dp["col_statut"])
            sc(ws.cell(row=2, column=dp["col_statut"],
                       value=f"Statut\n{jour_court}"),
               bold=True, bg="D9D9D9", size=9, fg="000000", wrap=True)
            sc(ws.cell(row=4, column=dp["col_statut"], value=""), bg="D9D9D9", border=True)

        # ── Données
        for i, club in enumerate(tous_clubs):
            row = DATA_START + i
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            sc(ws.cell(row=row, column=1, value=club), h="left")

            for dp in plan_dates:
                for cp in dp["cats"]:
                    cat       = cp["cat"]
                    club_data = groupes.get(dp["date"], {}).get(cat, {}).get(club, {})

                    if is_equipe:
                        nb_eq  = len(club_data.get("equipes", set())) if club_data else 0
                        tir_h  = club_data.get("tireurs_H", 0) if club_data else 0
                        tir_d  = club_data.get("tireurs_D", 0) if club_data else 0
                        tir_mx = club_data.get("tireurs_MX", 0) if club_data else 0
                        is_mx  = club_data.get("mixte", False) if club_data else False
                        if is_mx:
                            # Équipe mixte : afficher dans colonne H avec fond violet clair
                            for ci, val, bg in [
                                (cp["col_eq_h"], nb_eq or None,  "E8D5F5"),
                                (cp["col_ti_h"], tir_mx or None, "F5E8FF"),
                                (cp["col_eq_d"], None,           None),
                                (cp["col_ti_d"], None,           None),
                            ]:
                                sc(ws.cell(row=row, column=ci),
                                   val=val, bg=bg if val else None)
                        else:
                            for ci, val, bg in [
                                (cp["col_eq_h"], nb_eq if tir_h else None, COLORS["light_blue"]),
                                (cp["col_ti_h"], tir_h or None,            COLORS["light_h"]),
                                (cp["col_eq_d"], nb_eq if tir_d else None, COLORS["light_pink"]),
                                (cp["col_ti_d"], tir_d or None,            COLORS["light_d"]),
                            ]:
                                sc(ws.cell(row=row, column=ci),
                                   val=val, bg=bg if val else None)
                    else:
                        vh = club_data.get("H", 0) if club_data else 0
                        vd = club_data.get("D", 0) if club_data else 0
                        sc(ws.cell(row=row, column=cp["col_h"]),
                           val=vh or None,
                           bg=COLORS["light_blue"] if vh else None)
                        sc(ws.cell(row=row, column=cp["col_d"]),
                           val=vd or None,
                           bg=COLORS["light_pink"] if vd else None)

            # Formules totaux
            h_cols = [gcl(cp["col_eq_h"] if is_equipe else cp["col_h"])
                      for dp in plan_dates for cp in dp["cats"]]
            d_cols = [gcl(cp["col_eq_d"] if is_equipe else cp["col_d"])
                      for dp in plan_dates for cp in dp["cats"]]
            f_h = "+".join(f"{c}{row}" for c in h_cols) if h_cols else "0"
            f_d = "+".join(f"{c}{row}" for c in d_cols) if d_cols else "0"

            for ci, val in [
                (col_tot_h,   f"={f_h}"),
                (col_tot_d,   f"={f_d}"),
                (col_tot_clb, f"={gcl(col_tot_h)}{row}+{gcl(col_tot_d)}{row}"),
            ]:
                c = ws.cell(row=row, column=ci, value=val)
                c.font = mk_font(bold=True); c.alignment = mk_align()
                c.border = border_all(); c.fill = mk_fill(COLORS["green_bg"])

            # Besoin / Fournis / Statut par jour
            for dp in plan_dates:
                # Besoin (formule)
                if is_equipe:
                    h_j = "+".join(f"{gcl(cp['col_eq_h'])}{row}" for cp in dp["cats"])
                    d_j = "+".join(f"{gcl(cp['col_eq_d'])}{row}" for cp in dp["cats"])
                    t_j = f"{h_j}+{d_j}"
                    f_besoin = f"=IF({t_j}=0,0,IF({t_j}<=2,1,IF({t_j}<=4,2,3)))"
                else:
                    h_j = "+".join(f"{gcl(cp['col_h'])}{row}" for cp in dp["cats"])
                    d_j = "+".join(f"{gcl(cp['col_d'])}{row}" for cp in dp["cats"])
                    t_j = f"{h_j}+{d_j}"
                    f_besoin = f"=IF({t_j}<4,0,IF({t_j}<=8,1,2))"
                c = ws.cell(row=row, column=dp["col_arb"], value=f_besoin)
                c.font = mk_font(bold=True, size=10, color="7F6000")
                c.fill = mk_fill(BG_ARB); c.alignment = mk_align(); c.border = border_all()

                # Fournis (valeur réelle du club ce jour)
                nb_fournis = arb_par_club_date[dp["date"]].get(club, 0)
                c2 = ws.cell(row=row, column=dp["col_fournis"],
                             value=nb_fournis if nb_fournis else None)
                c2.font = mk_font(bold=True, size=10, color="1F4E79")
                c2.fill = mk_fill("D9E8F5"); c2.alignment = mk_align(); c2.border = border_all()

                # Statut : formule =IF(fournis>=besoin, "OK", "Manque "&(besoin-fournis))
                ca = gcl(dp["col_arb"]); cf = gcl(dp["col_fournis"])
                f_statut = (f'=IF({ca}{row}=0,"—",'
                            f'IF({cf}{row}>={ca}{row},"✓ OK",'
                            f'"▲ Manque "&({ca}{row}-{cf}{row})))')
                c3 = ws.cell(row=row, column=dp["col_statut"], value=f_statut)
                c3.font = mk_font(bold=True, size=10)
                c3.alignment = mk_align(); c3.border = border_all()
                # Mise en forme conditionnelle sur la cellule statut
                from openpyxl.formatting.rule import CellIsRule, FormulaRule
                from openpyxl.styles import PatternFill as _PF
                ref = f"{gcl(dp['col_statut'])}{row}"
                ws.conditional_formatting.add(ref, FormulaRule(
                    formula=[f'{ca}{row}=0'], fill=_PF("solid", fgColor="F2F2F2")))
                ws.conditional_formatting.add(ref, FormulaRule(
                    formula=[f'{cf}{row}>={ca}{row}',], fill=_PF("solid", fgColor="C6EFCE")))
                ws.conditional_formatting.add(ref, FormulaRule(
                    formula=[f'AND({ca}{row}>0,{cf}{row}<{ca}{row})'],
                    fill=_PF("solid", fgColor="FFC7CE")))

        # ── TOTAL GÉNÉRAL
        tot_row = DATA_START + len(tous_clubs)
        # Helper : retourne la formule SUM ou 0 si pas de données
        def _sum_or_zero(cl):
            if tot_row > DATA_START:
                return f"=SUM({cl}{DATA_START}:{cl}{tot_row-1})"
            return 0

        ws.merge_cells(start_row=tot_row, start_column=1,
                       end_row=tot_row, end_column=2)
        sc(ws.cell(row=tot_row, column=1, value="TOTAL GÉNÉRAL"),
           bold=True, bg=COLORS["navy"], fg=COLORS["white"], h="left")

        for dp in plan_dates:
            for cp in dp["cats"]:
                cols = ([cp["col_eq_h"], cp["col_ti_h"], cp["col_eq_d"], cp["col_ti_d"]]
                        if is_equipe else [cp["col_h"], cp["col_d"]])
                for ci in cols:
                    cl = gcl(ci)
                    c = ws.cell(row=tot_row, column=ci, value=_sum_or_zero(cl))
                    c.font = mk_font(bold=True, color=COLORS["white"])
                    c.fill = mk_fill(COLORS["navy"])
                    c.alignment = mk_align(); c.border = border_all()

            # Totaux arbitrage
            for ci, bg, fg in [
                (dp["col_arb"],     "FFF2CC", "7F6000"),
                (dp["col_fournis"], "D9E8F5", "1F4E79"),
            ]:
                cl = gcl(ci)
                c = ws.cell(row=tot_row, column=ci, value=_sum_or_zero(cl))
                c.font = mk_font(bold=True, color=fg)
                c.fill = mk_fill(bg); c.alignment = mk_align(); c.border = border_all()
            # Cellule statut global vide en total
            c = ws.cell(row=tot_row, column=dp["col_statut"], value="")
            c.fill = mk_fill("D9D9D9"); c.border = border_all()

        for ci in [col_tot_h, col_tot_d, col_tot_clb]:
            cl = gcl(ci)
            c = ws.cell(row=tot_row, column=ci, value=_sum_or_zero(cl))
            c.font = mk_font(bold=True, color=COLORS["white"])
            c.fill = mk_fill(COLORS["accent"])
            c.alignment = mk_align(); c.border = border_all()

        # ── Largeurs
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 2
        for dp in plan_dates:
            for cp in dp["cats"]:
                cols = ([cp["col_eq_h"], cp["col_ti_h"], cp["col_eq_d"], cp["col_ti_d"]]
                        if is_equipe else [cp["col_h"], cp["col_d"]])
                for ci in cols:
                    ws.column_dimensions[gcl(ci)].width = 7
            ws.column_dimensions[gcl(dp["col_arb"])].width = 9
            ws.column_dimensions[gcl(dp["col_fournis"])].width = 9
            ws.column_dimensions[gcl(dp["col_statut"])].width = 13
        for ci in [col_tot_h, col_tot_d, col_tot_clb]:
            ws.column_dimensions[gcl(ci)].width = 11

        ws.freeze_panes = ws.cell(row=DATA_START, column=3)

    # ── Feuille Financière

    def feuille_financiere(self, ws, groupes_indiv, groupes_equipe,
                           arbitres_all, dates_ordonnees, titre,
                           plages_arbitres=None):
        """Bilan financier : recettes par jour (indiv + équipes) + dépenses arbitres + solde."""

        C = COLORS
        col_lbl   = 1
        col_nb_h  = 2
        col_nb_d  = 3
        col_pu    = 4
        col_rec_h = 5
        col_rec_d = 6
        col_total = 7

        def ecrire_entetes(r):
            for ci, lbl in [
                (col_lbl,   "Désignation"),
                (col_nb_h,  "Nb H"),
                (col_nb_d,  "Nb D"),
                (col_pu,    "P.U. (€)"),
                (col_rec_h, "Recette H (€)"),
                (col_rec_d, "Recette D (€)"),
                (col_total, "Total (€)"),
            ]:
                c = ws.cell(row=r, column=ci, value=lbl)
                c.font = mk_font(bold=True, size=10, color=C["white"])
                c.fill = mk_fill(C["navy"]); c.alignment = mk_align(h="center")
                c.border = border_all()

        def style_row(r, bg=None):
            for ci in range(1, 8):
                c = ws.cell(row=r, column=ci)
                if bg:
                    c.fill = mk_fill(bg)
                c.border = border_all()
                c.alignment = mk_align(h="left" if ci == col_lbl else "center")
                c.font = mk_font(size=10)

        # Données par date
        data_par_date = {}
        for date_str in dates_ordonnees:
            d = {}
            for cat, clubs in groupes_indiv.get(date_str, {}).items():
                nb_h = sum(v.get("H", 0) for v in clubs.values())
                nb_d = sum(v.get("D", 0) for v in clubs.values())
                d.setdefault(cat, {"nb_indiv_H": 0, "nb_indiv_D": 0,
                                   "nb_eq_H": 0, "nb_eq_D": 0})
                d[cat]["nb_indiv_H"] += nb_h
                d[cat]["nb_indiv_D"] += nb_d
            for cat, clubs in groupes_equipe.get(date_str, {}).items():
                nb_eq_h = sum(len(v.get("equipes", set()))
                              for v in clubs.values() if v.get("tireurs_H", 0) > 0)
                nb_eq_d = sum(len(v.get("equipes", set()))
                              for v in clubs.values() if v.get("tireurs_D", 0) > 0)
                d.setdefault(cat, {"nb_indiv_H": 0, "nb_indiv_D": 0,
                                   "nb_eq_H": 0, "nb_eq_D": 0})
                d[cat]["nb_eq_H"] += nb_eq_h
                d[cat]["nb_eq_D"] += nb_eq_d
            data_par_date[date_str] = d

        # Arbitres par date
        arb_par_date = defaultdict(lambda: {"nb": 0, "cout": 0})
        seen_arb = defaultdict(set)
        bareme_dict = dict(BAREME_ARBITRES)
        for a in arbitres_all:
            date_src = a.get("date_source", "")
            lic      = a.get("licence", "").strip()
            if lic in seen_arb[date_src]:
                continue
            seen_arb[date_src].add(lic)
            cout = bareme_dict.get(a.get("categorie", "").strip(), 0)
            arb_par_date[date_src]["nb"]   += 1
            arb_par_date[date_src]["cout"] += cout

        row = 1

        # Titre
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f"Bilan Financier — {titre}")
        c.font = mk_font(bold=True, size=13, color=C["white"])
        c.fill = mk_fill(C["navy"]); c.alignment = mk_align(h="center")
        ws.row_dimensions[row].height = 26
        row += 2

        recette_rows = []
        depense_rows = []

        for date_str in dates_ordonnees:
            label_date = date_avec_jour(date_str).capitalize()
            d   = data_par_date.get(date_str, {})
            arb = arb_par_date.get(date_str, {"nb": 0, "cout": 0})

            # Bandeau date
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=f"  {label_date}")
            c.font = mk_font(bold=True, size=11, color=C["white"])
            c.fill = mk_fill(C["blue"]); c.alignment = mk_align(h="left")
            ws.row_dimensions[row].height = 20
            row += 1

            ecrire_entetes(row); row += 1

            # ── Recettes Indiv
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value="Droits d'engagement — Individuels")
            c.font = mk_font(bold=True, size=10)
            c.fill = mk_fill(C["fin_indiv"]); c.alignment = mk_align(h="left")
            row += 1

            indiv_rows = []
            for cat in self.CATS_INDIV:
                vals = d.get(cat)
                if not vals:
                    continue
                nb_h = vals.get("nb_indiv_H", 0)
                nb_d = vals.get("nb_indiv_D", 0)
                if nb_h == 0 and nb_d == 0:
                    continue
                pu  = self.TARIF_INDIV.get(cat, 15)
                lbl = self.CAT_LABEL_INDIV.get(cat, cat)
                r   = row
                ws.cell(row=r, column=col_lbl,   value=f"  {lbl}")
                ws.cell(row=r, column=col_nb_h,  value=nb_h or None)
                ws.cell(row=r, column=col_nb_d,  value=nb_d or None)
                ws.cell(row=r, column=col_pu,    value=pu)
                ws.cell(row=r, column=col_rec_h, value=f"={gcl(col_nb_h)}{r}*{gcl(col_pu)}{r}")
                ws.cell(row=r, column=col_rec_d, value=f"={gcl(col_nb_d)}{r}*{gcl(col_pu)}{r}")
                ws.cell(row=r, column=col_total, value=f"={gcl(col_rec_h)}{r}+{gcl(col_rec_d)}{r}")
                style_row(r)
                indiv_rows.append(r); row += 1

            # Sous-total Individuels — réservé, écrit après selon présence équipes
            sous_total_indiv_row = None
            if indiv_rows:
                sous_total_indiv_row = row
                row += 1

            # ── Recettes Équipe — calculer d'abord combien il y en a
            equipe_rows = []
            equipe_data = []  # stocker pour écriture après
            for cat in self.CATS_EQUIPE:
                vals = d.get(cat)
                if not vals: continue
                nb_h = vals.get("nb_eq_H", 0)
                nb_d = vals.get("nb_eq_D", 0)
                if nb_h == 0 and nb_d == 0: continue
                equipe_data.append((cat, nb_h, nb_d))

            # Écrire sous-total Individuels seulement s'il y a des équipes
            if sous_total_indiv_row is not None:
                if equipe_data:
                    ws.cell(row=sous_total_indiv_row, column=col_lbl,
                            value="Sous-total Individuels").font = mk_font(bold=True, size=10)
                    refs_i = "+".join(f"{gcl(col_total)}{r}" for r in indiv_rows)
                    ws.cell(row=sous_total_indiv_row, column=col_total,
                            value=f"={refs_i}").font = mk_font(bold=True, size=10)
                    style_row(sous_total_indiv_row, bg=C["green_bg"])
                else:
                    # Pas d'équipes → annuler la ligne réservée
                    row -= 1
                    sous_total_indiv_row = None

            # Bandeau + lignes équipes (seulement si équipes existent)
            if equipe_data:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                c = ws.cell(row=row, column=1, value="Droits d'engagement — Équipes")
                c.font = mk_font(bold=True, size=10)
                c.fill = mk_fill(C["fin_equipe"]); c.alignment = mk_align(h="left")
                row += 1

                for cat, nb_h, nb_d in equipe_data:
                    pu  = self.TARIF_EQUIPE.get(cat, 40)
                    lbl = self.CAT_LABEL_EQUIPE.get(cat, cat)
                    r   = row
                    ws.cell(row=r, column=col_lbl,   value=f"  {lbl}")
                    ws.cell(row=r, column=col_nb_h,  value=nb_h or None)
                    ws.cell(row=r, column=col_nb_d,  value=nb_d or None)
                    ws.cell(row=r, column=col_pu,    value=pu)
                    ws.cell(row=r, column=col_rec_h, value=f"={gcl(col_nb_h)}{r}*{gcl(col_pu)}{r}")
                    ws.cell(row=r, column=col_rec_d, value=f"={gcl(col_nb_d)}{r}*{gcl(col_pu)}{r}")
                    ws.cell(row=r, column=col_total, value=f"={gcl(col_rec_h)}{r}+{gcl(col_rec_d)}{r}")
                    style_row(r)
                    equipe_rows.append(r); row += 1

                ws.cell(row=row, column=col_lbl,
                        value="Sous-total Équipes").font = mk_font(bold=True, size=10)
                refs_e = "+".join(f"{gcl(col_total)}{r}" for r in equipe_rows)
                ws.cell(row=row, column=col_total,
                        value=f"={refs_e}").font = mk_font(bold=True, size=10)
                style_row(row, bg=C["green_bg"]); row += 1

            # Total recettes — somme uniquement les lignes de données (pas les sous-totaux)
            all_rec = indiv_rows + equipe_rows
            rec_tot = row
            ws.cell(row=rec_tot, column=col_lbl,
                    value="TOTAL RECETTES").font = mk_font(bold=True, size=11, color=C["white"])
            if all_rec:
                # SUMPRODUCT sur les lignes individuelles pour éviter d'inclure les sous-totaux
                refs = "+".join(f"{gcl(col_total)}{r}" for r in all_rec)
                ws.cell(row=rec_tot, column=col_total,
                        value=f"={refs}").font = mk_font(bold=True, size=11, color=C["white"])
            for ci in range(1, 8):
                c = ws.cell(row=rec_tot, column=ci)
                c.fill = mk_fill(C["accent"]); c.border = border_all()
                c.alignment = mk_align(h="left" if ci == col_lbl else "center")
                if ci not in [col_lbl, col_total]:
                    c.font = mk_font(color=C["white"])
            recette_rows.append(rec_tot); row += 1

            # ── Dépenses arbitres
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value="Dépenses — Arbitres")
            c.font = mk_font(bold=True, size=10)
            c.fill = mk_fill(C["fin_depense"]); c.alignment = mk_align(h="left")
            row += 1

            dep_row = row
            ws.cell(row=dep_row, column=col_lbl,   value="  Indemnités arbitres")
            if plages_arbitres and date_str in plages_arbitres:
                r1, r2 = plages_arbitres[date_str]
                ws.cell(row=dep_row, column=col_nb_h, value=f'=COUNTIF(Arbitres!G{r1}:G{r2},"Retenu")')
                ws.cell(row=dep_row, column=col_total, value=f'=SUMIFS(Arbitres!F{r1}:F{r2},Arbitres!G{r1}:G{r2},"Retenu")')
                ws.cell(row=dep_row, column=8, value="barème LREGE — mis à jour auto")
            else:
                ws.cell(row=dep_row, column=col_nb_h,  value=arb["nb"] or None)
                ws.cell(row=dep_row, column=col_total,  value=arb["cout"] or None)
                ws.cell(row=dep_row, column=8, value="barème LREGE")
            style_row(dep_row)
            depense_rows.append(dep_row); row += 1

            # ── Solde du jour
            solde_row = row
            ws.merge_cells(start_row=solde_row, start_column=1,
                           end_row=solde_row, end_column=6)
            ws.cell(row=solde_row, column=col_lbl,
                    value=f"SOLDE — {label_date}").font = mk_font(bold=True, size=11)
            ws.cell(row=solde_row, column=col_lbl).alignment = mk_align(h="left")
            c_solde = ws.cell(row=solde_row, column=col_total,
                              value=f"={gcl(col_total)}{rec_tot}-{gcl(col_total)}{dep_row}")
            c_solde.font = mk_font(bold=True, size=12)
            c_solde.alignment = mk_align(h="center")
            c_solde.border = border_all()
            for ci in range(1, 8):
                ws.cell(row=solde_row, column=ci).border = border_all()

            ref = f"{gcl(col_total)}{solde_row}"
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_pos"])))
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_neg"])))

            row += 2

        # ── Solde global (multi-jours)
        if len(dates_ordonnees) > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=col_lbl,
                    value="SOLDE GLOBAL").font = mk_font(bold=True, size=13, color=C["white"])
            ws.cell(row=row, column=col_lbl).fill = mk_fill(C["navy"])
            ws.cell(row=row, column=col_lbl).alignment = mk_align(h="left")

            rec_g = "+".join(f"{gcl(col_total)}{r}" for r in recette_rows)
            dep_g = "+".join(f"{gcl(col_total)}{r}" for r in depense_rows)
            c_g   = ws.cell(row=row, column=col_total,
                            value=f"=({rec_g})-({dep_g})")
            c_g.font = mk_font(bold=True, size=13, color=C["white"])
            c_g.fill = mk_fill(C["navy"]); c_g.alignment = mk_align(h="center")

            ref_g = f"{gcl(col_total)}{row}"
            ws.conditional_formatting.add(ref_g, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_pos"])))
            ws.conditional_formatting.add(ref_g, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_neg"])))

            for ci in range(1, 8):
                c = ws.cell(row=row, column=ci)
                c.border = border_all()
                if ci not in [col_lbl, col_total]:
                    c.fill = mk_fill(C["navy"])
            ws.row_dimensions[row].height = 24

        # Largeurs
        for col, w in [(1,34),(2,8),(3,8),(4,10),(5,14),(6,14),(7,14),(8,18)]:
            ws.column_dimensions[gcl(col)].width = w

        # Format monétaire
        for r in range(1, row + 2):
            for ci in [col_rec_h, col_rec_d, col_total]:
                ws.cell(row=r, column=ci).number_format = '#,##0 "€";-#,##0 "€"'
