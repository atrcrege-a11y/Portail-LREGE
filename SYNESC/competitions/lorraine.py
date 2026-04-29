"""
competitions/lorraine.py
Coupe de Lorraine — Individuels uniquement, M9 à M15, 3 armes séparées.
"""

import io
from openpyxl import Workbook

from core.parser import construire_donnees
from core.excel_commun import feuille_extranet, feuille_arbitres, feuille_recap_arbitres
from core.excel_base import sc, mk_font, mk_fill, mk_align, border_all, gcl
from core.config import BAREME_ARBITRES, COLORS
from core.parser import date_avec_jour
from collections import defaultdict
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as _PF


from .base import CompetitionBase


ARME_LABEL = {"F": "Fleuret", "E": "Épée", "S": "Sabre"}
ARME_ORDER = ["F", "E", "S"]   # ordre d'affichage des feuilles
ARME_COLORS = {
    "F": "2E6099",   # bleu  — Fleuret
    "E": "1F6B3A",   # vert  — Épée
    "S": "8B1A1A",   # rouge — Sabre
}


class Lorraine(CompetitionBase):
    """
    Coupe de Lorraine.
    - Individuel uniquement (pas d'équipe)
    - Catégories : M9, M11, M13, M15
    - 3 armes : Fleuret, Épée, Sabre → une feuille par arme
    - Bilan Financier global + par arme
    """

    CAT_MAP_INDIV = {
        "M9": "M9", "M11": "M11",
        "M13": "M13", "M15": "M15",
    }
    CAT_MAP_EQUIPE = {}   # pas d'équipe

    CATS_INDIV           = ["M9", "M11", "M13", "M15"]
    CATS_EQUIPE          = []
    CATS_EXTRANET_INDIV  = ["M9", "M11", "M13", "M15"]
    CATS_EXTRANET_EQUIPE = []

    CAT_LABEL_INDIV = {
        "M9": "M9", "M11": "M11",
        "M13": "M13", "M15": "M15",
    }
    CAT_LABEL_EQUIPE = {}

    TARIF_INDIV  = {"M9": 10, "M11": 10, "M13": 10, "M15": 10}
    TARIF_EQUIPE = {}

    NOM_FEUILLE_INDIV  = "Indiv Fleuret"   # sera remplacé dynamiquement
    NOM_FEUILLE_EQUIPE = ""                 # non utilisé

    # ── Pipeline principal (surcharge totale)

    def generer_excel(self, fichiers_list, titre_comp=""):
        """
        Génère le fichier Excel Lorraine :
          - Bilan Financier (en premier)
          - Épreuve Individuelle — Fleuret
          - Épreuve Individuelle — Épée
          - Épreuve Individuelle — Sabre
          - Indiv Extranet (toutes armes)
          - Arbitres
        """
        (groupes_indiv, groupes_equipe, arbitres_all,
         titre_xml, ligue_info, plage_dates,
         dates_ordonnees) = construire_donnees(
            fichiers_list,
            self.CAT_MAP_INDIV,
            self.CAT_MAP_EQUIPE,
        )

        titre_comp  = titre_comp or titre_xml or "Coupe de Lorraine"
        titre_excel = f"{titre_comp} — {plage_dates}" if plage_dates else titre_comp

        self._arbitres_all = arbitres_all
        groupes_par_arme = self._grouper_par_arme(fichiers_list)

        wb = Workbook()
        wb.remove(wb.active)

        # Arbitres en premier pour obtenir les vraies plages de lignes
        ws_arb = wb.create_sheet("Arbitres")
        plages = feuille_arbitres(ws_arb, arbitres_all, plage_dates=plage_dates,
                                  fichiers_list=fichiers_list,
                                  titre_comp=titre_excel)

        # Récap Arbitres juste après
        ws_recap = wb.create_sheet("Récap Arbitres")
        from core.config import NOMS_SUPERVISEURS_LORRAINE
        feuille_recap_arbitres(ws_recap, arbitres_all, dates_ordonnees,
                               titre_excel, plages, ws_arb_name="Arbitres",
                               responsables={
                                   "type": "superviseurs",
                                   "noms": NOMS_SUPERVISEURS_LORRAINE,
                                   "armes": ["Fleuret", "Épée", "Sabre"],
                               })

        # 1. Bilan Financier avec formules dynamiques
        ws_fin = wb.create_sheet("Bilan Financier")
        self.feuille_financiere(
            ws_fin, groupes_par_arme, arbitres_all,
            dates_ordonnees, titre_excel,
            plages_arbitres=plages,
        )

        # 2. Une feuille par arme
        for arme_code in ARME_ORDER:
            groupes_arme = groupes_par_arme.get(arme_code, {})
            if not groupes_arme:
                continue
            nom_feuille = f"Indiv {ARME_LABEL[arme_code]}"
            ws = wb.create_sheet(nom_feuille)
            self._feuille_indiv_arme(
                ws, groupes_arme, titre_excel,
                dates_ordonnees, arme_code,
            )

        # 3. Extranet Indiv (toutes armes confondues)
        ws_ext = wb.create_sheet("Indiv Extranet")
        feuille_extranet(
            ws_ext, fichiers_list,
            self.CATS_EXTRANET_INDIV, self.CAT_LABEL_INDIV,
            cat_map=self.CAT_MAP_INDIV,
            is_equipe=False,
        )

        # Déplacer Arbitres et Récap en fin de classeur
        n = len(wb.sheetnames)
        wb.move_sheet("Arbitres",       offset=n)
        wb.move_sheet("Récap Arbitres", offset=n)

        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.iterate = False
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Méthodes abstraites (non utilisées directement ici)

    def feuille_indiv(self, ws, groupes_indiv, titre, dates_ordonnees):
        pass   # remplacé par _feuille_indiv_arme

    def feuille_equipe(self, ws, groupes_equipe, titre, dates_ordonnees):
        pass   # pas d'équipe

    def feuille_financiere(self, ws, groupes_par_arme, arbitres_all,
                           dates_ordonnees, titre, plages_arbitres=None):
        """Bilan financier global + détail par arme et par jour."""
        C = COLORS
        col_lbl   = 1
        col_nb_h  = 2
        col_nb_d  = 3
        col_pu    = 4
        col_rec_h = 5
        col_rec_d = 6
        col_total = 7

        def ecrire_entetes(r, bg=C["navy"]):
            for ci, lbl in [
                (col_lbl,   "Désignation"),
                (col_nb_h,  "Nb H"),
                (col_nb_d,  "Nb D"),
                (col_pu,    "P.U. (€)"),
                (col_rec_h, "Recette H (€)"),
                (col_rec_d, "Recette D (€)"),
                (col_total, "Total (€)"),
            ]:
                c = ws.cell(row=r, column=ci, value=lbl)
                c.font = mk_font(bold=True, size=10, color=C["white"])
                c.fill = mk_fill(bg)
                c.alignment = mk_align(h="center"); c.border = border_all()

        def style_data(r):
            for ci in range(1, 8):
                c = ws.cell(row=r, column=ci)
                c.border = border_all()
                c.alignment = mk_align(h="left" if ci == col_lbl else "center")
                c.font = mk_font(size=10)

        bareme_dict = dict(BAREME_ARBITRES)

        # Arbitres par date
        arb_par_date = defaultdict(lambda: {"nb": 0, "cout": 0})
        seen_arb = defaultdict(set)
        for a in arbitres_all:
            date_src = a.get("date_source", "")
            lic = a.get("licence", "").strip()
            if lic in seen_arb[date_src]:
                continue
            seen_arb[date_src].add(lic)
            cout = bareme_dict.get(a.get("categorie", "").strip(), 0)
            arb_par_date[date_src]["nb"]   += 1
            arb_par_date[date_src]["cout"] += cout

        row = 1

        # Titre
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f"Bilan Financier — {titre}")
        c.font = mk_font(bold=True, size=13, color=C["white"])
        c.fill = mk_fill(C["navy"]); c.alignment = mk_align(h="center")
        ws.row_dimensions[row].height = 26
        row += 2

        recette_rows_global = []
        depense_rows_global = []

        for date_str in dates_ordonnees:
            label_date = date_avec_jour(date_str).capitalize()
            arb = arb_par_date.get(date_str, {"nb": 0, "cout": 0})

            # Bandeau date
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=f"  {label_date}")
            c.font = mk_font(bold=True, size=11, color=C["white"])
            c.fill = mk_fill(C["blue"]); c.alignment = mk_align(h="left")
            ws.row_dimensions[row].height = 20
            row += 1

            ecrire_entetes(row); row += 1

            arme_rows = []   # sous-totaux par arme pour ce jour

            for arme_code in ARME_ORDER:
                groupes_arme = groupes_par_arme.get(arme_code, {})
                groupes_jour = groupes_arme.get(date_str, {})
                if not groupes_jour:
                    continue

                arme_lbl   = ARME_LABEL[arme_code]
                arme_color = ARME_COLORS[arme_code]

                # Bandeau arme
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                c = ws.cell(row=row, column=1,
                            value=f"Droits d'engagement — {arme_lbl}")
                c.font = mk_font(bold=True, size=10, color=C["white"])
                c.fill = mk_fill(arme_color); c.alignment = mk_align(h="left")
                row += 1

                cat_rows = []
                for cat in self.CATS_INDIV:
                    clubs = groupes_jour.get(cat, {})
                    nb_h = sum(v.get("H", 0) for v in clubs.values())
                    nb_d = sum(v.get("D", 0) for v in clubs.values())
                    if nb_h == 0 and nb_d == 0:
                        continue
                    pu  = self.TARIF_INDIV.get(cat, 10)
                    lbl = self.CAT_LABEL_INDIV.get(cat, cat)
                    r   = row
                    ws.cell(row=r, column=col_lbl,   value=f"    {lbl}")
                    ws.cell(row=r, column=col_nb_h,  value=nb_h or None)
                    ws.cell(row=r, column=col_nb_d,  value=nb_d or None)
                    ws.cell(row=r, column=col_pu,    value=pu)
                    ws.cell(row=r, column=col_rec_h,
                            value=f"={gcl(col_nb_h)}{r}*{gcl(col_pu)}{r}")
                    ws.cell(row=r, column=col_rec_d,
                            value=f"={gcl(col_nb_d)}{r}*{gcl(col_pu)}{r}")
                    ws.cell(row=r, column=col_total,
                            value=f"={gcl(col_rec_h)}{r}+{gcl(col_rec_d)}{r}")
                    style_data(r)
                    cat_rows.append(r); row += 1

                # Sous-total arme
                if cat_rows:
                    ws.cell(row=row, column=col_lbl,
                            value=f"  Sous-total {arme_lbl}").font = mk_font(bold=True, size=10)
                    refs_cat = "+".join(f"{gcl(col_total)}{r}" for r in cat_rows)
                    ws.cell(row=row, column=col_total,
                            value=f"={refs_cat}").font = mk_font(bold=True, size=10)
                    for ci in range(1, 8):
                        c = ws.cell(row=row, column=ci)
                        c.fill = mk_fill(C["green_bg"]); c.border = border_all()
                        c.alignment = mk_align(h="left" if ci == col_lbl else "center")
                    arme_rows.append(row); row += 1

            # Total recettes du jour
            rec_tot = row
            ws.cell(row=rec_tot, column=col_lbl,
                    value="TOTAL RECETTES").font = mk_font(bold=True, size=11, color=C["white"])
            if arme_rows:
                refs_arme = "+".join(f"{gcl(col_total)}{r}" for r in arme_rows)
                ws.cell(row=rec_tot, column=col_total,
                        value=f"={refs_arme}").font = mk_font(bold=True, size=11, color=C["white"])
            for ci in range(1, 8):
                c = ws.cell(row=rec_tot, column=ci)
                c.fill = mk_fill(C["accent"]); c.border = border_all()
                c.alignment = mk_align(h="left" if ci == col_lbl else "center")
                if ci not in [col_lbl, col_total]:
                    c.font = mk_font(color=C["white"])
            recette_rows_global.append(rec_tot); row += 1

            # Dépenses arbitres
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value="Dépenses — Arbitres")
            c.font = mk_font(bold=True, size=10)
            c.fill = mk_fill(C["fin_depense"]); c.alignment = mk_align(h="left")
            row += 1

            dep_row = row
            ws.cell(row=dep_row, column=col_lbl,  value="  Indemnités arbitres")
            if plages_arbitres and date_str in plages_arbitres:
                r1, r2 = plages_arbitres[date_str]
                ws.cell(row=dep_row, column=col_nb_h, value=f'=COUNTIF(Arbitres!G{r1}:G{r2},"Retenu")')
                ws.cell(row=dep_row, column=col_total, value=f'=SUMIFS(Arbitres!F{r1}:F{r2},Arbitres!G{r1}:G{r2},"Retenu")')
                ws.cell(row=dep_row, column=8, value="barème LREGE — mis à jour auto")
            else:
                ws.cell(row=dep_row, column=col_nb_h, value=arb["nb"] or None)
                ws.cell(row=dep_row, column=col_total, value=arb["cout"] or None)
                ws.cell(row=dep_row, column=8, value="barème LREGE")
            ws.cell(row=dep_row, column=8,         value="barème LREGE")
            style_data(dep_row)
            depense_rows_global.append(dep_row); row += 1

            # Solde du jour
            solde_row = row
            ws.merge_cells(start_row=solde_row, start_column=1,
                           end_row=solde_row, end_column=6)
            ws.cell(row=solde_row, column=col_lbl,
                    value=f"SOLDE — {label_date}").font = mk_font(bold=True, size=11)
            ws.cell(row=solde_row, column=col_lbl).alignment = mk_align(h="left")
            c_s = ws.cell(row=solde_row, column=col_total,
                          value=f"={gcl(col_total)}{rec_tot}-{gcl(col_total)}{dep_row}")
            c_s.font = mk_font(bold=True, size=12); c_s.alignment = mk_align(h="center")
            c_s.border = border_all()
            for ci in range(1, 8):
                ws.cell(row=solde_row, column=ci).border = border_all()
            ref = f"{gcl(col_total)}{solde_row}"
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_pos"])))
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_neg"])))
            row += 2

        # Solde global (multi-jours)
        if len(dates_ordonnees) > 1 and recette_rows_global:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=col_lbl,
                    value="SOLDE GLOBAL").font = mk_font(bold=True, size=13, color=C["white"])
            ws.cell(row=row, column=col_lbl).fill = mk_fill(C["navy"])
            ws.cell(row=row, column=col_lbl).alignment = mk_align(h="left")
            rec_g = "+".join(f"{gcl(col_total)}{r}" for r in recette_rows_global)
            dep_g = "+".join(f"{gcl(col_total)}{r}" for r in depense_rows_global)
            c_g = ws.cell(row=row, column=col_total,
                          value=f"=({rec_g})-({dep_g})")
            c_g.font = mk_font(bold=True, size=13, color=C["white"])
            c_g.fill = mk_fill(C["navy"]); c_g.alignment = mk_align(h="center")
            ref_g = f"{gcl(col_total)}{row}"
            ws.conditional_formatting.add(ref_g, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_pos"])))
            ws.conditional_formatting.add(ref_g, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_neg"])))
            for ci in range(1, 8):
                c = ws.cell(row=row, column=ci)
                c.border = border_all()
                if ci not in [col_lbl, col_total]:
                    c.fill = mk_fill(C["navy"])
            ws.row_dimensions[row].height = 24

        for col, w in [(1,34),(2,8),(3,8),(4,10),(5,14),(6,14),(7,14),(8,18)]:
            ws.column_dimensions[gcl(col)].width = w
        for r in range(1, row + 2):
            for ci in [col_rec_h, col_rec_d, col_total]:
                ws.cell(row=r, column=ci).number_format = '#,##0 "€";-#,##0 "€"'

    # ── Feuille indiv par arme

    def _feuille_indiv_arme(self, ws, groupes_arme, titre, dates_ordonnees, arme_code):
        """Tableau récapitulatif indiv pour une arme donnée."""
        arme_lbl   = ARME_LABEL[arme_code]
        arme_color = ARME_COLORS[arme_code]

        # Catégories présentes (non vides)
        def cat_a_tireurs(date_str, cat):
            clubs = groupes_arme.get(date_str, {}).get(cat, {})
            return any(v.get("H", 0) + v.get("D", 0) > 0 for v in clubs.values())

        cats_par_date = {
            d: [c for c in self.CATS_INDIV if c in groupes_arme.get(d, {})
                and cat_a_tireurs(d, c)]
            for d in dates_ordonnees
        }

        tous_clubs = sorted({
            club
            for d in groupes_arme.values()
            for cat in d.values()
            for club in cat.keys()
        })

        # Plan colonnes
        col_idx = 3
        plan_dates = []
        for date_str in dates_ordonnees:
            cats = cats_par_date.get(date_str, [])
            if not cats:
                continue
            dp = {"date": date_str, "label": date_avec_jour(date_str),
                  "col_start": col_idx, "cats": []}
            for cat in cats:
                dp["cats"].append({"cat": cat, "col_h": col_idx, "col_d": col_idx + 1})
                col_idx += 2
            dp["col_end"] = col_idx - 1
            plan_dates.append(dp)

        if not plan_dates:
            ws.cell(row=1, column=1, value="Aucune donnée").font = mk_font(size=10)
            return

        col_tot_h   = col_idx;     col_idx += 1
        col_tot_d   = col_idx;     col_idx += 1
        col_tot_clb = col_idx;     col_idx += 1
        # Besoin / Fournis / Statut après les totaux — triplet par jour
        for dp in plan_dates:
            dp["col_arb"]     = col_idx;  col_idx += 1
            dp["col_fournis"] = col_idx;  col_idx += 1
            dp["col_statut"]  = col_idx;  col_idx += 1
        last_col = col_idx - 1

        DATA_START = 5

        # Pré-calcul : nb d'arbitres fournis par club et par jour
        arb_par_club_date = defaultdict(lambda: defaultdict(int))
        seen_arb = defaultdict(set)
        for a in (self._arbitres_all or []):
            date_src = a.get("date_source", "")
            club_a   = a.get("club", "").strip()
            lic      = a.get("licence", "").strip()
            if lic and lic not in seen_arb[(date_src, club_a)]:
                seen_arb[(date_src, club_a)].add(lic)
                arb_par_club_date[date_src][club_a] += 1

        # Ligne 1 : titre avec couleur arme
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        c = ws.cell(row=1, column=1, value=f"{titre} — {arme_lbl}")
        c.font = mk_font(bold=True, size=13, color="FFFFFF")
        c.fill = mk_fill(arme_color); c.alignment = mk_align(h="center")
        ws.row_dimensions[1].height = 26

        # Col A:B
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=2)
        sc(ws.cell(row=2, column=1, value="Club"),
           bold=True, bg=COLORS["navy"], fg=COLORS["white"])

        BG_ARB = "FFF2CC"
        for dp in plan_dates:
            ws.merge_cells(start_row=2, start_column=dp["col_start"],
                           end_row=2, end_column=dp["col_end"])
            c = ws.cell(row=2, column=dp["col_start"],
                        value=dp["label"].capitalize())
            c.font = mk_font(bold=True, size=10, color="FFFFFF")
            c.fill = mk_fill("1F3A6E"); c.alignment = mk_align(h="center")
            c.border = border_all()

            for cp in dp["cats"]:
                lbl = self.CAT_LABEL_INDIV.get(cp["cat"], cp["cat"])
                ws.merge_cells(start_row=3, start_column=cp["col_h"],
                               end_row=3, end_column=cp["col_d"])
                sc(ws.cell(row=3, column=cp["col_h"], value=lbl),
                   bold=True, bg=arme_color, fg="FFFFFF")
                sc(ws.cell(row=4, column=cp["col_h"], value="H"),
                   bold=True, bg=COLORS["light_blue"])
                sc(ws.cell(row=4, column=cp["col_d"], value="D"),
                   bold=True, bg=COLORS["light_pink"])

        for ci, lbl in [(col_tot_h, "Total H"), (col_tot_d, "Total D"),
                        (col_tot_clb, "Total Club")]:
            ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
            sc(ws.cell(row=2, column=ci, value=lbl),
               bold=True, bg=COLORS["accent"], fg=COLORS["white"], wrap=True)
            sc(ws.cell(row=4, column=ci, value=""), bg="EBD5A0")

        # En-têtes besoin / fournis / statut — après Total Club, triplet par jour
        BG_ARB = "FFF2CC"
        for dp in plan_dates:
            jour_court = dp["label"].split()[0].capitalize()
            ws.merge_cells(start_row=2, start_column=dp["col_arb"],
                           end_row=3, end_column=dp["col_arb"])
            sc(ws.cell(row=2, column=dp["col_arb"], value="Besoin\narb."),
               bold=True, bg=BG_ARB, size=9, fg="7F6000", wrap=True)
            sc(ws.cell(row=4, column=dp["col_arb"], value=""), bg=BG_ARB, border=True)
            ws.merge_cells(start_row=2, start_column=dp["col_fournis"],
                           end_row=3, end_column=dp["col_fournis"])
            sc(ws.cell(row=2, column=dp["col_fournis"],
                       value=f"Fournis\n{jour_court}"),
               bold=True, bg="D9E8F5", size=9, fg="1F4E79", wrap=True)
            sc(ws.cell(row=4, column=dp["col_fournis"], value=""), bg="D9E8F5", border=True)
            ws.merge_cells(start_row=2, start_column=dp["col_statut"],
                           end_row=3, end_column=dp["col_statut"])
            sc(ws.cell(row=2, column=dp["col_statut"],
                       value=f"Statut\n{jour_court}"),
               bold=True, bg="D9D9D9", size=9, fg="000000", wrap=True)
            sc(ws.cell(row=4, column=dp["col_statut"], value=""), bg="D9D9D9", border=True)

        # Données
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill as _PF

        for i, club in enumerate(tous_clubs):
            row = DATA_START + i
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            sc(ws.cell(row=row, column=1, value=club), h="left")

            for dp in plan_dates:
                for cp in dp["cats"]:
                    club_data = groupes_arme.get(dp["date"], {}).get(cp["cat"], {}).get(club, {})
                    vh = club_data.get("H", 0) if club_data else 0
                    vd = club_data.get("D", 0) if club_data else 0
                    sc(ws.cell(row=row, column=cp["col_h"]),
                       val=vh or None, bg=COLORS["light_blue"] if vh else None)
                    sc(ws.cell(row=row, column=cp["col_d"]),
                       val=vd or None, bg=COLORS["light_pink"] if vd else None)

                # Besoin arbitres
                h_j = "+".join(f"{gcl(cp['col_h'])}{row}" for cp in dp["cats"])
                d_j = "+".join(f"{gcl(cp['col_d'])}{row}" for cp in dp["cats"])
                t_j = f"{h_j}+{d_j}"
                f_besoin = f"=IF({t_j}<4,0,IF({t_j}<=8,1,2))"
                c = ws.cell(row=row, column=dp["col_arb"], value=f_besoin)
                c.font = mk_font(bold=True, size=10, color="7F6000")
                c.fill = mk_fill(BG_ARB); c.alignment = mk_align(); c.border = border_all()

                # Fournis
                nb_fournis = arb_par_club_date[dp["date"]].get(club, 0)
                c2 = ws.cell(row=row, column=dp["col_fournis"],
                             value=nb_fournis if nb_fournis else None)
                c2.font = mk_font(bold=True, size=10, color="1F4E79")
                c2.fill = mk_fill("D9E8F5"); c2.alignment = mk_align(); c2.border = border_all()

                # Statut
                ca = gcl(dp["col_arb"]); cf = gcl(dp["col_fournis"])
                f_statut = (f'=IF({ca}{row}=0,"—",'
                            f'IF({cf}{row}>={ca}{row},"✓ OK",'
                            f'"▲ Manque "&({ca}{row}-{cf}{row})))')
                c3 = ws.cell(row=row, column=dp["col_statut"], value=f_statut)
                c3.font = mk_font(bold=True, size=10)
                c3.alignment = mk_align(); c3.border = border_all()
                ref = f"{gcl(dp['col_statut'])}{row}"
                ws.conditional_formatting.add(ref, FormulaRule(
                    formula=[f'{ca}{row}=0'], fill=_PF("solid", fgColor="F2F2F2")))
                ws.conditional_formatting.add(ref, FormulaRule(
                    formula=[f'{cf}{row}>={ca}{row}'], fill=_PF("solid", fgColor="C6EFCE")))
                ws.conditional_formatting.add(ref, FormulaRule(
                    formula=[f'AND({ca}{row}>0,{cf}{row}<{ca}{row})'],
                    fill=_PF("solid", fgColor="FFC7CE")))

            h_cols = [gcl(cp["col_h"]) for dp in plan_dates for cp in dp["cats"]]
            d_cols = [gcl(cp["col_d"]) for dp in plan_dates for cp in dp["cats"]]
            f_h = "+".join(f"{c}{row}" for c in h_cols) if h_cols else "0"
            f_d = "+".join(f"{c}{row}" for c in d_cols) if d_cols else "0"
            for ci, val in [
                (col_tot_h,   f"={f_h}"),
                (col_tot_d,   f"={f_d}"),
                (col_tot_clb, f"={gcl(col_tot_h)}{row}+{gcl(col_tot_d)}{row}"),
            ]:
                c = ws.cell(row=row, column=ci, value=val)
                c.font = mk_font(bold=True); c.alignment = mk_align()
                c.border = border_all(); c.fill = mk_fill(COLORS["green_bg"])

        # Total général
        tot_row = DATA_START + len(tous_clubs)
        # Helper : retourne la formule SUM ou 0 si pas de données
        def _sum_or_zero(cl):
            if tot_row > DATA_START:
                return f"=SUM({cl}{DATA_START}:{cl}{tot_row-1})"
            return 0

        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=2)
        sc(ws.cell(row=tot_row, column=1, value="TOTAL GÉNÉRAL"),
           bold=True, bg=COLORS["navy"], fg=COLORS["white"], h="left")
        for dp in plan_dates:
            for cp in dp["cats"]:
                for ci in [cp["col_h"], cp["col_d"]]:
                    cl = gcl(ci)
                    c = ws.cell(row=tot_row, column=ci, value=_sum_or_zero(cl))
                    c.font = mk_font(bold=True, color="FFFFFF")
                    c.fill = mk_fill(COLORS["navy"])
                    c.alignment = mk_align(); c.border = border_all()
            for ci, bg, fg in [
                (dp["col_arb"],     "FFF2CC", "7F6000"),
                (dp["col_fournis"], "D9E8F5", "1F4E79"),
            ]:
                cl = gcl(ci)
                c = ws.cell(row=tot_row, column=ci, value=_sum_or_zero(cl))
                c.font = mk_font(bold=True, color=fg)
                c.fill = mk_fill(bg); c.alignment = mk_align(); c.border = border_all()
            c = ws.cell(row=tot_row, column=dp["col_statut"], value="")
            c.fill = mk_fill("D9D9D9"); c.border = border_all()
        for ci in [col_tot_h, col_tot_d, col_tot_clb]:
            cl = gcl(ci)
            c = ws.cell(row=tot_row, column=ci, value=_sum_or_zero(cl))
            c.font = mk_font(bold=True, color="FFFFFF")
            c.fill = mk_fill(COLORS["accent"])
            c.alignment = mk_align(); c.border = border_all()

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 2
        for dp in plan_dates:
            for cp in dp["cats"]:
                for ci in [cp["col_h"], cp["col_d"]]:
                    ws.column_dimensions[gcl(ci)].width = 7
            ws.column_dimensions[gcl(dp["col_arb"])].width = 9
            ws.column_dimensions[gcl(dp["col_fournis"])].width = 9
            ws.column_dimensions[gcl(dp["col_statut"])].width = 13
        for ci in [col_tot_h, col_tot_d, col_tot_clb]:
            ws.column_dimensions[gcl(ci)].width = 11
        ws.freeze_panes = ws.cell(row=DATA_START, column=3)

    # ── Helper : grouper les tireurs par arme

    def _grouper_par_arme(self, fichiers_list):
        """
        Retourne {arme_code: {date: {cat: {club: {"H": n, "D": n}}}}}
        Uniquement les fichiers indiv (Type=I).
        """
        result = defaultdict(lambda: defaultdict(
            lambda: defaultdict(lambda: defaultdict(lambda: {"H": 0, "D": 0}))
        ))
        for meta, tireurs, _ in fichiers_list:
            if meta["type"] == "E":
                continue
            arme     = meta["arme"]
            cat_raw  = meta["categorie"].upper()
            cat_key  = self.CAT_MAP_INDIV.get(cat_raw)
            date_src = meta.get("date_debut") or meta.get("date", "")
            if not cat_key:
                continue
            for t in tireurs:
                club = t["club"].strip()
                sexe = "H" if t["sexe"].upper() == "M" else "D"
                result[arme][date_src][cat_key][club][sexe] += 1
        return result
