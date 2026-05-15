"""
sabre_laser/parseurs.py — Import des classements Sabre Laser.

Formats supportés : PDF (FFE liste), XLSX/XLS, HTML (extranet), FFF (Engarde/WIN).
Tous retournent un DataFrame avec les colonnes standard défines dans config.py.
"""
import re
import pandas as pd

from .config import (
    COL_RANG, COL_NOM, COL_PRENOM, COL_ADHERENT,
    COL_REGION, COL_CLUB, COL_GRAND_EST, COL_NOTE,
    COL_PARTICIPANTS, COL_SOUS_CAT, detecter_format,
)

_GE_KEYWORDS = {"grand est", "grand-est", "ges"}


def _est_grand_est(region: str) -> bool:
    if not region or not isinstance(region, str):
        return False
    return any(k in region.lower() for k in _GE_KEYWORDS)


def _normaliser(df: pd.DataFrame) -> pd.DataFrame:
    """Normalise et complète les colonnes du DataFrame."""
    # Dédupliquer les colonnes au cas où (garde la première occurrence)
    df = df.loc[:, ~df.columns.duplicated()]

    for col in [COL_RANG, COL_NOM, COL_PRENOM, COL_ADHERENT, COL_REGION, COL_CLUB]:
        if col not in df.columns:
            df[col] = ""

    df[COL_NOM]    = df[COL_NOM].astype(str).str.upper().str.strip()
    df[COL_PRENOM] = df[COL_PRENOM].astype(str).str.strip()
    df[COL_REGION] = df[COL_REGION].astype(str).str.strip()
    df[COL_CLUB]   = df[COL_CLUB].astype(str).str.strip()
    df[COL_GRAND_EST] = df[COL_REGION].apply(_est_grand_est)

    # Rang numérique (COL_RANG est forcément une Series après dédupplication)
    rang_col = df[COL_RANG]
    if isinstance(rang_col, pd.DataFrame):
        rang_col = rang_col.iloc[:, 0]  # prendre la première colonne si encore doublon
    df[COL_RANG] = pd.to_numeric(rang_col, errors="coerce").fillna(0).astype(int)
    # Supprimer les lignes avec rang=0 (valeurs non numériques = métadonnées parasites)
    df = df[df[COL_RANG] > 0]
    return df


# ── PDF ───────────────────────────────────────────────────────────────

def lire_pdf(path: str) -> pd.DataFrame:
    """
    Parse un PDF FFE liste des qualifiés (Sabre Laser ou escrime olympique).
    Détecte dynamiquement les positions x des colonnes depuis la ligne d'en-tête.
    Compatible avec tous les formats FFE : avec ou sans colonne Adhérent/Année.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber requis : pip install pdfplumber")

    rows = []
    rang = 0

    def words_in(ws, x_min, x_max):
        return " ".join(w["text"] for w in ws if w["x0"] >= x_min and w["x0"] < x_max)

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            lines = {}
            for w in words:
                y = round(w["top"] / 4) * 4
                lines.setdefault(y, []).append(w)

            in_list = False
            # Positions x détectées dynamiquement depuis l'en-tête
            x_nom = x_prenom = x_mid = x_region = x_club = None

            for y in sorted(lines):
                ws = sorted(lines[y], key=lambda w: w["x0"])
                row_text = " ".join(w["text"] for w in ws)

                # Détecter et calibrer depuis la ligne d'en-tête
                if not in_list and "Nom" in row_text and ("Ligue" in row_text or "Club" in row_text or "Prénom" in row_text):
                    # Trouver la position x de chaque colonne
                    for w in ws:
                        t = w["text"].strip()
                        if t == "Nom":           x_nom    = w["x0"]
                        elif t == "Prénom":      x_prenom = w["x0"]
                        elif t in ("Adhérent", "Licencié", "N° licence"): x_mid = w["x0"]
                        elif "Ligue" in t or "Région" in t: x_region = w["x0"]
                        elif t == "Club":        x_club   = w["x0"]
                    in_list = True
                    continue

                if in_list and ("Version du" in row_text or "[Règlement" in row_text
                                or row_text.strip().startswith("*")):
                    in_list = False
                    x_nom = x_prenom = x_mid = x_region = x_club = None
                    continue
                if not in_list:
                    continue

                # Extraction par positions dynamiques avec fallback sur positions fixes
                _x_nom    = x_nom    if x_nom    is not None else 19
                _x_prenom = x_prenom if x_prenom is not None else _x_nom + 94
                _x_mid    = x_mid    if x_mid    is not None else _x_prenom + 100
                _x_region = x_region if x_region is not None else _x_mid + 80
                _x_club   = x_club   if x_club   is not None else _x_region + 130

                nom    = words_in(ws, _x_nom,    _x_prenom - 1).strip()
                prenom = words_in(ws, _x_prenom, _x_mid - 1).strip()
                mid    = words_in(ws, _x_mid,    _x_region - 1).strip()
                region = words_in(ws, _x_region, _x_club - 1).strip()
                club   = words_in(ws, _x_club,   9999).strip()

                # mid peut contenir : adhérent seul, année+adhérent, ou début de région
                adh = ""
                m_adh = re.search(r"(\d{5,6})", mid)
                if m_adh:
                    adh = m_adh.group(1)
                    # Extraire aussi l'année du prénom si elle s'y est glissée
                    prenom = re.sub(r"\s*\d{4}\s*$", "", prenom).strip()
                else:
                    # Pas d'adhérent détecté → mid fait partie de la région
                    region = (mid + " " + region).strip()

                if not nom or len(nom) < 2:
                    continue
                # Ignorer les lignes de métadonnées (nom = chiffre ou symbole)
                if not re.search(r"[A-Za-z\xc0-\xff]", nom):
                    continue

                rang += 1
                rows.append({
                    COL_RANG: rang, COL_NOM: nom, COL_PRENOM: prenom,
                    COL_ADHERENT: adh, COL_REGION: region, COL_CLUB: club,
                })

    return _normaliser(pd.DataFrame(rows)) if rows else pd.DataFrame()


# ── XLSX / XLS ────────────────────────────────────────────────────────

def lire_xlsx(path: str) -> pd.DataFrame:
    """
    Parse un fichier Excel FFE, extranet ou classement régional.
    Gère les fichiers avec en-tête décalée et colonnes Rang multiples.
    """
    try:
        df_scan = pd.read_excel(path, header=None, dtype=str)
    except Exception as e:
        raise ValueError(f"Impossible de lire le fichier Excel : {e}")

    # Trouver la ligne d'en-tête (contient "Nom" et "Club" ou "Prénom")
    header_row = None
    for i, row in df_scan.iterrows():
        # Nettoyage : supprimer les espaces insécables et normaliser
        vals = [str(v).strip().replace("\xa0", "").replace("\u00a0", "").lower()
                for v in row if pd.notna(v) and str(v).strip() not in ("", "nan")]
        # "nom" EXACT (pas "nombre", "nomade"...) : cellule = exactement "nom"
        has_nom    = any(v in ("nom",) for v in vals)
        has_prenom = any("prénom" in v or "prenom" in v for v in vals)
        has_club   = any("club" in v for v in vals)
        if has_nom and (has_prenom or has_club):
            header_row = i
            break

    if header_row is None:
        df_raw = pd.read_excel(path, dtype=str)
    else:
        df_raw = pd.read_excel(path, header=header_row, dtype=str)

    df_raw.columns = [str(c).strip().replace("\xa0", "").replace("\u00a0", "") for c in df_raw.columns]
    df_raw = df_raw.dropna(how="all")

    # Mapping souple — une seule colonne cible par rôle
    # Priorités : "Place" > "Rang" pour le classement général
    col_map = {}
    mapped = set()
    for col in df_raw.columns:
        cl = col.lower().strip()
        if "nom" in cl and "prénom" not in cl and "prenom" not in cl and COL_NOM not in mapped:
            col_map[col] = COL_NOM;      mapped.add(COL_NOM)
        elif ("prénom" in cl or "prenom" in cl) and COL_PRENOM not in mapped:
            col_map[col] = COL_PRENOM;   mapped.add(COL_PRENOM)
        elif ("adhér" in cl or "licen" in cl) and COL_ADHERENT not in mapped:
            col_map[col] = COL_ADHERENT; mapped.add(COL_ADHERENT)
        elif ("ligue" in cl or "région" in cl or "region" in cl) and COL_REGION not in mapped:
            col_map[col] = COL_REGION;   mapped.add(COL_REGION)
        elif "club" in cl and COL_CLUB not in mapped:
            col_map[col] = COL_CLUB;     mapped.add(COL_CLUB)
        elif "place" in cl and COL_RANG not in mapped:
            col_map[col] = COL_RANG;     mapped.add(COL_RANG)
        elif cl == "rang" and COL_RANG not in mapped:
            # Seulement si pas de colonne "Place" (sinon c'est un rang par compétition)
            col_map[col] = COL_RANG;     mapped.add(COL_RANG)
        # "rang.1", "rang.2", "pts", etc. → ignorés

    # Renommer PUIS supprimer les colonnes non mappées qui ont le même nom cible
    # (ex: colonne "Rang" originale qui n'a pas été mappée mais porte le même nom)
    cols_cibles = set(col_map.values())
    df = df_raw[[c for c in df_raw.columns if c in col_map or col_map.get(c) not in cols_cibles]].copy()
    df = df.rename(columns=col_map)

    # Garder uniquement la PREMIÈRE occurrence de chaque colonne cible
    seen_cols = set()
    cols_garder = []
    for c in df.columns:
        if c not in seen_cols:
            cols_garder.append(c)
            seen_cols.add(c)
    df = df[cols_garder]

    # Garder uniquement les colonnes standard
    std_cols = [c for c in [COL_RANG, COL_NOM, COL_PRENOM, COL_ADHERENT, COL_REGION, COL_CLUB]
                if c in df.columns]
    df = df[std_cols].copy()

    # Si pas de rang → numéroter
    if COL_RANG not in df.columns:
        df.insert(0, COL_RANG, range(1, len(df) + 1))

    # Si pas de région → classement GE (tous sont GE)
    if COL_REGION not in df.columns:
        df[COL_REGION] = "GES - GRAND EST"

    # Supprimer lignes sans nom valide (lettres) — ex: ":", chiffres
    # Le filtre rang (rang > 0) est fait dans _normaliser après dédupplication
    if COL_NOM in df.columns:
        import re as _re
        def _nom_valide(v):
            s = str(v).strip()
            return bool(s) and s not in ("nan", "") and bool(_re.search(r"[A-Za-z\xc0-\xff]", s))
        df = df[df[COL_NOM].apply(_nom_valide)]

    return _normaliser(df)


# ── HTML ──────────────────────────────────────────────────────────────

def lire_html(path: str) -> pd.DataFrame:
    """
    Parse un export HTML de l'extranet FFE.
    Cherche le premier tableau avec Nom/Prénom/Club.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise ImportError("beautifulsoup4 requis : pip install beautifulsoup4")

    with open(path, encoding="utf-8", errors="replace") as f:
        content = f.read()

    soup = BeautifulSoup(content, "lxml")
    tables = soup.find_all("table")

    for table in tables:
        headers = [th.get_text(strip=True).lower() for th in table.find_all("th")]
        if not headers:
            # Chercher en-tête dans première ligne <tr>
            first_row = table.find("tr")
            if first_row:
                headers = [td.get_text(strip=True).lower() for td in first_row.find_all(["th", "td"])]

        if not any("nom" in h for h in headers):
            continue

        # Mapper les colonnes
        col_map = {}
        for i, h in enumerate(headers):
            if "nom" in h and "prénom" not in h and "prenom" not in h:
                col_map[i] = COL_NOM
            elif "prénom" in h or "prenom" in h:
                col_map[i] = COL_PRENOM
            elif "adhér" in h or "licen" in h:
                col_map[i] = COL_ADHERENT
            elif "ligue" in h or "région" in h or "region" in h:
                col_map[i] = COL_REGION
            elif "club" in h:
                col_map[i] = COL_CLUB
            elif "rang" in h or "class" in h or "place" in h:
                col_map[i] = COL_RANG

        rows = []
        rang = 0
        for tr in table.find_all("tr")[1:]:
            cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if not cells:
                continue
            rang += 1
            row = {COL_RANG: rang}
            for i, val in enumerate(cells):
                if i in col_map:
                    row[col_map[i]] = val
            rows.append(row)

        if rows:
            return _normaliser(pd.DataFrame(rows))

    return pd.DataFrame()


# ── FFF ───────────────────────────────────────────────────────────────

def lire_fff(path: str) -> pd.DataFrame:
    """
    Parse un fichier FFF Engarde/WIN (classement individuel).
    Format : NOM;PRENOM;GENRE;NAT;ADHERENT;...;LIGUE;CLUB;RANG;...
    Structure variable selon l'export — détection par position des champs connus.
    """
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception:
        with open(path, encoding="latin-1", errors="replace") as f:
            content = f.read()

    lines = content.strip().splitlines()
    if not lines:
        return pd.DataFrame()

    # Ignorer les lignes d'en-tête (commencent par #, X, ou sont trop courtes)
    data_lines = [l for l in lines if l.strip() and not l.startswith("#")]

    rows = []
    rang = 0

    for line in data_lines:
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 4:
            continue

        # Format classement individuel FFE standard :
        # 0:NOM 1:PRENOM 2:CLUB 3:LIGUE 4:NAT 5:ADHERENT 6:... dernier:RANG
        # Ou format Engarde : 0:NOM 1:PRENOM 2:NATION 3:CLUB 4:RANG ...

        nom    = parts[0].upper().strip()
        prenom = parts[1].strip() if len(parts) > 1 else ""

        if not nom or len(nom) < 2:
            continue

        # Détecter le rang (dernier champ numérique ou champ "rang")
        rang_val = 0
        for p in reversed(parts):
            if re.match(r"^\d+$", p.strip()):
                rang_val = int(p.strip())
                break
        if not rang_val:
            rang += 1
            rang_val = rang

        # Chercher club et ligue dans les autres champs
        club   = parts[2].strip() if len(parts) > 2 else ""
        region = parts[3].strip() if len(parts) > 3 else ""
        adh    = parts[5].strip() if len(parts) > 5 and re.match(r"\d{5,6}", parts[5]) else ""

        rows.append({
            COL_RANG: rang_val, COL_NOM: nom, COL_PRENOM: prenom,
            COL_ADHERENT: adh, COL_REGION: region, COL_CLUB: club,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values(COL_RANG).reset_index(drop=True)
    return _normaliser(df)




# ── Épreuve Technique ────────────────────────────────────────────────

def lire_et(path: str) -> pd.DataFrame:
    """
    Parse un classement Épreuve Technique.
    Supporte deux formats :
      - Tabulaire XLSX : Rang | Nom | Prénom | Club | Note  (format principal)
      - Texte : "1er: NOM Prénom - score"  (format legacy)
    Tous les tireurs sont GE (classement régional Grand Est).
    """
    ext = path.lower().rsplit('.', 1)[-1] if '.' in path else ''
    rows = []

    if ext in ('xlsx', 'xls'):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        # Chercher la feuille avec des données
        ws = None
        for sh in wb.sheetnames:
            s = wb[sh]
            if s.max_row > 1 and s.max_column > 1:
                ws = s
                break
        if ws is None:
            return pd.DataFrame()

        # Lire toutes les lignes en texte
        raw_rows = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() if v is not None else '' for v in row]
            raw_rows.append(vals)

        # Détecter la ligne d'en-tête (contient "Rang" ou "Nom")
        header_idx = None
        col_rang = col_nom = col_prenom = col_club = col_note = None
        for i, row in enumerate(raw_rows):
            row_lower = [v.lower() for v in row]
            if 'rang' in row_lower or 'nom' in row_lower:
                header_idx = i
                for j, v in enumerate(row_lower):
                    if v == 'rang' or v == 'place':    col_rang   = j
                    elif v == 'nom':                   col_nom    = j
                    elif 'prénom' in v or 'prenom' in v: col_prenom = j
                    elif 'club' in v:                  col_club   = j
                    elif 'note' in v or 'score' in v:  col_note   = j
                break

        if header_idx is not None and col_nom is not None:
            # Format tabulaire
            for row in raw_rows[header_idx + 1:]:
                rang_val = row[col_rang] if col_rang is not None else ''
                if not rang_val or not rang_val.isdigit():
                    continue
                rang  = int(rang_val)
                nom   = row[col_nom].upper().strip() if col_nom is not None else ''
                prenom = row[col_prenom].strip() if col_prenom is not None else ''
                club  = row[col_club].strip() if col_club is not None else ''
                note  = row[col_note].strip() if col_note is not None else ''
                if not nom or len(nom) < 2:
                    continue
                rows.append({
                    COL_RANG: rang, COL_NOM: nom, COL_PRENOM: prenom,
                    COL_ADHERENT: '', COL_REGION: 'GES - GRAND EST',
                    COL_CLUB: club, COL_NOTE: note, COL_GRAND_EST: True,
                })
        else:
            # Format texte legacy dans les cellules
            all_text = ' '.join(v for row in raw_rows for v in row if v)
            lines = all_text.split('  ')
            for line in lines:
                m = re.match(r'(\d+)\w*\s*:\s*(.+?)\s*[-–]\s*([\d.,]+)', line.strip())
                if not m:
                    continue
                rang  = int(m.group(1))
                label = m.group(2).strip()
                score = m.group(3).replace(',', '.')
                parts = label.split()
                nom_parts    = [p for p in parts if p.isupper()]
                prenom_parts = [p for p in parts if not p.isupper()]
                nom    = ' '.join(nom_parts) if nom_parts else parts[0].upper()
                prenom = ' '.join(prenom_parts) if prenom_parts else ''
                rows.append({
                    COL_RANG: rang, COL_NOM: nom, COL_PRENOM: prenom,
                    COL_ADHERENT: '', COL_REGION: 'GES - GRAND EST',
                    COL_CLUB: '', COL_NOTE: score, COL_GRAND_EST: True,
                })

    elif ext == 'pdf':
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            lines = []
            for page in pdf.pages:
                lines += (page.extract_text() or '').splitlines()
        for line in lines:
            m = re.match(r'(\d+)\w*\s*:\s*(.+?)\s*[-–]\s*([\d.,]+)', line.strip())
            if not m:
                continue
            rang  = int(m.group(1))
            label = m.group(2).strip()
            score = m.group(3)
            parts = label.split()
            nom_parts    = [p for p in parts if p.isupper()]
            prenom_parts = [p for p in parts if not p.isupper()]
            rows.append({
                COL_RANG: rang,
                COL_NOM: ' '.join(nom_parts) if nom_parts else parts[0].upper(),
                COL_PRENOM: ' '.join(prenom_parts) if prenom_parts else '',
                COL_ADHERENT: '', COL_REGION: 'GES - GRAND EST',
                COL_CLUB: '', COL_NOTE: score, COL_GRAND_EST: True,
            })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ── Chorégraphie ──────────────────────────────────────────────────────

def _normaliser_participant(p: str) -> str:
    """Normalise un participant : 'NOM,Prenom' → 'Prenom NOM'."""
    p = p.strip()
    if ',' in p:
        parts = p.split(',', 1)
        return parts[1].strip() + ' ' + parts[0].strip()
    return p


def _lire_feuille_chore(ws, sous_cat: str) -> list:
    """Lit une feuille chorégraphie et retourne la liste des groupes."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = []
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if s and s.lower() not in ('nan', 'none'):
                vals.append(s)
        if not vals:
            continue

        # Ligne de données : col0 = rang (entier), col1 = participants
        if not re.match(r'^\d+$', vals[0]):
            continue
        rang = int(vals[0])
        if len(vals) < 2:
            continue

        participants = vals[1].strip()
        club = vals[2].strip() if len(vals) > 2 else ''

        # Ignorer les lignes d'en-tête déguisées
        if not re.search(r'[A-Za-zÀ-ÿ]', participants):
            continue

        # Séparer et normaliser les participants
        parts_raw   = re.split(r'\s*/\s*', participants)
        parts_clean = [_normaliser_participant(p) for p in parts_raw
                       if p.strip() and p.strip() != '/']
        participants_clean = ' / '.join(parts_clean)
        if not participants_clean:
            continue

        # Score : dernière colonne numérique significative
        score = ''
        for v in reversed(vals[3:]):
            v_clean = v.replace(',', '.').strip()
            try:
                float(v_clean)
                score = v_clean
                break
            except ValueError:
                continue

        rows.append({
            COL_RANG:         rang,
            COL_PARTICIPANTS: participants_clean,
            COL_CLUB:         club,
            COL_SOUS_CAT:     sous_cat,
            COL_REGION:       'GES - GRAND EST',
            COL_GRAND_EST:    True,
            COL_NOTE:         score,
        })
    return rows


def lire_chore(path: str) -> pd.DataFrame:
    """
    Parse un classement Chorégraphie multi-feuilles XLSX.
    Format : 1 feuille par sous-catégorie (DUEL, BATAILLE, ENSEMBLE).
    Chaque ligne : rang | "Participant1 / Participant2" | club | scores...
    Retourne un DataFrame avec COL_SOUS_CAT, COL_PARTICIPANTS, COL_CLUB, COL_RANG.
    """
    ext = path.lower().rsplit('.', 1)[-1] if '.' in path else ''
    if ext not in ('xlsx', 'xls'):
        raise ValueError(f"Format {ext} non supporté pour Chorégraphie (XLSX attendu)")

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # Mapping feuille → sous-catégorie
    FEUILLES_MAP = {
        'duel':     'Duel',
        'duels':    'Duel',
        'bataille': 'Bataille',
        'batailles': 'Bataille',
        'ensemble': 'Ensemble',
        'ensembles': 'Ensemble',
    }

    rows = []
    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().lower()
        sous_cat = None
        for mot, sc in FEUILLES_MAP.items():
            if mot in key:
                sous_cat = sc
                break
        if sous_cat is None:
            continue  # feuille non reconnue

        ws = wb[sheet_name]
        rows.extend(_lire_feuille_chore(ws, sous_cat))

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ── Markdown BellePoule ───────────────────────────────────────────────

def lire_md(path: str) -> pd.DataFrame:
    """
    Parse un classement BellePoule exporté en Markdown (.md).
    Format réel : tout le classement sur une seule ligne par page, précédé de
    la ligne d'en-tête 'place nom prénom club CID région nation'.
    Structure : rang NOM Prénom CLUB_MOT1 [CLUB_MOT2 ...] rang NOM Prénom ...
    Sans colonne région dans le contenu (classement national mixte).
    """
    import unicodedata

    rows = []
    seen_rangs = set()

    with open(path, encoding="utf-8", errors="replace") as f:
        content = f.read()

    # Extraire les blocs de données : après chaque ligne d'en-tête
    # On cherche les zones contenant 'place nom' puis les données qui suivent
    blocs = re.split(r'place\s+nom\s+pr[eé]nom\s+club', content, flags=re.IGNORECASE)

    for bloc in blocs[1:]:  # ignorer ce qui précède la première en-tête
        # Nettoyer : retirer les ## et sauts de ligne
        texte = re.sub(r'##[^\n]*', ' ', bloc)
        texte = re.sub(r'\s+', ' ', texte).strip()
        # Retirer "CID région nation" résiduel en début
        texte = re.sub(r'^CID\s+r[eé]gion\s+nation\s*', '', texte, flags=re.IGNORECASE).strip()

        # Découper sur les frontières "rang NOM_MAJUSCULES"
        # Pattern : entier + au moins un token tout-majuscules
        _PAT = r'(\d+)\s+((?:[A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ][A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ\-]*(?:\s+(?=[A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ][A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ\-]))?)+)\s+'
        segments = re.split(_PAT, texte)
        # segments = [avant, rang1, NOM1, reste1, rang2, NOM2, reste2, ...]
        i = 1
        while i + 2 < len(segments):
            rang_str = segments[i].strip()
            nom      = segments[i + 1].strip()
            reste    = segments[i + 2].strip()
            i += 3

            if not rang_str.isdigit():
                continue
            rang = int(rang_str)
            if rang in seen_rangs or rang == 0:
                continue

            # Dans "reste" : prénom(s) puis club (commence au premier token tout-majuscule)
            reste_tokens = reste.split()
            prenom_parts = []
            club_parts   = []
            for j, tok in enumerate(reste_tokens):
                is_upper = (tok == tok.upper() and len(tok) > 1
                            and re.sub(r"[^A-ZÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ\-]", "", tok) == tok)
                if prenom_parts and is_upper:
                    club_parts = reste_tokens[j:]
                    break
                prenom_parts.append(tok)
            else:
                club_parts = []

            prenom = ' '.join(prenom_parts).strip()
            club   = ' '.join(club_parts).strip()

            seen_rangs.add(rang)
            rows.append({
                COL_RANG:     rang,
                COL_NOM:      nom,
                COL_PRENOM:   prenom,
                COL_ADHERENT: "",
                COL_REGION:   "",
                COL_CLUB:     club,
            })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows).sort_values(COL_RANG).reset_index(drop=True)
    # Tous les tireurs de ce classement = pas de filtre GE (classement national)
    # Grand_Est = False par défaut (pas d'info région)
    df[COL_GRAND_EST] = False
    return df


# ── Dispatcher enrichi ────────────────────────────────────────────────

def lire_classement(path: str, format_hint: str = None, parseur_hint: str = None) -> pd.DataFrame:
    """
    Dispatcher principal.
    parseur_hint : 'et' | 'chore' | None (autodétection format standard)
    """
    import os
    fmt = format_hint or detecter_format(os.path.basename(path))

    # Parseurs spéciaux
    if parseur_hint == 'et':
        return lire_et(path)
    if parseur_hint == 'chore':
        return lire_chore(path)

    parseurs = {
        "pdf":  lire_pdf,
        "xlsx": lire_xlsx,
        "xls":  lire_xlsx,
        "html": lire_html,
        "htm":  lire_html,
        "fff":  lire_fff,
        "md":   lire_md,
    }
    fn = parseurs.get(fmt)
    if fn is None:
        raise ValueError(f"Format non supporté : {fmt}")
    return fn(path)

