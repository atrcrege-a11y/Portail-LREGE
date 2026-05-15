"""
sabre_laser/generateur.py — Génération Excel Sabre Laser.

Deux modes :
  - mode="separes"    → un Workbook par discipline (liste de wb)
  - mode="multi"      → un Workbook avec un onglet par discipline
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .config import get_palette as _get_palette

# Palette par défaut (fallback)
PALETTE_SL = {
    "titre":      "4B0082",
    "section1":   "6A0DAD",
    "section2":   "8B5CF6",
    "section3":   "C4B5FD",
    "alerte":     "F3E8FF",
    "entete_col": "EDE9FE",
    "fond_ligne": "F9F5FF",
    "bord":       "C4B5FD",
}

def _palette(disc_id: str = "") -> dict:
    """Retourne la palette de la discipline, ou le défaut."""
    return _get_palette(disc_id) if disc_id else PALETTE_SL


def _thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value="", bold=False, size=10, color="000000",
          bg=None, align_h="left", wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=color,
                       italic=italic)
    c.alignment = Alignment(horizontal=align_h, vertical="center",
                            wrap_text=wrap)
    c.border    = _thin_border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def _fusionner(ws, row, col_debut, col_fin, value="", bold=False, size=11,
               font_color="FFFFFF", bg=None, align_h="center"):
    col_l = get_column_letter(col_fin)
    ws.merge_cells(f"{get_column_letter(col_debut)}{row}:{col_l}{row}")
    c = ws.cell(row=row, column=col_debut, value=value)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=font_color)
    c.alignment = Alignment(horizontal=align_h, vertical="center",
                            wrap_text=True)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[row].height = 22
    return c


def _entete_feuille(ws, meta: dict, nb_cols: int = 5):
    """Construit l'en-tête commune (4 lignes + séparateur)."""
    col_l = get_column_letter(nb_cols)

    # Largeurs colonnes
    if nb_cols == 4:
        largeurs = {1: 8, 2: 50, 3: 22, 4: 20}
    else:
        largeurs = {1: 8, 2: 28, 3: 20, 4: 30, 5: 22}
    for i in range(1, nb_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = largeurs.get(i, 20)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    disc_id = meta.get("disc_id", "")
    pal   = _palette(disc_id)
    disc  = meta.get("discipline", "")
    date  = meta.get("date", "")
    lieu  = meta.get("lieu", "")
    mail       = meta.get("mail_retour", "")
    date_lim   = meta.get("date_limite_retour", "")
    date_extra = meta.get("date_engagement_extranet", "")

    # L1 : "SABRE LASER"
    _fusionner(ws, 1, 1, nb_cols, "SABRE LASER", bold=True, size=12,
               bg=pal["titre"], font_color="FFFFFF")

    # L2 : Date · Lieu
    date_lieu = f"{date}  ·  {lieu}" if date and lieu else date or lieu or ""
    _fusionner(ws, 2, 1, nb_cols, date_lieu, bold=False, size=10,
               bg=pal["alerte"], font_color="1B3F7A")

    # L3 : Discipline
    _fusionner(ws, 3, 1, nb_cols, disc.upper(), bold=True, size=12,
               bg=pal["titre"], font_color="FFFFFF")

    # L4 : Infos retour (mail, date limite, extranet)
    infos = []
    if mail:       infos.append(f"Retour : {mail}")
    if date_lim:   infos.append(f"Date limite : {date_lim}")
    if date_extra: infos.append(f"Extranet : {date_extra}")
    info_txt = "   |   ".join(infos) if infos else ""
    ws.merge_cells(f"A4:{col_l}4")
    c = ws.cell(row=4, column=1, value=info_txt)
    c.font      = Font(name="Calibri", size=8, italic=True, color="1B3F7A")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    c.fill      = PatternFill("solid", fgColor=pal.get("alerte", "F3E8FF"))
    ws.row_dimensions[4].height = 14

    # L5 : séparateur vide
    ws.row_dimensions[5].height = 6
    sep_color = pal.get("entete_col", "EDE9FE")
    for c in range(1, nb_cols + 1):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=sep_color)

    return 6  # première ligne disponible


def _entete_colonnes(ws, ligne: int, nb_cols: int = 5, pal: dict = None, mode_groupe: bool = False):
    """En-tête des colonnes tireurs."""
    pal = pal or PALETTE_SL
    if mode_groupe:
        # Choré : Rang | Participants | Club | Participation
        headers = ["Rang", "Participants", "Club", "Participation\nOui / Non"]
    else:
        headers = ["Rang", "Nom", "Prénom", "Club", "Participation\nOui / Non"]
    for i, h in enumerate(headers[:nb_cols], 1):
        c = ws.cell(row=ligne, column=i, value=h)
        c.font      = Font(name="Calibri", size=9, bold=True, color="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", fgColor=pal["entete_col"])
        c.border    = _thin_border()
    ws.row_dimensions[ligne].height = 18
    return ligne + 1


def _bloc_section(ws, ligne: int, section: dict, nb_cols: int = 5, pal: dict = None) -> int:
    """Affiche un bloc : titre + textes + liste de tireurs."""
    pal       = pal or PALETTE_SL
    label     = section.get("label", "")
    couleur   = section.get("couleur", pal.get("section1", PALETTE_SL["section1"]))
    textes    = section.get("textes", [])
    tireurs   = section.get("tireurs", [])
    avec_part = section.get("avec_participation", True)
    # mode_groupe = True pour la Chorégraphie (groupes au lieu de Nom/Prénom)
    mode_groupe = section.get("mode_groupe", False)
    # En mode groupe : 4 colonnes (Rang|Participants|Club|Participation)
    if mode_groupe:
        nb_cols = 4

    # Titre section
    _fusionner(ws, ligne, 1, nb_cols, f"  {label}", bold=True, size=11,
               bg=couleur, font_color="FFFFFF", align_h="left")
    ligne += 1

    # Textes critères
    for txt in textes:
        ws.merge_cells(f"A{ligne}:{get_column_letter(nb_cols)}{ligne}")
        c = ws.cell(row=ligne, column=1, value=f"  ▸  {txt}")
        c.font      = Font(name="Calibri", size=9, italic=True, color="1B3F7A")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill      = PatternFill("solid", fgColor=PALETTE_SL["alerte"])
        c.border    = _thin_border()
        ws.row_dimensions[ligne].height = 15
        ligne += 1
    if textes:
        ligne += 1

    if not tireurs:
        return ligne

    # En-tête colonnes
    ligne = _entete_colonnes(ws, ligne, nb_cols, pal=pal, mode_groupe=mode_groupe)

    # Lignes tireurs
    for i, t in enumerate(tireurs):
        bg = "FFFFFF" if i % 2 == 0 else pal.get("fond_ligne", "F9F5FF")
        note = t.get("note", "")
        note_extra = t.get("note_extra", "")

        if mode_groupe:
            # Chorégraphie : Rang | Participants (large) | Club | Participation
            participants = t.get("nom", "")  # stocké dans "nom" pour les groupes
            club = t.get("club", "")
            ws.row_dimensions[ligne].height = max(18, 14 * max(1, participants.count("/") + 1))
            _cell(ws, ligne, 1, t.get("rang", ""), align_h="center", bg=bg, size=9)
            _cell(ws, ligne, 2, participants, bold=True, bg=bg, size=9, wrap=True)
            _cell(ws, ligne, 3, club, bg=bg, size=9, wrap=True)
            _cell(ws, ligne, 4, "", bg="FFFFFF", size=9, align_h="center")
        else:
            # Standard : Rang | Nom | Prénom | Club | Participation
            ws.row_dimensions[ligne].height = 18
            _cell(ws, ligne, 1, t.get("rang", ""),   align_h="center", bg=bg, size=9)
            _cell(ws, ligne, 2, t.get("nom", ""),    bold=True,         bg=bg, size=9)
            _cell(ws, ligne, 3, t.get("prenom", ""),                    bg=bg, size=9)
            club = t.get("club", "")
            _cell(ws, ligne, 4, club,                                   bg=bg, size=9, wrap=True)
            if nb_cols >= 5:
                _cell(ws, ligne, 5, "", bg="FFFFFF", size=9, align_h="center")
        ligne += 1

    return ligne + 1  # espace entre sections


def _generer_feuille(ws, data: dict, nb_cols: int = 5):
    """Génère une feuille complète à partir des données de sélection."""
    disc_id    = data.get("meta", {}).get("disc_id", "")
    pal        = _palette(disc_id)
    has_part   = any(s.get("avec_participation", True) for s in data.get("sections", []))
    is_chore   = any(s.get("mode_groupe", False) for s in data.get("sections", []))
    # Choré : 4 cols (Rang|Participants|Club|Participation)
    # Autres : 5 cols (Rang|Nom|Prénom|Club|Participation)
    effective_cols = 4 if is_chore else (5 if has_part else nb_cols)
    ligne = _entete_feuille(ws, data["meta"], effective_cols)
    for section in data.get("sections", []):
        ligne = _bloc_section(ws, ligne, section, effective_cols, pal=pal)


# ── API publique ──────────────────────────────────────────────────────

def generer_docs_separes(selections: list[dict]) -> list[tuple[str, Workbook]]:
    """
    Génère un Workbook séparé par sélection.
    selections : liste de dict {"nom_feuille": str, "data": dict}
    Retourne : liste de (nom_fichier_suggéré, workbook)
    """
    resultats = []
    for sel in selections:
        wb = Workbook()
        ws = wb.active
        ws.title = sel.get("nom_feuille", "Sélection")[:31]
        _generer_feuille(ws, sel["data"])
        resultats.append((sel.get("nom_fichier", "selection_sl.xlsx"), wb))
    return resultats


def generer_doc_multi(selections: list[dict], titre_wb: str = "Sélections Sabre Laser") -> Workbook:
    """
    Génère un seul Workbook avec un onglet par sélection.
    selections : liste de dict {"nom_feuille": str, "data": dict}
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sel in selections:
        nom = sel.get("nom_feuille", "Sélection")[:31]
        ws  = wb.create_sheet(title=nom)
        _generer_feuille(ws, sel["data"])

    return wb
