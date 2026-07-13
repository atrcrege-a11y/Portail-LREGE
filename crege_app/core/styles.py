"""
core/styles.py — Palette, constantes visuelles et fonctions de style Excel.
Palette arme × catégorie : get_palette_arme(cat_id, arme) -> dict
"""
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection

# ── Constantes rétrocompatibles ───────────────────────────────────────
COULEUR_TITRE        = "1B3F7A"
COULEUR_N1           = "1B3F7A"
COULEUR_N2           = "2563A8"
COULEUR_N3           = "4A86C8"
COULEUR_EQUIPE       = "1B3F7A"
COULEUR_REMPLACANT   = "3A5A8A"
COULEUR_ENTETE_COL   = "EBF2FA"
COULEUR_LIGNE_PAIRE  = "F5F9FF"
COULEUR_ALERTE       = "F0F5FF"
COULEUR_BLANC        = "FFFFFF"
COULEUR_GE1          = "1B3F7A"
COULEUR_GE2          = "2563A8"
COULEUR_GE3          = "4A86C8"
COULEUR_REMPL_EQ     = "5C5C5C"

_thin = Side(style="thin",   color="CCCCCC")
_med  = Side(style="medium", color="AAAAAA")
BORDURE_FINE   = Border(left=_thin, right=_thin, top=_thin,   bottom=_thin)
BORDURE_MEDIUM = Border(left=_med,  right=_med,  top=_med,    bottom=_med)

LABELS_ARB = {
    "regional":           "Régional",
    "regional_formation": "Régional en formation",
    "national":           "National",
    "national_formation": "National en formation",
    "international":      "International",
}

# ── Palette arme × catégorie ──────────────────────────────────────────
# Clés : titre · n1 · n2 · n3 · entete_col · ligne_paire · alerte · bordure
# n1 = bandeau section principale, n2 = en-tête colonnes, n3 = sous-section
# bordure = couleur des Side() fins sur les cellules données
_PA = {
    "M13": {
        "F": dict(titre="3949AB", n1="5C6BC0", n2="7986CB", n3="9FA8DA", entete_col="7986CB", ligne_paire="C5CAE9", alerte="E8EAF6", bordure="9FA8DA"),
        "E": dict(titre="6A1B9A", n1="8E24AA", n2="AB47BC", n3="CE93D8", entete_col="AB47BC", ligne_paire="E1BEE7", alerte="F3E5F5", bordure="CE93D8"),
        "S": dict(titre="C2185B", n1="E91E63", n2="EC407A", n3="F48FB1", entete_col="EC407A", ligne_paire="F8BBD9", alerte="FCE4EC", bordure="F48FB1"),
    },
    "M15": {
        "F": dict(titre="00695C", n1="00897B", n2="26A69A", n3="80CBC4", entete_col="26A69A", ligne_paire="B2DFDB", alerte="E0F2F1", bordure="80CBC4"),
        "E": dict(titre="2E7D32", n1="43A047", n2="66BB6A", n3="A5D6A7", entete_col="66BB6A", ligne_paire="C8E6C9", alerte="E8F5E9", bordure="A5D6A7"),
        "S": dict(titre="558B2F", n1="7CB342", n2="9CCC65", n3="C5E1A5", entete_col="9CCC65", ligne_paire="DCEDC8", alerte="F1F8E9", bordure="C5E1A5"),
    },
    "M17": {
        "F": dict(titre="1A237E", n1="283593", n2="3949AB", n3="7986CB", entete_col="3949AB", ligne_paire="C5CAE9", alerte="E8EAF6", bordure="9FA8DA"),
        "E": dict(titre="33691E", n1="558B2F", n2="689F38", n3="C5E1A5", entete_col="689F38", ligne_paire="DCEDC8", alerte="F1F8E9", bordure="C5E1A5"),
        "S": dict(titre="7B0000", n1="B71C1C", n2="E53935", n3="EF9A9A", entete_col="E53935", ligne_paire="FFCDD2", alerte="FFEBEE", bordure="EF9A9A"),
    },
    "M20": {
        "F": dict(titre="1565C0", n1="1976D2", n2="42A5F5", n3="90CAF9", entete_col="42A5F5", ligne_paire="BBDEFB", alerte="E3F2FD", bordure="90CAF9"),
        "E": dict(titre="7A6000", n1="A08000", n2="C4A010", n3="E8D44D", entete_col="C4A010", ligne_paire="FFF0A0", alerte="FFFDF0", bordure="E8D44D"),
        "S": dict(titre="BF360C", n1="E64A19", n2="FF7043", n3="FFAB91", entete_col="FF7043", ligne_paire="FFCCBC", alerte="FBE9E7", bordure="FFAB91"),
    },
    "Seniors": {
        "F": dict(titre="003580", n1="1565C0", n2="1976D2", n3="90CAF9", entete_col="1976D2", ligne_paire="BBDEFB", alerte="E3F2FD", bordure="90CAF9"),
        "E": dict(titre="1B5E20", n1="2E7D32", n2="388E3C", n3="A5D6A7", entete_col="388E3C", ligne_paire="C8E6C9", alerte="E8F5E9", bordure="A5D6A7"),
        "S": dict(titre="7B0000", n1="B71C1C", n2="C62828", n3="EF9A9A", entete_col="C62828", ligne_paire="FFCDD2", alerte="FFEBEE", bordure="EF9A9A"),
    },
    "V1": {
        "F": dict(titre="0277BD", n1="0288D1", n2="29B6F6", n3="81D4FA", entete_col="29B6F6", ligne_paire="B3E5FC", alerte="E1F5FE", bordure="81D4FA"),
        "E": dict(titre="00838F", n1="00ACC1", n2="26C6DA", n3="80DEEA", entete_col="26C6DA", ligne_paire="B2EBF2", alerte="E0F7FA", bordure="80DEEA"),
        "S": dict(titre="922B21", n1="C0392B", n2="E74C3C", n3="F1948A", entete_col="E74C3C", ligne_paire="FADBD8", alerte="FDEDEC", bordure="F1948A"),
    },
    "V2": {
        "F": dict(titre="0277BD", n1="0288D1", n2="29B6F6", n3="81D4FA", entete_col="29B6F6", ligne_paire="B3E5FC", alerte="E1F5FE", bordure="81D4FA"),
        "E": dict(titre="00838F", n1="00ACC1", n2="26C6DA", n3="80DEEA", entete_col="26C6DA", ligne_paire="B2EBF2", alerte="E0F7FA", bordure="80DEEA"),
        "S": dict(titre="922B21", n1="C0392B", n2="E74C3C", n3="F1948A", entete_col="E74C3C", ligne_paire="FADBD8", alerte="FDEDEC", bordure="F1948A"),
    },
    "V3": {
        "F": dict(titre="1976D2", n1="42A5F5", n2="64B5F6", n3="90CAF9", entete_col="64B5F6", ligne_paire="BBDEFB", alerte="E3F2FD", bordure="90CAF9"),
        "E": dict(titre="1B7A3C", n1="27AE60", n2="52BE80", n3="82E0AA", entete_col="52BE80", ligne_paire="ABEBC6", alerte="EAFAF1", bordure="82E0AA"),
        "S": dict(titre="7D3C98", n1="9B59B6", n2="AF7AC5", n3="D2B4DE", entete_col="AF7AC5", ligne_paire="E8DAEF", alerte="F5EEF8", bordure="D2B4DE"),
    },
    "V4": {
        "F": dict(titre="1976D2", n1="42A5F5", n2="64B5F6", n3="90CAF9", entete_col="64B5F6", ligne_paire="BBDEFB", alerte="E3F2FD", bordure="90CAF9"),
        "E": dict(titre="1B7A3C", n1="27AE60", n2="52BE80", n3="82E0AA", entete_col="52BE80", ligne_paire="ABEBC6", alerte="EAFAF1", bordure="82E0AA"),
        "S": dict(titre="7D3C98", n1="9B59B6", n2="AF7AC5", n3="D2B4DE", entete_col="AF7AC5", ligne_paire="E8DAEF", alerte="F5EEF8", bordure="D2B4DE"),
    },
    # Groupes vétérans F/S équipes
    "VET": {
        "F": dict(titre="0277BD", n1="0288D1", n2="29B6F6", n3="81D4FA", entete_col="29B6F6", ligne_paire="B3E5FC", alerte="E1F5FE", bordure="81D4FA"),
        "E": dict(titre="00838F", n1="00ACC1", n2="26C6DA", n3="80DEEA", entete_col="26C6DA", ligne_paire="B2EBF2", alerte="E0F7FA", bordure="80DEEA"),
        "S": dict(titre="922B21", n1="C0392B", n2="E74C3C", n3="F1948A", entete_col="E74C3C", ligne_paire="FADBD8", alerte="FDEDEC", bordure="F1948A"),
    },
    "GVET": {
        "F": dict(titre="1976D2", n1="42A5F5", n2="64B5F6", n3="90CAF9", entete_col="64B5F6", ligne_paire="BBDEFB", alerte="E3F2FD", bordure="90CAF9"),
        "E": dict(titre="1B7A3C", n1="27AE60", n2="52BE80", n3="82E0AA", entete_col="52BE80", ligne_paire="ABEBC6", alerte="EAFAF1", bordure="82E0AA"),
        "S": dict(titre="7D3C98", n1="9B59B6", n2="AF7AC5", n3="D2B4DE", entete_col="AF7AC5", ligne_paire="E8DAEF", alerte="F5EEF8", bordure="D2B4DE"),
    },
}

# Ancienne palette (rétrocompat) — utilisée en fallback
PALETTES = {
    "M13":     _PA["M13"]["E"],
    "M15":     _PA["M15"]["E"],
    "M17":     _PA["M17"]["S"],
    "M20":     _PA["M20"]["E"],
    "Seniors": _PA["Seniors"]["E"],
    "V1":      _PA["V1"]["E"],
    "V2":      _PA["V2"]["E"],
    "V3":      _PA["V3"]["E"],
    "V4":      _PA["V4"]["E"],
}
PALETTE_DEFAUT = PALETTES["Seniors"]


def get_palette(cat_id: str, arme: str = "E") -> dict:
    """Retourne la palette couleur pour une catégorie et une arme."""
    cat = _PA.get(cat_id)
    if cat is None:
        return PALETTE_DEFAUT
    return cat.get(arme, cat.get("E", PALETTE_DEFAUT))


# Alias explicite
get_palette_arme = get_palette


def make_border(color: str, style: str = "thin") -> Border:
    """Crée un objet Border openpyxl avec la couleur donnée."""
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def make_border_h(color: str) -> Border:
    """Bordure fine uniquement haut+bas (pas de bordure latérale)."""
    s = Side(style="thin", color=color)
    n = Side(style=None)
    return Border(top=s, bottom=s, left=n, right=n)


# ── Fonctions de style ────────────────────────────────────────────────

def style_cellule(ws, cell_ref, valeur=None, bold=False, italic=False,
                  font_size=10, font_color="000000", bg_color=None,
                  align_h="left", align_v="center", wrap=False, border=None):
    cell = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
    if valeur is not None:
        cell.value = valeur
    cell.font = Font(name="Calibri", size=font_size, bold=bold,
                     italic=italic, color=font_color)
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if border is not None:
        cell.border = border


def fusionner_style(ws, debut, fin, valeur, bold=False, font_size=11,
                    font_color="FFFFFF", bg_color=COULEUR_TITRE, align_h="center"):
    from openpyxl.utils import column_index_from_string, get_column_letter
    ws.merge_cells(f"{debut}:{fin}")
    cell = ws[debut]
    style_cellule(ws, cell, valeur, bold=bold, font_size=font_size,
                  font_color=font_color, bg_color=bg_color,
                  align_h=align_h, align_v="center", border=None)
    row_num   = int(''.join(filter(str.isdigit, debut)))
    col_start = column_index_from_string(''.join(filter(str.isalpha, debut)))
    col_end   = column_index_from_string(''.join(filter(str.isalpha, fin)))
    no_border = Border()
    fill      = PatternFill("solid", fgColor=bg_color)
    for col_i in range(col_start + 1, col_end + 1):
        ws.cell(row=row_num, column=col_i).border = no_border
        ws.cell(row=row_num, column=col_i).fill   = fill
