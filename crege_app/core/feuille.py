"""
core/feuille.py — Composants de feuille Excel communs à toutes les catégories.
Entêtes, lignes info, lignes tireurs, arbitrage.
"""
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .styles import (
    style_cellule, fusionner_style, BORDURE_FINE,
    COULEUR_TITRE, COULEUR_N1, COULEUR_N2, COULEUR_N3,
    COULEUR_ENTETE_COL, COULEUR_LIGNE_PAIRE, COULEUR_BLANC,
    COULEUR_ALERTE, LABELS_ARB, get_palette,
)


# ─────────────────────────────────────────────
# LIGNES INFO (en-tête du document)
# ─────────────────────────────────────────────

def info_row(ws, ligne, icone, label, valeur, col_fin, h=18, bg=None):
    """Ligne info pleine largeur : icone + label + valeur."""
    thin = Side(style="thin", color="C5D8F0")
    ws.row_dimensions[ligne].height = h
    ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
    cell = ws[f"A{ligne}"]
    cell.value     = f"{icone}  {label} :    {valeur}"
    cell.font      = Font(name="Calibri", size=9, color="1B3F7A")
    cell.fill      = PatternFill("solid", fgColor=bg or COULEUR_ENTETE_COL)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    cell.border    = Border(bottom=thin)
    return ligne + 1


def ligne_arbitre(ws, ligne, col_fin="F", h=22, label="Arbitre", bg_lbl=None):
    """Ligne arbitre saisissable : label | Nom Prénom | Club | Niveau (dropdown)."""
    ws.row_dimensions[ligne].height = h
    thin      = Side(style="thin", color="BBCFE0")
    bord_full = Border(top=thin, bottom=thin, left=thin, right=thin)
    bord_r    = Border(top=thin, bottom=thin, right=thin)
    fill_lbl  = PatternFill("solid", fgColor=bg_lbl or COULEUR_ENTETE_COL)
    fill_w    = PatternFill("solid", fgColor="FFFFFF")
    font_lbl  = Font(name="Calibri", size=9, bold=True, color="1B3F7A")
    font_noir = Font(name="Calibri", size=9, color="000000")
    align_c   = Alignment(horizontal="left", vertical="center")
    no_lock   = Protection(locked=False)
    lock      = Protection(locked=True)

    def _s(col, val, fill, font, prot, bord):
        c = ws.cell(row=ligne, column=col)
        c.value = val; c.fill = fill; c.font = font
        c.alignment = align_c; c.border = bord; c.protection = prot

    _s(1, f"👤  {label}", fill_lbl, font_lbl, lock,    bord_full)
    _s(2, "Nom  Prénom",  fill_w,   font_noir, no_lock, bord_full)
    _s(3, "",             fill_w,   font_noir, no_lock, bord_r)
    _s(4, "Club",         fill_w,   font_noir, no_lock, bord_full)
    _s(5, "Cliquer pour choisir le niveau ▼", fill_w, font_noir, no_lock, bord_full)
    _s(6, "",             fill_w,   font_noir, no_lock, bord_r)

    niveaux = '"Régional en formation,Régional,National en formation,National,International"'
    dv = DataValidation(type="list", formula1=niveaux, allow_blank=True,
                        showDropDown=False, showErrorMessage=False)
    dv.sqref = f"E{ligne}"
    ws.add_data_validation(dv)
    return ligne + 1


# ─────────────────────────────────────────────
# ENTÊTE DU DOCUMENT
# ─────────────────────────────────────────────

def init_feuille(ws, meta, nb_cols=6):
    """Construit l'entête commune : titre, compétition, discipline, infos pratiques."""
    col_fin = get_column_letter(nb_cols)
    cat_id  = meta.get("cat_id", "")
    pal     = get_palette(cat_id)
    # Surcharger les couleurs globales avec la palette de la catégorie
    C_TITRE      = pal["titre"]
    C_N1         = pal["n1"]
    C_ALERTE     = pal["alerte"]
    C_ENTETE_COL = pal["entete_col"]
    largeurs = {1: 16, 2: 24, 3: 18, 4: 36, 5: 20, 6: 28, 7: 28}
    for i in range(1, nb_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = largeurs.get(i, 20)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ligne = 1

    # ── Ligne 1 : Compétition + catégorie ──────────────────────────────
    competition = meta.get("competition", "Championnat de France")
    cat_label   = meta.get("cat_label", "")
    # Éviter redondance "VÉTÉRANS Vétérans V1" → extraire juste le numéro pour vétérans
    comp_upper = competition.upper()
    if cat_label and "ÉTÉRAN" in comp_upper:
        # Garder uniquement le suffixe (V1, V2, V3, V4)
        cat_suffix = cat_label.split()[-1] if cat_label else ""
        titre_comp = f"{comp_upper} {cat_suffix}".strip()
    else:
        titre_comp = f"{comp_upper} {cat_label.upper()}".strip()
    ws.row_dimensions[ligne].height = 22
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    titre_comp, bold=True, font_size=12, bg_color=C_TITRE,
                    font_color="FFFFFF", align_h="center")
    ligne += 1

    # ── Ligne 2 : Date · Lieu (fond rose pâle) ───────────────────────
    date_str  = meta.get("date", "")
    lieu      = meta.get("lieu", "")
    date_lieu = f"{date_str}  ·  {lieu}" if date_str and lieu else date_str or lieu
    ws.row_dimensions[ligne].height = 16
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    date_lieu, bold=False, font_size=10,
                    bg_color=pal.get("entete_col", "FAD0D0"), font_color="1B3F7A", align_h="center")
    ligne += 1

    # ── Ligne 3 : Discipline ─────────────────────────────────────────
    discipline = meta.get("discipline", "")
    ws.row_dimensions[ligne].height = 22
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    discipline.upper(), bold=True, font_size=12, bg_color=C_TITRE,
                    font_color="FFFFFF", align_h="center")
    ligne += 1

    # Séparateur
    ligne += 1

    # ── Infos pratiques ──
    mail_retour   = meta.get("mail_retour", "administration@crege.fr")
    date_limite   = meta.get("date_limite_retour", "")
    date_extranet = meta.get("date_engagement_extranet", "")
    arb_cfg       = meta.get("arbitrage_config", {})
    arb_source    = arb_cfg.get("source", "")

    # Texte arbitrage
    if not arb_source or arb_source == "aucun":
        arb_txt = "Pas d'arbitre requis"
    else:
        src      = "par le club du tireur" if arb_source == "club" else "par la ligue"
        s1       = arb_cfg.get("seuil1", arb_cfg.get("ratio", 4))
        use2     = arb_cfg.get("use_seuil2", False)
        s2       = arb_cfg.get("seuil2", 9)
        arb_txt  = (f"Fourni {src} — "
                    f"1 arbitre à partir de {s1} tireurs engagés (H+D confondus)")
        if use2 and s2:
            arb_txt += (f", 2 arbitres à partir de {s2} tireurs engagés (H+D confondus)")

    # Phrase engagement (M13 spécifique)
    is_m13 = "M13" in meta.get("discipline", "").upper()
    thin_c = Side(style="thin", color="C5D8F0")
    ws.row_dimensions[ligne].height = 14 if not is_m13 else 22
    ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
    c = ws[f"A{ligne}"]
    c.value = ("La LREGE engage les tireurs. Les équipes sont à l'initiative des clubs "
               "ainsi que l'inscription sur l'extranet."
               if is_m13 else
               "Le LREGE engage les tireurs et équipes qualifiés.")
    c.font      = Font(name="Calibri", size=9, color="444444", italic=True)
    c.fill      = PatternFill("solid", fgColor=pal.get("alerte", "F0F4F9"))
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = Border()
    ligne += 1

    # Date limite
    if date_limite:
        ligne = info_row(ws, ligne, "📧", "Confirmations avant le", date_limite, col_fin, bg=C_ENTETE_COL)

    # Mail retour
    ws.row_dimensions[ligne].height = 18
    ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
    c2 = ws[f"A{ligne}"]
    c2.value = f"  ↳  {mail_retour}"
    c2.font  = Font(name="Calibri", size=9, color="1F6391", italic=True)
    c2.fill  = PatternFill("solid", fgColor="FFFFFF")
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border    = Border(bottom=thin_c)
    ligne += 1

    # Date extranet
    if date_extranet:
        ligne = info_row(ws, ligne, "🖥", "Clôture extranet", date_extranet, col_fin, bg=C_ENTETE_COL)

    # Arbitrage
    ligne = info_row(ws, ligne, "⚖️", "Arbitrage", arb_txt, col_fin, bg=C_ENTETE_COL)

    # Lignes arbitres saisissables
    if arb_source and arb_source != "aucun":
        ligne = ligne_arbitre(ws, ligne, col_fin, label="Arbitre 1", bg_lbl=C_ENTETE_COL)
        ligne = ligne_arbitre(ws, ligne, col_fin, label="Arbitre 2", bg_lbl=C_ENTETE_COL)

    ligne += 1
    return ligne


# ─────────────────────────────────────────────
# ENTÊTE COLONNES TIREURS
# ─────────────────────────────────────────────

def entete_tireurs(ws, ligne, col_fin="F", avec_participation=True, avec_taille=False):
    """Ligne d'entête des colonnes tireurs."""
    nb_cols = ord(col_fin) - ord("A") + 1
    hdrs = ["Rang / Classement", "Nom", "Prénom", "Club"]
    if avec_participation: hdrs.append("Participation\nOui / Non")
    if avec_taille:        hdrs.append("Taille Veste")
    elif avec_participation: hdrs.append("Remarque")

    ws.row_dimensions[ligne].height = 30
    for i, h in enumerate(hdrs, 1):
        align = "center" if i in (1, 5, 6) else "left"
        style_cellule(ws, f"{get_column_letter(i)}{ligne}", h,
                      bold=True, font_size=9, bg_color=COULEUR_ENTETE_COL,
                      font_color="000000", align_h=align,
                      border=BORDURE_FINE, wrap=True)
    return ligne + 1


# ─────────────────────────────────────────────
# LIGNE TIREUR
# ─────────────────────────────────────────────

def ligne_tireur(ws, ligne, tireur, bg, col_fin="F",
                 avec_participation=True, avec_taille=False):
    """Affiche une ligne de tireur avec rang, nom, prénom, club."""
    club = tireur.get("club", "")
    note = tireur.get("note", "")
    if len(club) > 40 or len(note) > 50:
        ws.row_dimensions[ligne].height = 35
    elif len(club) > 28 or len(note) > 25:
        ws.row_dimensions[ligne].height = 28
    else:
        ws.row_dimensions[ligne].height = 20

    champs = [
        (tireur.get("rang",   ""), "center", False, 10, False),
        (tireur.get("nom",    ""), "left",   True,  10, False),
        (tireur.get("prenom", ""), "left",   False, 10, False),
        (club,                     "left",   False,  9, True),
    ]
    for j, (val, align, bold, sz, wrap) in enumerate(champs, 1):
        style_cellule(ws, f"{get_column_letter(j)}{ligne}", val,
                      bold=bold, font_size=sz, align_h=align,
                      border=BORDURE_FINE, bg_color=bg, wrap=wrap)

    if avec_participation:
        note = tireur.get("note", "")
        # E = Participation (case vide à remplir), F = Remarque (note)
        style_cellule(ws, f"E{ligne}", "", align_h="center",
                      border=BORDURE_FINE, bg_color=COULEUR_BLANC)
        if avec_taille:
            style_cellule(ws, f"F{ligne}", "", align_h="center",
                          border=BORDURE_FINE, bg_color=COULEUR_BLANC)
        else:
            style_cellule(ws, f"F{ligne}", note, font_size=9, italic=bool(note),
                          font_color="666666" if note else "000000",
                          border=BORDURE_FINE, bg_color=bg, wrap=True)


# ── Marqueur de version du format Excel ──────────────────────────────────────
# Écrit dans les propriétés du classeur (keywords). Contrôlé par SuiviGE à la
# lecture : absent = accepté (fichiers antérieurs), différent = erreur.
EXCEL_FORMAT_VERSION = "SELECGE_XLSX_V1"


def marquer_version(wb):
    """Appose le marqueur de version de format sur un classeur openpyxl."""
    wb.properties.keywords = EXCEL_FORMAT_VERSION
    return wb
