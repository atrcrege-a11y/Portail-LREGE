"""
generateur/excel.py — Point d'entrée de la génération Excel.
Dispatch selon le format (jeunes, seniors, equipes_m15).
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..core.feuille import init_feuille
from ..core.styles import (
    style_cellule, fusionner_style, BORDURE_FINE,
    COULEUR_N1, COULEUR_N2, COULEUR_ALERTE,
    COULEUR_LIGNE_PAIRE, COULEUR_BLANC, COULEUR_EQUIPE,
)
from .sections import bloc_section
from .equipes  import remplir_feuille_equipes


def generer_feuille_simple(ws, data):
    """Remplit une feuille existante — format jeunes ou seniors."""
    fmt     = data.get("format", "seniors")
    col_fin = "F"
    ligne   = init_feuille(ws, data["meta"], nb_cols=6)

    # M15 individuel : ajouter la colonne taille de veste
    is_m15 = "M15" in data.get("meta", {}).get("discipline", "").upper()
    if is_m15:
        for section in data["sections"]:
            section["avec_taille"] = True
            for ss in section.get("sous_sections", []):
                ss["avec_taille"] = True

    for section in data["sections"]:
        ligne = bloc_section(ws, ligne, section, col_fin)


def generer_multi_genres(data_h, data_d):
    """
    Génère un classeur 2 feuilles (Hommes + Dames).
    data_d peut être None pour un classeur 1 feuille.
    """
    wb = Workbook()

    ws_h = wb.active
    ws_h.title = "Hommes"
    generer_feuille_simple(ws_h, data_h)

    if data_d:
        ws_d = wb.create_sheet("Dames")
        generer_feuille_simple(ws_d, data_d)

    return wb


def generer_equipes_m15(data_h, data_d):
    """
    Génère un classeur équipes M15 : 2 feuilles (Hommes + Dames).
    """
    wb = Workbook()

    ws_h = wb.active
    ws_h.title = "Hommes"
    remplir_feuille_equipes(ws_h, data_h)

    if data_d:
        ws_d = wb.create_sheet("Dames")
        remplir_feuille_equipes(ws_d, data_d)

    return wb
