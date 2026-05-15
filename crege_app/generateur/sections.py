"""
generateur/sections.py — Rendu des sections de tireurs dans une feuille Excel.
"""
from openpyxl.utils import get_column_letter
from ..core.styles import (
    style_cellule, fusionner_style, BORDURE_FINE,
    COULEUR_N1, COULEUR_N2, COULEUR_N3, COULEUR_ALERTE,
    COULEUR_LIGNE_PAIRE, COULEUR_BLANC, COULEUR_REMPLACANT,
)
from ..core.feuille import entete_tireurs, ligne_tireur


def bloc_section(ws, ligne, section, col_fin="F"):
    """Affiche un bloc complet : titre + textes + tireurs (+ sous-sections)."""
    label        = section.get("label", "")
    couleur      = section.get("couleur", COULEUR_N1)
    textes       = section.get("textes", [])
    tireurs      = section.get("tireurs", [])
    avec_part    = section.get("avec_participation", True)
    avec_taille  = section.get("avec_taille", False)
    texte_pied   = section.get("texte_pied", "")
    sous_sections= section.get("sous_sections", [])

    # Titre
    ws.row_dimensions[ligne].height = 20
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    f"  {label}", bold=True, font_size=11,
                    bg_color=couleur, align_h="left")
    ligne += 1

    # Textes critères
    for texte in textes:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {texte}", italic=True,
                      font_size=9, font_color="1B3F7A", bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1
    if textes:
        ligne += 1

    if sous_sections:
        for ss in sous_sections:
            ws.row_dimensions[ligne].height = 15
            fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                            f"  — {ss['label']}", bold=True, font_size=10,
                            bg_color=ss.get("couleur", COULEUR_N3), align_h="left")
            ligne += 1
            for t in ss.get("textes", []):
                ws.row_dimensions[ligne].height = 13
                ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
                style_cellule(ws, f"A{ligne}", f"    {t}", italic=True,
                              font_size=9, font_color="444444", bg_color=COULEUR_ALERTE, wrap=True)
                ligne += 1
            if ss.get("tireurs"):
                ligne = entete_tireurs(ws, ligne, col_fin, avec_part, avec_taille)
                for i, t in enumerate(ss["tireurs"]):
                    bg = COULEUR_LIGNE_PAIRE if i % 2 == 0 else COULEUR_BLANC
                    ligne_tireur(ws, ligne, t, bg, col_fin, avec_part, avec_taille)
                    ligne += 1
            ligne += 1
    else:
        if tireurs:
            ligne = entete_tireurs(ws, ligne, col_fin, avec_part, avec_taille)
            for i, t in enumerate(tireurs):
                bg = COULEUR_LIGNE_PAIRE if i % 2 == 0 else COULEUR_BLANC
                ligne_tireur(ws, ligne, t, bg, col_fin, avec_part, avec_taille)
                ligne += 1
        else:
            ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
            style_cellule(ws, f"A{ligne}", "  (aucun tireur pour cette section)",
                          italic=True, font_color="888888")
            ligne += 1

    # Texte pied (après les tireurs)
    if texte_pied:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {texte_pied}", italic=True,
                      font_size=9, font_color="1B3F7A", bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

    return ligne + 1
