"""
generateur/equipes.py — Rendu des feuilles équipes M15.
"""
from openpyxl.utils import get_column_letter
from ..core.styles import (
    style_cellule, fusionner_style, BORDURE_FINE, get_palette, make_border,
    COULEUR_GE1, COULEUR_GE2, COULEUR_GE3, COULEUR_REMPL_EQ,
    COULEUR_ENTETE_COL, COULEUR_LIGNE_PAIRE, COULEUR_BLANC, COULEUR_ALERTE,
)
from ..core.feuille import init_feuille


def remplir_feuille_equipes(ws, data):
    """Remplit une feuille avec les équipes M15 (GE1/GE2/GE3 + remplaçants)."""
    nb_cols = 6
    col_fin = get_column_letter(nb_cols)
    ligne   = init_feuille(ws, data["meta"], nb_cols)

    arme_id = data.get("meta", {}).get("arme_id", "E")
    pal     = get_palette("M15", arme_id)
    couleurs_eq = [pal["n1"], pal["n2"], pal["n3"]]
    labels = {
        1: "Rang / Classement", 2: "Nom", 3: "Prénom",
        4: "Club", 5: "Participation\nOui / Non", 6: "Taille de veste"
    }

    for idx, eq in enumerate(data.get("equipes", [])):
        couleur = couleurs_eq[idx] if idx < len(couleurs_eq) else COULEUR_GE3
        numero  = eq.get("numero", idx + 1)
        label   = eq.get("label", f"Grand Est {numero}")
        critere = eq.get("critere", "")
        tireurs = eq.get("tireurs", [])

        # Titre équipe
        ws.row_dimensions[ligne].height = 20
        fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                        f"  ÉQUIPE {label.upper()}",
                        bold=True, font_size=11, bg_color=couleur, align_h="left")
        ligne += 1

        # Critère
        if critere:
            ws.row_dimensions[ligne].height = 15
            ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
            style_cellule(ws, f"A{ligne}", f"  ▸  {critere}",
                          italic=True, font_size=9, font_color="1B3F7A",
                          bg_color=COULEUR_ALERTE, wrap=True)
            ligne += 1

        # Entête colonnes
        ws.row_dimensions[ligne].height = 30
        for i in range(1, nb_cols + 1):
            align = "center" if i in (1, 5, 6) else "left"
            style_cellule(ws, f"{get_column_letter(i)}{ligne}", labels.get(i, ""),
                          bold=True, font_size=9, bg_color=COULEUR_ENTETE_COL,
                          font_color="1B3F7A", align_h=align,
                          border=BORDURE_FINE, wrap=True)
        ligne += 1

        # Tireurs
        for i, t in enumerate(tireurs):
            bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
            ws.row_dimensions[ligne].height = 22
            club = t.get("club", "")
            if len(club) > 32:
                ws.row_dimensions[ligne].height = 34
            for j, (val, align, bold, sz, wrap) in enumerate([
                (t.get("rang",""),   "center", False, 10, False),
                (t.get("nom",""),    "left",   True,  10, False),
                (t.get("prenom",""), "left",   False, 10, False),
                (club,               "left",   False,  9, True),
            ], 1):
                style_cellule(ws, f"{get_column_letter(j)}{ligne}", val,
                              bold=bold, font_size=sz, align_h=align,
                              align_v="center", wrap=wrap,
                              border=BORDURE_FINE, bg_color=bg)
            style_cellule(ws, f"E{ligne}", "", align_h="center",
                          border=BORDURE_FINE, bg_color=COULEUR_BLANC)
            style_cellule(ws, f"F{ligne}", "", align_h="center",
                          border=BORDURE_FINE, bg_color=COULEUR_BLANC)
            ligne += 1

        # Mention désistement GE1
        if numero == 1:
            ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
            ws.row_dimensions[ligne].height = 15
            style_cellule(ws, f"A{ligne}",
                          "  En cas de désistement d'un tireur de l'équipe n°1, "
                          "il sera fait appel au suivant du classement national.",
                          italic=True, font_size=9, font_color="1B3F7A",
                          bg_color=COULEUR_ALERTE)
            ligne += 1

        ligne += 1  # espace entre équipes

    # Remplaçants
    remplacants = data.get("remplacants", [])
    if remplacants:
        ws.row_dimensions[ligne].height = 18
        fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                        "  TIREURS REMPLAÇANTS", bold=True, font_size=11,
                        bg_color=COULEUR_REMPL_EQ, align_h="left")
        ligne += 1
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", "  ▸  Appelés dans l'ordre en cas de désistement",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

        # Entête
        ws.row_dimensions[ligne].height = 30
        for i in range(1, nb_cols + 1):
            align = "center" if i in (1, 5, 6) else "left"
            style_cellule(ws, f"{get_column_letter(i)}{ligne}", labels.get(i, ""),
                          bold=True, font_size=9, bg_color=COULEUR_ENTETE_COL,
                          font_color="1B3F7A", align_h=align,
                          border=BORDURE_FINE, wrap=True)
        ligne += 1

        for i, t in enumerate(remplacants):
            bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
            ws.row_dimensions[ligne].height = 17
            club = t.get("club", "")
            for j, (val, align, bold, sz, wrap) in enumerate([
                (t.get("rang",""),   "center", False, 10, False),
                (t.get("nom",""),    "left",   True,  10, False),
                (t.get("prenom",""), "left",   False, 10, False),
                (club,               "left",   False,  9, True),
            ], 1):
                style_cellule(ws, f"{get_column_letter(j)}{ligne}", val,
                              bold=bold, font_size=sz, align_h=align,
                              align_v="center", wrap=wrap,
                              border=BORDURE_FINE, bg_color=bg)
            style_cellule(ws, f"E{ligne}", "", align_h="center",
                          border=BORDURE_FINE, bg_color=COULEUR_BLANC)
            style_cellule(ws, f"F{ligne}", "", align_h="center",
                          border=BORDURE_FINE, bg_color=COULEUR_BLANC)
            ligne += 1

        # Phrase pied
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        ws.row_dimensions[ligne].height = 15
        style_cellule(ws, f"A{ligne}",
                      "  et suivant dans l'ordre du classement régional",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE)
        ligne += 1
