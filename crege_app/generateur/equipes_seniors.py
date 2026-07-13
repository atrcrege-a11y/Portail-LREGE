"""
generateur/equipes_seniors.py — Feuille Excel équipes M17→Vétérans.

Structure : 2 feuilles (Hommes + Dames), chacune avec :
  - Section N1/N2 : 8 premières équipes ½ finale FFE
  - Section N3    : équipes quota LREGE ou mention "open"
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from ..core.styles import (
    style_cellule, fusionner_style, BORDURE_FINE,
    COULEUR_ENTETE_COL, COULEUR_LIGNE_PAIRE, COULEUR_BLANC,
    COULEUR_ALERTE, get_palette,
)
from ..core.feuille import marquer_version, init_feuille


NB_COLS = 5

# Labels groupe vétérans selon cat_id
def _label_groupe_vet(cat_id: str) -> str:
    """Retourne 'Vétérans' pour V1/V2, 'Grands Vétérans' pour V3/V4."""
    if cat_id in ("V1", "V2"):
        return "Vétérans"
    if cat_id in ("V3", "V4"):
        return "Grands Vétérans"
    return cat_id
LABELS_COLS = {
    1: "Rang / Classement",
    2: "Nom de l'équipe",
    3: "Club / Comité",
    4: "Participation\nOui / Non",
    5: "Observations",
}



def _lignes_composition(ws, ligne, taille_equipe):
    """Insère N lignes Nom/Prénom sous une ligne d'équipe."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    bg = PatternFill("solid", fgColor="F0F4FB")
    thin = Side(style="thin", color="D0DCF0")
    bord = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, taille_equipe + 1):
        ws.row_dimensions[ligne].height = 16
        # Col A : numéro tireur
        c = ws.cell(row=ligne, column=1, value=f"  {i}")
        c.font      = Font(name="Calibri", size=9, color="1F3864", bold=False)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = bord
        c.fill      = bg
        # Col B : Nom (placeholder visible, noir)
        c2 = ws.cell(row=ligne, column=2, value="Nom")
        c2.font      = Font(name="Calibri", size=9, color="1F3864")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border    = bord
        c2.fill      = PatternFill("solid", fgColor="FFFFFF")
        # Col C : Prénom (placeholder visible, noir)
        c3 = ws.cell(row=ligne, column=3, value="Prénom")
        c3.font      = Font(name="Calibri", size=9, color="1F3864")
        c3.alignment = Alignment(horizontal="left", vertical="center")
        c3.border    = bord
        c3.fill      = PatternFill("solid", fgColor="FFFFFF")
        # Col D-E : vides
        for col in [4, 5]:
            cx = ws.cell(row=ligne, column=col)
            cx.border = bord
            cx.fill   = bg
        ligne += 1
    return ligne


def _entete_colonnes(ws, ligne, pal):
    ws.row_dimensions[ligne].height = 28
    for i in range(1, NB_COLS + 1):
        align = "center" if i in (1, 4) else "left"
        style_cellule(ws, f"{get_column_letter(i)}{ligne}", LABELS_COLS.get(i, ""),
                      bold=True, font_size=9, bg_color=pal["entete_col"],
                      font_color="1B3F7A", align_h=align,
                      border=BORDURE_FINE, wrap=True)
    return ligne + 1


def _ligne_equipe(ws, ligne, rang, nom_equipe, club, note="", bg=None):
    bg = bg or COULEUR_BLANC
    ws.row_dimensions[ligne].height = 22
    # Formater le rang : si numérique, préfixer avec "N°"
    import re as _re
    rang_fmt = f"N° {rang}" if rang and _re.match(r"^\d+$", str(rang)) else str(rang)
    vals = [
        (rang_fmt,    "center", False, 10, False),
        (nom_equipe,  "left",   True,  10, False),
        (club,        "left",   False,  9, True),
        ("",          "center", False,  9, False),
        (note,        "left",   False,  9, True),
    ]
    for j, (val, align, bold, sz, wrap) in enumerate(vals, 1):
        style_cellule(ws, f"{get_column_letter(j)}{ligne}", val,
                      bold=bold, font_size=sz, align_h=align,
                      align_v="center", wrap=wrap,
                      border=BORDURE_FINE, bg_color=bg)
    return ligne + 1


def _section_n1n2(ws, ligne, equipes, pal, col_fin, cat_id="", taille_equipe=4, nb_n1n2=8):
    """Section N1/N2 : équipes qualifiées de la ½ finale FFE."""
    ws.row_dimensions[ligne].height = 20
    groupe = f" ({_label_groupe_vet(cat_id)})" if cat_id in ("V1","V2","V3","V4") else ""
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    f"  ÉQUIPES GRAND EST QUALIFIÉES N1/N2{groupe} — ½ FINALE NATIONALE FFE",
                    bold=True, font_size=11, bg_color=pal["n1"], align_h="left")
    ligne += 1

    for txt in [
        f"Équipes Grand Est qualifiées parmi les {nb_n1n2} premières de la ½ finale nationale",
        "Le classement initial est établi en fonction du résultat de la ½ finale nationale",
        "En cas de désistement d'une équipe qualifiée FFE : non remplacée par une équipe LREGE",
    ]:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {txt}",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

    if not equipes:
        ligne = _entete_colonnes(ws, ligne, pal)
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", "  Aucune équipe Grand Est qualifiée en N1/N2",
                      italic=True, font_size=9, font_color="6B7280", bg_color=COULEUR_BLANC)
        return ligne + 2

    # EH/FH Senior (nb_n1n2=16) : scinder en titre (1-8) et maintien (9-16)
    if nb_n1n2 == 16:
        titre    = [e for e in equipes if int(e.get("rang", 99)) <= 8]
        maintien = [e for e in equipes if int(e.get("rang", 99)) >  8]

        for groupe_label, groupe_equipes in [
            ("— Pour le titre (rangs 1 à 8)", titre),
            ("— Pour le maintien (rangs 9 à 16)", maintien),
        ]:
            if not groupe_equipes:
                continue
            # Sous-header
            ws.row_dimensions[ligne].height = 16
            ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
            style_cellule(ws, f"A{ligne}", f"  {groupe_label}",
                          bold=True, font_size=10, font_color="1B3F7A",
                          bg_color=pal["ligne_paire"], align_h="left")
            ligne += 1
            ligne = _entete_colonnes(ws, ligne, pal)
            for i, eq in enumerate(groupe_equipes):
                bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
                ligne = _ligne_equipe(ws, ligne, eq.get("rang", ""), eq.get("nom_equipe", ""), eq.get("club", ""), bg=bg)
                ligne = _lignes_composition(ws, ligne, taille_equipe)
        return ligne + 1

    # Cas standard (8 équipes)
    ligne = _entete_colonnes(ws, ligne, pal)
    for i, eq in enumerate(equipes):
        bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
        ligne = _ligne_equipe(ws, ligne, eq.get("rang", f"{i+1}"), eq.get("nom_equipe", ""), eq.get("club", ""), bg=bg)
        ligne = _lignes_composition(ws, ligne, taille_equipe)

    return ligne + 1


def _section_n3_ffe(ws, ligne, equipes, pal, col_fin, taille_equipe=4):
    """Section N3 FFE : équipes rangs 9-12 qualifiées par la FFE."""
    if not equipes:
        return ligne  # Pas d'équipe GE en N3 FFE — section masquée
    nb_lignes_vides = 0  # uniquement les équipes GE présentes
    ws.row_dimensions[ligne].height = 20
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    "  ÉQUIPES QUALIFIÉES N3 — GRAND EST (rangs 9 à 12, sélection nationale FFE)",
                    bold=True, font_size=11, bg_color=pal["n2"], align_h="left")
    ligne += 1

    for txt in [
        "Équipes Grand Est classées de la 9e à la 12e place de la ½ finale nationale",
        "Qualifiées directement par la FFE — pas de remplacement par quota LREGE",
    ]:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {txt}",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

    ligne = _entete_colonnes(ws, ligne, pal)

    nb_affiches = len(equipes)
    for i, eq in enumerate(equipes):
        bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
        ligne = _ligne_equipe(ws, ligne,
                              eq.get("rang", str(i + 9)),
                              eq.get("nom_equipe", ""),
                              eq.get("club", ""), bg=bg)
        ligne = _lignes_composition(ws, ligne, taille_equipe)
    return ligne + 1


def _section_n3(ws, ligne, equipes, quota, mode_n3, pal, col_fin, cat_id="", taille_equipe=4, nb_open_n3=0, label_n3="N3"):
    """Section N3 (ou N2 open pour SD) : quota LREGE ou open."""
    ws.row_dimensions[ligne].height = 20
    groupe = f" {_label_groupe_vet(cat_id)}" if cat_id in ("V1","V2","V3","V4") else ""
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    f"  ÉQUIPES QUALIFIÉES {label_n3}{groupe} — QUOTA LREGE GRAND EST",
                    bold=True, font_size=11, bg_color=pal["n2"], align_h="left")
    ligne += 1

    is_vet = cat_id in ("V1", "V2", "V3", "V4")
    if mode_n3 == "open":
        if is_vet:
            groupe = "Vétérans (V1+V2)" if cat_id in ("V1","V2") else "Grands Vétérans (V3+V4)"
            txt_regle = f"Épreuve open — {groupe} — Pas de quota régional"
            txt_info  = "Merci de contacter le secrétariat pour la composition de votre équipe"
        else:
            txt_regle = "Épreuve open — pas de quota régional LREGE pour cette catégorie/arme/genre"
            txt_info  = "Merci de vous inscrire auprès du secrétariat"
    else:
        txt_regle = (f"Quota LREGE Grand Est : {quota} équipe{'s' if quota > 1 else ''} "
                     f"qualifiée{'s' if quota > 1 else ''} dans l'ordre du classement du championnat régional GE")
        txt_info  = "En cas de désistement : l'équipe suivante du classement régional est appelée"

    for txt in [txt_regle, txt_info]:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {txt}",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

    if mode_n3 == "open":
        if nb_open_n3 > 0:
            # Cases d'inscription avec mention secrétariat
            ligne = _entete_colonnes(ws, ligne, pal)
            for i in range(nb_open_n3):
                bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
                ligne = _ligne_equipe(ws, ligne, f"{i+1}", "", "", bg=bg)
                ligne = _lignes_composition(ws, ligne, taille_equipe)
        else:
            ws.row_dimensions[ligne].height = 15
            ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
            style_cellule(ws, f"A{ligne}", "  Épreuve open — pas de liste LREGE",
                          italic=True, font_size=10, font_color="6B7280",
                          bg_color=COULEUR_BLANC)
        return ligne + 2
        txt_regle = (f"Quota LREGE Grand Est : {quota} équipe{'s' if quota > 1 else ''} "
                     f"qualifiée{'s' if quota > 1 else ''} dans l'ordre du classement du championnat régional GE")
        txt_info  = "En cas de désistement : l'équipe suivante du classement régional est appelée"

    for txt in [txt_regle, txt_info]:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {txt}",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

    if mode_n3 == "open":
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", "  Épreuve open — pas de liste LREGE",
                      italic=True, font_size=10, font_color="6B7280",
                      bg_color=COULEUR_BLANC)
        return ligne + 2

    ligne = _entete_colonnes(ws, ligne, pal)

    nb = max(quota, len(equipes))
    for i in range(nb):
        bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
        if i < len(equipes):
            eq = equipes[i]
            ligne = _ligne_equipe(ws, ligne,
                                  eq.get("rang", f"{i+1}"),
                                  eq.get("nom_equipe", ""),
                                  eq.get("club", ""), bg=bg)
        else:
            ligne = _ligne_equipe(ws, ligne, f"{i+1}", "", "", bg=bg)
        ligne = _lignes_composition(ws, ligne, taille_equipe)

    # Pied
    ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
    ws.row_dimensions[ligne].height = 15
    style_cellule(ws, f"A{ligne}",
                  "  et suivant(es) dans l'ordre du classement du championnat régional Grand Est",
                  italic=True, font_size=9, font_color="1B3F7A", bg_color=COULEUR_ALERTE)
    return ligne + 2


def _section_remplacants(ws, ligne, remplacants, pal, col_fin, cat_id="", nb_lignes_vides=3, taille_equipe=4):
    """Section remplaçants équipes. Affiche toujours au moins nb_lignes_vides lignes vides."""
    ws.row_dimensions[ligne].height = 20
    groupe = f" {_label_groupe_vet(cat_id)}" if cat_id in ("V1","V2","V3","V4") else ""
    fusionner_style(ws, f"A{ligne}", f"{col_fin}{ligne}",
                    f"  ÉQUIPES REMPLAÇANTES{groupe}",
                    bold=True, font_size=11, bg_color=pal["n3"], align_h="left")
    ligne += 1

    for txt in [
        "Appelées dans l'ordre en cas de désistement d'une équipe qualifiée LREGE",
        "En cas de désistement : le classement régional Grand Est sert de liste de référence",
    ]:
        ws.row_dimensions[ligne].height = 15
        ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
        style_cellule(ws, f"A{ligne}", f"  ▸  {txt}",
                      italic=True, font_size=9, font_color="1B3F7A",
                      bg_color=COULEUR_ALERTE, wrap=True)
        ligne += 1

    ligne = _entete_colonnes(ws, ligne, pal)

    nb_affiches = len(remplacants)
    for i, eq in enumerate(remplacants):
        bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
        ligne = _ligne_equipe(ws, ligne,
                              eq.get("rang", f"{i+1}"),
                              eq.get("nom_equipe", ""),
                              eq.get("club", ""), bg=bg)
        ligne = _lignes_composition(ws, ligne, taille_equipe)

    # Lignes vides supplémentaires si moins de nb_lignes_vides remplaçants
    for i in range(nb_affiches, max(nb_affiches, nb_lignes_vides)):
        bg = pal["ligne_paire"] if i % 2 == 0 else COULEUR_BLANC
        ligne = _ligne_equipe(ws, ligne, str(i + 1), "", "", bg=bg)
        ligne = _lignes_composition(ws, ligne, taille_equipe)

    # Pied
    ws.merge_cells(f"A{ligne}:{col_fin}{ligne}")
    ws.row_dimensions[ligne].height = 15
    style_cellule(ws, f"A{ligne}",
                  "  et suivante(s) dans l'ordre du classement du championnat régional Grand Est",
                  italic=True, font_size=9, font_color="1B3F7A", bg_color=COULEUR_ALERTE)
    return ligne + 2


def _remplir_feuille_equipes_seniors(ws, data, genre):
    """Remplit une feuille équipes (H ou D) pour M17→Vétérans."""
    g      = genre.lower()
    cat_id = data["meta"].get("cat_id", "Seniors")
    pal    = get_palette(cat_id)
    col_fin = get_column_letter(NB_COLS)

    # En-tête
    meta_genre = dict(data["meta"])
    disc = meta_genre.get("discipline", "")
    meta_genre["discipline"] = f"{disc} {'HOMMES' if genre=='H' else 'DAMES'}".strip()
    ligne = init_feuille(ws, meta_genre, NB_COLS)

    eq_n1n2  = data.get(f"equipes_n1n2_{g}", [])
    eq_n3    = data.get(f"equipes_n3_{g}",   [])
    quota_n3 = data.get(f"quota_n3_{g}", 0)
    mode_n3  = data.get(f"mode_n3_{g}", "quota")
    nb_open_n3 = int(data.get(f"nb_open_n3_{g}", data.get("nb_open_n3", 0)))
    remplacants = data.get(f"remplacants_{g}", [])
    eq_n3_ffe  = data.get(f"equipes_n3_ffe_{g}", [])
    # Taille d'équipe : 5 pour Vétérans, 4 sinon (modifiable via l'interface)
    taille_defaut = 5 if cat_id in ("V1", "V2", "V3", "V4") else 4
    taille_equipe = int(data.get("taille_equipe", taille_defaut))

    # Seniors EH et FH : 16 équipes N1+N2 (1-8 titre, 9-16 maintien)
    # Tous les autres : 8 équipes
    nb_n1n2 = 16 if (cat_id == "Seniors" and arme_id in ("E","F") and genre == "H") else 8
    arme_id = data["meta"].get("arme_id", "E")

    ligne = _section_n1n2(ws, ligne, eq_n1n2, pal, col_fin, cat_id, taille_equipe, nb_n1n2)
    ligne = _section_n3_ffe(ws, ligne, eq_n3_ffe, pal, col_fin, taille_equipe)
    # SD Seniors : N1 sur liste FFE, N2 open (pas de N3)
    # → label "N2" pour la section open
    label_n3 = "N2" if (cat_id == "Seniors" and arme_id == "S" and genre == "D") else "N3"

    ligne = _section_n3(ws, ligne, eq_n3, quota_n3, mode_n3, pal, col_fin, cat_id, taille_equipe, nb_open_n3, label_n3)
    if mode_n3 == "quota":
        ligne = _section_remplacants(ws, ligne, remplacants, pal, col_fin, cat_id, taille_equipe=taille_equipe)

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.sheet_view.showGridLines = False


def generer_equipes_seniors(data: dict) -> Workbook:
    """Génère un classeur Excel 2 feuilles (Hommes + Dames) pour équipes M17→Vétérans."""
    wb = marquer_version(Workbook())
    wb.remove(wb.active)

    ws_h = wb.create_sheet("Hommes")
    _remplir_feuille_equipes_seniors(ws_h, data, "H")

    ws_d = wb.create_sheet("Dames")
    _remplir_feuille_equipes_seniors(ws_d, data, "D")

    return wb
