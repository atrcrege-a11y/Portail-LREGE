"""
categories/jeunes.py — Catégories jeunes : M13, M15, M17, M20.

Règles métier :
  - M13 : quota fédéral et quota national optionnels (OFF par défaut)
           texte engagement spécifique
  - M15 : 3 compétitions (FDJ / Zone / 1/2 Finale), équipes GE1/GE2/GE3
  - M17/M20 : CDF, nationalité française obligatoire
"""
from .base import BaseCategorie
from ..core.utils import filtrer_df, est_grand_est, COL_RANG, COL_NOM, COL_PRENOM, COL_CLUB, COL_REGION


# ── Couleurs sous-sections ───────────────────────────────────────────
from ..core.styles import COULEUR_N1, COULEUR_N2, COULEUR_N3, get_palette


def _construire_jeunes(df_national, df_regional, config: dict) -> dict:
    """
    Logique commune de construction pour M13/M15/M17/M20.

    Options M13 :
        use_fed (bool) : inclure le quota fédéral (filtre top-N national)
        use_nat (bool) : inclure le quota LREGE national
    """
    quota_fed   = config.get("quota_federal", 40)    # seuil N1 (32 pour M17/M20)
    quota_cn    = config.get("quota_crege_nat", 2)   # places LREGE sur classement national
    quota_cr    = config.get("quota_crege_reg", 2)   # places LREGE sur classement régional
    nb_rempl    = config.get("nb_remplacants", 10)
    # CDF M17→M23 : nationalité française obligatoire (sélectionnables EDF)
    filtre_fr   = config.get("nationalite_francaise", True)
    cat_id      = config.get("cat_id", "")
    # Wild Cards DTN : M17 et M20 uniquement (pas M13/M15)
    nb_wc       = config.get("nb_wildcards", 4) if cat_id in ("M17", "M20") else 0
    # Seuil N1 = toujours 32 pour M17/M20 (indépendant du quota LREGE total)
    n1_seuil    = 32 if cat_id in ("M17", "M20", "M23") else quota_fed
    use_fed     = quota_fed > 0  # True si section N1 active
    use_nat     = quota_cn > 0
    pal       = get_palette(cat_id, config.get("arme_id", "E"))

    import pandas as pd
    _COLS = [COL_RANG, COL_NOM, COL_PRENOM, COL_CLUB, COL_REGION]
    _df_empty = pd.DataFrame(columns=_COLS)
    df_nat = filtrer_df(df_national, filtre_fr) if df_national is not None else _df_empty
    df_reg = filtrer_df(df_regional, filtre_fr) if df_regional is not None else _df_empty

    def _norm(s):
        import unicodedata, re as _re
        s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()
        return _re.sub(r"[^A-Z]", "", s)

    noms_sel = set(config.get("noms_exclus", set()))  # déjà normalisés

    def _in_sel(nom, prenom):
        return (_norm(nom), _norm(prenom)) in noms_sel

    def _add_sel(nom, prenom):
        noms_sel.add((_norm(nom), _norm(prenom)))

    # ── Quota Fédéral ──────────────────────────────────────────────
    if use_fed:
        ge_top = df_nat[
            (df_nat[COL_RANG] <= n1_seuil) &
            df_nat[COL_REGION].apply(est_grand_est)
        ].copy()
        tireurs_fed = [
            {"rang": f"CL NAT {r[COL_RANG]}", "nom": r[COL_NOM],
             "prenom": r[COL_PRENOM], "club": r[COL_CLUB], "note": ""}
            for _, r in ge_top.iterrows()
        ]
        for _, r in ge_top.iterrows():
            _add_sel(r[COL_NOM], r[COL_PRENOM])
        # Wild Cards DTN : section dédiée si GE présent dans la zone WC (rangs 33-36)
        if nb_wc > 0:
            ge_zone_wc = df_nat[
                (df_nat[COL_RANG] > n1_seuil) &
                (df_nat[COL_RANG] <= n1_seuil + nb_wc) &
                df_nat[COL_REGION].apply(est_grand_est)
            ]
            if ge_zone_wc.empty:
                nb_wc = 0  # pas de WC pertinente pour le GE
        # tireurs_wc : liste des tireurs GE en zone WC + lignes vides
        tireurs_wc = []
        if nb_wc > 0:
            for _, r in ge_zone_wc.iterrows():
                tireurs_wc.append({
                    "rang": f"CL NAT {r[COL_RANG]}", "nom": r[COL_NOM],
                    "prenom": r[COL_PRENOM], "club": r[COL_CLUB],
                    "note": "Possible maintien en N1, en attente de validation"
                })
            # Uniquement les tireurs GE réels — pas de lignes vides
    else:
        tireurs_fed = []

    # ── Quota LREGE national ────────────────────────────────────────
    if use_nat:
        nb_wc_cfg = config.get("nb_wildcards", 4) if cat_id in ("M17", "M20", "M23") else 0
        tireurs_cn = []
        # Quota LREGE national = UNIQUEMENT les tireurs GE en zone WC (rangs 33-36)
        # → note "Montée en N1 possible..."
        if nb_wc > 0 and use_fed:
            for _, r in ge_zone_wc.iterrows():
                tireurs_cn.append({
                    "rang": f"CL NAT {r[COL_RANG]}", "nom": r[COL_NOM],
                    "prenom": r[COL_PRENOM], "club": r[COL_CLUB],
                    "note": "Montée en N1 possible, en attente de l\'attribution des Wild Cards"
                })
                _add_sel(r[COL_NOM], r[COL_PRENOM])
        # Compléter jusqu'à quota_cn si pas assez de candidats WC
        seuil_nat = (n1_seuil + nb_wc_cfg) if (use_fed and config.get("quota_federal", 0) > 0) else 0
        places_restantes = max(0, quota_cn - len(tireurs_cn))
        if places_restantes > 0:
            ge_nat_extra = df_nat[
                (df_nat[COL_RANG] > seuil_nat) &
                df_nat[COL_REGION].apply(est_grand_est) &
                ~df_nat.apply(lambda r: _in_sel(r[COL_NOM], r[COL_PRENOM]), axis=1)
            ].head(places_restantes).copy()
            for _, r in ge_nat_extra.iterrows():
                tireurs_cn.append({
                    "rang": f"CL NAT {r[COL_RANG]}", "nom": r[COL_NOM],
                    "prenom": r[COL_PRENOM], "club": r[COL_CLUB], "note": ""
                })
                _add_sel(r[COL_NOM], r[COL_PRENOM])
        # Mémoriser les GE au-delà de la zone WC pour les intégrer au quota régional
        ge_nat_hors_wc = df_nat[
            (df_nat[COL_RANG] > seuil_nat) &
            df_nat[COL_REGION].apply(est_grand_est) &
            ~df_nat.apply(lambda r: _in_sel(r[COL_NOM], r[COL_PRENOM]), axis=1)
        ].copy()
    else:
        tireurs_cn = []
        seuil_nat = 0
        import pandas as pd
        ge_nat_hors_wc = df_nat.iloc[0:0] if df_nat is not None else pd.DataFrame()

    # ── Quota LREGE régional ────────────────────────────────────────
    # quota_cr = classement régional GE pur (CL GE X)
    # Les GE rang 37+ du national complètent quota_cn (déjà traité ci-dessus)
    reg_restants = df_reg[
        ~df_reg.apply(
            lambda r: _in_sel(r[COL_NOM], r[COL_PRENOM]), axis=1)
    ]
    tireurs_cr = []
    for _, r in reg_restants.head(quota_cr).iterrows():
        tireurs_cr.append({
            "rang": f"CL GE {r[COL_RANG]}", "nom": r[COL_NOM],
            "prenom": r[COL_PRENOM], "club": r[COL_CLUB], "note": ""
        })
        _add_sel(r[COL_NOM], r[COL_PRENOM])

    # ── Remplaçants ─────────────────────────────────────────────────
    rempl_df = df_reg[
        ~df_reg.apply(
            lambda r: _in_sel(r[COL_NOM], r[COL_PRENOM]), axis=1)
    ].head(nb_rempl)
    tireurs_rempl = []
    for i, (_, r) in enumerate(rempl_df.iterrows()):
        # 1er remplaçant : note "rattrapage possible" si zone WC active
        note = ""
        if i == 0 and nb_wc > 0 and use_fed and not ge_zone_wc.empty:
            note = "Rattrapage possible en attente de l\'attribution des Wild Cards"
        tireurs_rempl.append({
            "rang": f"CL GE {r[COL_RANG]}", "nom": r[COL_NOM],
            "prenom": r[COL_PRENOM], "club": r[COL_CLUB], "note": note
        })

    # ── Textes ──────────────────────────────────────────────────────
    qf_display   = config.get("quota_federal_display", quota_fed)
    # Texte N1 : toujours "Les 32 premiers" pour M17/M20, sinon quota_fed
    n1_display = n1_seuil if cat_id in ("M17", "M20", "M23") else qf_display
    textes_fed_default = [
        f"Les {n1_display} premiers du classement national hors DOM/TOM"
        if cat_id == "M15" else
        f"Les {n1_display} premiers tireurs sélectionnables en équipe de France "
        f"du classement national",
    ]
    if nb_wc > 0 and use_fed and not ge_zone_wc.empty:
        textes_fed_default.append(
            f"+ jusqu'à {nb_wc} Wild Card{'s' if nb_wc > 1 else ''} attribuée{'s' if nb_wc > 1 else ''} par la DTN — "
            f"voir section quota LREGE pour les candidats GE"
        )
    textes_fed   = config.get("textes_federal", textes_fed_default)
    qcn_eff      = quota_cn if use_nat else 0
    quota_lrege_total = qcn_eff + quota_cr
    # Label du niveau selon la catégorie et l'arme (N2 ou N3)
    niveau_lrege = config.get("niveau_lrege", "")  # ex: "N2", "N3" passé depuis _build_cfg
    label_niveau = f" — {niveau_lrege}" if niveau_lrege else ""
    textes_crege = config.get("textes_crege", [
        f"Quota LREGE Grand Est{label_niveau} : {quota_lrege_total} place{'s' if quota_lrege_total > 1 else ''} "
        f"({'⅓' if qcn_eff > 0 else '0'} classement FFE + {'⅔' if quota_cr > 0 else '0'} classement régional GE)"
    ])

    # ── Sections dynamiques ─────────────────────────────────────────
    sections = []
    if use_fed:
        sections.append({
            "label": "TIREURS QUALIFIÉS — N1 (QUOTA FÉDÉRAL)",
            "couleur": pal["n1"],
            "textes": textes_fed,
            "tireurs": tireurs_fed,
            "avec_participation": True,
        })
    # Les tireurs WC sont désormais intégrés dans le quota LREGE national (N2/N3)

    sous_sections = []
    if use_nat and quota_cn > 0:
        nb_wc_in_cn = sum(1 for t in tireurs_cn if "Wild Card" in t.get("note","") or "Montée en N1" in t.get("note",""))
        label_cn_extra = f" (dont {nb_wc_in_cn} candidats WC)" if nb_wc_in_cn > 0 else ""
        sous_sections.append({
            "label": f"Sur classement national (LREGE) : {quota_cn} "
                     f"{'place' if quota_cn <= 1 else 'places'}{label_cn_extra}",
            "couleur": pal["n3"], "textes": [], "tireurs": tireurs_cn,
        })
    sous_sections.append({
        "label": f"Sur classement régional (LREGE) : {quota_cr} "
                 f"{'place' if quota_cr <= 1 else 'places'}",
        "couleur": pal["n3"], "textes": [], "tireurs": tireurs_cr,
    })

    sections.append({
        "label": f"TIREURS QUALIFIÉS — QUOTA LREGE{(' — ' + niveau_lrege) if niveau_lrege else ''}",
        "couleur": pal["n2"],
        "textes": textes_crege,
        "tireurs": [],
        "avec_participation": True,
        "sous_sections": sous_sections,
    })

    sections.append({
        "label": "TIREURS REMPLAÇANTS",
        "couleur": "4A4A4A",
        "textes": ["Les remplaçants sont appelés dans l'ordre en cas de désistement"],
        "texte_pied": "et suivant dans l'ordre du classement régional",
        "tireurs": tireurs_rempl,
        "avec_participation": True,
    })

    return {
        "format": "jeunes",
        "meta": {
            "region": "Grand Est",
            "cat_id":                   cat_id,
            "cat_label":                config.get("cat_label", cat_id),
            "competition":              config.get("competition", "Fête des Jeunes"),
            "date":                     config.get("date", ""),
            "lieu":                     config.get("lieu", ""),
            "discipline":               config.get("discipline", ""),
            "mail_retour":              config.get("mail_retour", "administration@crege.fr"),
            "date_limite_retour":       config.get("date_limite_retour", ""),
            "date_engagement_extranet": config.get("date_engagement_extranet", ""),
            "arbitrage_config":         config.get("arbitrage_config", {}),
        },
        "sections": sections,
        "equipes_numerotees": [],
    }



# ─────────────────────────────────────────────
# Open circuit (M20 Sabre, M17 Sabre, M23 Sabre)
# Pas de quota LREGE : uniquement les GES dans les listes PDF FFE N1 + N2
# ─────────────────────────────────────────────

def _construire_jeunes_open_circuit(df_n1, df_n2, df_nat, config: dict) -> dict:
    """
    Mode open_circuit : compétition ouverte sous condition de participation nationale.
    La LREGE ne dispose d'aucun quota propre.
    On extrait les GES présents dans les listes FFE N1 et N2.
    Le rang affiché est celui du classement national (df_nat) si fourni, sinon
    l'ordre d'apparition dans le PDF.

    df_n1  : DataFrame issu du PDF N1 FFE (tous les tireurs, GES filtrés ici)
    df_n2  : DataFrame issu du PDF N2 FFE (idem)
    df_nat : DataFrame classement national (pour les rangs) — peut être None
    """
    cat_id = config.get("cat_id", "")
    pal    = get_palette(cat_id, config.get("arme_id", "E"))

    # Index classement national : (NOM_NORM, PRENOM_NORM) → rang
    import unicodedata, re as _re
    def _norm(s):
        s = unicodedata.normalize("NFKD", str(s).upper()).encode("ascii", "ignore").decode()
        return _re.sub(r"[^A-Z]", "", s)

    nat_idx = {}
    if df_nat is not None and not df_nat.empty:
        for _, r in df_nat.iterrows():
            k = (_norm(r.get(COL_NOM, "")), _norm(r.get(COL_PRENOM, "")))
            if k[0]:
                nat_idx[k] = int(r.get(COL_RANG, 0))

    def _rang_nat(nom, prenom, fallback):
        k = (_norm(nom), _norm(prenom))
        return nat_idx.get(k, fallback)

    def _tireurs_ges(df, fallback_start=1):
        """Filtre les GES d'un DataFrame PDF et retourne la liste tireurs."""
        if df is None or df.empty:
            return []
        ge = df[df[COL_REGION].apply(est_grand_est)].copy()
        # Réinitialiser le rang séquentiellement dans ce niveau
        ge = ge.reset_index(drop=True)
        result = []
        for i, (_, r) in enumerate(ge.iterrows(), fallback_start):
            try:
                rang_pdf = int(r.get(COL_RANG) or i)
            except (ValueError, TypeError):
                rang_pdf = i
            rang_affiche = _rang_nat(r[COL_NOM], r[COL_PRENOM], rang_pdf) or rang_pdf
            result.append({
                "rang":   f"CL NAT {rang_affiche}",
                "nom":    str(r[COL_NOM]),
                "prenom": str(r.get(COL_PRENOM, "")),
                "club":   str(r.get(COL_CLUB, "")),
                "note":   "",
            })
        def _sort_key(t):
            try: return int(str(t["rang"]).split()[-1])
            except: return 9999
        result.sort(key=_sort_key)
        return result

    tireurs_n1_all = _tireurs_ges(df_n1, fallback_start=1)
    tireurs_n2 = _tireurs_ges(df_n2, fallback_start=1)

    sections = []

    if tireurs_n1_all:
        sections.append({
            "label":   "TIREURS QUALIFIÉS — N1 (LISTE NATIONALE FFE)",
            "couleur": pal["n1"],
            "textes":  [
                "Les 32 premiers du classement national FFE + Wild Cards intégrées",
                "Compétition ouverte — sélectionnables en équipe de France",
            ],
            "tireurs": tireurs_n1_all,
            "avec_participation": True,
        })

    if tireurs_n2:
        sections.append({
            "label": "TIREURS QUALIFIÉS — N2 (LISTE NATIONALE FFE)",
            "couleur": pal["n2"],
            "textes": [
                "Compétition open — sous condition d'avoir participé à au moins une épreuve nationale",
            ],
            "tireurs": tireurs_n2,
            "avec_participation": True,
        })

    return {
        "format": "jeunes",
        "meta": {
            "region":                   "Grand Est",
            "cat_id":                   cat_id,
            "cat_label":                config.get("cat_label", cat_id),
            "competition":              config.get("competition", "Championnat de France"),
            "date":                     config.get("date", ""),
            "lieu":                     config.get("lieu", ""),
            "discipline":               config.get("discipline", ""),
            "mail_retour":              config.get("mail_retour", "administration@crege.fr"),
            "date_limite_retour":       config.get("date_limite_retour", ""),
            "date_engagement_extranet": config.get("date_engagement_extranet", ""),
            "arbitrage_config":         config.get("arbitrage_config", {}),
        },
        "sections": sections,
        "equipes_numerotees": [],
    }



# ─────────────────────────────────────────────
# FH M17/M20 Fleuret/Épée : N1+N2 PDF FFE, N3 quotas LREGE
# ─────────────────────────────────────────────

def _construire_jeunes_ffe_n1n2_n3quota(df_n1, df_n2, df_nat, df_reg, config: dict) -> dict:
    """
    N1 : liste PDF FFE (32 + WC éventuelles)
    N2 : liste PDF FFE (32 suivants + WC éventuelles)
    N3 : quotas LREGE standard (comme _construire_jeunes mais uniquement la section N3)
    """
    from ..core.utils import filtrer_df
    filtre_fr = config.get("nationalite_francaise", True)
    if df_reg is not None: df_reg = filtrer_df(df_reg, filtre_fr)
    cat_id = config.get("cat_id", "")
    pal    = get_palette(cat_id, config.get("arme_id", "E"))
    import unicodedata, re as _re
    def _norm(s):
        s = unicodedata.normalize("NFKD", str(s).upper()).encode("ascii", "ignore").decode()
        return _re.sub(r"[^A-Z]", "", s)

    nat_idx = {}
    if df_nat is not None and not df_nat.empty:
        for _, r in df_nat.iterrows():
            k = (_norm(r.get(COL_NOM, "")), _norm(r.get(COL_PRENOM, "")))
            if k[0]:
                nat_idx[k] = int(r.get(COL_RANG, 0))

    def _rang_nat(nom, prenom, fallback):
        return nat_idx.get((_norm(nom), _norm(prenom)), fallback)

    def _tireurs_niveau(df, label_rang="CL NAT"):
        """Filtre GES d'un df PDF, retourne liste triée par rang."""
        if df is None or df.empty:
            return []
        ge = df[df[COL_REGION].apply(est_grand_est)].copy()
        result = []
        for i, (_, r) in enumerate(ge.iterrows(), 1):
            try:
                rang_pdf = int(r.get(COL_RANG) or i)
            except (ValueError, TypeError):
                rang_pdf = i
            rang_aff = _rang_nat(r[COL_NOM], r[COL_PRENOM], rang_pdf) or rang_pdf
            result.append({
                "rang":   f"{label_rang} {rang_aff}",
                "nom":    str(r[COL_NOM]),
                "prenom": str(r.get(COL_PRENOM, "")),
                "club":   str(r.get(COL_CLUB, "")),
                "note":   "",
            })
        result.sort(key=lambda t: int(str(t["rang"]).split()[-1]) if str(t["rang"]).split()[-1].isdigit() else 9999)
        return result

    tireurs_n1 = _tireurs_niveau(df_n1)
    tireurs_n2 = _tireurs_niveau(df_n2)

    # Noms qualifiés N1 pour exclusion des sections N2/N3
    import unicodedata as _ud2, re as _re2
    def _norm2(s):
        s = _ud2.normalize("NFKD", str(s).upper()).encode("ascii", "ignore").decode()
        return _re2.sub(r"[^A-Z]", "", s)

    sections = []

    if tireurs_n1:
        sections.append({
            "label":   "TIREURS QUALIFIÉS — N1 (LISTE NATIONALE FFE)",
            "couleur": pal["n1"],
            "textes":  ["Les 36 premiers du classement national FFE (Wild Cards intégrées)"],
            "tireurs": tireurs_n1,
            "avec_participation": True,
        })
    if tireurs_n2:
        sections.append({
            "label":   "TIREURS QUALIFIÉS — N2 (LISTE NATIONALE FFE)",
            "couleur": pal["n2"],
            "textes":  ["Les 36 premiers du classement FFE non qualifiés en N1 (Wild Cards intégrées)"],
            "tireurs": tireurs_n2,
            "avec_participation": True,
        })

    # N3 : quotas LREGE — on délègue à _construire_jeunes en mode N3 uniquement
    # On injecte un config modifié pour ne générer que la section N3
    cfg_n3 = {**config, "quota_federal": 0, "nb_wildcards": 0}
    # Noms déjà qualifiés N1+N2 pour exclusion
    noms_deja_qualifies = {
        (_norm(t["nom"]), _norm(t.get("prenom", "")))
        for t in tireurs_n1 + tireurs_n2
    }
    cfg_n3["noms_exclus"] = noms_deja_qualifies
    # ── df_nat pour N3 : classement national réel ou synthétique depuis PDF FFE ──
    # Si pas de classement national Excel importé, reconstruire depuis df_n1 + df_n2
    if df_nat is not None and not df_nat.empty:
        df_nat_n3 = df_nat
    else:
        import pandas as pd
        frames = [f for f in [df_n1, df_n2] if f is not None and not f.empty]
        df_nat_n3 = pd.concat(frames, ignore_index=True) if frames else None

    data_n3 = _construire_jeunes(df_nat_n3, df_reg, cfg_n3) if df_reg is not None else None
    if data_n3:
        for sec in data_n3.get("sections", []):
            # Relabelliser N1→N3 et filtrer les déjà qualifiés
            sec["label"] = sec["label"].replace("N1", "N3").replace("QUOTA FÉDÉRAL", "QUOTA N3")
            sec["couleur"] = pal.get("n3", pal["n2"])
            # Exclure les tireurs déjà en N1/N2 (dans tireurs directs ET sous_sections)
            sec["tireurs"] = [
                t for t in sec.get("tireurs", [])
                if (_norm(t.get("nom","")), _norm(t.get("prenom",""))) not in noms_deja_qualifies
            ]
            for ss in sec.get("sous_sections", []):
                ss["tireurs"] = [
                    t for t in ss.get("tireurs", [])
                    if (_norm(t.get("nom","")), _norm(t.get("prenom",""))) not in noms_deja_qualifies
                ]
            # Inclure si tireurs directs OU sous_sections non vides
            has_content = bool(sec["tireurs"]) or any(ss.get("tireurs") for ss in sec.get("sous_sections", []))
            if has_content:
                sections.append(sec)

    return {
        "format": "jeunes",
        "meta": {
            "region":                   "Grand Est",
            "cat_id":                   cat_id,
            "cat_label":                config.get("cat_label", cat_id),
            "competition":              config.get("competition", "Championnat de France"),
            "date":                     config.get("date", ""),
            "lieu":                     config.get("lieu", ""),
            "discipline":               config.get("discipline", ""),
            "mail_retour":              config.get("mail_retour", "administration@crege.fr"),
            "date_limite_retour":       config.get("date_limite_retour", ""),
            "date_engagement_extranet": config.get("date_engagement_extranet", ""),
            "arbitrage_config":         config.get("arbitrage_config", {}),
        },
        "sections": sections,
        "equipes_numerotees": [],
    }


# ─────────────────────────────────────────────
# FD M17/M20 Fleuret/Épée : N1 PDF FFE, N2 quotas LREGE
# ─────────────────────────────────────────────

def _construire_jeunes_n1_ffe_n2_quota(df_n1, df_nat, df_reg, config: dict) -> dict:
    """
    N1 : liste PDF FFE (32 + WC éventuelles)
    N2 : quotas LREGE standard (comme _construire_jeunes)
    """
    from ..core.utils import filtrer_df
    filtre_fr = config.get("nationalite_francaise", True)
    if df_reg is not None: df_reg = filtrer_df(df_reg, filtre_fr)
    cat_id = config.get("cat_id", "")
    pal    = get_palette(cat_id, config.get("arme_id", "E"))
    import unicodedata as _ud3, re as _re3
    def _norm(s):
        s = _ud3.normalize("NFKD", str(s).upper()).encode("ascii", "ignore").decode()
        return _re3.sub(r"[^A-Z]", "", s)

    nat_idx = {}
    if df_nat is not None and not df_nat.empty:
        for _, r in df_nat.iterrows():
            k = (_norm(r.get(COL_NOM, "")), _norm(r.get(COL_PRENOM, "")))
            if k[0]:
                nat_idx[k] = int(r.get(COL_RANG, 0))

    def _rang_nat(nom, prenom, fallback):
        return nat_idx.get((_norm(nom), _norm(prenom)), fallback)

    def _tireurs_ges_pdf(df):
        if df is None or df.empty:
            return []
        ge = df[df[COL_REGION].apply(est_grand_est)].copy()
        result = []
        for i, (_, r) in enumerate(ge.iterrows(), 1):
            try:
                rang_pdf = int(r.get(COL_RANG) or i)
            except (ValueError, TypeError):
                rang_pdf = i
            rang_aff = _rang_nat(r[COL_NOM], r[COL_PRENOM], rang_pdf) or rang_pdf
            result.append({
                "rang":   f"CL NAT {rang_aff}",
                "nom":    str(r[COL_NOM]),
                "prenom": str(r.get(COL_PRENOM, "")),
                "club":   str(r.get(COL_CLUB, "")),
                "note":   "",
            })
        result.sort(key=lambda t: int(str(t["rang"]).split()[-1]) if str(t["rang"]).split()[-1].isdigit() else 9999)
        return result

    tireurs_n1 = _tireurs_ges_pdf(df_n1)

    sections = []

    if tireurs_n1:
        sections.append({
            "label":   "TIREUSES QUALIFIÉES — N1 (LISTE NATIONALE FFE)",
            "couleur": pal["n1"],
            "textes":  ["Les 36 premières du classement national FFE (Wild Cards intégrées)"],
            "tireurs": tireurs_n1,
            "avec_participation": True,
        })

    # N2 : quotas LREGE — on délègue à _construire_jeunes
    noms_n1 = {
        (_norm(t["nom"]), _norm(t.get("prenom", "")))
        for t in tireurs_n1
    }
    cfg_n2 = {**config, "quota_federal": 0, "nb_wildcards": 0, "noms_exclus": noms_n1}
    # df_nat pour N2 : classement national réel ou synthétique depuis PDF FFE N1
    if df_nat is not None and not df_nat.empty:
        df_nat_n2 = df_nat
    else:
        import pandas as pd
        df_nat_n2 = df_n1 if df_n1 is not None and not df_n1.empty else None
    data_n2 = _construire_jeunes(df_nat_n2, df_reg, cfg_n2) if df_reg is not None else None
    if data_n2:
        for sec in data_n2.get("sections", []):
            sec["label"] = sec["label"].replace("N1", "N2").replace("QUOTA FÉDÉRAL", "QUOTA N2")
            sec["couleur"] = pal.get("n2", pal["n1"])
            sec["tireurs"] = [
                t for t in sec.get("tireurs", [])
                if (_norm(t.get("nom","")), _norm(t.get("prenom",""))) not in noms_n1
            ]
            for ss in sec.get("sous_sections", []):
                ss["tireurs"] = [
                    t for t in ss.get("tireurs", [])
                    if (_norm(t.get("nom","")), _norm(t.get("prenom",""))) not in noms_n1
                ]
            has_content = bool(sec["tireurs"]) or any(ss.get("tireurs") for ss in sec.get("sous_sections", []))
            if has_content:
                sections.append(sec)

    return {
        "format": "jeunes",
        "meta": {
            "region":                   "Grand Est",
            "cat_id":                   cat_id,
            "cat_label":                config.get("cat_label", cat_id),
            "competition":              config.get("competition", "Championnat de France"),
            "date":                     config.get("date", ""),
            "lieu":                     config.get("lieu", ""),
            "discipline":               config.get("discipline", ""),
            "mail_retour":              config.get("mail_retour", "administration@crege.fr"),
            "date_limite_retour":       config.get("date_limite_retour", ""),
            "date_engagement_extranet": config.get("date_engagement_extranet", ""),
            "arbitrage_config":         config.get("arbitrage_config", {}),
        },
        "sections": sections,
        "equipes_numerotees": [],
    }

class M13(BaseCategorie):
    """
    M13 — Challenge de France.
    Quota fédéral et quota national désactivés par défaut.
    Texte engagement spécifique (équipes à l'initiative des clubs).
    """
    CAT_ID         = "M13"
    LABEL          = "M13"
    FORMAT         = "jeunes"
    COMPETITIONS   = ["Challenge de France"]
    NATIONALITE_FR = False

    # Options M13 désactivées par défaut
    USE_QUOTA_FED = False
    USE_QUOTA_NAT = False

    def construire(self, df_national, df_regional, config: dict) -> dict:
        # Appliquer les options M13
        use_fed = config.get("m13_use_quota_fed", self.USE_QUOTA_FED)
        use_nat = config.get("m13_use_quota_nat", self.USE_QUOTA_NAT)

        if not use_fed:
            config = {**config, "quota_federal": 0, "textes_federal": []}
        if not use_nat:
            config = {**config, "quota_crege_nat": 0}

        # Recalculer textes_crege après application des options
        qcr = config.get("quota_crege_reg", 2)
        qcn = config.get("quota_crege_nat", 0) if use_nat else 0
        config = {**config, "textes_crege": [
            f"Quota LREGE : {qcn + qcr} places réparties de la façon suivante :"
        ]}

        # M13 : si pas de classement national fourni, utiliser le régional
        df_nat_eff = df_national if df_national is not None else df_regional
        return _construire_jeunes(df_nat_eff, df_regional, config)

    def generer(self, data: dict):
        from ..generateur.excel import generer_feuille_simple
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Sélection"
        generer_feuille_simple(ws, data)
        return wb


# ─────────────────────────────────────────────
# M15
# ─────────────────────────────────────────────

class M15(BaseCategorie):
    """
    M15 — Fête des Jeunes / Épreuve de zone / 1/2 Finale.
    Mode individuel ou équipes GE1/GE2/GE3.
    """
    CAT_ID         = "M15"
    LABEL          = "M15"
    FORMAT         = "jeunes"
    COMPETITIONS   = ["Fête des Jeunes", "Épreuve de zone", "1/2 Finale"]
    NATIONALITE_FR = False

    def construire(self, df_national, df_regional, config: dict) -> dict:
        return _construire_jeunes(df_national, df_regional, config)

    def generer(self, data: dict):
        from ..generateur.excel import generer_feuille_simple
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Sélection"
        generer_feuille_simple(ws, data)
        return wb


# ─────────────────────────────────────────────
# M17
# ─────────────────────────────────────────────

class M17(BaseCategorie):
    CAT_ID         = "M17"
    LABEL          = "M17"
    FORMAT         = "jeunes"
    COMPETITIONS   = ["Championnat de France"]
    NATIONALITE_FR = True

    def construire(self, df_national, df_regional, config: dict) -> dict:
        return _construire_jeunes(df_national, df_regional, config)

    def generer(self, data: dict):
        from ..generateur.excel import generer_feuille_simple
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Sélection"
        generer_feuille_simple(ws, data)
        return wb


# ─────────────────────────────────────────────
# M20
# ─────────────────────────────────────────────

class M20(BaseCategorie):
    CAT_ID         = "M20"
    LABEL          = "M20"
    FORMAT         = "jeunes"
    COMPETITIONS   = ["Championnat de France"]
    NATIONALITE_FR = True

    def construire(self, df_national, df_regional, config: dict) -> dict:
        return _construire_jeunes(df_national, df_regional, config)

    def generer(self, data: dict):
        from ..generateur.excel import generer_feuille_simple
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Sélection"
        generer_feuille_simple(ws, data)
        return wb
