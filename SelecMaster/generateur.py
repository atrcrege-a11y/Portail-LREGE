"""
generateur.py — Génération Excel sélection Master Grand Est M11/M13
Calqué exactement sur SelecGE (couleurs, tailles, hauteurs, largeurs mesurées)
"""
import io, datetime
from openpyxl import Workbook

EXCEL_FORMAT_VERSION = "SELECMASTER_V1"   # doit correspondre à SuiviMaster/suivi.py
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palettes par arme (R1-R4) ────────────────────────────────────────────────
PALETTES_ARME = {
    "Épée":    {"titre": "1E6B3A", "clair": "D5EFDF", "info": "EAF7EE",
                "section": "27AE60", "sub": "52BE80", "pale": "F0FAF3",
                "pale2": "F5FBF7", "alt": "F0FAF3"},
    "Fleuret": {"titre": "1B3F7A", "clair": "D6E8F7", "info": "EBF2FA",
                "section": "2563A8", "sub": "4A86C8", "pale": "F0F4F9",
                "pale2": "F0F5FF", "alt": "F5F9FF"},
    "Sabre":   {"titre": "7B0C0C", "clair": "FADBD8", "info": "FDEDEC",
                "section": "C0392B", "sub": "E74C3C", "pale": "FEF5F5",
                "pale2": "FEF9F9", "alt": "FEF5F5"},
}
# Couleurs neutres (non liées à l'arme)
C_NOIR_SECT    = "4A4A4A"   # titre section remplaçants
# Couleurs statut tireurs
C_SEL          = "E2EFDA"   # vert sélectionné
C_REM          = "FCE4D6"   # orange remplaçant
C_NON          = "FFCCCC"   # rouge non sélectionnable
C_BLANC        = "FFFFFF"
C_JAUNE        = "FFF2CC"   # alerte M11 double
C_BLEU_M13OK   = "EBF3FF"   # M11 autorisé M13

NIVEAUX_ARB = '"Formation Régionale,Régionale,Formation Nationale,National,International"'
OUI_NON     = '"Oui,Non"'
MOIS_FR  = ["Janvier","Février","Mars","Avril","Mai","Juin",
            "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
JOURS_FR = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def _s(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _b(color="CCCCCC"):
    s = _s(color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _f(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, italic=False, size=9, color="000000"):
    return Font(bold=bold, italic=italic, size=size, color=color, name="Calibri")

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _cell(ws, r, c, val="", bold=False, italic=False, size=9, color="000000",
          bg=None, align="left", wrap=False, border_color="CCCCCC"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _font(bold, italic, size, color)
    if bg:
        cell.fill = _f(bg)
    cell.alignment = _align(align, wrap)
    cell.border = _b(border_color)
    return cell

def _row(ws, r, c1, c2, val="", bold=False, italic=False, size=9, color="000000",
         bg=None, align="left", height=None, wrap=False):
    """Ligne fusionnée c1→c2."""
    if c1 < c2:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=val)
    cell.font = _font(bold, italic, size, color)
    if bg:
        cell.fill = _f(bg)
    cell.alignment = _align(align, wrap)
    cell.border = _b("CCCCCC")
    if height:
        ws.row_dimensions[r].height = height
    return cell

def _date_longue(ds):
    if not ds:
        return "à préciser"
    try:
        p = ds.split("/")
        d = datetime.date(int(p[2]), int(p[1]), int(p[0]))
        return f"{JOURS_FR[d.weekday()]} {d.day} {MOIS_FR[d.month-1]} {d.year}"
    except Exception:
        return ds

def _saison():
    t = datetime.date.today()
    a = t.year if t.month >= 9 else t.year - 1
    return f"{a}-{a+1}"

# ── Feuille ───────────────────────────────────────────────────────────────────
def _feuille(ws, sel, genre_label, date_confirmation, date_extranet, tireurs_autre):
    arme      = sel.get("arme", "")
    categorie = sel.get("categorie", "")
    territoire= sel.get("territoire", "")
    quota     = sel.get("quota", 5)
    cond_part = sel.get("condition_participations", False)

    # Palette couleur selon arme
    P = PALETTES_ARME.get(arme, PALETTES_ARME["Fleuret"])
    C_BLEU_TITRE   = P["titre"]
    C_BLEU_CLAIR   = P["clair"]
    C_BLEU_INFO    = P["info"]
    C_BLEU_SECTION = P["section"]
    C_BLEU_SUB     = P["sub"]
    C_BLEU_PALE    = P["pale"]
    C_BLEU_PALE2   = P["pale2"]
    C_TIREUR_ALT   = P["alt"]

    # Largeurs exactes SelecGE : A=16, B=24, C=18, D=36, E=20, F=28
    for col, w in zip("ABCDEF", [16, 24, 18, 36, 20, 28]):
        ws.column_dimensions[col].width = w

    NC = 6  # nombre de colonnes
    r = 1

    # R1 — "LREGE Grand Est — Sélection régionale"
    _row(ws, r, 1, NC, "LREGE Grand Est — Sélection régionale",
         bold=True, size=13, color="FFFFFF", bg=C_BLEU_TITRE,
         align="center", height=26)
    r += 1

    # R2 — "MASTER GRAND EST"
    _row(ws, r, 1, NC, "MASTER GRAND EST",
         bold=True, size=11, color="FFFFFF", bg=C_BLEU_TITRE,
         align="center", height=20)
    r += 1

    # R3 — Territoire · Saison
    _row(ws, r, 1, NC, f"Territoire : {territoire}  ·  Saison {_saison()}",
         size=10, color=C_BLEU_TITRE, bg=C_BLEU_CLAIR,
         align="center", height=16)
    r += 1

    # R4 — Arme / Cat / Genre
    _row(ws, r, 1, NC, f"{arme.upper()} H/D {categorie} {genre_label.upper()}",
         bold=True, size=12, color="FFFFFF", bg=C_BLEU_TITRE,
         align="center", height=22)
    r += 1

    # R5 — note générale
    _row(ws, r, 1, NC,
         "La LREGE engage les tireurs. Les équipes sont à l'initiative des clubs ainsi que l'inscription sur l'extranet.",
         italic=True, size=9, color="444444", bg="F0F4F9",
         align="left", height=22)
    r += 1

    # R6 — confirmations
    _row(ws, r, 1, NC,
         f"📧  Confirmations avant le :    {_date_longue(date_confirmation)}",
         size=9, color=C_BLEU_TITRE, bg=C_BLEU_INFO,
         align="left", height=18)
    r += 1

    # R7 — mails (italic, bleu lien)
    _row(ws, r, 1, NC,
         "  ↳  administration@crege.fr, copie atrcrege@gmail.com",
         italic=True, size=9, color="1F6391", bg=C_BLANC,
         align="left", height=18)
    r += 1

    # R8 — clôture extranet
    _row(ws, r, 1, NC,
         f"🖥  Clôture extranet :    {_date_longue(date_extranet)}",
         size=9, color=C_BLEU_TITRE, bg=C_BLEU_INFO,
         align="left", height=18)
    r += 1

    # R9 — arbitrage
    _row(ws, r, 1, NC,
         "⚖️  Arbitrage :    Fourni par le club du tireur — 1 arbitre à partir d'un tireur sélectionné",
         size=9, color=C_BLEU_TITRE, bg=C_BLEU_INFO,
         align="left", height=18)
    r += 1

    # ── Ligne arbitre (unique, comme SelecGE) ────────────────────────────────
    dv_arb = DataValidation(type="list", formula1=NIVEAUX_ARB,
                            allow_blank=True, showDropDown=False, showErrorMessage=False)
    ws.add_data_validation(dv_arb)

    _cell(ws, r, 1, "👤  Arbitre 1", bold=True, size=9,
          color=C_BLEU_TITRE, bg=C_BLEU_INFO)
    _cell(ws, r, 2, "Nom  Prénom", size=9, color="000000", bg=C_BLANC)
    _cell(ws, r, 3, "", bg=C_BLANC)
    _cell(ws, r, 4, "Club", size=9, color="000000", bg=C_BLANC)
    c_niv = _cell(ws, r, 5, "Cliquer pour choisir le niveau ▼",
                  size=9, color="000000", bg=C_BLANC)
    dv_arb.add(c_niv)
    _cell(ws, r, 6, "", bg=C_BLANC)
    ws.row_dimensions[r].height = 22
    r += 1

    r += 1  # ligne vide (R11 dans modèle = ligne vide entre arbitres et section)

    # ── Section sélectionnés ──────────────────────────────────────────────────
    dv_yn = DataValidation(type="list", formula1=OUI_NON,
                           allow_blank=True, showDropDown=False, showErrorMessage=False)
    ws.add_data_validation(dv_yn)

    tireurs = sel.get("tireurs", [])
    selectionnes = [t for t in tireurs if t["statut"] == "selectionne"]
    remplacants  = [t for t in tireurs if t["statut"] == "remplacant"]
    non_sel      = [t for t in tireurs if t["statut"] == "non_selectionnable"]

    # Titre section sélectionnés
    _row(ws, r, 1, NC, f"  TIREURS SÉLECTIONNÉS — QUOTA MASTER",
         bold=True, size=11, color="FFFFFF", bg=C_BLEU_SECTION,
         align="left", height=20)
    r += 1

    # Note quota
    _row(ws, r, 1, NC,
         f"  ▸  Quota Master : {quota} places",
         italic=True, size=9, color=C_BLEU_TITRE, bg=C_BLEU_PALE,
         align="left", height=15)
    r += 1

    # Note condition participations (épée/fleuret)
    if cond_part:
        _row(ws, r, 1, NC,
             f"  — Condition ({arme}) : ≥ 3 épreuves du territoire pour être sélectionnable",
             bold=True, size=10, color="FFFFFF", bg=C_BLEU_SUB,
             align="left", height=15)
        r += 1

    # En-tête colonnes
    entetes = ["Rang / Classement", "Nom", "Prénom", "Club",
               "Participation\nOui / Non", "Remarque"]
    aligns  = ["center", "left", "left", "left", "center", "center"]
    for c, (h, a) in enumerate(zip(entetes, aligns), 1):
        cell = _cell(ws, r, c, h, bold=True, size=9,
                     color="000000", bg=C_BLEU_INFO, align=a, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

    # Tireurs sélectionnés (lignes alternées F5F9FF / FFFFFF)
    for i, t in enumerate(selectionnes):
        bg = C_JAUNE if t.get("alerte_m11") == "double" else              C_BLEU_M13OK if t.get("alerte_m11") == "m13only" else C_SEL

        remarque = ""
        if t.get("alerte_m11") == "double":    remarque = "⚠ Double qualif. M11+M13 — choix obligatoire"
        elif t.get("alerte_m11") == "m13only": remarque = "ℹ M11 autorisé (qualifié M13 uniquement)"

        _cell(ws, r, 1, f"M {i+1}",          size=10, color="000000", bg=bg, align="center")
        _cell(ws, r, 2, t["nom"],             bold=True, size=10, color="000000", bg=bg)
        _cell(ws, r, 3, t["prenom"],          size=10, color="000000", bg=bg)
        _cell(ws, r, 4, t["club"],            size=9,  color="000000", bg=bg)
        c_yn = _cell(ws, r, 5, None,          size=9,  color="000000", bg=bg, align="center")
        dv_yn.add(c_yn)
        _cell(ws, r, 6, remarque,             size=9,  color="000000", bg=bg)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1  # ligne vide

    # ── Section remplaçants ───────────────────────────────────────────────────
    if remplacants:
        _row(ws, r, 1, NC, "  TIREURS REMPLAÇANTS",
             bold=True, size=11, color="FFFFFF", bg=C_NOIR_SECT,
             align="left", height=20)
        r += 1

        _row(ws, r, 1, NC,
             "  ▸  Les remplaçants sont appelés dans l'ordre en cas de désistement",
             italic=True, size=9, color=C_BLEU_TITRE, bg=C_BLEU_PALE2,
             align="left", height=15)
        r += 1

        # En-tête colonnes (identique)
        for c, (h, a) in enumerate(zip(entetes, aligns), 1):
            _cell(ws, r, c, h, bold=True, size=9,
                  color="000000", bg=C_BLEU_INFO, align=a, wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1

        for i, t in enumerate(remplacants):
            rang_abs = len(selectionnes) + i + 1
            bg = C_JAUNE if t.get("alerte_m11") == "double" else                  C_BLEU_M13OK if t.get("alerte_m11") == "m13only" else C_REM

            remarque = ""
            if t.get("alerte_m11") == "double":    remarque = "⚠ Double qualif. M11+M13 — choix obligatoire"
            elif t.get("alerte_m11") == "m13only": remarque = "ℹ M11 autorisé (qualifié M13 uniquement)"

            _cell(ws, r, 1, f"M {rang_abs}",  size=10, color="000000", bg=bg, align="center")
            _cell(ws, r, 2, t["nom"],          bold=True, size=10, color="000000", bg=bg)
            _cell(ws, r, 3, t["prenom"],       size=10, color="000000", bg=bg)
            _cell(ws, r, 4, t["club"],         size=9,  color="000000", bg=bg)
            c_yn = _cell(ws, r, 5, None,       size=9,  color="000000", bg=bg, align="center")
            dv_yn.add(c_yn)
            _cell(ws, r, 6, remarque,          size=9,  color="000000", bg=bg)
            ws.row_dimensions[r].height = 20
            r += 1

        _row(ws, r, 1, NC,
             "  ▸  et suivant dans l'ordre du classement territorial",
             italic=True, size=9, color=C_BLEU_TITRE, bg=C_BLEU_PALE2,
             align="left", height=15)
        r += 1

    # ── Non sélectionnables (si épée/fleuret) ────────────────────────────────
    if non_sel and cond_part:
        r += 1
        _row(ws, r, 1, NC,
             f"  NON SÉLECTIONNABLES — Moins de 3 épreuves du territoire ({arme})",
             bold=True, size=10, color="FFFFFF", bg="595959",
             align="left", height=18)
        r += 1

        for i, t in enumerate(non_sel):
            bg = C_NON
            _cell(ws, r, 1, t["place"],         size=10, color="888888", bg=bg, align="center")
            _cell(ws, r, 2, t["nom"],            bold=True, size=10, color="888888", bg=bg)
            _cell(ws, r, 3, t["prenom"],         size=10, color="888888", bg=bg)
            _cell(ws, r, 4, t["club"],           size=9,  color="888888", bg=bg)
            _cell(ws, r, 5, f"{t['participations']}/3 épreuves",
                                                 size=9,  color="888888", bg=bg, align="center")
            _cell(ws, r, 6, "",                  bg=bg)
            ws.row_dimensions[r].height = 16
            r += 1

    ws.freeze_panes = "A6"


# ── Point d'entrée ────────────────────────────────────────────────────────────
def generer_excel(selection_h, selection_d,
                  date_retour=None,
                  date_confirmation=None,
                  date_extranet=None):
    if date_retour and not date_confirmation:
        date_confirmation = date_retour

    wb = Workbook()
    # Marqueur de version du format — contrôlé par SuiviMaster à la lecture
    wb.properties.keywords = EXCEL_FORMAT_VERSION
    wb.remove(wb.active)

    for genre_label, sel, autre in [
        ("Hommes", selection_h, selection_d),
        ("Dames",  selection_d, selection_h),
    ]:
        if not sel or not sel.get("tireurs"):
            continue
        ws = wb.create_sheet(title=genre_label)
        tireurs_autre = (autre or {}).get("tireurs", [])
        _feuille(ws, sel, genre_label, date_confirmation, date_extranet, tireurs_autre)

    if not wb.sheetnames:
        raise ValueError("Aucune donnée à générer.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
