"""
suivi.py — Suivi des confirmations, remplaçants et arbitres Master Grand Est
Persistance disque via pickle. Une session = une arme + catégorie.
"""
import os, pickle, datetime
from collections import defaultdict

_SUIVI_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".suivi_selecmaster.pkl")

# ── Structure d'un suivi ──────────────────────────────────────────────────────
# {
#   "arme": str,
#   "categorie": str,
#   "created_at": datetime,
#   "updated_at": datetime,
#   "territoires": {
#     "Alsace": {
#       "tireurs": [
#         {
#           "rang": int, "nom": str, "prenom": str, "club": str,
#           "genre": str,  # H ou D
#           "statut": str,  # "selectionne" | "remplacant" | "non_selectionnable"
#           "confirmation": str,  # "oui" | "non" | "attente"
#           "confirmation_at": datetime | None,
#           "note": str,
#           "appele": bool,  # remplaçant officiellement appelé
#         }
#       ],
#       "arbitre": {
#         "nom": str, "prenom": str, "niveau": str,
#         "club": str, "valide": bool, "note": str
#       }
#     }
#   },
#   "audit": [ {"ts": datetime, "action": str, "detail": str} ]
# }

_suivi = {}  # clé : "arme_categorie" ex: "Épée_M13"


def _cle(arme, categorie):
    return f"{arme}_{categorie}"


def _sauvegarder():
    try:
        with open(_SUIVI_FILE, "wb") as f:
            pickle.dump(_suivi, f)
    except Exception as e:
        print(f"[SUIVI] Erreur sauvegarde : {e}")


def _charger():
    global _suivi
    try:
        if os.path.isfile(_SUIVI_FILE):
            with open(_SUIVI_FILE, "rb") as f:
                _suivi = pickle.load(f)
    except Exception as e:
        print(f"[SUIVI] Erreur chargement : {e}")
        _suivi = {}


def _audit(cle, action, detail=""):
    if cle in _suivi:
        _suivi[cle]["audit"].append({
            "ts": datetime.datetime.now(),
            "action": action,
            "detail": detail,
        })


def _tireur_key(t):
    return (t["nom"].strip().upper(), t["prenom"].strip().upper(), t["genre"])


# ── API publique ──────────────────────────────────────────────────────────────

def initialiser_suivi(arme, categorie, territoire, tireurs_h, tireurs_d):
    """
    Initialise ou met à jour le suivi pour un territoire.
    Préserve les confirmations existantes si le tireur est déjà connu.
    """
    cle = _cle(arme, categorie)
    if cle not in _suivi:
        _suivi[cle] = {
            "arme": arme,
            "categorie": categorie,
            "created_at": datetime.datetime.now(),
            "updated_at": datetime.datetime.now(),
            "territoires": {},
            "audit": [],
        }

    # Construire liste tireurs avec préservation confirmations existantes
    existants = {}
    if territoire in _suivi[cle]["territoires"]:
        for t in _suivi[cle]["territoires"][territoire].get("tireurs", []):
            existants[_tireur_key(t)] = t

    tireurs = []
    rang_sel_h = rang_sel_d = rang_rem_h = rang_rem_d = 0

    for genre, liste in [("H", tireurs_h), ("D", tireurs_d)]:
        for t in liste:
            if t["statut"] == "selectionne":
                if genre == "H": rang_sel_h += 1
                else: rang_sel_d += 1
                rang = rang_sel_h if genre == "H" else rang_sel_d
            elif t["statut"] == "remplacant":
                if genre == "H": rang_rem_h += 1
                else: rang_rem_d += 1
                rang = rang_rem_h if genre == "H" else rang_rem_d
            else:
                rang = t.get("place", 0)

            key = (t["nom"].strip().upper(), t["prenom"].strip().upper(), genre)
            existant = existants.get(key, {})

            tireurs.append({
                "rang":           rang,
                "nom":            t["nom"],
                "prenom":         t["prenom"],
                "club":           t["club"],
                "genre":          genre,
                "statut":         t["statut"],
                "alerte_m11":     t.get("alerte_m11"),
                "confirmation":   existant.get("confirmation", "attente"),
                "confirmation_at":existant.get("confirmation_at"),
                "appele":         existant.get("appele", False),
                "note":           existant.get("note", ""),
            })

    # Préserver arbitre existant
    arbitre_existant = _suivi[cle]["territoires"].get(territoire, {}).get("arbitre", {
        "nom": "", "prenom": "", "niveau": "", "club": "", "valide": False, "note": ""
    })

    _suivi[cle]["territoire_organisateur"] = _suivi[cle].get("territoire_organisateur", "")
    _suivi[cle]["territoires"][territoire] = {
        "tireurs": tireurs,
        "arbitre": arbitre_existant,
    }
    _suivi[cle]["updated_at"] = datetime.datetime.now()
    _audit(cle, "IMPORT", f"Territoire {territoire} importé — {len(tireurs)} tireurs")
    _sauvegarder()


def get_suivi(arme, categorie):
    """Retourne le suivi complet pour une arme/catégorie."""
    return _suivi.get(_cle(arme, categorie))


def get_tous_suivis():
    """Retourne la liste de tous les suivis actifs."""
    return {k: {
        "arme": v["arme"],
        "categorie": v["categorie"],
        "updated_at": v["updated_at"].strftime("%d/%m/%Y %H:%M"),
        "nb_territoires": len(v["territoires"]),
    } for k, v in _suivi.items()}


def mettre_a_jour_confirmation(arme, categorie, territoire, nom, prenom, genre, confirmation, note=""):
    """
    Met à jour la confirmation d'un tireur.
    confirmation : "oui" | "non" | "attente"
    """
    VALEURS_OK = {"oui", "non", "attente"}
    if confirmation not in VALEURS_OK:
        raise ValueError(f"Confirmation invalide : {confirmation}. Valeurs acceptées : {VALEURS_OK}")

    cle = _cle(arme, categorie)
    if cle not in _suivi:
        raise KeyError(f"Aucun suivi pour {arme} {categorie}")
    if territoire not in _suivi[cle]["territoires"]:
        raise KeyError(f"Territoire {territoire} non chargé")

    tireurs = _suivi[cle]["territoires"][territoire]["tireurs"]
    key = (nom.strip().upper(), prenom.strip().upper(), genre)
    trouve = False

    for t in tireurs:
        if _tireur_key(t) == key:
            ancien = t["confirmation"]
            t["confirmation"]    = confirmation
            t["confirmation_at"] = datetime.datetime.now()
            t["note"]            = note
            trouve = True
            _audit(cle, "CONFIRMATION",
                   f"{nom} {prenom} ({genre}) {territoire} : {ancien} → {confirmation}"
                   + (f" | note: {note}" if note else ""))
            break

    if not trouve:
        raise KeyError(f"Tireur {nom} {prenom} ({genre}) introuvable dans {territoire}")

    _suivi[cle]["updated_at"] = datetime.datetime.now()
    _sauvegarder()


def marquer_appele(arme, categorie, territoire, nom, prenom, genre, appele=True):
    """Marque un remplaçant comme officiellement appelé."""
    cle = _cle(arme, categorie)
    tireurs = _suivi[cle]["territoires"][territoire]["tireurs"]
    key = (nom.strip().upper(), prenom.strip().upper(), genre)

    for t in tireurs:
        if _tireur_key(t) == key:
            t["appele"] = appele
            _audit(cle, "APPEL_REMPLACANT",
                   f"{nom} {prenom} ({genre}) {territoire} : appele={appele}")
            break

    _suivi[cle]["updated_at"] = datetime.datetime.now()
    _sauvegarder()


def mettre_a_jour_arbitre(arme, categorie, territoire, nom, prenom, niveau, club, valide=False, note=""):
    """Met à jour l'arbitre proposé par un club/territoire."""
    NIVEAUX_OK = {"Formation Régionale", "Régionale", "Formation Nationale",
                  "National", "International", ""}
    if niveau and niveau not in NIVEAUX_OK:
        raise ValueError(f"Niveau invalide : {niveau}")

    cle = _cle(arme, categorie)
    if cle not in _suivi:
        raise KeyError(f"Aucun suivi pour {arme} {categorie}")

    ancien = _suivi[cle]["territoires"].get(territoire, {}).get("arbitre", {})
    _suivi[cle]["territoires"].setdefault(territoire, {"tireurs": [], "arbitre": {}})
    _suivi[cle]["territoires"][territoire]["arbitre"] = {
        "nom": nom, "prenom": prenom, "niveau": niveau,
        "club": club, "valide": valide, "note": note,
    }
    _audit(cle, "ARBITRE",
           f"{territoire} : {ancien.get('nom','')} → {nom} {prenom} ({niveau})"
           + (" ✓ validé" if valide else ""))

    _suivi[cle]["updated_at"] = datetime.datetime.now()
    _sauvegarder()


def importer_excel_retour(arme, categorie, territoire, contenu_bytes):
    """
    Importe un Excel retourné par un club.
    Lit la colonne Participation (Oui/Non) et le bloc arbitre.
    Retourne un dict résumé des modifications.
    """
    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(contenu_bytes), data_only=True)
    modifications = {"confirmations": [], "arbitre": None, "erreurs": []}

    cle = _cle(arme, categorie)
    if cle not in _suivi or territoire not in _suivi[cle]["territoires"]:
        raise KeyError(f"Suivi {arme} {categorie} {territoire} non initialisé")

    tireurs_suivi = {
        _tireur_key(t): t
        for t in _suivi[cle]["territoires"][territoire]["tireurs"]
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        genre = "H" if sheet_name == "Hommes" else "D" if sheet_name == "Dames" else None
        if not genre:
            continue

        arbitre_data = {"nom": "", "prenom": "", "club": "", "niveau": ""}
        in_arbitre = False

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue

            val0 = str(row[0]).strip()

            # Détecter ligne arbitre
            if "Arbitre" in val0:
                in_arbitre = True
                # Col B=nom prénom, Col D=club, Col E=niveau
                nom_prenom = str(row[1]).strip() if row[1] else ""
                club_arb   = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                niveau_arb = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                # Nettoyer "Cliquer pour choisir le niveau ▼"
                if "Cliquer" in niveau_arb:
                    niveau_arb = ""
                # Séparer Nom Prénom si possible (prénom en italique dans Excel, on prend tout)
                arbitre_data = {
                    "nom": nom_prenom, "prenom": "", "club": club_arb, "niveau": niveau_arb
                }
                continue

            # Détecter ligne tireur : col A = "M 1", "M 2"...
            if val0.startswith("M ") and len(val0) <= 5:
                nom    = str(row[1]).strip() if row[1] else ""
                prenom = str(row[2]).strip() if row[2] else ""
                conf   = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ""

                if conf in ("oui", "non"):
                    key = (nom.upper(), prenom.upper(), genre)
                    if key in tireurs_suivi:
                        ancien = tireurs_suivi[key]["confirmation"]
                        tireurs_suivi[key]["confirmation"]    = conf
                        tireurs_suivi[key]["confirmation_at"] = datetime.datetime.now()
                        modifications["confirmations"].append({
                            "nom": nom, "prenom": prenom, "genre": genre,
                            "ancien": ancien, "nouveau": conf,
                        })
                    else:
                        modifications["erreurs"].append(
                            f"Tireur introuvable : {nom} {prenom} ({genre})"
                        )

        if arbitre_data["nom"] and arbitre_data["nom"] not in ("Nom  Prénom", ""):
            modifications["arbitre"] = arbitre_data
            _suivi[cle]["territoires"][territoire]["arbitre"] = {
                "nom": arbitre_data["nom"], "prenom": arbitre_data["prenom"],
                "niveau": arbitre_data["niveau"], "club": arbitre_data["club"],
                "valide": False, "note": "Importé automatiquement",
            }

    nb = len(modifications["confirmations"])
    _audit(cle, "IMPORT_EXCEL",
           f"{territoire} : {nb} confirmation(s) importée(s)"
           + (f" | arbitre: {modifications['arbitre']}" if modifications["arbitre"] else "")
           + (f" | erreurs: {modifications['erreurs']}" if modifications["erreurs"] else ""))

    _suivi[cle]["updated_at"] = datetime.datetime.now()
    _sauvegarder()
    return modifications


ORDRE_PIOCHE = {
    "Champagne-Ardenne": ["Champagne-Ardenne", "Lorraine",          "Alsace"],
    "Lorraine":          ["Lorraine",          "Alsace",            "Champagne-Ardenne"],
    "Alsace":            ["Alsace",            "Champagne-Ardenne", "Lorraine"],
}


def get_ordre_territoires(territoire_organisateur):
    """Retourne l'ordre de sollicitation des territoires selon l'organisateur."""
    return ORDRE_PIOCHE.get(territoire_organisateur,
                            ["Champagne-Ardenne", "Lorraine", "Alsace"])


EFFECTIF_MAX = 16  # max H+D confondus par arme/catégorie


def get_remplacants_a_appeler(arme, categorie, territoire_organisateur=None):
    """
    Retourne les remplaçants à appeler par territoire.
    Contrainte : effectif total confirmés (oui + appelés) ≤ 16 H+D confondus.
    Si remplaçants d'un territoire épuisés → pioche dans l'ordre inter-territorial.
    """
    cle = _cle(arme, categorie)
    if cle not in _suivi:
        return {}

    territoires_data = _suivi[cle]["territoires"]
    orga = territoire_organisateur or _suivi[cle].get("territoire_organisateur", "Champagne-Ardenne")
    ordre = get_ordre_territoires(orga)

    # ── Calculer l'effectif actuel confirmé (oui + remplaçants appelés) ──────
    def compter_effectif_actuel():
        total = 0
        for td in territoires_data.values():
            for t in td["tireurs"]:
                if t["statut"] == "selectionne" and t["confirmation"] == "oui":
                    total += 1
                elif t["statut"] == "remplacant" and t["appele"] and t["confirmation"] != "non":
                    total += 1
        return total

    resultats = {}

    for territoire_source in ordre:
        if territoire_source not in territoires_data:
            continue
        tireurs = territoires_data[territoire_source]["tireurs"]
        a_appeler = []

        for genre in ["H", "D"]:
            sel         = [t for t in tireurs if t["genre"] == genre and t["statut"] == "selectionne"]
            rem_propres = [t for t in tireurs if t["genre"] == genre and t["statut"] == "remplacant"]

            refus_count   = sum(1 for t in sel if t["confirmation"] == "non")
            # Appelés ET confirmés (pas refus) — les appelés qui disent Non libèrent une place
            appeles_ok    = sum(1 for t in rem_propres if t["appele"] and t["confirmation"] != "non")
            appeles_non   = sum(1 for t in rem_propres if t["appele"] and t["confirmation"] == "non")
            dispo_propres = [t for t in rem_propres if not t["appele"]]
            # Manquants = refus sélectionnés + refus remplaçants appelés - remplaçants OK
            manquants     = refus_count + appeles_non - appeles_ok
            appeles       = appeles_ok  # pour le calcul de position

            if manquants <= 0:
                continue

            # Vérifier la contrainte effectif max avant chaque appel
            def places_disponibles():
                return max(0, EFFECTIF_MAX - compter_effectif_actuel())

            # Remplir avec remplaçants du territoire source
            for i in range(min(manquants, len(dispo_propres))):
                if places_disponibles() <= 0:
                    break
                a_appeler.append({
                    "genre":             genre,
                    "remplacant":        dispo_propres[i],
                    "territoire_rem":    territoire_source,
                    "position":          appeles + i + 1,
                    "inter_territorial": False,
                })
            manquants_restants = manquants - min(manquants, len(dispo_propres))

            # Pioche inter-territoriale si remplaçants propres épuisés
            if manquants_restants > 0:
                for territoire_autre in ordre:
                    if territoire_autre == territoire_source:
                        continue
                    if territoire_autre not in territoires_data:
                        continue
                    if places_disponibles() <= 0:
                        break
                    rem_autres = [
                        t for t in territoires_data[territoire_autre]["tireurs"]
                        if t["genre"] == genre
                        and t["statut"] == "remplacant"
                        and not t["appele"]
                    ]
                    for i in range(min(manquants_restants, len(rem_autres))):
                        if places_disponibles() <= 0:
                            break
                        a_appeler.append({
                            "genre":             genre,
                            "remplacant":        rem_autres[i],
                            "territoire_rem":    territoire_autre,
                            "position":          appeles + len(dispo_propres) + i + 1,
                            "inter_territorial": True,
                        })
                    manquants_restants -= min(manquants_restants, len(rem_autres))
                    if manquants_restants <= 0:
                        break

        if a_appeler:
            resultats[territoire_source] = a_appeler

    return resultats


def get_audit(arme, categorie):
    cle = _cle(arme, categorie)
    if cle not in _suivi:
        return []
    return [
        {
            "ts":     e["ts"].strftime("%d/%m/%Y %H:%M:%S"),
            "action": e["action"],
            "detail": e["detail"],
        }
        for e in reversed(_suivi[cle]["audit"])
    ]


def supprimer_suivi(arme, categorie):
    cle = _cle(arme, categorie)
    if cle in _suivi:
        del _suivi[cle]
        _sauvegarder()


def get_stats(arme, categorie):
    """Retourne les statistiques consolidées tous territoires."""
    cle = _cle(arme, categorie)
    if cle not in _suivi:
        return None

    stats = {
        "arme": arme, "categorie": categorie,
        "territoires": {},
        "total": {"sel": 0, "oui": 0, "non": 0, "attente": 0,
                  "arbitres_proposes": 0, "arbitres_valides": 0},
    }

    for territoire, data in _suivi[cle]["territoires"].items():
        t_stats = {"H": {}, "D": {}}
        for genre in ["H", "D"]:
            tireurs = [t for t in data["tireurs"]
                       if t["genre"] == genre and t["statut"] == "selectionne"]
            rem     = [t for t in data["tireurs"]
                       if t["genre"] == genre and t["statut"] == "remplacant"]
            t_stats[genre] = {
                "sel":     len(tireurs),
                "oui":     sum(1 for t in tireurs if t["confirmation"] == "oui"),
                "non":     sum(1 for t in tireurs if t["confirmation"] == "non"),
                "attente": sum(1 for t in tireurs if t["confirmation"] == "attente"),
                "rem":     len(rem),
                "rem_appeles": sum(1 for t in rem if t["appele"]),
                "rem_confirmes": sum(1 for t in rem if t["appele"] and t["confirmation"] == "oui"),
            }
            stats["total"]["sel"]     += t_stats[genre]["sel"]
            stats["total"]["oui"]     += t_stats[genre]["oui"]
            stats["total"]["non"]     += t_stats[genre]["non"]
            stats["total"]["attente"] += t_stats[genre]["attente"]

        arb = data.get("arbitre", {})
        a_props = bool(arb.get("nom") and arb["nom"] not in ("", "Nom  Prénom"))
        a_val   = a_props and arb.get("valide", False)
        stats["total"]["arbitres_proposes"] += int(a_props)
        stats["total"]["arbitres_valides"]  += int(a_val)

        stats["territoires"][territoire] = {
            "H": t_stats["H"],
            "D": t_stats["D"],
            "arbitre": arb,
            "arbitre_propose": a_props,
            "arbitre_valide": a_val,
        }

    return stats


def rafraichir_alertes_m11(arme, categorie, noms_m11_par_genre, flag_m13_par_genre):
    """
    Recalcule alerte_m11 des tireurs d'un suivi DÉJÀ créé, à partir des données
    déjà importées (cache), SANS réimport.
      noms_m11_par_genre : {"H": set((NOM, PRENOM)), "D": set(...)} présents au classement M11
      flag_m13_par_genre : {"H": {(NOM, PRENOM): bool}, "D": {...}} est_m11_dans_m13
    Présence dans le classement M11 → "double" ; sinon M11 par l'âge → "m13only" ; sinon None.
    Retourne {"ok": True, "modifications": n} ou {"erreur": "..."}.
    """
    cle = _cle(arme, categorie)
    if cle not in _suivi:
        return {"erreur": f"Aucun suivi pour {arme} {categorie}"}

    modifs = 0
    for territoire, td in _suivi[cle]["territoires"].items():
        for t in td["tireurs"]:
            g = t["genre"]
            key = (t["nom"].strip().upper(), t["prenom"].strip().upper())
            if key in noms_m11_par_genre.get(g, set()):
                nouvelle = "double"
            elif flag_m13_par_genre.get(g, {}).get(key, False):
                nouvelle = "m13only"
            else:
                nouvelle = None
            if t.get("alerte_m11") != nouvelle:
                t["alerte_m11"] = nouvelle
                modifs += 1

    _suivi[cle]["updated_at"] = datetime.datetime.now()
    _audit(cle, "RAFRAICHIR_M11", f"{modifs} alerte(s) M11 recalculée(s) depuis le cache")
    _sauvegarder()
    return {"ok": True, "modifications": modifs}


# Chargement au démarrage
_charger()
