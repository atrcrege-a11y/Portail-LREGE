"""Exports PDF, Excel, iCal pour le module Calendrier LREGE"""
import json, io
from datetime import datetime, date

# ── iCal ──────────────────────────────────────────────────────────────────────
def build_ical(events):
    from icalendar import Calendar, Event as IEvent
    cal = Calendar()
    cal.add('prodid', '-//LREGE//Calendrier//FR')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('x-wr-calname', 'Calendrier LREGE')
    cal.add('x-wr-timezone', 'Europe/Paris')

    NIVEAUX = {'international':'🌍','national':'🇫🇷','zone':'🗺','regional':'🏅','departemental':'📍','club':'🏠'}
    TYPE_FR = {'competition':'Compétition','stage':'Stage','formation_arbitrage':'Formation arbitrage',
               'formation_animateur':'Formation animateur','formation_cadre':'Formation cadre'}

    for e in events:
        ev = IEvent()
        emoji = NIVEAUX.get(e.get('niveau',''),'')
        ev.add('summary', f"{emoji} {e['intitule']}")
        d0 = datetime.strptime(e['date_debut'], '%Y-%m-%d').date()
        ev.add('dtstart', d0)
        if e.get('date_fin') and e['date_fin'] != e['date_debut']:
            d1 = datetime.strptime(e['date_fin'], '%Y-%m-%d').date()
            from datetime import timedelta
            ev.add('dtend', d1 + timedelta(days=1))
        else:
            from datetime import timedelta
            ev.add('dtend', d0 + timedelta(days=1))
        desc = f"Niveau: {e.get('niveau_raw') or e.get('niveau','')}\n"
        desc += f"Type: {TYPE_FR.get(e.get('type_evenement','competition'),'')}\n"
        if e.get('arme'): desc += f"Arme: {e['arme']} {e.get('sexe','')}\n"
        if e.get('categories'): desc += f"Catégories: {', '.join(e['categories'])}\n"
        if e.get('notes'): desc += f"Notes: {e['notes']}\n"
        if e.get('url'): desc += f"Lien: {e['url']}"
        ev.add('description', desc)
        ev.add('location', e.get('lieu',''))
        ev.add('uid', e['id'] + '@lrege.fr')
        cal.add_component(ev)
    return cal.to_ical()

# ── Excel ─────────────────────────────────────────────────────────────────────
def build_excel(events):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BLEU = '1F3864'; BLEU_CLAIR = '2E75B6'; OR = 'C9A84C'
    NIVEAUX_FR = {'international':'International','national':'National','zone':'Zone',
                  'regional':'Régional','departemental':'Départemental','club':'Club'}
    TYPE_FR = {'competition':'Compétition','stage':'Stage','formation_arbitrage':'Formation arbitrage',
               'formation_animateur':'Formation animateur','formation_cadre':'Formation cadre'}
    COULEURS_NIV = {'international':'FFF3CD','national':'CFE2FF','zone':'E2D9F3',
                    'regional':'D1E7DD','departemental':'F8D7DA','club':'E2E3E5'}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Calendrier LREGE'

    # En-tête titre
    ws.merge_cells('A1:J1')
    ws['A1'] = f'Calendrier LREGE — Export du {datetime.now().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=BLEU)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # En-têtes colonnes
    headers = ['Date début','Date fin','Type','Niveau','Intitulé','Lieu','Arme','Catégories','Grand Est','Statut']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=BLEU_CLAIR)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 20

    thin = Border(
        left=Side(style='thin', color='D5DCE4'), right=Side(style='thin', color='D5DCE4'),
        top=Side(style='thin', color='D5DCE4'), bottom=Side(style='thin', color='D5DCE4')
    )

    for i, e in enumerate(events, 3):
        niv = e.get('niveau','')
        fill_color = COULEURS_NIV.get(niv, 'FFFFFF')
        row_data = [
            e.get('date_debut',''), e.get('date_fin',''),
            TYPE_FR.get(e.get('type_evenement','competition'),'Compétition'),
            NIVEAUX_FR.get(niv, niv),
            e.get('intitule',''), e.get('lieu',''),
            f"{e.get('arme','')} {e.get('sexe','')}".strip(),
            ', '.join(e.get('categories',[])),
            'Oui' if e.get('grand_est') else '',
            e.get('statut',''),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=fill_color)
            c.border = thin
            c.alignment = Alignment(vertical='center', wrap_text=(col==5))
        ws.row_dimensions[i].height = 16

    # Largeurs colonnes
    widths = [12,12,16,14,42,22,14,28,10,22]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Feuille légende
    ws2 = wb.create_sheet('Légende')
    ws2['A1'] = 'Légende des couleurs'; ws2['A1'].font = Font(bold=True, size=11)
    legends = [('International','FFF3CD'),('National','CFE2FF'),('Zone','E2D9F3'),
               ('Régional','D1E7DD'),('Départemental','F8D7DA'),('Club','E2E3E5')]
    for i,(label,color) in enumerate(legends, 2):
        ws2.cell(row=i, column=1, value=label).font = Font(name='Arial', size=10)
        ws2.cell(row=i, column=2).fill = PatternFill('solid', fgColor=color)
        ws2.cell(row=i, column=2, value='          ')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

# ── PDF ───────────────────────────────────────────────────────────────────────
def build_pdf(events):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    BLEU = colors.HexColor('#1F3864'); BLEU_CLAIR = colors.HexColor('#2E75B6')
    OR = colors.HexColor('#C9A84C')
    NIVEAUX_FR = {'international':'Intl','national':'Nat.','zone':'Zone',
                  'regional':'Rég.','departemental':'Dép.','club':'Club'}
    TYPE_FR = {'competition':'Compét.','stage':'Stage','formation_arbitrage':'Form. arb.',
               'formation_animateur':'Form. anim.','formation_cadre':'Form. cadre'}
    NIV_COLORS = {
        'international': colors.HexColor('#FFF3CD'),
        'national': colors.HexColor('#CFE2FF'),
        'zone': colors.HexColor('#E2D9F3'),
        'regional': colors.HexColor('#D1E7DD'),
        'departemental': colors.HexColor('#F8D7DA'),
        'club': colors.HexColor('#E2E3E5'),
    }

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    # Titre
    title_style = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=16,
        textColor=colors.white, backColor=BLEU, alignment=TA_CENTER, spaceAfter=4,
        borderPadding=(8,8,8,8))
    sub_style = ParagraphStyle('sub', fontName='Helvetica', fontSize=9,
        textColor=colors.HexColor('#6B7A99'), alignment=TA_CENTER, spaceAfter=10)

    story.append(Paragraph('Calendrier LREGE', title_style))
    story.append(Paragraph(f"Export du {datetime.now().strftime('%d/%m/%Y %H:%M')} — {len(events)} événement(s)", sub_style))
    story.append(Spacer(1, 4*mm))

    # Grouper par mois
    MOIS_FR = ['','Janvier','Février','Mars','Avril','Mai','Juin',
               'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

    month_style = ParagraphStyle('month', fontName='Helvetica-Bold', fontSize=10,
        textColor=BLEU_CLAIR, spaceAfter=2, spaceBefore=6)
    cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=7.5, leading=10)
    bold_style = ParagraphStyle('bold', fontName='Helvetica-Bold', fontSize=7.5, leading=10)

    by_month = {}
    for e in events:
        y, m, _ = e['date_debut'].split('-')
        key = (int(y), int(m))
        by_month.setdefault(key, []).append(e)

    for (y, m) in sorted(by_month):
        story.append(Paragraph(f"{MOIS_FR[m]} {y}", month_style))
        story.append(HRFlowable(width='100%', thickness=0.5, color=BLEU_CLAIR, spaceAfter=2))

        col_widths = [20*mm, 14*mm, 15*mm, 14*mm, 62*mm, 30*mm, 20*mm]
        data = [[
            Paragraph('<b>Date</b>', bold_style),
            Paragraph('<b>Niveau</b>', bold_style),
            Paragraph('<b>Type</b>', bold_style),
            Paragraph('<b>Arme</b>', bold_style),
            Paragraph('<b>Intitulé</b>', bold_style),
            Paragraph('<b>Lieu</b>', bold_style),
            Paragraph('<b>Catégories</b>', bold_style),
        ]]
        row_colors = [colors.HexColor('#E8ECF4')]

        for e in by_month[(y, m)]:
            d0 = e['date_debut'][8:] + '/' + e['date_debut'][5:7]
            df = e.get('date_fin','')
            if df and df != e['date_debut']:
                d0 += '→' + df[8:] + '/' + df[5:7]
            niv = e.get('niveau','')
            cats = ', '.join(e.get('categories',[])[:4])
            arme_txt = e.get('arme','') + (' '+e.get('sexe','') if e.get('sexe') else '')
            data.append([
                Paragraph(d0, cell_style),
                Paragraph(NIVEAUX_FR.get(niv, niv), cell_style),
                Paragraph(TYPE_FR.get(e.get('type_evenement','competition'),''), cell_style),
                Paragraph(arme_txt, cell_style),
                Paragraph(e.get('intitule',''), cell_style),
                Paragraph(e.get('lieu',''), cell_style),
                Paragraph(cats, cell_style),
            ])
            row_colors.append(NIV_COLORS.get(niv, colors.white))

        t = Table(data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E8ECF4')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 7.5),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#D5DCE4')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ]
        for i, c in enumerate(row_colors[1:], 1):
            style_cmds.append(('BACKGROUND', (0,i), (-1,i), c))
        t.setStyle(TableStyle(style_cmds))
        story.append(t)
        story.append(Spacer(1, 3*mm))

    doc.build(story)
    buf.seek(0)
    return buf.read()
