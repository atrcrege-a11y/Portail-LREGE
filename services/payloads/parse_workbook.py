"""
services/payloads/parse_workbook.py — Parser générique des classeurs
Excel générés par SelecGE (openpyxl Workbook, en mémoire ou fichier).

Le parser est PILOTÉ PAR LE CONTENU (bannières de section, lignes
d'en-tête de colonnes), pas par des positions de lignes fixes : il
tolère les variations d'en-tête de document (ancien layout « GRAND EST — »
3 lignes vs layout refonte « CHAMPIONNAT … / date · lieu / discipline »).

Sortie par feuille : liste de sections
    {"label": str brut, "mode": "tireur"|"tireur_sans_rang"|"equipe",
     "rows": [dict]}
rows tireur           : {rang, nom, prenom, club}
rows tireur_sans_rang : {nom, prenom, club}          (listes open F/S)
rows equipe           : {rang, nom_equipe, club}     (club = Ligue ou Club)
"""
import re
import unicodedata


def _norm(s):
    """Majuscules sans accents, espaces normalisés."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().upper()


def _val(cell):
    v = cell.value
    return str(v).strip() if v is not None else ""


# Bannières de section reconnues (texte normalisé, sans accents).
_SECTION_PREFIXES = (
    "TIREURS QUALIFIES",     # N1 FFE / QUOTA LREGE (indiv epee)
    "TIREURS REMPLACANTS",
    "TIREURS GRAND EST",     # open avec liste (epee D/V4, F/S reference)
    "TIREUSES GRAND EST",
    "EPREUVE OPEN",          # bandeau open (avec ou sans liste)
    "EQUIPES ",              # equipes vet (N1/N2, N3, open, F/S)
)
_RE_SOUS_SECTION = re.compile(r"^—?\s*SUR CLASSEMENT (NATIONAL|REGIONAL)")


def _est_banniere(txt_norm):
    if _RE_SOUS_SECTION.match(txt_norm):
        return True
    return any(txt_norm.startswith(p) for p in _SECTION_PREFIXES)


def _mode_entete(vals):
    """Détecte une ligne d'en-tête de colonnes ; retourne le mode ou None."""
    a, b = _norm(vals[0]), _norm(vals[1])
    if a.startswith("RANG"):
        if b.startswith("EQUIPE"):
            return "equipe"
        return "tireur"
    if a == "NOM" and b.startswith("PRENOM"):
        return "tireur_sans_rang"
    if a == "N°" or a == "N" or a.startswith("N°"):
        # équipes F/S open (placeholders) : "N° | Nom de l'equipe | Club"
        if "EQUIPE" in b or "NOM" in b:
            return "equipe"
    return None


def _parse_feuille(ws, max_col=6):
    sections = []
    courante = None          # section en cours
    quota_parent = None      # label de la section quota (pour sous-sections)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
        vals = [_val(c) for c in row]
        if not any(vals):
            continue
        a_brut = vals[0]
        a_norm = _norm(a_brut)

        # Lignes info "▸ ..." : ignorées
        if a_brut.lstrip().startswith("▸"):
            continue

        # Bannière de section ?
        if a_norm and not any(vals[1:]) and _est_banniere(a_norm):
            label = re.sub(r"\s+", " ", a_brut).strip()
            if _RE_SOUS_SECTION.match(a_norm):
                # Sous-section d'un quota : label composé
                parent = quota_parent or (courante["label"] if courante else "")
                label = f"{parent} · {label.lstrip('— ').strip()}"
            elif a_norm.startswith("TIREURS QUALIFIES") and "QUOTA" in a_norm:
                quota_parent = label
            else:
                quota_parent = None
            courante = {"label": label, "mode": None, "rows": []}
            sections.append(courante)
            continue

        # En-tête de colonnes ?
        mode = _mode_entete(vals)
        if mode:
            if courante is None:   # en-tête sans bannière : feuille inattendue
                courante = {"label": "", "mode": None, "rows": []}
                sections.append(courante)
            courante["mode"] = mode
            continue

        # Ligne de données ?
        if courante is None or courante["mode"] is None:
            continue   # en-tête de document (lignes 1-9) ou bruit

        m = courante["mode"]
        if m == "tireur":
            nom = vals[1]
            if not nom:
                continue
            courante["rows"].append({
                "rang": vals[0], "nom": nom, "prenom": vals[2], "club": vals[3],
            })
        elif m == "tireur_sans_rang":
            nom = vals[0]
            if not nom:
                continue
            courante["rows"].append({
                "nom": nom, "prenom": vals[1], "club": vals[2],
            })
        elif m == "equipe":
            nom_eq = vals[1]
            # Placeholders F/S : lignes composition "1 | Nom | Prenom" et
            # "Eq. N" sans nom d'équipe → ignorés.
            if not nom_eq or _norm(nom_eq) in ("NOM", "NOM PRENOM"):
                continue
            courante["rows"].append({
                "rang": vals[0], "nom_equipe": nom_eq, "club": vals[2],
            })

    return sections


def parse_workbook(wb, max_col=6):
    """Parse toutes les feuilles d'un Workbook openpyxl (ou chemin xlsx).

    Retourne {nom_feuille: [sections]}.
    """
    if isinstance(wb, str):
        from openpyxl import load_workbook
        wb = load_workbook(wb, data_only=True)
    return {ws.title: _parse_feuille(ws, max_col=max_col) for ws in wb.worksheets}
