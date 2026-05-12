"""
core/excel_commun.py
Feuilles Excel communes à toutes les compétitions :
  - Extranet (indiv et équipe)
  - Arbitres
"""

from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as _PF

from .excel_base import sc, mk_font, mk_fill, mk_align, border_all, gcl
from .config import BAREME_ARBITRES
from .parser import date_avec_jour, parse_date_key


# ── Feuille Extranet (indiv ou équipe)

ARMES_LABEL = {"F": "Fleuret", "E": "Épée", "S": "Sabre"}
ARMES_ORDER = ["F", "E", "S"]
ARMES_COLOR = {"F": "DDEEFF", "E": "DDEEDD", "S": "FFEECC"}

def feuille_extranet(ws, fichiers_list, categories, cat_label,
                     cat_map, is_equipe=False, par_arme=False):
    """
    Génère une feuille Extranet : liste des tireurs par catégorie/sexe.
    Si par_arme=True, regroupe d'abord par arme (bandeau), puis par catégorie.

    Paramètres :
      categories : liste ordonnée des clés catégories à afficher
      cat_label  : {cat_key: libellé}
      cat_map    : {cat_xml → cat_key} pour normaliser les catégories du XML
      par_arme   : si True, section par arme avant les catégories
    """
    # data_by_arme_cat[arme_code][cat_key] = {"H": [...], "D": [...]}
    data_by_arme_cat = defaultdict(lambda: defaultdict(lambda: {"H": [], "D": []}))
    data_by_cat      = defaultdict(lambda: {"H": [], "D": []})  # fallback sans arme

    for meta, tireurs, arbitres in fichiers_list:
        cat_raw    = meta["categorie"].upper()
        type_c     = meta["type"]
        arme_code  = meta.get("arme", "?")
        sexe_epreuve = meta.get("sexe", "").upper()
        if is_equipe and type_c != "E":
            continue
        if not is_equipe and type_c == "E":
            continue
        cat_key = cat_map.get(cat_raw)
        if cat_key is None:
            continue
        for t in tireurs:
            sexe = "H" if t["sexe"].upper() == "M" else "D"
            # Mixte : on met tous dans H pour l'extranet
            if sexe_epreuve == "MF" and is_equipe:
                sexe = "H"
            if par_arme:
                data_by_arme_cat[arme_code][cat_key][sexe].append(t)
            else:
                data_by_cat[cat_key][sexe].append(t)

    GAP = 5; COL_H = 1; COL_D = 6; MAX_R = 35
    current_row = 1

    if par_arme:
        # Regroupement arme → catégorie
        has_data = any(
            data_by_arme_cat[a][c]["H"] or data_by_arme_cat[a][c]["D"]
            for a in ARMES_ORDER for c in categories
        )
        if not has_data:
            ws.cell(row=1, column=1, value="Aucune donnée").font = mk_font(size=10)
            return

        for arme_code in ARMES_ORDER:
            arme_data = data_by_arme_cat[arme_code]
            cats_actives = [c for c in categories
                            if arme_data[c]["H"] or arme_data[c]["D"]]
            if not cats_actives:
                continue

            # Bandeau arme
            arme_lbl = ARMES_LABEL.get(arme_code, arme_code)
            arme_bg  = ARMES_COLOR.get(arme_code, "EEEEEE")
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=GAP * 2)
            c = ws.cell(row=current_row, column=1, value=f"── {arme_lbl.upper()} ──")
            c.font = mk_font(bold=True, size=12, color="FFFFFF")
            c.fill = mk_fill("1F4E79")
            c.alignment = mk_align(h="center")
            c.border = border_all()
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            current_row = _ecrire_cats_extranet(
                ws, cats_actives, arme_data, cat_label,
                current_row, COL_H, COL_D, GAP, arme_bg
            )
            current_row += 1  # ligne vide entre armes

    else:
        # Sans regroupement par arme (Grand Est, Lorraine)
        cats_actives = [c for c in categories
                        if data_by_cat[c]["H"] or data_by_cat[c]["D"]]
        if not cats_actives:
            ws.cell(row=1, column=1, value="Aucune donnée").font = mk_font(size=10)
            return
        current_row = _ecrire_cats_extranet(
            ws, cats_actives, data_by_cat, cat_label,
            current_row, COL_H, COL_D, GAP
        )

    for ci, w in enumerate([14, 16, 26, 28, 12], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        ws.column_dimensions[get_column_letter(ci + GAP)].width = w


def _ecrire_cats_extranet(ws, cats_actives, data, cat_label,
                          current_row, COL_H, COL_D, GAP, header_bg="BDD7EE"):
    """Écrit les blocs catégorie/H/D dans la feuille Extranet."""
    for cat in cats_actives:
        lh = f"{cat_label.get(cat, cat)} HOMMES"
        ld = f"{cat_label.get(cat, cat)} DAMES"

        for col_start, lbl in [(COL_H, lh), (COL_D, ld)]:
            ws.merge_cells(start_row=current_row, start_column=col_start,
                           end_row=current_row, end_column=col_start + GAP - 1)
            c = ws.cell(row=current_row, column=col_start, value=lbl)
            c.font = mk_font(bold=True); c.fill = mk_fill(header_bg)
            c.alignment = mk_align(h="center"); c.border = border_all()

        th = sorted(data[cat]["H"], key=lambda t: t["club"])
        td = sorted(data[cat]["D"], key=lambda t: t["club"])
        max_t = max(len(th), len(td), 1)

        for i in range(max_t):
            r = current_row + 1 + i
            for col_start, tlist in [(COL_H, th), (COL_D, td)]:
                if i < len(tlist):
                    t = tlist[i]
                    vals = [t["region"], t["ligue"],
                            f"{t['dept']} {t['club']}",
                            f"{t['nom']} {t['prenom']}", t["licence"]]
                    for j, v in enumerate(vals):
                        c = ws.cell(row=r, column=col_start + j, value=v)
                        c.font = mk_font(size=10); c.border = border_all()
                        c.alignment = mk_align(h="left" if j in [2, 3] else "center")
                else:
                    for j in range(GAP):
                        ws.cell(row=r, column=col_start + j).border = border_all()

        current_row += max_t + 2

    return current_row


# ── Feuille Arbitres

def feuille_arbitres(ws, arbitres_list, plage_dates="", fichiers_list=None, titre_comp=""):
    """
    Génère la feuille Arbitres avec sections par jour,
    dédoublonnage par licence, niveau depuis le XML, VLOOKUP sur le barème.
    Colonne Statut avec liste déroulante Retenu/Libéré.
    Retourne un dict {date: {row_start, row_end}} pour que le bilan puisse
    créer des formules COUNTIFS/SUMIFS liées au statut.
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as _PF

    # Dédoublonnage et regroupement par date
    groupes        = defaultdict(list)
    seen_par_jour  = defaultdict(set)

    for a in arbitres_list:
        date_src = a.get("date_source", "")
        lic      = a.get("licence", "").strip()
        if lic in seen_par_jour[date_src]:
            continue
        seen_par_jour[date_src].add(lic)
        groupes[date_src].append(a)

    dates_triees = sorted(groupes.keys(), key=parse_date_key)
    multi_jours  = len(dates_triees) > 1

    # Col layout : 1=Club, 2=Nom, 3=Licence, 4=Niveau, 5=Arme, 6=Tarif, 7=Statut
    # Barème en J:K (col 10:11)
    COL_STATUT = 7
    COL_BAREME_NV  = 10
    COL_BAREME_TAR = 11

    # En-têtes colonnes
    headers = ["Club", "Nom Prénom", "Licence", "Niveau", "Arme", "Tarif", "Statut"]
    for j, h in enumerate(headers, start=1):
        sc(ws.cell(row=1, column=j, value=h), bold=True, bg="1F4E79", fg="FFFFFF")

    # Barème (colonnes J:K)
    sc(ws.cell(row=1, column=COL_BAREME_NV,  value="Barème"),     bold=True, bg="D9D9D9", border=False)
    sc(ws.cell(row=2, column=COL_BAREME_NV,  value="Niveau"),     bold=True, bg="D9D9D9", border=False)
    sc(ws.cell(row=2, column=COL_BAREME_TAR, value="Indemnités"), bold=True, bg="D9D9D9", border=False)
    for i, (nv, tarif) in enumerate(BAREME_ARBITRES, start=3):
        ws.cell(row=i, column=COL_BAREME_NV,  value=nv).font    = mk_font(size=10)
        ws.cell(row=i, column=COL_BAREME_TAR, value=tarif).font = mk_font(size=10)

    bareme_end = 2 + len(BAREME_ARBITRES)

    # Validation liste déroulante pour Statut
    dv = DataValidation(
        type="list",
        formula1='"Retenu,Libéré"',
        allow_blank=True,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv)

    current_row = 2
    # Stocker les plages de données arbitres par date {date: (first_row, last_row)}
    plages_arbitres = {}

    for date_str in dates_triees:
        arbitres_groupe = sorted(groupes[date_str], key=lambda a: a.get("club", ""))

        label = (f"Arbitres du {date_avec_jour(date_str)}"
                 if date_str else f"Arbitres — {plage_dates}" if plage_dates else "Arbitres")

        ws.merge_cells(f"A{current_row}:G{current_row}")
        c = ws.cell(row=current_row, column=1, value=label)
        c.font      = mk_font(bold=True, size=11)
        c.fill      = mk_fill("BDD7EE")
        c.alignment = mk_align(h="left")
        c.border    = border_all()
        current_row += 1
        first_data_row = current_row

        for a in arbitres_groupe:
            nom_prn  = f"{a.get('nom', '')} {a.get('prenom', '')}".strip()
            niveau   = a.get("categorie", "").strip()
            arme_lbl = a.get("arme_label", a.get("arme", ""))
            # Tarif : valeur directe depuis le barème Python (pas de VLOOKUP)
            # → permet au SUMPRODUCT du bilan de fonctionner sans dépendance croisée
            tarif = dict(BAREME_ARBITRES).get(niveau, 0)

            vals = [
                a.get("club", ""), nom_prn, a.get("licence", ""),
                niveau, arme_lbl,
                tarif,      # valeur directe, recalculable
                "Retenu",   # Statut par défaut
            ]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=current_row, column=j, value=v)
                c.font      = mk_font(size=10)
                c.border    = border_all()
                c.alignment = mk_align(h="left" if j in [1, 2] else "center")
            if niveau:
                ws.cell(row=current_row, column=4).fill = mk_fill("EBF3E8")
            if arme_lbl:
                arme_bg = {"Fleuret": "EEF5FF", "Épée": "EEFFF0", "Sabre": "FFF5EE"}.get(arme_lbl, "FFFFFF")
                ws.cell(row=current_row, column=5).fill = mk_fill(arme_bg)

            # Ajouter la validation dropdown sur la cellule Statut
            dv.add(ws.cell(row=current_row, column=COL_STATUT))
            current_row += 1

        plages_arbitres[date_str] = (first_data_row, current_row - 1)

        if multi_jours and date_str != dates_triees[-1]:
            current_row += 1

    # Mise en forme conditionnelle sur la colonne Statut
    statut_range = f"G2:G{current_row}"
    ws.conditional_formatting.add(statut_range, CellIsRule(
        operator="equal", formula=['"Retenu"'],
        fill=_PF("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(statut_range, CellIsRule(
        operator="equal", formula=['"Libéré"'],
        fill=_PF("solid", fgColor="FFC7CE")))

    # ── Section Arbitres-Tireurs
    if fichiers_list:
        licences_tireurs = {}
        for meta, tireurs, _ in fichiers_list:
            for t in tireurs:
                lic = t.get("licence", "").strip()
                if lic:
                    licences_tireurs[lic] = {
                        "nom":    t.get("nom", ""),
                        "prenom": t.get("prenom", ""),
                        "club":   t.get("club", ""),
                    }

        licences_arbitres = {}
        for a in arbitres_list:
            lic = a.get("licence", "").strip()
            if lic and lic not in licences_arbitres:
                licences_arbitres[lic] = a

        arb_tireurs = [
            (lic, licences_tireurs[lic], licences_arbitres[lic])
            for lic in licences_tireurs
            if lic in licences_arbitres
        ]
        arb_tireurs.sort(key=lambda x: x[1]["club"])

        current_row += 1

        ws.merge_cells(f"A{current_row}:G{current_row}")
        c = ws.cell(row=current_row, column=1,
                    value=f"⚠ Arbitres-Tireurs — {len(arb_tireurs)} détecté(s)")
        c.font      = mk_font(bold=True, size=11, color="7F3F00")
        c.fill      = mk_fill("FFE5CC")
        c.alignment = mk_align(h="left")
        c.border    = border_all()
        current_row += 1

        if arb_tireurs:
            for j, h in enumerate(["Club", "Nom Prénom", "Licence"], start=1):
                sc(ws.cell(row=current_row, column=j, value=h),
                   bold=True, bg="FFC89A", fg="7F3F00", size=10)
            current_row += 1

            for lic, tireur, arbitre in arb_tireurs:
                nom_prn = f"{tireur['nom']} {tireur['prenom']}".strip()
                for j, v in enumerate([tireur["club"], nom_prn, lic], start=1):
                    c = ws.cell(row=current_row, column=j, value=v)
                    c.font      = mk_font(size=10, bold=True, color="7F3F00")
                    c.fill      = mk_fill("FFF3E8")
                    c.border    = border_all()
                    c.alignment = mk_align(h="left" if j in [1, 2] else "center")
                current_row += 1
        else:
            c = ws.cell(row=current_row, column=1, value="Aucun arbitre-tireur détecté.")
            c.font = mk_font(size=10, color="7F3F00")
            c.alignment = mk_align(h="left")

    for col, w in [(1, 26), (2, 24), (3, 12), (4, 10), (5, 10), (6, 10), (7, 12),
                   (8, 2), (9, 2), (10, 12), (11, 12)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.sheet_view.showGridLines = False
    return plages_arbitres


# ── Feuille Récap Arbitres

def feuille_recap_arbitres(ws, arbitres_list, dates_ordonnees, titre_comp,
                            plages_arbitres, ws_arb_name="Arbitres",
                            responsables=None):
    """
    Feuille récapitulative des arbitres retenus/libérés.
    Formules liées à la colonne Statut de la feuille Arbitres.
    plages_arbitres : {date: (first_row, last_row)} retourné par feuille_arbitres

    responsables : dict décrivant les responsables à afficher
      Grand Est / Alsace : {"type": "cra", "noms": ["Nom1", "Nom2", ...]}
      Lorraine :           {"type": "superviseurs", "noms": ["Nom1", ...],
                            "armes": ["Fleuret", "Épée", "Sabre"]}
    """
    from openpyxl.utils import get_column_letter as gcl2
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as _PF2
    from openpyxl.worksheet.datavalidation import DataValidation

    REF = f"Arbitres"
    BAREME_END = 2 + len(BAREME_ARBITRES)

    row = 1

    # ── Titre
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value=titre_comp or "Récapitulatif Arbitres")
    c.font = mk_font(bold=True, size=13, color="FFFFFF")
    c.fill = mk_fill("1F4E79"); c.alignment = mk_align(h="center"); c.border = border_all()
    ws.row_dimensions[row].height = 26
    row += 1

    # ── Section Responsables
    resp = responsables or {}
    noms = resp.get("noms", [])
    noms_formula = ",".join(noms) if noms else ""

    if resp.get("type") == "cra":
        # Grand Est / Alsace : CRA + éventuellement un superviseur
        def _cell_responsable(label, fill_col, noms_f, row_n):
            ws.merge_cells(f"A{row_n}:B{row_n}")
            c = ws.cell(row=row_n, column=1, value=label)
            c.font = mk_font(bold=True, size=10, color="FFFFFF")
            c.fill = mk_fill("2E6099"); c.alignment = mk_align(h="left"); c.border = border_all()
            ws.merge_cells(f"C{row_n}:F{row_n}")
            cell = ws.cell(row=row_n, column=3, value="")
            cell.font = mk_font(size=11, bold=True); cell.border = border_all()
            cell.alignment = mk_align(h="left"); cell.fill = mk_fill(fill_col)
            if noms_f:
                dv = DataValidation(type="list", formula1=f'"{noms_f}"',
                                    allow_blank=True, showErrorMessage=False)
                ws.add_data_validation(dv); dv.add(cell)
            ws.row_dimensions[row_n].height = 18
            return cell

        _cell_responsable("Responsable CRA", "EEF5FF", noms_formula, row)
        row += 1

        # Superviseur Grand Est (liste séparée)
        noms_sup_ge = resp.get("noms_superviseur", [])
        noms_sup_formula = ",".join(noms_sup_ge) if noms_sup_ge else ""
        _cell_responsable("Superviseur", "FFF5EE", noms_sup_formula, row)
        row += 2

    elif resp.get("type") == "superviseurs":
        # Lorraine : 1 cellule par arme
        armes = resp.get("armes", ["Fleuret", "Épée", "Sabre"])
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="Superviseurs par arme")
        c.font = mk_font(bold=True, size=10, color="FFFFFF")
        c.fill = mk_fill("2E6099"); c.alignment = mk_align(h="left"); c.border = border_all()
        row += 1
        ARME_COLORS = {"Fleuret": "DDEEFF", "Épée": "DDFFEE", "Sabre": "FFEECC"}
        for arme in armes:
            ws.merge_cells(f"A{row}:B{row}")
            c = ws.cell(row=row, column=1, value=f"  Superviseur {arme}")
            c.font = mk_font(bold=True, size=10)
            c.fill = mk_fill(ARME_COLORS.get(arme, "EEEEEE"))
            c.alignment = mk_align(h="left"); c.border = border_all()
            ws.merge_cells(f"C{row}:F{row}")
            cell_sup = ws.cell(row=row, column=3, value="")
            cell_sup.font = mk_font(size=11, bold=True)
            cell_sup.border = border_all(); cell_sup.alignment = mk_align(h="left")
            cell_sup.fill = mk_fill(ARME_COLORS.get(arme, "EEEEEE"))
            if noms_formula:
                dv = DataValidation(type="list", formula1=f'"{noms_formula}"',
                                    allow_blank=True, showErrorMessage=False)
                ws.add_data_validation(dv); dv.add(cell_sup)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    for date_str in dates_ordonnees:
        if date_str not in plages_arbitres:
            continue
        r1, r2 = plages_arbitres[date_str]
        label = date_avec_jour(date_str).capitalize()

        # ── Bandeau date
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"  {label}")
        c.font = mk_font(bold=True, size=11, color="FFFFFF")
        c.fill = mk_fill("2E6099"); c.alignment = mk_align(h="left"); c.border = border_all()
        ws.row_dimensions[row].height = 20
        row += 1

        # ── En-têtes colonnes
        for ci, lbl in [(1,"Club"),(2,"Nom Prénom"),(3,"Licence"),(4,"Niveau"),(5,"Arme"),(6,"Tarif")]:
            sc(ws.cell(row=row, column=ci, value=lbl), bold=True, bg="1F4E79", fg="FFFFFF", size=9)
        row += 1

        # ── RETENUS
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="✅  Arbitres RETENUS")
        c.font = mk_font(bold=True, size=10, color="276221")
        c.fill = mk_fill("C6EFCE"); c.alignment = mk_align(h="left"); c.border = border_all()
        row += 1

        for src_row in range(r1, r2 + 1):
            for ci, col_src in enumerate([1,2,3,4,5,6], start=1):
                f = f'=IF({REF}!G{src_row}="Retenu",{REF}!{gcl2(col_src)}{src_row},"")'
                cell = ws.cell(row=row, column=ci, value=f)
                cell.font = mk_font(size=10); cell.border = border_all()
                cell.alignment = mk_align(h="left" if ci <= 2 else "center")
            # Fond vert clair sur les retenus
            ws.conditional_formatting.add(
                f"A{row}:F{row}",
                CellIsRule(operator="notEqual", formula=['""'],
                           fill=_PF2("solid", fgColor="EBF5E8"))
            )
            row += 1

        # Ligne total retenus — utilise col F (Tarif) directement, pas de VLOOKUP imbriqué
        nb_ret   = f'=COUNTIF({REF}!G{r1}:G{r2},"Retenu")'
        cout_ret = f'=SUMIFS({REF}!F{r1}:F{r2},{REF}!G{r1}:G{r2},"Retenu")'
        sc(ws.cell(row=row, column=1, value=nb_ret),   bold=True, bg="C6EFCE", fg="276221")
        sc(ws.cell(row=row, column=2, value="arbitre(s) retenu(s)"), bg="C6EFCE", fg="276221", h="left")
        sc(ws.cell(row=row, column=3, value="Coût :"), bg="C6EFCE", fg="276221")
        cell_cout = ws.cell(row=row, column=4, value=cout_ret)
        cell_cout.font = mk_font(bold=True, size=10, color="276221")
        cell_cout.fill = mk_fill("C6EFCE"); cell_cout.border = border_all()
        cell_cout.number_format = '#,##0 "€"'
        cell_cout.alignment = mk_align(h="center")
        for ci in [5,6]: sc(ws.cell(row=row, column=ci), bg="C6EFCE")
        row += 1

        # ── LIBÉRÉS
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="❌  Arbitres LIBÉRÉS")
        c.font = mk_font(bold=True, size=10, color="9C0006")
        c.fill = mk_fill("FFC7CE"); c.alignment = mk_align(h="left"); c.border = border_all()
        row += 1

        for src_row in range(r1, r2 + 1):
            for ci, col_src in enumerate([1,2,3,4,5,6], start=1):
                f = f'=IF({REF}!G{src_row}="Libéré",{REF}!{gcl2(col_src)}{src_row},"")'
                cell = ws.cell(row=row, column=ci, value=f)
                cell.font = mk_font(size=10); cell.border = border_all()
                cell.alignment = mk_align(h="left" if ci <= 2 else "center")
            ws.conditional_formatting.add(
                f"A{row}:F{row}",
                CellIsRule(operator="notEqual", formula=['""'],
                           fill=_PF2("solid", fgColor="FFE8E8"))
            )
            row += 1

        # Ligne total libérés
        nb_lib = f'=COUNTIF({REF}!G{r1}:G{r2},"Libéré")'
        sc(ws.cell(row=row, column=1, value=nb_lib),   bold=True, bg="FFC7CE", fg="9C0006")
        sc(ws.cell(row=row, column=2, value="arbitre(s) libéré(s)"), bg="FFC7CE", fg="9C0006", h="left")
        for ci in [3,4,5,6]: sc(ws.cell(row=row, column=ci), bg="FFC7CE")
        row += 2

    for col, w in [(1,26),(2,24),(3,10),(4,14),(5,10),(6,12)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.sheet_view.showGridLines = False
