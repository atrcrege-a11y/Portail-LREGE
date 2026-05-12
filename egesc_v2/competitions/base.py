"""
competitions/base.py
Classe abstraite définissant le contrat de toute compétition.
Chaque compétition doit implémenter les méthodes abstraites.
"""

from abc import ABC, abstractmethod
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage  # conservé pour compatibilité future

from core.parser import construire_donnees, date_avec_jour
from core.excel_commun import feuille_extranet, feuille_arbitres, feuille_recap_arbitres

class CompetitionBase(ABC):
    """
    Contrat de base pour toutes les compétitions.

    Attributs à définir dans chaque sous-classe :
      CAT_MAP_INDIV      : {cat_xml: cat_key}
      CAT_MAP_EQUIPE     : {cat_xml: cat_key}
      CATS_INDIV         : [cat_key, ...] ordre d'affichage indiv
      CATS_EQUIPE        : [cat_key, ...] ordre d'affichage équipe
      CATS_EXTRANET_INDIV: [cat_key, ...] pour la feuille Extranet indiv
      CATS_EXTRANET_EQUIPE:[cat_key, ...] pour la feuille Extranet équipe
      CAT_LABEL_INDIV    : {cat_key: libellé}
      CAT_LABEL_EQUIPE   : {cat_key: libellé}
      TARIF_INDIV        : {cat_key: prix €}
      TARIF_EQUIPE       : {cat_key: prix €}
      NOM_FEUILLE_INDIV  : str
      NOM_FEUILLE_EQUIPE : str
    """

    # ── À définir dans chaque sous-classe
    CAT_MAP_INDIV       = {}
    CAT_MAP_EQUIPE      = {}
    CATS_INDIV          = []
    CATS_EQUIPE         = []
    CATS_EXTRANET_INDIV = []
    CATS_EXTRANET_EQUIPE= []
    CAT_LABEL_INDIV     = {}
    CAT_LABEL_EQUIPE    = {}
    TARIF_INDIV         = {}
    TARIF_EQUIPE        = {}
    NOM_FEUILLE_INDIV   = "Épreuve Individuelle"
    NOM_FEUILLE_EQUIPE  = "Épreuve Équipe"

    # ── Pipeline principal

    def generer_excel(self, fichiers_list, titre_comp=""):
        """
        Point d'entrée principal.
        Retourne un BytesIO contenant le fichier Excel.
        """
        (groupes_indiv, groupes_equipe, arbitres_all,
         titre_xml, ligue_info, plage_dates,
         dates_ordonnees) = construire_donnees(
            fichiers_list,
            self.CAT_MAP_INDIV,
            self.CAT_MAP_EQUIPE,
        )

        titre_comp  = titre_comp or titre_xml or "EGESC"
        titre_excel = f"{titre_comp} — {plage_dates}" if plage_dates else titre_comp
        self._arbitres_all = arbitres_all

        wb = Workbook()
        wb.remove(wb.active)

        # Arbitres en premier pour obtenir les vraies plages de lignes
        ws_arb = wb.create_sheet("Arbitres")
        plages = feuille_arbitres(ws_arb, arbitres_all, plage_dates=plage_dates,
                                  fichiers_list=fichiers_list,
                                  titre_comp=titre_excel)

        # Récap Arbitres juste après
        ws_recap = wb.create_sheet("Récap Arbitres")
        from core.config import NOMS_CRA, NOMS_SUPERVISEURS_GRAND_EST
        feuille_recap_arbitres(ws_recap, arbitres_all, dates_ordonnees,
                               titre_excel, plages, ws_arb_name="Arbitres",
                               responsables={"type": "cra", "noms": NOMS_CRA,
                                             "noms_superviseur": NOMS_SUPERVISEURS_GRAND_EST})

        # 1. Bilan Financier avec formules dynamiques
        ws_fin = wb.create_sheet("Bilan Financier")
        self.feuille_financiere(
            ws_fin, groupes_indiv, groupes_equipe,
            arbitres_all, dates_ordonnees, titre_excel,
            plages_arbitres=plages,
        )

        # Détecter s'il y a des équipes dans cette compétition
        has_equipes = any(
            cats for cats in groupes_equipe.values()
            if any(cats.values())
        ) if groupes_equipe else False

        # 2. Épreuve Individuelle
        ws_indiv = wb.create_sheet(self.NOM_FEUILLE_INDIV)
        self.feuille_indiv(ws_indiv, groupes_indiv, titre_excel, dates_ordonnees)

        # 3. Épreuve Équipe — uniquement si équipes présentes
        if has_equipes:
            ws_equipe = wb.create_sheet(self.NOM_FEUILLE_EQUIPE)
            self.feuille_equipe(ws_equipe, groupes_equipe, titre_excel, dates_ordonnees)

        # 4. Extranet Indiv
        ws_ext_i = wb.create_sheet("Indiv Extranet")
        feuille_extranet(
            ws_ext_i, fichiers_list,
            self.CATS_EXTRANET_INDIV, self.CAT_LABEL_INDIV,
            cat_map=self.CAT_MAP_INDIV, is_equipe=False,
        )

        # 5. Extranet Équipe — uniquement si équipes présentes
        if has_equipes:
            ws_ext_e = wb.create_sheet("Equipe Extranet")
            feuille_extranet(
                ws_ext_e, fichiers_list,
                self.CATS_EXTRANET_EQUIPE, self.CAT_LABEL_EQUIPE,
                cat_map=self.CAT_MAP_EQUIPE, is_equipe=True,
            )

        # Déplacer Arbitres et Récap en fin de classeur
        n = len(wb.sheetnames)
        wb.move_sheet("Arbitres",       offset=n)
        wb.move_sheet("Récap Arbitres", offset=n)

        # Forcer Excel à recalculer sans itération (élimine faux positif référence circulaire)
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.iterate = False

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Méthodes abstraites à implémenter dans chaque compétition

    @abstractmethod
    def feuille_indiv(self, ws, groupes_indiv, titre, dates_ordonnees):
        """Génère la feuille des épreuves individuelles."""
        ...

    @abstractmethod
    def feuille_equipe(self, ws, groupes_equipe, titre, dates_ordonnees):
        """Génère la feuille des épreuves par équipes."""
        ...

    @abstractmethod
    def feuille_financiere(self, ws, groupes_indiv, groupes_equipe,
                           arbitres_all, dates_ordonnees, titre):
        """Génère la feuille du bilan financier."""
        ...
