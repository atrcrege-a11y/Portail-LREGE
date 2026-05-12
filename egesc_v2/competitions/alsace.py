"""
competitions/alsace.py
Compétition Championnat d'Alsace.
Hérite de GrandEst — ajoute M9, M11 en indiv et M9, M11, M13 en équipe.
Gestion par arme (Fleuret / Épée / Sabre) avec feuilles séparées.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from core.parser import construire_donnees
from core.excel_commun import feuille_extranet, feuille_arbitres, feuille_recap_arbitres
from .grand_est import GrandEst

ARMES = [("F", "Fleuret"), ("E", "Épée"), ("S", "Sabre")]


class Alsace(GrandEst):
    """
    Championnat d'Alsace.
    Même structure que Grand Est avec M9 et M11 en plus (indiv + équipe),
    et M13 en équipe (inexistant en Grand Est).
    Gestion par arme : une feuille Indiv et Équipe par arme.
    """

    # ── Indiv : ajoute M9 et M11
    CAT_MAP_INDIV = {
        "M9": "M9", "M11": "M11",
        **GrandEst.CAT_MAP_INDIV,
    }

    # ── Équipe : ajoute M9, M11, M13
    CAT_MAP_EQUIPE = {
        "M9": "M9", "M11": "M11", "M13": "M13",
        **GrandEst.CAT_MAP_EQUIPE,
    }

    CATS_INDIV  = ["M9", "M11"] + GrandEst.CATS_INDIV
    CATS_EQUIPE = ["M9", "M11", "M13"] + GrandEst.CATS_EQUIPE

    CATS_EXTRANET_INDIV  = ["M9", "M11"] + GrandEst.CATS_EXTRANET_INDIV
    CATS_EXTRANET_EQUIPE = ["M9", "M11", "M13"] + GrandEst.CATS_EXTRANET_EQUIPE

    CAT_LABEL_INDIV = {
        "M9": "M9", "M11": "M11",
        **GrandEst.CAT_LABEL_INDIV,
    }

    CAT_LABEL_EQUIPE = {
        "M9": "M9", "M11": "M11", "M13": "M13",
        **GrandEst.CAT_LABEL_EQUIPE,
    }

    TARIF_INDIV = {
        "M9": 10, "M11": 10,
        **GrandEst.TARIF_INDIV,
    }

    TARIF_EQUIPE = {
        "M9": 30, "M11": 30, "M13": 30,
        **GrandEst.TARIF_EQUIPE,
    }

    def feuille_financiere(self, ws, groupes_indiv, groupes_equipe,
                           arbitres_all, dates_ordonnees, titre,
                           plages_arbitres=None):
        """
        Bilan financier Alsace — recettes groupées par arme puis catégorie.
        Les clés dans groupes_equipe sont "CAT|ARME" (ex: "M13|F").
        """
        from core.excel_base import sc as _sc, mk_font, mk_fill, mk_align, border_all, gcl
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill as _PF
        from collections import defaultdict
        from core.config import BAREME_ARBITRES, COLORS as C
        from core.parser import date_avec_jour

        ARMES = [("F", "Fleuret", "DDEEFF", "2E6099"),
                 ("E", "Épée",    "DDFFEE", "1F7A4D"),
                 ("S", "Sabre",   "FFEECC", "C87020")]

        col_lbl   = 1
        col_nb_h  = 2
        col_nb_d  = 3
        col_nb_mx = 4
        col_pu    = 5
        col_rec_h = 6
        col_rec_d = 7
        col_total = 8

        def style_row(r, bg=None):
            for ci in range(1, 9):
                cell = ws.cell(row=r, column=ci)
                if bg: cell.fill = mk_fill(bg)
                cell.border = border_all()
                cell.alignment = mk_align(h="left" if ci == col_lbl else "center")
                cell.font = mk_font(size=10)

        def ecrire_entetes(r):
            for ci, lbl in [
                (col_lbl,   "Désignation"),
                (col_nb_h,  "Eq.H"),
                (col_nb_d,  "Eq.D"),
                (col_nb_mx, "Eq.MX"),
                (col_pu,    "P.U. (€)"),
                (col_rec_h, "Rec.H (€)"),
                (col_rec_d, "Rec.D (€)"),
                (col_total, "Total (€)"),
            ]:
                cell = ws.cell(row=r, column=ci, value=lbl)
                cell.font = mk_font(bold=True, size=9, color=C["white"])
                cell.fill = mk_fill(C["navy"])
                cell.alignment = mk_align(h="center")
                cell.border = border_all()

        # Arbitres par date (même logique que Grand Est)
        arb_par_date = defaultdict(lambda: {"nb": 0, "cout": 0})
        seen_arb = defaultdict(set)
        bareme_dict = dict(BAREME_ARBITRES)
        for a in arbitres_all:
            date_src = a.get("date_source", "")
            lic      = a.get("licence", "").strip()
            if lic in seen_arb[date_src]:
                continue
            seen_arb[date_src].add(lic)
            cout = bareme_dict.get(a.get("categorie", "").strip(), 0)
            arb_par_date[date_src]["nb"]   += 1
            arb_par_date[date_src]["cout"] += cout

        row = 1

        # Titre
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"Bilan Financier — {titre}")
        cell.font = mk_font(bold=True, size=13, color=C["white"])
        cell.fill = mk_fill(C["navy"]); cell.alignment = mk_align(h="center")
        ws.row_dimensions[row].height = 26
        row += 2

        recette_rows = []
        depense_rows = []

        for date_str in dates_ordonnees:
            label_date = date_avec_jour(date_str).capitalize()
            arb        = arb_par_date.get(date_str, {"nb": 0, "cout": 0})

            # Bandeau date
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=f"  {label_date}")
            cell.font = mk_font(bold=True, size=11, color=C["white"])
            cell.fill = mk_fill(C["blue"]); cell.alignment = mk_align(h="left")
            ws.row_dimensions[row].height = 20
            row += 1
            ecrire_entetes(row); row += 1

            # ── Recettes Indiv (sans arme pour l'instant — s'il y a des indiv)
            indiv_rows = []
            _sous_total_indiv_row = None
            cats_indiv = groupes_indiv.get(date_str, {})
            has_indiv = any(
                v.get("H", 0) + v.get("D", 0) > 0
                for clubs in cats_indiv.values()
                for v in clubs.values()
            )
            if has_indiv:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                cell = ws.cell(row=row, column=1, value="Droits d'engagement — Individuels")
                cell.font = mk_font(bold=True, size=10)
                cell.fill = mk_fill(C["fin_indiv"]); cell.alignment = mk_align(h="left")
                row += 1
                for cat in self.CATS_INDIV:
                    clubs = cats_indiv.get(cat, {})
                    nb_h = sum(v.get("H", 0) for v in clubs.values())
                    nb_d = sum(v.get("D", 0) for v in clubs.values())
                    if not nb_h and not nb_d: continue
                    pu  = self.TARIF_INDIV.get(cat, 15)
                    lbl = self.CAT_LABEL_INDIV.get(cat, cat)
                    r   = row
                    ws.cell(row=r, column=col_lbl,   value=f"  {lbl}")
                    ws.cell(row=r, column=col_nb_h,  value=nb_h or None)
                    ws.cell(row=r, column=col_nb_d,  value=nb_d or None)
                    ws.cell(row=r, column=col_pu,    value=pu)
                    ws.cell(row=r, column=col_rec_h, value=f"={gcl(col_nb_h)}{r}*{gcl(col_pu)}{r}")
                    ws.cell(row=r, column=col_rec_d, value=f"={gcl(col_nb_d)}{r}*{gcl(col_pu)}{r}")
                    ws.cell(row=r, column=col_total, value=f"={gcl(col_rec_h)}{r}+{gcl(col_rec_d)}{r}")
                    style_row(r); indiv_rows.append(r); row += 1
                if indiv_rows:
                    _sous_total_indiv_row = row
                    ws.cell(row=row, column=col_lbl, value="Sous-total Individuels").font = mk_font(bold=True, size=10)
                    refs_i = "+".join(f"{gcl(col_total)}{r}" for r in indiv_rows)
                    ws.cell(row=row, column=col_total,
                            value=f"={refs_i}").font = mk_font(bold=True, size=10)
                    style_row(row, bg=C["green_bg"]); row += 1

            # ── Recettes Équipes par arme
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value="Droits d'engagement — Équipes")
            cell.font = mk_font(bold=True, size=10)
            cell.fill = mk_fill(C["fin_equipe"]); cell.alignment = mk_align(h="left")
            row += 1

            equipe_rows = []
            cats_equipe = groupes_equipe.get(date_str, {})

            for arme_code, arme_lbl, arme_bg, arme_hbg in ARMES:
                cats_arme = {
                    k[:-2]: v for k, v in cats_equipe.items()
                    if k.endswith(f"|{arme_code}")
                }
                if not cats_arme: continue

                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                cell = ws.cell(row=row, column=1, value=f"  {arme_lbl}")
                cell.font = mk_font(bold=True, size=10, color=C["white"])
                cell.fill = mk_fill(arme_hbg); cell.alignment = mk_align(h="left")
                cell.border = border_all()
                row += 1

                for cat in self.CATS_EQUIPE:
                    clubs = cats_arme.get(cat)
                    if not clubs: continue
                    is_mx    = any(v.get("mixte", False) for v in clubs.values())
                    nb_eq_h  = sum(len(v.get("equipes", set()))
                                   for v in clubs.values() if v.get("tireurs_H", 0) > 0) if not is_mx else 0
                    nb_eq_d  = sum(len(v.get("equipes", set()))
                                   for v in clubs.values() if v.get("tireurs_D", 0) > 0) if not is_mx else 0
                    nb_eq_mx = sum(len(v.get("equipes", set()))
                                   for v in clubs.values() if v.get("tireurs_MX", 0) > 0) if is_mx else 0
                    if not nb_eq_h and not nb_eq_d and not nb_eq_mx: continue
                    pu  = self.TARIF_EQUIPE.get(cat, 30)
                    lbl = self.CAT_LABEL_EQUIPE.get(cat, cat)
                    r   = row
                    ws.cell(row=r, column=col_lbl,   value=f"    {lbl}")
                    ws.cell(row=r, column=col_nb_h,  value=nb_eq_h or None)
                    ws.cell(row=r, column=col_nb_d,  value=nb_eq_d or None)
                    ws.cell(row=r, column=col_nb_mx, value=nb_eq_mx or None)
                    ws.cell(row=r, column=col_pu,    value=pu)
                    if is_mx:
                        ws.cell(row=r, column=col_rec_h, value=None)
                        ws.cell(row=r, column=col_rec_d, value=None)
                        ws.cell(row=r, column=col_total, value=f"={gcl(col_nb_mx)}{r}*{gcl(col_pu)}{r}")
                    else:
                        ws.cell(row=r, column=col_rec_h, value=f"={gcl(col_nb_h)}{r}*{gcl(col_pu)}{r}")
                        ws.cell(row=r, column=col_rec_d, value=f"={gcl(col_nb_d)}{r}*{gcl(col_pu)}{r}")
                        ws.cell(row=r, column=col_total, value=f"={gcl(col_rec_h)}{r}+{gcl(col_rec_d)}{r}")
                    style_row(r, bg=arme_bg); equipe_rows.append(r); row += 1

            _sous_total_eq_row = None
            if equipe_rows:
                _sous_total_eq_row = row
                ws.cell(row=row, column=col_lbl, value="Sous-total Équipes").font = mk_font(bold=True, size=10)
                refs_e = "+".join(f"{gcl(col_total)}{r}" for r in equipe_rows)
                ws.cell(row=row, column=col_total,
                        value=f"={refs_e}").font = mk_font(bold=True, size=10)
                style_row(row, bg=C["green_bg"]); row += 1

            # Total recettes — somme des sous-totaux uniquement (pas des lignes détail)
            rec_tot = row
            ws.cell(row=rec_tot, column=col_lbl, value="TOTAL RECETTES").font = mk_font(bold=True, size=11, color=C["white"])
            if _sous_total_indiv_row and _sous_total_eq_row:
                formula = f"={gcl(col_total)}{_sous_total_indiv_row}+{gcl(col_total)}{_sous_total_eq_row}"
            elif _sous_total_indiv_row:
                formula = f"={gcl(col_total)}{_sous_total_indiv_row}"
            elif _sous_total_eq_row:
                formula = f"={gcl(col_total)}{_sous_total_eq_row}"
            else:
                formula = "0"
            ws.cell(row=rec_tot, column=col_total, value=formula).font = mk_font(bold=True, size=11, color=C["white"])
            for ci in range(1, 9):
                cell = ws.cell(row=rec_tot, column=ci)
                cell.fill = mk_fill(C["accent"]); cell.border = border_all()
                cell.alignment = mk_align(h="left" if ci == col_lbl else "center")
                if ci not in [col_lbl, col_total]:
                    cell.font = mk_font(color=C["white"])
            recette_rows.append(rec_tot); row += 1

            # ── Dépenses arbitres
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value="Dépenses — Arbitres")
            cell.font = mk_font(bold=True, size=10)
            cell.fill = mk_fill(C["fin_depense"]); cell.alignment = mk_align(h="left")
            row += 1

            dep_row = row
            ws.cell(row=dep_row, column=col_lbl, value="  Indemnités arbitres")

            # Formules dynamiques liées au Statut (col G) et Tarif (col F) de la feuille Arbitres
            if plages_arbitres and date_str in plages_arbitres:
                r1, r2 = plages_arbitres[date_str]
                # Nb retenus : COUNTIF sur colonne G
                nb_formula   = f'=COUNTIF(Arbitres!G{r1}:G{r2},"Retenu")'
                cout_formula = f'=SUMIFS(Arbitres!F{r1}:F{r2},Arbitres!G{r1}:G{r2},"Retenu")'
                ws.cell(row=dep_row, column=col_nb_h, value=nb_formula)
                ws.cell(row=dep_row, column=col_total, value=cout_formula)
                ws.cell(row=dep_row, column=9, value="barème LREGE — mis à jour auto")
            else:
                ws.cell(row=dep_row, column=col_nb_h, value=arb["nb"] or None)
                ws.cell(row=dep_row, column=col_total, value=arb["cout"] or None)
                ws.cell(row=dep_row, column=9, value="barème LREGE")

            style_row(dep_row)
            depense_rows.append(dep_row); row += 1

            # ── Solde du jour
            solde_row = row
            ws.merge_cells(start_row=solde_row, start_column=1, end_row=solde_row, end_column=7)
            ws.cell(row=solde_row, column=col_lbl,
                    value=f"SOLDE — {label_date}").font = mk_font(bold=True, size=11)
            ws.cell(row=solde_row, column=col_lbl).alignment = mk_align(h="left")
            c_solde = ws.cell(row=solde_row, column=col_total,
                              value=f"={gcl(col_total)}{rec_tot}-{gcl(col_total)}{dep_row}")
            c_solde.font = mk_font(bold=True, size=12)
            c_solde.alignment = mk_align(h="center")
            c_solde.border = border_all()
            for ci in range(1, 9):
                ws.cell(row=solde_row, column=ci).border = border_all()

            ref = f"{gcl(col_total)}{solde_row}"
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_pos"])))
            ws.conditional_formatting.add(ref, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_neg"])))
            row += 2

        # ── Solde global
        if len(dates_ordonnees) > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(row=row, column=col_lbl, value="SOLDE GLOBAL").font = mk_font(bold=True, size=13, color=C["white"])
            ws.cell(row=row, column=col_lbl).fill = mk_fill(C["navy"])
            ws.cell(row=row, column=col_lbl).alignment = mk_align(h="left")
            rec_g = "+".join(f"{gcl(col_total)}{r}" for r in recette_rows)
            dep_g = "+".join(f"{gcl(col_total)}{r}" for r in depense_rows)
            c_g   = ws.cell(row=row, column=col_total, value=f"=({rec_g})-({dep_g})")
            c_g.font = mk_font(bold=True, size=13, color=C["white"])
            c_g.fill = mk_fill(C["navy"]); c_g.alignment = mk_align(h="center")
            for ci in range(1, 9):
                ws.cell(row=row, column=ci).border = border_all()
            ref_g = f"{gcl(col_total)}{row}"
            ws.conditional_formatting.add(ref_g, CellIsRule(
                operator="greaterThanOrEqual", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_pos"])))
            ws.conditional_formatting.add(ref_g, CellIsRule(
                operator="lessThan", formula=["0"],
                fill=_PF("solid", fgColor=C["solde_neg"])))

        for ci, w in [(1,32),(2,7),(3,7),(4,7),(5,9),(6,12),(7,12),(8,12)]:
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.sheet_view.showGridLines = False

    def generer_excel(self, fichiers_list, titre_comp=""):
        """
        Surcharge : génère des feuilles séparées par arme (F / E / S).
        """
        # Construire les données avec clés "CAT|ARME"
        (groupes_indiv, groupes_equipe, arbitres_all,
         titre_xml, ligue_info, plage_dates,
         dates_ordonnees) = construire_donnees(
            fichiers_list,
            self.CAT_MAP_INDIV,
            self.CAT_MAP_EQUIPE,
            par_arme=True,
        )

        titre_comp  = titre_comp or titre_xml or "EGESC"
        titre_excel = f"{titre_comp} — {plage_dates}" if plage_dates else titre_comp
        self._arbitres_all = arbitres_all

        wb = Workbook()
        wb.remove(wb.active)

        # Créer Arbitres en premier pour obtenir les vraies plages de lignes
        # nécessaires aux formules COUNTIF/SUMPRODUCT du Bilan Financier
        ws_arb = wb.create_sheet("Arbitres")
        plages = feuille_arbitres(ws_arb, arbitres_all,
                                  plage_dates=plage_dates,
                                  fichiers_list=fichiers_list,
                                  titre_comp=titre_excel)

        # Récap Arbitres juste après
        ws_recap = wb.create_sheet("Récap Arbitres")
        from core.config import NOMS_CRA
        feuille_recap_arbitres(ws_recap, arbitres_all, dates_ordonnees,
                               titre_excel, plages, ws_arb_name="Arbitres",
                               responsables={"type": "cra", "noms": NOMS_CRA})

        # 1. Bilan Financier avec formules dynamiques liées aux vraies plages
        (groupes_indiv_bilan, _, _, _, _, _, _) = construire_donnees(
            fichiers_list,
            self.CAT_MAP_INDIV,
            self.CAT_MAP_EQUIPE,
            par_arme=False,
        )
        ws_fin = wb.create_sheet("Bilan Financier")
        self.feuille_financiere(
            ws_fin, groupes_indiv_bilan, groupes_equipe,
            arbitres_all, dates_ordonnees, titre_excel,
            plages_arbitres=plages,
        )

        # 2. Feuilles Indiv par arme
        for arme_code, arme_lbl in ARMES:
            gi_arme = self._filtrer_par_arme(groupes_indiv, arme_code)
            if not self._a_des_donnees(gi_arme):
                continue
            ws = wb.create_sheet(f"Indiv {arme_lbl}")
            self.feuille_indiv(ws, gi_arme, f"{titre_excel} — {arme_lbl}", dates_ordonnees)

        # 3. Feuilles Équipe par arme
        for arme_code, arme_lbl in ARMES:
            ge_arme = self._filtrer_par_arme(groupes_equipe, arme_code)
            if not self._a_des_donnees_equipe(ge_arme):
                continue
            ws = wb.create_sheet(f"Équipe {arme_lbl}")
            self.feuille_equipe(ws, ge_arme, f"{titre_excel} — {arme_lbl}", dates_ordonnees)

        # 4. Extranet avec regroupement par arme
        ws_ext_i = wb.create_sheet("Indiv Extranet")
        feuille_extranet(
            ws_ext_i, fichiers_list,
            self.CATS_EXTRANET_INDIV, self.CAT_LABEL_INDIV,
            cat_map=self.CAT_MAP_INDIV, is_equipe=False, par_arme=True,
        )
        ws_ext_e = wb.create_sheet("Equipe Extranet")
        feuille_extranet(
            ws_ext_e, fichiers_list,
            self.CATS_EXTRANET_EQUIPE, self.CAT_LABEL_EQUIPE,
            cat_map=self.CAT_MAP_EQUIPE, is_equipe=True, par_arme=True,
        )

        # Déplacer Arbitres et Récap en toute fin du classeur
        # Les formules restent valides car elles référencent le NOM de feuille
        n = len(wb.sheetnames)
        wb.move_sheet("Arbitres",       offset=n)
        wb.move_sheet("Récap Arbitres", offset=n)

        from io import BytesIO
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.iterate = False
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Helpers

    def _filtrer_par_arme(self, groupes, arme_code):
        """
        Extrait les données d'une arme spécifique depuis les groupes par arme.
        Retourne un dict avec les clés sans le suffixe |ARME.
        """
        from collections import defaultdict
        result = defaultdict(lambda: defaultdict(dict))
        for date_str, cats in groupes.items():
            for cat_key, clubs in cats.items():
                if cat_key.endswith(f"|{arme_code}"):
                    cat_clean = cat_key[: -len(f"|{arme_code}")]
                    result[date_str][cat_clean] = clubs
        return result

    def _a_des_donnees(self, groupes):
        return any(
            v.get("H", 0) + v.get("D", 0) > 0
            for cats in groupes.values()
            for clubs in cats.values()
            for v in clubs.values()
        )

    def _a_des_donnees_equipe(self, groupes):
        return any(
            v.get("tireurs_H", 0) + v.get("tireurs_D", 0) + v.get("tireurs_MX", 0) > 0
            for cats in groupes.values()
            for clubs in cats.values()
            for v in clubs.values()
        )
