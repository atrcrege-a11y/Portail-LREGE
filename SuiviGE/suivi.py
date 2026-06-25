"""
suivi.py — SuiviGE
Suivi des confirmations CDF individuel + équipes, toutes catégories.
Autonome de SelecGE — lit les Excel générés.
"""
import os, pickle, datetime, io, re
from openpyxl import load_workbook

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
_SUIVI_FILE = os.path.join(BASE_DIR, ".suivi_ge.pkl")

CONFIRMATIONS_OK = {"oui", "non", "attente"}

# ── Constantes format Excel SelecGE ───────────────────────────────────────────
# Correspondent exactement aux labels générés par SelecGE/generateur.py
EXCEL_PREFIX_INDIV   = r"^CL\s+(NAT|GE)\s+\d+"   # ex: "CL NAT 1", "CL GE 3"
EXCEL_PREFIX_EQUIPE  = r"^N°\s*\d+"                # ex: "N° 1", "N°2"
EXCEL_PREFIX_COMPO   = r"^\d{1,2}$"                 # ex: "1", "2" (après strip des espaces)
EXCEL_LABEL_QUALIFIE      = "QUALIFIÉ"
EXCEL_LABEL_REMPLACANT    = "REMPLAÇANT"
EXCEL_LABEL_QUALIFIEE_EQ  = "QUALIFIÉE"
EXCEL_ARMES_MAP = {
    "FLEURET": "Fleuret",
    "ÉPÉE": "Épée", "EPEE": "Épée",
    "SABRE": "Sabre",
}
EXCEL_CATS = ["M13", "M15", "M17", "M20", "M23", "SENIORS", "VÉTÉRANS", "VETERANS"]


class ExcelParseError(ValueError):
    """Erreur de parsing Excel SelecGE avec message lisible."""
    def __init__(self, message, hint=""):
        super().__init__(message)
        self.hint = hint
    def __str__(self):
        return f"{super().__str__()} — {self.hint}" if self.hint else super().__str__()


DB_VERSION = 2

# ── Persistance ───────────────────────────────────────────────────────────────
_db = {}  # clé : "type|arme|categorie|genre" — NB: "__version__" est aussi dans _db


def _cle(type_, arme, categorie, genre):
    return f"{type_}|{arme}|{categorie}|{genre}"


def _sauvegarder():
    try:
        _db["__version__"] = DB_VERSION
        with open(_SUIVI_FILE, "wb") as f:
            pickle.dump(_db, f)
    except Exception as e:
        print(f"[SUIVI] Erreur sauvegarde : {e}")


def _migrer(db_brut):
    version = db_brut.get("__version__", 1)
    if version == DB_VERSION:
        return db_brut
    print(f"[SUIVI] Migration base v{version} -> v{DB_VERSION}")
    migre = {k: v for k, v in db_brut.items() if k != "__version__"}
    # v1 -> v2 : garantir champs manquants sur tireurs et équipes
    for cle, session in migre.items():
        if not isinstance(session, dict):
            continue
        for t in session.get("tireurs", []):
            t.setdefault("note", "")
            t.setdefault("confirmation_at", None)
        for eq in session.get("equipes", []):
            eq.setdefault("note", "")
            eq.setdefault("confirmation_at", None)
            eq.setdefault("composition", [])
    migre["__version__"] = DB_VERSION
    return migre


def _charger():
    global _db
    try:
        if os.path.isfile(_SUIVI_FILE):
            with open(_SUIVI_FILE, "rb") as f:
                db_brut = pickle.load(f)
            _db = _migrer(db_brut)
            if db_brut.get("__version__", 1) != DB_VERSION:
                _sauvegarder()
        else:
            _db = {"__version__": DB_VERSION}
    except Exception as e:
        print(f"[SUIVI] Erreur chargement : {e}")
        _db = {"__version__": DB_VERSION}


def _audit(cle, action, detail=""):
    if cle in _db:
        _db[cle]["audit"].append({
            "ts": datetime.datetime.now(),
            "action": action, "detail": detail,
        })


# ── Parseurs Excel SelecGE ────────────────────────────────────────────────────

def _extraire_meta_titre(ws):
    """Extrait compétition, date/lieu, arme/genre/catégorie depuis les 3 premières lignes."""
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    competition = str(rows[0][0] or "").strip() if rows else ""
    date_lieu   = str(rows[1][0] or "").strip() if len(rows) > 1 else ""
    ligne3      = str(rows[2][0] or "").strip() if len(rows) > 2 else ""

    arme = categorie = genre = None
    for label, val in EXCEL_ARMES_MAP.items():
        if label in ligne3.upper():
            arme = val
    for g in ["HOMMES", "DAMES"]:
        if g in ligne3.upper():
            genre = "H" if g == "HOMMES" else "D"
    for cat in EXCEL_CATS:
        if cat in ligne3.upper() or cat in competition.upper():
            categorie = cat.replace("VÉTÉRANS", "Vétérans").replace(
                "VETERANS", "Vétérans").replace("SENIORS", "Seniors")

    return competition, date_lieu, arme, categorie, genre


def lire_excel_indiv(contenu_bytes):
    """
    Parse un Excel SelecGE individuel (feuilles Hommes / Dames).
    Retourne liste de dicts par feuille.
    """
    wb  = load_workbook(io.BytesIO(contenu_bytes), data_only=True)
    resultats = []

    for sheet_name in wb.sheetnames:
        genre_sheet = "H" if sheet_name == "Hommes" else "D" if sheet_name == "Dames" else None
        if not genre_sheet:
            continue

        ws = wb[sheet_name]
        competition, date_lieu, arme, categorie, genre = _extraire_meta_titre(ws)
        genre = genre or genre_sheet

        tireurs  = []
        section  = "qualifie"   # qualifie | remplacant
        arbitres = []

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            val0 = str(row[0]).strip()
            val0u = val0.upper()

            # Détecter sections
            if EXCEL_LABEL_REMPLACANT in val0u or "REMPLACANT" in val0u:
                section = "remplacant"
                continue
            if EXCEL_LABEL_QUALIFIE in val0u or "QUALIFIE" in val0u:
                section = "qualifie"
                continue
            if "Rang / Classement" in val0:
                continue
            if "Arbitre" in val0:
                nom_arb  = str(row[1]).strip() if row[1] else ""
                club_arb = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                niv_arb  = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                if nom_arb and "Nom" not in nom_arb and "Cliquer" not in (niv_arb or ""):
                    arbitres.append({"nom": nom_arb, "club": club_arb, "niveau": niv_arb})
                continue

            # Ligne tireur : col A = "CL NAT X" ou "CL GE X"
            if re.match(EXCEL_PREFIX_INDIV, val0):
                nom    = str(row[1]).strip() if row[1] else ""
                prenom = str(row[2]).strip() if row[2] else ""
                club   = str(row[3]).strip() if row[3] else ""
                conf   = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ""
                if conf not in ("oui", "non"):
                    conf = "attente"

                tireurs.append({
                    "rang":           val0,
                    "nom":            nom,
                    "prenom":         prenom,
                    "club":           club,
                    "statut":         section,
                    "confirmation":   conf,
                    "confirmation_at": None,
                    "note":           "",
                })

        resultats.append({
            "genre":       genre,
            "competition": competition,
            "date_lieu":   date_lieu,
            "arme":        arme,
            "categorie":   categorie,
            "tireurs":     tireurs,
            "arbitres":    arbitres,
        })

    return resultats


def lire_excel_equipes(contenu_bytes):
    """
    Parse un Excel SelecGE équipes (feuilles Hommes / Dames).
    Retourne liste de dicts par feuille.
    """
    wb  = load_workbook(io.BytesIO(contenu_bytes), data_only=True)
    resultats = []

    for sheet_name in wb.sheetnames:
        genre_sheet = "H" if sheet_name == "Hommes" else "D" if sheet_name == "Dames" else None
        if not genre_sheet:
            continue

        ws = wb[sheet_name]
        competition, date_lieu, arme, categorie, genre = _extraire_meta_titre(ws)
        genre = genre or genre_sheet

        equipes  = []
        equipe_courante = None
        section  = "qualifiee"

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            val0  = str(row[0]).strip()
            val0u = val0.upper()

            if "Rang / Classement" in val0:
                continue
            if EXCEL_LABEL_QUALIFIEE_EQ in val0u or "QUALIFIEE" in val0u:
                section = "qualifiee"
                continue

            # Ligne équipe : col A = "N° X"
            if re.match(EXCEL_PREFIX_EQUIPE, val0):
                if equipe_courante:
                    equipes.append(equipe_courante)
                nom_eq  = str(row[1]).strip() if row[1] and str(row[1]).strip() else ""
                club_eq = str(row[2]).strip() if row[2] else ""
                conf    = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ""
                if conf not in ("oui", "non"):
                    conf = "attente"
                equipe_courante = {
                    "rang":           val0,
                    "nom":            nom_eq,
                    "club":           club_eq,
                    "statut":         section,
                    "confirmation":   conf,
                    "confirmation_at": None,
                    "note":           "",
                    "composition":    [],
                }
                continue

            # Ligne composition : col A = "  1", "  2"...
            if equipe_courante and re.match(EXCEL_PREFIX_COMPO, val0):
                nom_t    = str(row[1]).strip() if row[1] else ""
                prenom_t = str(row[2]).strip() if row[2] else ""
                if nom_t and nom_t != "Nom":
                    equipe_courante["composition"].append({
                        "nom": nom_t, "prenom": prenom_t
                    })

        if equipe_courante:
            equipes.append(equipe_courante)

        resultats.append({
            "genre":       genre,
            "competition": competition,
            "date_lieu":   date_lieu,
            "arme":        arme,
            "categorie":   categorie,
            "equipes":     equipes,
        })

    return resultats


# ── Initialisation suivi ──────────────────────────────────────────────────────

def initialiser_indiv(contenu_bytes):
    """Initialise le suivi individuel depuis un Excel SelecGE."""
    feuilles = lire_excel_indiv(contenu_bytes)
    cles_init = []

    for f in feuilles:
        arme, cat, genre = f["arme"], f["categorie"], f["genre"]
        if not all([arme, cat, genre]):
            manquants = [n for n, v in [("arme", arme), ("catégorie", cat), ("genre", genre)] if not v]
            print(f"[SUIVI] Feuille ignorée — champs non détectés : {manquants}")
            continue
        cle = _cle("indiv", arme, cat, genre)

        # Préserver confirmations existantes
        existants = {}
        if cle in _db:
            for t in _db[cle].get("tireurs", []):
                existants[(t["nom"].upper(), t["prenom"].upper())] = t

        nouveaux = []
        for t in f["tireurs"]:
            key = (t["nom"].upper(), t["prenom"].upper())
            ex  = existants.get(key, {})
            nouveaux.append({
                **t,
                "confirmation":    ex.get("confirmation", t["confirmation"]),
                "confirmation_at": ex.get("confirmation_at"),
                "note":            ex.get("note", ""),
            })

        _db[cle] = {
            "type":        "indiv",
            "arme":        arme,
            "categorie":   cat,
            "genre":       genre,
            "competition": f["competition"],
            "date_lieu":   f["date_lieu"],
            "tireurs":     nouveaux,
            "arbitres":    f.get("arbitres", []),
            "created_at":  _db.get(cle, {}).get("created_at", datetime.datetime.now()),
            "updated_at":  datetime.datetime.now(),
            "audit":       _db.get(cle, {}).get("audit", []),
        }
        _audit(cle, "IMPORT",
               f"{arme} {cat} {genre} — {len(nouveaux)} tireurs "
               f"({sum(1 for t in nouveaux if t['statut']=='qualifie')} qual. / "
               f"{sum(1 for t in nouveaux if t['statut']=='remplacant')} remplac.)")
        cles_init.append(cle)

    _sauvegarder()
    return cles_init


def initialiser_equipes(contenu_bytes):
    """Initialise le suivi équipes depuis un Excel SelecGE."""
    feuilles = lire_excel_equipes(contenu_bytes)
    cles_init = []

    for f in feuilles:
        arme, cat, genre = f["arme"], f["categorie"], f["genre"]
        if not all([arme, cat, genre]):
            manquants = [n for n, v in [("arme", arme), ("catégorie", cat), ("genre", genre)] if not v]
            print(f"[SUIVI] Feuille équipes ignorée — champs non détectés : {manquants}")
            continue
        cle = _cle("equipe", arme, cat, genre)

        existants = {}
        if cle in _db:
            for eq in _db[cle].get("equipes", []):
                existants[eq["rang"]] = eq

        nouveaux = []
        for eq in f["equipes"]:
            ex = existants.get(eq["rang"], {})
            nouveaux.append({
                **eq,
                "confirmation":    ex.get("confirmation", eq["confirmation"]),
                "confirmation_at": ex.get("confirmation_at"),
                "note":            ex.get("note", ""),
                "composition":     ex.get("composition") or eq.get("composition", []),
            })

        _db[cle] = {
            "type":        "equipe",
            "arme":        arme,
            "categorie":   cat,
            "genre":       genre,
            "competition": f["competition"],
            "date_lieu":   f["date_lieu"],
            "equipes":     nouveaux,
            "created_at":  _db.get(cle, {}).get("created_at", datetime.datetime.now()),
            "updated_at":  datetime.datetime.now(),
            "audit":       _db.get(cle, {}).get("audit", []),
        }
        _audit(cle, "IMPORT",
               f"Équipes {arme} {cat} {genre} — {len(nouveaux)} équipe(s)")
        cles_init.append(cle)

    _sauvegarder()
    return cles_init


# ── Import retour club ────────────────────────────────────────────────────────

def importer_retour_indiv(contenu_bytes):
    """Parse un Excel individuel retourné avec col E (Oui/Non) remplie."""
    feuilles = lire_excel_indiv(contenu_bytes)
    modifs   = []

    for f in feuilles:
        arme, cat, genre = f["arme"], f["categorie"], f["genre"]
        if not all([arme, cat, genre]):
            continue
        cle = _cle("indiv", arme, cat, genre)
        if cle not in _db:
            modifs.append({"erreur": f"Suivi {arme} {cat} {genre} non initialisé"})
            continue

        idx = {(t["nom"].upper(), t["prenom"].upper()): t
               for t in _db[cle]["tireurs"]}
        nb  = 0
        for t in f["tireurs"]:
            if t["confirmation"] == "attente":
                continue
            key = (t["nom"].upper(), t["prenom"].upper())
            if key in idx:
                ancien = idx[key]["confirmation"]
                idx[key]["confirmation"]    = t["confirmation"]
                idx[key]["confirmation_at"] = datetime.datetime.now()
                nb += 1
                _audit(cle, "RETOUR",
                       f"{t['nom']} {t['prenom']} : {ancien} → {t['confirmation']}")

        _db[cle]["updated_at"] = datetime.datetime.now()
        modifs.append({"arme": arme, "categorie": cat, "genre": genre, "nb": nb})

    _sauvegarder()
    return modifs


def importer_retour_equipes(contenu_bytes):
    """Parse un Excel équipes retourné avec col D (Oui/Non) remplie."""
    feuilles = lire_excel_equipes(contenu_bytes)
    modifs   = []

    for f in feuilles:
        arme, cat, genre = f["arme"], f["categorie"], f["genre"]
        if not all([arme, cat, genre]):
            manquants = [n for n, v in [("arme", arme), ("catégorie", cat), ("genre", genre)] if not v]
            print(f"[SUIVI] Feuille équipes ignorée — champs non détectés : {manquants}")
            continue
        cle = _cle("equipe", arme, cat, genre)
        if cle not in _db:
            modifs.append({"erreur": f"Suivi équipes {arme} {cat} {genre} non initialisé"})
            continue

        idx = {eq["rang"]: eq for eq in _db[cle]["equipes"]}
        nb  = 0
        for eq in f["equipes"]:
            if eq["confirmation"] == "attente":
                continue
            if eq["rang"] in idx:
                ancien = idx[eq["rang"]]["confirmation"]
                idx[eq["rang"]]["confirmation"]    = eq["confirmation"]
                idx[eq["rang"]]["confirmation_at"] = datetime.datetime.now()
                # Mise à jour composition si fournie
                if eq.get("composition"):
                    idx[eq["rang"]]["composition"] = eq["composition"]
                nb += 1
                _audit(cle, "RETOUR",
                       f"Équipe {eq['rang']} ({eq['nom']}) : {ancien} → {eq['confirmation']}")

        _db[cle]["updated_at"] = datetime.datetime.now()
        modifs.append({"arme": arme, "categorie": cat, "genre": genre, "nb": nb})

    _sauvegarder()
    return modifs


# ── Actions manuelles ─────────────────────────────────────────────────────────

def maj_confirmation_tireur(arme, categorie, genre, nom, prenom, confirmation, note=""):
    if confirmation not in CONFIRMATIONS_OK:
        raise ValueError(f"Valeur invalide : {confirmation}")
    cle = _cle("indiv", arme, categorie, genre)
    if cle not in _db:
        raise KeyError(f"Suivi {arme} {categorie} {genre} introuvable")
    key = (nom.strip().upper(), prenom.strip().upper())
    for t in _db[cle]["tireurs"]:
        if (t["nom"].upper(), t["prenom"].upper()) == key:
            ancien = t["confirmation"]
            t["confirmation"]    = confirmation
            t["confirmation_at"] = datetime.datetime.now()
            t["note"]            = note
            _audit(cle, "CONFIRMATION",
                   f"{nom} {prenom} : {ancien} → {confirmation}"
                   + (f" | {note}" if note else ""))
            _db[cle]["updated_at"] = datetime.datetime.now()
            _sauvegarder()
            return
    raise KeyError(f"Tireur {nom} {prenom} introuvable")


def maj_confirmation_equipe(arme, categorie, genre, rang, confirmation, note="",
                             composition=None):
    if confirmation not in CONFIRMATIONS_OK:
        raise ValueError(f"Valeur invalide : {confirmation}")
    cle = _cle("equipe", arme, categorie, genre)
    if cle not in _db:
        raise KeyError(f"Suivi équipes {arme} {categorie} {genre} introuvable")
    for eq in _db[cle]["equipes"]:
        if eq["rang"] == rang:
            ancien = eq["confirmation"]
            eq["confirmation"]    = confirmation
            eq["confirmation_at"] = datetime.datetime.now()
            eq["note"]            = note
            if composition:
                eq["composition"] = composition
            _audit(cle, "CONFIRMATION_EQ",
                   f"Équipe {rang} ({eq['nom']}) : {ancien} → {confirmation}"
                   + (f" | {note}" if note else ""))
            _db[cle]["updated_at"] = datetime.datetime.now()
            _sauvegarder()
            return
    raise KeyError(f"Équipe {rang} introuvable")


def maj_composition_equipe(arme, categorie, genre, rang, composition):
    """Met à jour la composition d'une équipe."""
    cle = _cle("equipe", arme, categorie, genre)
    if cle not in _db:
        raise KeyError(f"Suivi équipes {arme} {categorie} {genre} introuvable")
    for eq in _db[cle]["equipes"]:
        if eq["rang"] == rang:
            eq["composition"] = composition
            _audit(cle, "COMPOSITION",
                   f"Équipe {rang} ({eq['nom']}) : {len(composition)} membre(s)")
            _db[cle]["updated_at"] = datetime.datetime.now()
            _sauvegarder()
            return
    raise KeyError(f"Équipe {rang} introuvable")


# ── Getters ───────────────────────────────────────────────────────────────────

def get_tous_suivis():
    res = {}
    for k, v in _db.items():
        if k == "__version__":
            continue
        res[k] = {
            "type":       v["type"],
            "arme":       v["arme"],
            "categorie":  v["categorie"],
            "genre":      v["genre"],
            "competition":v.get("competition", ""),
            "updated_at": v["updated_at"].strftime("%d/%m/%Y %H:%M"),
        }
    return res


def get_detail(arme, categorie, genre, type_="indiv"):
    import copy
    cle = _cle(type_, arme, categorie, genre)
    if cle not in _db:
        return None
    d = copy.deepcopy(_db[cle])
    d["created_at"] = d["created_at"].strftime("%d/%m/%Y %H:%M")
    d["updated_at"] = d["updated_at"].strftime("%d/%m/%Y %H:%M")
    # Sérialiser datetimes
    items = d.get("tireurs") or d.get("equipes") or []
    for item in items:
        if item.get("confirmation_at"):
            item["confirmation_at"] = item["confirmation_at"].strftime("%d/%m/%Y %H:%M")
    return d


def get_stats(arme, categorie, genre, type_="indiv"):
    cle = _cle(type_, arme, categorie, genre)
    if cle not in _db:
        return None
    d = _db[cle]

    if type_ == "indiv":
        tireurs = d["tireurs"]
        qual    = [t for t in tireurs if t["statut"] == "qualifie"]
        rem     = [t for t in tireurs if t["statut"] == "remplacant"]
        return {
            "arme": arme, "categorie": categorie, "genre": genre,
            "competition": d.get("competition", ""),
            "date_lieu":   d.get("date_lieu", ""),
            "updated_at":  d["updated_at"].strftime("%d/%m/%Y %H:%M"),
            "qualifies":   {"total": len(qual),
                            "oui":     sum(1 for t in qual if t["confirmation"] == "oui"),
                            "non":     sum(1 for t in qual if t["confirmation"] == "non"),
                            "attente": sum(1 for t in qual if t["confirmation"] == "attente")},
            "remplacants": {"total": len(rem),
                            "oui":     sum(1 for t in rem if t["confirmation"] == "oui"),
                            "non":     sum(1 for t in rem if t["confirmation"] == "non"),
                            "attente": sum(1 for t in rem if t["confirmation"] == "attente")},
        }
    else:
        equipes = d["equipes"]
        return {
            "arme": arme, "categorie": categorie, "genre": genre,
            "competition": d.get("competition", ""),
            "date_lieu":   d.get("date_lieu", ""),
            "updated_at":  d["updated_at"].strftime("%d/%m/%Y %H:%M"),
            "equipes":     {"total":    len(equipes),
                            "oui":      sum(1 for e in equipes if e["confirmation"] == "oui"),
                            "non":      sum(1 for e in equipes if e["confirmation"] == "non"),
                            "attente":  sum(1 for e in equipes if e["confirmation"] == "attente"),
                            "avec_compo": sum(1 for e in equipes if e.get("composition"))},
        }


def get_remplacants_a_appeler(arme, categorie, genre):
    """Retourne les remplaçants à appeler (1 par refus non comblé)."""
    cle = _cle("indiv", arme, categorie, genre)
    if cle not in _db:
        return []
    tireurs = _db[cle]["tireurs"]
    qual    = [t for t in tireurs if t["statut"] == "qualifie"]
    rem     = [t for t in tireurs if t["statut"] == "remplacant"]

    refus   = sum(1 for t in qual if t["confirmation"] == "non")
    app_ok  = sum(1 for t in rem if t["confirmation"] == "oui")
    app_non = sum(1 for t in rem if t["confirmation"] == "non")
    manque  = refus + app_non - app_ok
    dispo   = [t for t in rem if t["confirmation"] == "attente"]

    return dispo[:max(0, manque)]


def get_audit(arme, categorie, genre, type_="indiv"):
    cle = _cle(type_, arme, categorie, genre)
    if cle not in _db:
        return []
    return [
        {"ts": e["ts"].strftime("%d/%m/%Y %H:%M:%S"),
         "action": e["action"], "detail": e["detail"]}
        for e in reversed(_db[cle]["audit"])
    ]


def supprimer(arme, categorie, genre, type_="indiv"):
    cle = _cle(type_, arme, categorie, genre)
    if cle in _db:
        del _db[cle]
        _sauvegarder()


_charger()


# ── Pont plateforme de confirmation (option A : genre combiné 'HD') ─────────────
# La plateforme en ligne fusionne Hommes+Dames en une compétition genre='HD'.
# SuiviGE range donc ces données sous une clé '...|HD' (distinct des imports Excel
# par genre). Périmètre : INDIVIDUEL. Les équipes (granularité différente) = TODO.

_ARME_PF = {"epee": "Épée", "fleuret": "Fleuret", "sabre": "Sabre",
            "e": "Épée", "f": "Fleuret", "s": "Sabre"}


def _arme_plateforme(a):
    return _ARME_PF.get((a or "").strip().lower(), a)


def _cat_plateforme(cat):
    u = (cat or "").strip().upper()
    if u in ("V1", "V2", "V3", "V4", "VETERANS", "VÉTÉRANS"):
        return "Vétérans"
    if u in ("SENIOR", "SENIORS"):
        return "Seniors"
    return (cat or "").strip()


def _confirmation_pf(att):
    """Statut SuiviGE depuis un attendu plateforme (saisi/present)."""
    if not att.get("saisi"):
        return "attente"
    return "oui" if att.get("present") else "non"


def mapper_plateforme(competitions):
    """JSON /api/suivi → dict {cle: entry} pour les compétitions INDIVIDUELLES.

    Pure (pas de réseau, pas d'écriture _db). Genre forcé à 'HD' (option A).
    """
    entries = {}
    for c in competitions or []:
        if c.get("format") != "individuel":
            continue  # équipes = hors périmètre v1
        arme = _arme_plateforme(c.get("arme"))
        cat = _cat_plateforme(c.get("categorie"))
        if not arme or not cat:
            continue
        cle = _cle("indiv", arme, cat, "HD")
        tireurs = []
        for k in c.get("clubs", []):
            club = k.get("club", "")
            for at in k.get("attendus", []):
                section = (at.get("section") or "")
                statut = "remplacant" if "REMPLA" in section.upper() else "qualifie"
                rang = at.get("rang")
                tireurs.append({
                    "rang":            str(rang) if rang is not None else "",
                    "nom":             at.get("nom", ""),
                    "prenom":          at.get("prenom", ""),
                    "club":            club,
                    "statut":          statut,
                    "confirmation":    _confirmation_pf(at),
                    "confirmation_at": None,
                    "note":            "",
                })
        entries[cle] = {
            "type": "indiv", "arme": arme, "categorie": cat, "genre": "HD",
            "competition": c.get("nom", ""),
            "date_lieu": f"{c.get('date','')} {c.get('lieu','')}".strip(),
            "tireurs": tireurs, "arbitres": [],
        }
    return entries


def importer_depuis_plateforme(base_url=None, token=None, timeout=20):
    """Récupère /api/suivi/<token> et remplit le suivi INDIVIDUEL (genre HD).

    base_url/token : sinon env PLATEFORME_URL / PLATEFORME_TOKEN.
    Retourne un résumé {ok, cles, nb_competitions, nb_tireurs} ou lève ValueError.
    """
    import json as _json
    import urllib.request as _u
    import urllib.error as _ue

    base_url = (base_url or os.environ.get("PLATEFORME_URL", "")).rstrip("/")
    token = token or os.environ.get("PLATEFORME_TOKEN", "")
    if not base_url or not token:
        raise ValueError("PLATEFORME_URL ou PLATEFORME_TOKEN non configuré")

    url = f"{base_url}/api/suivi/{token}"
    try:
        with _u.urlopen(url, timeout=timeout) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
    except _ue.HTTPError as e:
        raise ValueError(f"Plateforme HTTP {e.code}")
    except _ue.URLError as e:
        raise ValueError(f"Connexion impossible : {e.reason}")

    entries = mapper_plateforme(data.get("competitions", []))
    nb_tireurs = 0
    for cle, entry in entries.items():
        existant = _db.get(cle, {})
        entry["created_at"] = existant.get("created_at", datetime.datetime.now())
        entry["updated_at"] = datetime.datetime.now()
        entry["audit"] = existant.get("audit", [])
        _db[cle] = entry
        _audit(cle, "IMPORT PLATEFORME",
               f"{entry['arme']} {entry['categorie']} HD — {len(entry['tireurs'])} tireurs")
        nb_tireurs += len(entry["tireurs"])
    _sauvegarder()
    return {"ok": True, "cles": list(entries.keys()),
            "nb_competitions": len(entries), "nb_tireurs": nb_tireurs}
