"""Fixtures Excel SelecMaster synthétiques pour les tests SuiviMaster."""
import io
import openpyxl
from openpyxl.styles import PatternFill


def excel_selecmaster(arme="Épée", categorie="M13",
                      territoire="Lorraine",
                      selectionnes_h=None, remplacants_h=None,
                      selectionnes_d=None, remplacants_d=None):
    """
    Génère un Excel SelecMaster minimal en mémoire.
    Structure : feuilles Hommes + Dames avec sections SÉLECTIONNÉS / REMPLAÇANTS.
    """
    if selectionnes_h is None:
        selectionnes_h = [
            {"nom": "MARTIN", "prenom": "Jean", "club": "NANCY CE"},
            {"nom": "DUPONT", "prenom": "Paul", "club": "METZ ESC"},
        ]
    if remplacants_h is None:
        remplacants_h = [{"nom": "KLEIN", "prenom": "Marc", "club": "EPINAL ESC"}]
    if selectionnes_d is None:
        selectionnes_d = [{"nom": "WEBER", "prenom": "Anna", "club": "NANCY CE"}]
    if remplacants_d is None:
        remplacants_d = []

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for genre_label, sel, rem in [
        ("Hommes", selectionnes_h, remplacants_h),
        ("Dames",  selectionnes_d, remplacants_d),
    ]:
        ws = wb.create_sheet(genre_label)
        row = 1

        # En-tête avec arme + catégorie (comme generateur.py)
        ws.cell(row=row, column=1, value=f"LREGE Grand Est — Sélection {arme} {categorie}")
        row += 1
        ws.cell(row=row, column=1, value=f"MASTER GRAND EST")
        row += 2

        # Section SÉLECTIONNÉS
        ws.cell(row=row, column=1, value="SÉLECTIONNÉS")
        row += 1
        # En-tête colonnes
        for col, h in enumerate(["Rang", "Nom", "Prénom", "Club", "Participation", "Remarque"], 1):
            ws.cell(row=row, column=col, value=h)
        row += 1

        for i, t in enumerate(sel, 1):
            ws.cell(row=row, column=1, value=f"M {i}")
            ws.cell(row=row, column=2, value=t["nom"])
            ws.cell(row=row, column=3, value=t["prenom"])
            ws.cell(row=row, column=4, value=t["club"])
            row += 1

        row += 1  # ligne vide

        # Section REMPLAÇANTS
        ws.cell(row=row, column=1, value="REMPLAÇANTS")
        row += 1
        for col, h in enumerate(["Rang", "Nom", "Prénom", "Club", "Participation", "Remarque"], 1):
            ws.cell(row=row, column=col, value=h)
        row += 1

        for i, t in enumerate(rem, 1):
            ws.cell(row=row, column=1, value=f"M {i}")
            ws.cell(row=row, column=2, value=t["nom"])
            ws.cell(row=row, column=3, value=t["prenom"])
            ws.cell(row=row, column=4, value=t["club"])
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_invalide():
    """Excel qui ne ressemble pas à un export SelecMaster."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Contenu quelconque"
    ws["A2"] = "Pas de tireur"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_retour(arme="Épée", categorie="M13",
                 confirmations=None):
    """
    Excel retourné par un club avec colonne Participation remplie.
    confirmations = [{"nom", "prenom", "genre", "conf"}]
    """
    confirmations = confirmations or [
        {"nom": "MARTIN", "prenom": "Jean", "genre": "H", "conf": "oui"},
        {"nom": "DUPONT", "prenom": "Paul", "genre": "H", "conf": "non"},
    ]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Regrouper par genre
    par_genre = {"H": [], "D": []}
    for c in confirmations:
        par_genre[c["genre"]].append(c)

    for genre_label, genre_code in [("Hommes", "H"), ("Dames", "D")]:
        ws = wb.create_sheet(genre_label)
        ws["A1"] = "SÉLECTIONNÉS"
        for i, c in enumerate(par_genre[genre_code], 1):
            ws.cell(row=i+1, column=1, value=f"M {i}")
            ws.cell(row=i+1, column=2, value=c["nom"])
            ws.cell(row=i+1, column=3, value=c["prenom"])
            ws.cell(row=i+1, column=4, value="CLUB")
            ws.cell(row=i+1, column=5, value=c["conf"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
