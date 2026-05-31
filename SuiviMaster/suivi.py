"""
suivi.py — Logique métier SuiviMaster
Gestion des confirmations de participation Master Grand Est M11/M13
Autonome de SelecMaster — lit les Excel générés.
"""
import os, pickle, datetime, io
from collections import defaultdict
from openpyxl import load_workbook

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
_SUIVI_FILE = os.path.join(BASE_DIR, ".suivi_master.pkl")

EFFECTIF_MAX = 16

# ── Constantes format Excel SelecMaster ───────────────────────────────────────
# Ces constantes doivent correspondre exactement aux labels générés par
# SelecMaster/generateur.py. Si le format Excel change, modifier ici uniquement.
EXCEL_FORMAT_VERSION   = "SELECMASTER_V1"   # marqueur dans cellule cachée (à venir)
EXCEL_PREFIX_RANG      = "M "               # préfixe des lignes tireurs : "M 1", "M 2"
EXCEL_RANG_MAX_LEN     = 5                  # longueur max du préfixe + numéro
EXCEL_LABEL_SELECTIONNES    = "SÉLECTIONNÉS"
EXCEL_LABEL_REMPLACANTS     = "REMPLAÇANTS"
EXCEL_LABEL_NON_SELECTIONNABLE = "NON SÉLECTIONNABLE"
EXCEL_ARMES_MAP = {
    "ÉPÉE": "Épée", "EPEE": "Épée",
    "FLEURET": "Fleuret",
    "SABRE": "Sabre",
}

class ExcelParseError(ValueError):
    """Erreur de parsing Excel SelecMaster avec message lisible."""
    def __init__(self, message, hint=""):
        super().__init__(message)
        self.hint = hint
    def __str__(self):
        return f"{super().__str__()} — {self.hint}" if self.hint else super().__str__()

ORDRE_PIOCHE = {
    "Champagne-Ardenne": ["Champagne-Ardenne", "Lorraine",           "Alsace"],
    "Lorraine":          ["Lorraine",           "Alsace",            "Champagne-Ardenne"],
    "Alsace":            ["Alsace",             "Champagne-Ardenne", "Lorraine"],
}

CONFIRMATIONS_OK = {"oui", "non", "attente"}
NIVEAUX_ARB_OK   = {"Formation Régionale", "Régionale", "Formation Nationale",
                    "National", "International", ""}

# ── Persistance ───────────────────────────────────────────────────────────────
_db = {}           # clé : "arme|categorie"
DB_VERSION = 2     # Incrémenté à chaque changement de structure _db


def _cle(arme, categorie):
    return f"{arme}|{categorie}"


def _sauvegarder():
    try:
        _db["__version__"] = DB_VERSION
        with open(_SUIVI_FILE, "wb") as f:
            pickle.dump(_db, f)
    except Exception as e:
        print(f"[SUIVI] Erreur sauvegarde : {e}")


def _migrer(db_ancien):
    """
    Migre un _db d'une version antérieure vers DB_VERSION.
    Retourne le dict migré. Tolère les clés manquantes.
    """
    version = db_ancien.get("__version__", 1)
    if version == DB_VERSION:
        return db_ancien

    print(f"[SUIVI] Migration base v{version} → v{DB_VERSION}")
    migre = {k: v for k, v in db_ancien.items() if k != "__version__"}

    # v1 → v2 : garantir que chaque tireur a les champs appele + note
    for cle, session in migre.items():
        if not isinstance(session, dict):
            continue
        for territoire, tireurs in session.get("territoires", {}).items():
            for t in tireurs:
                t.setdefault("appele", False)
                t.setdefault("note", "")
                t.setdefault("confirmation_at", None)

    migre["__version__"] = DB_VERSION
    return migre


def _charger():
    global _db
    try:
        if os.path.isfile(_SUIVI_FILE):
            with open(_SUIVI_FILE, "rb") as f:
                db_brut = pickle.load(f)
            _db = _migrer(db_brut)
            # Resauvegarder si migration a eu lieu
            if db_brut.get("__version__", 1) != DB_VERSION:
                _sauvegarder()
        else:
            _db = {"__version__": DB_VERSION}
    except Exception as e:
        print(f"[SUIVI] Erreur chargement : {e}")
        _db = {"__version__": DB_VERSION}


def _audit(cle, action, detail=""):
    if cle in _db:
        _db[cle]["audit"].append({
            "ts":     datetime.datetime.now(),
            "action": action,
            "detail": detail,
        })


def _cle_tireur(t):
    return (t["nom"].strip().upper(), t["prenom"].strip().upper(), t["genre"])


# ── Lecture Excel SelecMaster ─────────────────────────────────────────────────
def lire_excel_selecmaster(contenu_bytes, territoire, territoire_organisateur):
    """
    Parse un Excel généré par SelecMaster (feuilles Hommes / Dames).
    Extrait la liste des sélectionnés et remplaçants.
    Retourne : {"arme", "categorie", "tireurs": [...]}
    """
    wb = load_workbook(io.BytesIO(contenu_bytes), data_only=True)
    tireurs = []
    arme = categorie = None

    for sheet_name in wb.sheetnames:
        genre = "H" if sheet_name == "Hommes" else "D" if sheet_name == "Dames" else None
        if not genre:
            continue

        ws = wb[sheet_name]
        statut_courant = None
        rang_sel = rang_rem = 0

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            val0 = str(row[0]).strip()

            # Détecter l'arme/catégorie depuis le titre (R1 ou R2)
            if arme is None:
                for cell in row:
                    if cell and isinstance(cell, str):
                        s = cell.upper()
                        for label, val in EXCEL_ARMES_MAP.items():
                            if label in s:
                                arme = val
                        for cat in ["M11", "M13"]:
                            if cat in s:
                                categorie = cat

            # Détecter section
            val0_up = val0.upper()
            if EXCEL_LABEL_SELECTIONNES in val0_up or "SELECTIONNES" in val0_up:
                statut_courant = "selectionne"
                rang_sel = rang_rem = 0
                continue
            if EXCEL_LABEL_REMPLACANTS in val0_up or "REMPLACANTS" in val0_up:
                statut_courant = "remplacant"
                continue
            if EXCEL_LABEL_NON_SELECTIONNABLE in val0_up:
                statut_courant = None
                continue

            # Ligne tireur : col A = "M 1", "M 2"...
            if statut_courant and val0.startswith(EXCEL_PREFIX_RANG) and len(val0) <= EXCEL_RANG_MAX_LEN:
                nom    = str(row[1]).strip() if row[1] else ""
                prenom = str(row[2]).strip() if row[2] else ""
                club   = str(row[3]).strip() if row[3] else ""

                if not nom:
                    continue

                if statut_courant == "selectionne":
                    rang_sel += 1
                    rang = rang_sel
                else:
                    rang_rem += 1
                    rang = rang_rem

                tireurs.append({
                    "rang":           rang,
                    "nom":            nom,
                    "prenom":         prenom,
                    "club":           club,
                    "genre":          genre,
                    "statut":         statut_courant,
                    "territoire":     territoire,
                    "confirmation":   "attente",
                    "confirmation_at": None,
                    "appele":         False,
                    "note":           "",
                })

    # Validation : au moins un tireur doit avoir été extrait
    if not tireurs:
        raise ExcelParseError(
            "Aucun tireur extrait du fichier Excel.",
            hint=(
                f"Vérifiez que le fichier est bien généré par SelecMaster "
                f"(colonnes attendues : rang 'M 1', sections '{EXCEL_LABEL_SELECTIONNES}' / '{EXCEL_LABEL_REMPLACANTS}')."
            )
        )

    if not arme:
        raise ExcelParseError(
            "Impossible de détecter l'arme dans le fichier Excel.",
            hint="Le titre doit contenir 'Épée', 'Fleuret' ou 'Sabre'."
        )

    return {
        "arme":      arme,
        "categorie": categorie or "Inconnue",
        "tireurs":   tireurs,
    }


# ── API publique ──────────────────────────────────────────────────────────────
def initialiser_depuis_excel(contenu_bytes, territoire, territoire_organisateur):
    """
    Initialise ou met à jour le suivi depuis un Excel SelecMaster.
    Préserve les confirmations existantes.
    """
    parsed = lire_excel_selecmaster(contenu_bytes, territoire, territoire_organisateur)
    arme, categorie = parsed["arme"], parsed["categorie"]
    cle = _cle(arme, categorie)

    if cle not in _db:
        _db[cle] = {
            "arme":                  arme,
            "categorie":             categorie,
            "territoire_organisateur": territoire_organisateur,
            "created_at":            datetime.datetime.now(),
            "updated_at":            datetime.datetime.now(),
            "territoires":           {},
            "audit":                 [],
        }

    # Mettre à jour territoire_organisateur si fourni
    if territoire_organisateur:
        _db[cle]["territoire_organisateur"] = territoire_organisateur

    # Préserver confirmations existantes
    existants = {}
    if territoire in _db[cle]["territoires"]:
        for t in _db[cle]["territoires"][territoire]:
            existants[_cle_tireur(t)] = t

    nouveaux = []
    for t in parsed["tireurs"]:
        key = _cle_tireur(t)
        ex  = existants.get(key, {})
        nouveaux.append({
            **t,
            "confirmation":    ex.get("confirmation", "attente"),
            "confirmation_at": ex.get("confirmation_at"),
            "appele":          ex.get("appele", False),
            "note":            ex.get("note", ""),
        })

    _db[cle]["territoires"][territoire] = nouveaux
    _db[cle]["updated_at"] = datetime.datetime.now()
    _audit(cle, "IMPORT_EXCEL",
           f"{territoire} — {len(nouveaux)} tireurs "
           f"({sum(1 for t in nouveaux if t['statut']=='selectionne')} sél. / "
           f"{sum(1 for t in nouveaux if t['statut']=='remplacant')} remplac.)")
    _sauvegarder()
    return {"arme": arme, "categorie": categorie, "nb_tireurs": len(nouveaux)}


def importer_excel_retour(arme, categorie, territoire, contenu_bytes):
    """
    Parse un Excel retourné par un club (colonne Participation remplie).
    Met à jour les confirmations.
    """
    cle = _cle(arme, categorie)
    if cle not in _db or territoire not in _db[cle]["territoires"]:
        raise KeyError(f"Suivi {arme} {categorie} / {territoire} non initialisé")

    wb = load_workbook(io.BytesIO(contenu_bytes), data_only=True)
    tireurs_idx = {_cle_tireur(t): t for t in _db[cle]["territoires"][territoire]}
    modifs = {"confirmations": [], "erreurs": []}

    for sheet_name in wb.sheetnames:
        genre = "H" if sheet_name == "Hommes" else "D" if sheet_name == "Dames" else None
        if not genre:
            continue

        for row in wb[sheet_name].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            val0 = str(row[0]).strip()
            if not (val0.startswith(EXCEL_PREFIX_RANG) and len(val0) <= EXCEL_RANG_MAX_LEN):
                continue

            nom    = str(row[1]).strip() if row[1] else ""
            prenom = str(row[2]).strip() if row[2] else ""
            conf   = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ""

            if conf not in ("oui", "non"):
                continue

            key = (nom.upper(), prenom.upper(), genre)
            if key not in tireurs_idx:
                modifs["erreurs"].append(f"Tireur introuvable : {nom} {prenom} ({genre})")
                continue

            ancien = tireurs_idx[key]["confirmation"]
            tireurs_idx[key]["confirmation"]    = conf
            tireurs_idx[key]["confirmation_at"] = datetime.datetime.now()
            modifs["confirmations"].append({
                "nom": nom, "prenom": prenom, "genre": genre,
                "ancien": ancien, "nouveau": conf,
            })

    nb = len(modifs["confirmations"])
    _audit(cle, "IMPORT_RETOUR",
           f"{territoire} : {nb} confirmation(s) importée(s)"
           + (f" | {len(modifs['erreurs'])} erreur(s)" if modifs["erreurs"] else ""))
    _db[cle]["updated_at"] = datetime.datetime.now()
    _sauvegarder()
    return modifs


def mettre_a_jour_confirmation(arme, categorie, territoire, nom, prenom, genre,
                                confirmation, note=""):
    if confirmation not in CONFIRMATIONS_OK:
        raise ValueError(f"Valeur invalide : {confirmation}. Accepté : {CONFIRMATIONS_OK}")

    cle = _cle(arme, categorie)
    if cle not in _db:
        raise KeyError(f"Aucun suivi pour {arme} {categorie}")
    if territoire not in _db[cle]["territoires"]:
        raise KeyError(f"Territoire {territoire} non chargé")

    key = (nom.strip().upper(), prenom.strip().upper(), genre)
    for t in _db[cle]["territoires"][territoire]:
        if _cle_tireur(t) == key:
            ancien = t["confirmation"]
            t["confirmation"]    = confirmation
            t["confirmation_at"] = datetime.datetime.now()
            t["note"]            = note
            _audit(cle, "CONFIRMATION",
                   f"{nom} {prenom} ({genre}) {territoire} : {ancien} → {confirmation}"
                   + (f" | {note}" if note else ""))
            _db[cle]["updated_at"] = datetime.datetime.now()
            _sauvegarder()
            return
    raise KeyError(f"Tireur {nom} {prenom} ({genre}) introuvable dans {territoire}")


def marquer_appele(arme, categorie, territoire, nom, prenom, genre, appele=True):
    cle = _cle(arme, categorie)
    key = (nom.strip().upper(), prenom.strip().upper(), genre)
    for t in _db[cle]["territoires"][territoire]:
        if _cle_tireur(t) == key:
            t["appele"] = appele
            _audit(cle, "APPEL",
                   f"{nom} {prenom} ({genre}) {territoire} : appele={appele}")
            _db[cle]["updated_at"] = datetime.datetime.now()
            _sauvegarder()
            return
    raise KeyError(f"Tireur {nom} {prenom} ({genre}) introuvable")


def get_stats(arme, categorie):
    cle = _cle(arme, categorie)
    if cle not in _db:
        return None

    data   = _db[cle]
    orga   = data.get("territoire_organisateur", "")
    terrs  = data["territoires"]
    stats  = {
        "arme":      arme, "categorie": categorie,
        "territoire_organisateur": orga,
        "updated_at": data["updated_at"].strftime("%d/%m/%Y %H:%M"),
        "total":     {"sel": 0, "oui": 0, "non": 0, "attente": 0, "rem_appeles": 0},
        "territoires": {},
    }

    for terr, tireurs in terrs.items():
        ts = {}
        for genre in ["H", "D"]:
            sel = [t for t in tireurs if t["genre"] == genre and t["statut"] == "selectionne"]
            rem = [t for t in tireurs if t["genre"] == genre and t["statut"] == "remplacant"]
            ts[genre] = {
                "sel":           len(sel),
                "oui":           sum(1 for t in sel if t["confirmation"] == "oui"),
                "non":           sum(1 for t in sel if t["confirmation"] == "non"),
                "attente":       sum(1 for t in sel if t["confirmation"] == "attente"),
                "rem":           len(rem),
                "rem_appeles":   sum(1 for t in rem if t["appele"]),
                "rem_confirmes": sum(1 for t in rem if t["appele"] and t["confirmation"] == "oui"),
            }
            stats["total"]["sel"]         += ts[genre]["sel"]
            stats["total"]["oui"]         += ts[genre]["oui"]
            stats["total"]["non"]         += ts[genre]["non"]
            stats["total"]["attente"]     += ts[genre]["attente"]
            stats["total"]["rem_appeles"] += ts[genre]["rem_appeles"]
        stats["territoires"][terr] = ts

    return stats


def get_detail(arme, categorie):
    cle = _cle(arme, categorie)
    if cle not in _db:
        return None

    import copy
    data = copy.deepcopy(_db[cle])
    for terr, tireurs in data["territoires"].items():
        for t in tireurs:
            if t.get("confirmation_at"):
                t["confirmation_at"] = t["confirmation_at"].strftime("%d/%m/%Y %H:%M")
    data["created_at"] = data["created_at"].strftime("%d/%m/%Y %H:%M")
    data["updated_at"] = data["updated_at"].strftime("%d/%m/%Y %H:%M")
    return data


def get_remplacants_a_appeler(arme, categorie):
    cle = _cle(arme, categorie)
    if cle not in _db:
        return {}

    data  = _db[cle]
    terrs = data["territoires"]
    orga  = data.get("territoire_organisateur", "Champagne-Ardenne")
    ordre = ORDRE_PIOCHE.get(orga, list(terrs.keys()))

    def effectif_actuel():
        n = 0
        for td in terrs.values():
            for t in td:
                if t["statut"] == "selectionne" and t["confirmation"] == "oui":
                    n += 1
                elif t["statut"] == "remplacant" and t["appele"] and t["confirmation"] != "non":
                    n += 1
        return n

    resultats = {}

    for terr_src in ordre:
        if terr_src not in terrs:
            continue
        tireurs = terrs[terr_src]
        a_appeler = []

        for genre in ["H", "D"]:
            sel         = [t for t in tireurs if t["genre"] == genre and t["statut"] == "selectionne"]
            rem_propres = [t for t in tireurs if t["genre"] == genre and t["statut"] == "remplacant"]

            refus_count   = sum(1 for t in sel if t["confirmation"] == "non")
            appeles_ok    = sum(1 for t in rem_propres if t["appele"] and t["confirmation"] != "non")
            appeles_non   = sum(1 for t in rem_propres if t["appele"] and t["confirmation"] == "non")
            dispo_propres = [t for t in rem_propres if not t["appele"]]
            manquants     = refus_count + appeles_non - appeles_ok

            if manquants <= 0:
                continue

            # Remplaçants propres
            for i in range(min(manquants, len(dispo_propres))):
                if effectif_actuel() >= EFFECTIF_MAX:
                    break
                a_appeler.append({
                    "genre":             genre,
                    "remplacant":        dispo_propres[i],
                    "territoire_rem":    terr_src,
                    "position":          appeles_ok + i + 1,
                    "inter_territorial": False,
                })

            manquants_restants = manquants - min(manquants, len(dispo_propres))

            # Pioche inter-territoriale
            if manquants_restants > 0:
                for terr_autre in ordre:
                    if terr_autre == terr_src or terr_autre not in terrs:
                        continue
                    if effectif_actuel() >= EFFECTIF_MAX:
                        break
                    rem_autres = [
                        t for t in terrs[terr_autre]
                        if t["genre"] == genre
                        and t["statut"] == "remplacant"
                        and not t["appele"]
                    ]
                    for i in range(min(manquants_restants, len(rem_autres))):
                        if effectif_actuel() >= EFFECTIF_MAX:
                            break
                        a_appeler.append({
                            "genre":             genre,
                            "remplacant":        rem_autres[i],
                            "territoire_rem":    terr_autre,
                            "position":          appeles_ok + len(dispo_propres) + i + 1,
                            "inter_territorial": True,
                        })
                    manquants_restants -= min(manquants_restants, len(rem_autres))
                    if manquants_restants <= 0:
                        break

        if a_appeler:
            resultats[terr_src] = a_appeler

    return resultats


def get_audit(arme, categorie):
    cle = _cle(arme, categorie)
    if cle not in _db:
        return []
    return [
        {"ts": e["ts"].strftime("%d/%m/%Y %H:%M:%S"), "action": e["action"], "detail": e["detail"]}
        for e in reversed(_db[cle]["audit"])
    ]


def get_tous_suivis():
    return {
        k: {
            "arme":       v["arme"],
            "categorie":  v["categorie"],
            "updated_at": v["updated_at"].strftime("%d/%m/%Y %H:%M"),
            "nb_terr":    len(v["territoires"]),
        }
        for k, v in _db.items()
    }


def supprimer_suivi(arme, categorie):
    cle = _cle(arme, categorie)
    if cle in _db:
        del _db[cle]
        _sauvegarder()


# Chargement au démarrage
_charger()


# ══════════════════════════════════════════════════════════════════════════════
# MODULE ARBITRES
# Stockage par arme (M11+M13 confondus). Max 8 retenus par arme.
# ══════════════════════════════════════════════════════════════════════════════

ARMES_VALIDES    = {"Épée", "Fleuret", "Sabre"}
STATUTS_ARB      = {"propose", "retenu", "libere"}
NIVEAUX_ARB_VALS = {"Formation Régionale", "Régionale", "Formation Nationale",
                    "National", "International", "Nationale"}
MAX_RETENUS      = 8
_ARB_FILE        = os.path.join(BASE_DIR, ".arbitres_master.pkl")

_arbitres = {}  # clé : arme → liste d'arbitres


def _sauvegarder_arb():
    try:
        with open(_ARB_FILE, "wb") as f:
            pickle.dump(_arbitres, f)
    except Exception as e:
        print(f"[ARB] Erreur sauvegarde : {e}")


def _charger_arb():
    global _arbitres
    try:
        if os.path.isfile(_ARB_FILE):
            with open(_ARB_FILE, "rb") as f:
                _arbitres = pickle.load(f)
    except Exception as e:
        print(f"[ARB] Erreur chargement : {e}")
        _arbitres = {}


def _arb_id():
    """Génère un ID unique pour un arbitre."""
    import uuid
    return str(uuid.uuid4())[:8]


def _valider_arme(arme):
    if arme not in ARMES_VALIDES:
        raise ValueError(f"Arme invalide : {arme}. Valeurs : {ARMES_VALIDES}")


def ajouter_arbitre(arme, nom, prenom, niveau, club, territoire, note=""):
    """Ajoute un arbitre proposé pour une arme."""
    _valider_arme(arme)
    nom, prenom = nom.strip(), prenom.strip()
    if not nom:
        raise ValueError("Le nom est obligatoire")
    if niveau and niveau not in NIVEAUX_ARB_VALS:
        raise ValueError(f"Niveau invalide : {niveau}")

    _arbitres.setdefault(arme, [])

    # Contrôle doublon (même nom+prénom+club)
    for a in _arbitres[arme]:
        if (a["nom"].upper() == nom.upper()
                and a["prenom"].upper() == prenom.upper()
                and a["club"].upper() == club.upper()):
            raise ValueError(f"Arbitre déjà enregistré : {nom} {prenom} ({club})")

    arbitre = {
        "id":         _arb_id(),
        "nom":        nom,
        "prenom":     prenom,
        "niveau":     niveau,
        "club":       club,
        "territoire": territoire,
        "statut":     "propose",
        "note":       note,
        "created_at": datetime.datetime.now(),
        "updated_at": datetime.datetime.now(),
    }
    _arbitres[arme].append(arbitre)
    _sauvegarder_arb()
    return arbitre


def maj_statut_arbitre(arme, arb_id, statut, note=""):
    """Change le statut d'un arbitre (propose → retenu | libere)."""
    _valider_arme(arme)
    if statut not in STATUTS_ARB:
        raise ValueError(f"Statut invalide : {statut}. Valeurs : {STATUTS_ARB}")

    # Contrôle quota retenus
    if statut == "retenu":
        nb_retenus = sum(1 for a in _arbitres.get(arme, [])
                         if a["statut"] == "retenu" and a["id"] != arb_id)
        if nb_retenus >= MAX_RETENUS:
            raise ValueError(f"Quota atteint : {MAX_RETENUS} arbitres déjà retenus pour {arme}")

    for a in _arbitres.get(arme, []):
        if a["id"] == arb_id:
            a["statut"]     = statut
            a["note"]       = note or a["note"]
            a["updated_at"] = datetime.datetime.now()
            _sauvegarder_arb()
            return a

    raise KeyError(f"Arbitre {arb_id} introuvable pour {arme}")


def supprimer_arbitre(arme, arb_id):
    _valider_arme(arme)
    avant = len(_arbitres.get(arme, []))
    _arbitres[arme] = [a for a in _arbitres.get(arme, []) if a["id"] != arb_id]
    if len(_arbitres[arme]) == avant:
        raise KeyError(f"Arbitre {arb_id} introuvable")
    _sauvegarder_arb()


def importer_arbitres_excel(arme, territoire, contenu_bytes):
    """
    Importe les arbitres depuis un Excel SelecMaster retourné.
    Lit la ligne 'Arbitre 1' dans les feuilles Hommes/Dames.
    """
    _valider_arme(arme)
    wb = load_workbook(io.BytesIO(contenu_bytes), data_only=True)
    importes = []
    erreurs  = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in ("Hommes", "Dames"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            val0 = str(row[0]).strip()
            if "Arbitre" not in val0:
                continue
            # Col B = Nom Prénom, Col D = Club, Col E ou F = Niveau
            nom_prenom = str(row[1]).strip() if row[1] else ""
            club_arb   = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            niveau_arb = str(row[5]).strip() if len(row) > 5 and row[5] else ""

            if "Cliquer" in niveau_arb or not nom_prenom or nom_prenom == "Nom  Prénom":
                continue
            if niveau_arb not in NIVEAUX_ARB_VALS:
                niveau_arb = ""

            try:
                a = ajouter_arbitre(arme, nom_prenom, "", niveau_arb, club_arb, territoire)
                importes.append(a)
            except ValueError as e:
                erreurs.append(str(e))

    return {"importes": len(importes), "erreurs": erreurs}


def get_arbitres(arme=None):
    """Retourne les arbitres, filtrés par arme si précisé."""
    if arme:
        _valider_arme(arme)
        return _arbitres.get(arme, [])
    return _arbitres


def get_stats_arbitres():
    """Retourne les stats par arme."""
    stats = {}
    for arme in ARMES_VALIDES:
        arbs = _arbitres.get(arme, [])
        stats[arme] = {
            "total":    len(arbs),
            "propose":  sum(1 for a in arbs if a["statut"] == "propose"),
            "retenu":   sum(1 for a in arbs if a["statut"] == "retenu"),
            "libere":   sum(1 for a in arbs if a["statut"] == "libere"),
            "quota_ok": sum(1 for a in arbs if a["statut"] == "retenu") >= MAX_RETENUS,
        }
    return stats


def get_arbitres_serialisable(arme=None):
    """Version sérialisable (datetimes → strings)."""
    def _ser(a):
        return {**a,
                "created_at": a["created_at"].strftime("%d/%m/%Y %H:%M"),
                "updated_at": a["updated_at"].strftime("%d/%m/%Y %H:%M")}
    if arme:
        return [_ser(a) for a in _arbitres.get(arme, [])]
    return {k: [_ser(a) for a in v] for k, v in _arbitres.items()}


def supprimer_tous_arbitres(arme):
    _valider_arme(arme)
    _arbitres[arme] = []
    _sauvegarder_arb()


_charger_arb()
